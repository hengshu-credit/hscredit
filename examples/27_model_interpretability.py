"""HSCredit 模型解释、原因码、反事实和 Excel 报告完整样例。"""

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split

from hscredit.core.models import CounterfactualExplainer, ModelExplainer
from hscredit.report import ModelReport

FEATURES = ["衡枢鉴真分老客版", "近六个月非银多头机构数", "青云24"]
TARGET = "FPD"


def run(input_path: Path, output_path: Path, max_samples: int, bootstrap: int) -> None:
    """执行确定性的模型解释工作流并写出报告。"""
    frame = pd.read_excel(input_path)
    missing = [column for column in FEATURES + [TARGET] if column not in frame.columns]
    if missing:
        raise ValueError(f"输入文件缺少必要字段: {missing}")
    data = frame[FEATURES + [TARGET]].copy()
    for column in FEATURES + [TARGET]:
        data[column] = pd.to_numeric(data[column], errors="coerce")
    data = data.dropna(subset=[TARGET])
    data[FEATURES] = data[FEATURES].replace([np.inf, -np.inf], np.nan)
    data[FEATURES] = data[FEATURES].fillna(data[FEATURES].median(numeric_only=True)).fillna(0)
    data[TARGET] = data[TARGET].astype(int)
    if data[TARGET].nunique() != 2:
        raise ValueError("FPD 必须同时包含好、坏两类样本")

    X_train, X_test, y_train, y_test = train_test_split(data[FEATURES], data[TARGET], test_size=0.3, stratify=data[TARGET], random_state=42)
    model = RandomForestClassifier(n_estimators=40, max_depth=5, min_samples_leaf=3, random_state=42)
    model.fit(X_train, y_train)

    explainer = ModelExplainer(model, background_data=X_train, random_state=42)
    result = explainer.explain(X_test, max_samples=max_samples)
    global_report = explainer.get_global_report(result)
    sample_report = explainer.get_sample_report(result, sample_id=result.sample_ids[0], top_n=3)
    reason_codes = explainer.get_reason_codes(result, keep=3)
    print("\n全局SHAP重要性：")
    print(global_report.head().to_string(index=False))
    print("\n首个样本贡献：")
    print(sample_report.to_string(index=False))
    print("\n不利原因码：")
    print(reason_codes.head(6).to_string(index=False))

    counter = CounterfactualExplainer(
        model,
        X_train,
        constraints={"衡枢鉴真分老客版": {"mutable": False}},
        random_state=42,
    )
    current_probability = float(model.predict_proba(X_test.iloc[[0]])[0, 1])
    target_probability = max(0.0, current_probability - 0.10)
    counterfactual = counter.generate(X_test.iloc[[0]], target_probability=target_probability, max_changes=2, top_n=3)
    print("\n反事实结果（模型条件下的非因果建议）：")
    print(counterfactual.to_string(index=False))

    report = ModelReport(
        model,
        X_train=X_train,
        y_train=y_train,
        X_test=X_test,
        y_test=y_test,
        n_jobs=1,
        explain_config={
            "enabled": True,
            "data": X_test,
            "background_data": X_train,
            "max_samples": max_samples,
            "n_bootstrap": bootstrap,
            "random_state": 42,
        },
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report.to_excel(output_path, with_plots=False)
    sheets = load_workbook(output_path, read_only=True).sheetnames
    if "7-模型解释" not in sheets:
        raise RuntimeError("模型解释工作表写入失败")
    print(f"\n模型解释报告已生成: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="生成 HSCredit 模型解释报告")
    parser.add_argument("--input", type=Path, default=Path("examples/hscredit_yyp.xlsx"), help="输入 xlsx 文件")
    parser.add_argument("--output", type=Path, default=Path("examples/model_interpretability_report.xlsx"), help="输出报告")
    parser.add_argument("--max-samples", type=int, default=100, help="最大解释样本数")
    parser.add_argument("--bootstrap", type=int, default=20, help="解释稳定性 Bootstrap 次数")
    args = parser.parse_args()
    run(args.input, args.output, args.max_samples, args.bootstrap)


if __name__ == "__main__":
    main()
