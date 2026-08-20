"""hsreport 完整 Excel 报告操作适配器。"""

from collections import OrderedDict
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import load_workbook

from ...report import auto_feature_analysis, auto_model_report, swap_out_report
from ..artifacts import summarize_dataframe
from ..errors import SkillExecutionError
from ..registry import OperationSpec
from .binning import _data, _parameters


_FEATURE_SUMMARY_COLUMNS = (
    "特征名",
    "字段类型",
    "样本数",
    "缺失率",
    "唯一值数",
    "IV",
    "KS",
    "趋势",
    "PSI",
)


def _stage_workbook(context) -> Path:
    return context.artifacts.stage_path(f"{context.request.output.name}.xlsx")


def _validate_and_publish_workbook(context, staged: Path) -> None:
    if not staged.is_file():
        raise SkillExecutionError(code="ARTIFACT_WRITE_FAILED", message=f"报告未生成 Excel 文件：{staged.name}")
    try:
        workbook = load_workbook(staged, read_only=True, data_only=False)
        if not workbook.sheetnames:
            raise ValueError("工作簿没有工作表")
        workbook.close()
    except Exception as exc:
        raise SkillExecutionError(
            code="ARTIFACT_WRITE_FAILED",
            message=f"报告 Excel 无法重新打开：{exc}",
            cause=exc,
        ) from exc
    context.artifacts.publish(
        staged,
        f"{context.request.output.name}.xlsx",
        artifact_type="excel",
    )


def _publish_report_images(context, image_dir: Path) -> None:
    if not image_dir.is_dir():
        return
    for path in sorted(image_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".png", ".svg"}:
            continue
        relative = path.relative_to(image_dir)
        context.artifacts.publish(
            path,
            str(Path(f"{context.request.output.name}-assets") / relative),
            artifact_type="image",
        )


def _as_parameter_list(value: Any) -> list:
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return list(value)
    return [value]


def _label_combinations(parameters: Mapping[str, Any]) -> list:
    overdue_fields = _as_parameter_list(parameters.get("overdue"))
    dpds = _as_parameter_list(parameters.get("dpds"))
    return [{"overdue": overdue, "dpd": dpd} for overdue in overdue_fields for dpd in dpds]


def _extract_feature_summary(staged: Path) -> pd.DataFrame:
    """从报告工作簿提取“变量综合统计”的关键列。"""
    workbook = None
    try:
        workbook = load_workbook(staged, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            title_row = next(
                (
                    row_number
                    for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1)
                    if any(isinstance(value, str) and value.endswith("变量综合统计") for value in values)
                ),
                None,
            )
            if title_row is None:
                continue

            header_row = None
            header_positions = {}
            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=title_row + 1,
                    max_row=min(title_row + 5, worksheet.max_row),
                    values_only=True,
                ),
                start=title_row + 1,
            ):
                if "特征名" not in values:
                    continue
                header_row = row_number
                header_positions = {value: index for index, value in enumerate(values) if isinstance(value, str)}
                break
            if header_row is None:
                continue

            selected_columns = [column for column in _FEATURE_SUMMARY_COLUMNS if column in header_positions]
            records = []
            feature_position = header_positions["特征名"]
            for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(value is not None for value in values):
                    if records:
                        break
                    continue
                feature_name = values[feature_position] if feature_position < len(values) else None
                if feature_name is None:
                    break
                records.append(
                    {
                        column: values[header_positions[column]] if header_positions[column] < len(values) else None
                        for column in selected_columns
                    }
                )
            return pd.DataFrame(records, columns=selected_columns)
    except Exception as exc:
        raise SkillExecutionError(
            code="ARTIFACT_WRITE_FAILED",
            message=f"无法读取报告中的变量综合统计：{exc}",
            cause=exc,
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()

    raise SkillExecutionError(
        code="ARTIFACT_WRITE_FAILED",
        message="报告中缺少变量综合统计",
    )


def _auto_feature_analysis(context) -> dict:
    params = _parameters(
        auto_feature_analysis,
        context.request.parameters,
        reserved=("data", "excel_writer", "output_dir"),
    )
    staged = _stage_workbook(context)
    image_dir = context.artifacts.stage_path("report-images/.anchor").parent
    end_row, end_col = auto_feature_analysis(
        _data(context),
        excel_writer=str(staged),
        output_dir=str(image_dir),
        **params,
    )
    feature_summary = _extract_feature_summary(staged)
    _validate_and_publish_workbook(context, staged)
    _publish_report_images(context, image_dir)
    return {
        "summary": {
            "end_row": int(end_row),
            "end_col": int(end_col),
            "label_combinations": _label_combinations(params),
            "feature_summary": summarize_dataframe(feature_summary),
        }
    }


def _resolve_model(context):
    spec = context.request.inputs.get("model")
    if spec is None:
        raise SkillExecutionError(code="SCHEMA_INVALID", message="缺少 inputs.model", field="inputs.model")
    return context.resolver.resolve(spec)


def _resolve_datasets(context, mapping: Any) -> Mapping[str, Any]:
    if not isinstance(mapping, Mapping) or not mapping:
        raise SkillExecutionError(
            code="SCHEMA_INVALID",
            message="parameters.datasets 必须是数据集名称到 inputs 键的非空映射",
            field="parameters.datasets",
        )
    resolved = OrderedDict()
    for dataset_name, input_name in mapping.items():
        if not isinstance(input_name, str) or input_name not in context.request.inputs:
            raise SkillExecutionError(
                code="SCHEMA_INVALID",
                message=f"数据集“{dataset_name}”引用了不存在的 inputs 键“{input_name}”",
                field=f"parameters.datasets.{dataset_name}",
            )
        value = context.resolver.resolve(context.request.inputs[input_name])
        if not isinstance(value, (pd.DataFrame, tuple, list)):
            raise SkillExecutionError(
                code="SCHEMA_INVALID",
                message=f"数据集“{dataset_name}”必须解析为 DataFrame 或 (X, y)",
                field=f"inputs.{input_name}",
            )
        resolved[str(dataset_name)] = value
    return resolved


def _auto_model_report(context) -> dict:
    params = _parameters(
        auto_model_report,
        context.request.parameters,
        reserved=("model", "excel_path"),
    )
    dataset_mapping = params.pop("datasets", None)
    datasets = _resolve_datasets(context, dataset_mapping)
    staged = _stage_workbook(context)
    auto_model_report(
        _resolve_model(context),
        datasets=datasets,
        excel_path=str(staged),
        **params,
    )
    _validate_and_publish_workbook(context, staged)
    return {"summary": {"datasets": list(datasets), "dataset_count": len(datasets)}}


def _swap_out_report(context) -> dict:
    params = _parameters(
        swap_out_report,
        context.request.parameters,
        reserved=("data", "save"),
    )
    rules = params.get("rules")
    rule_count = 1 if isinstance(rules, str) else len(rules or [])
    if rule_count == 0:
        raise SkillExecutionError(code="SCHEMA_INVALID", message="parameters.rules 不能为空", field="parameters.rules")
    staged = _stage_workbook(context)
    swap_out_report(_data(context), save=str(staged), **params)
    _validate_and_publish_workbook(context, staged)
    return {"summary": {"rules": int(rule_count)}}


def register_report_operations(registry) -> None:
    """登记 hsreport 的三个完整报告操作。"""
    handlers = OrderedDict(
        [
            ("auto_feature_analysis", _auto_feature_analysis),
            ("auto_model_report", _auto_model_report),
            ("swap_out_report", _swap_out_report),
        ]
    )
    for name, handler in handlers.items():
        registry.register(OperationSpec("hsreport", name, handler, extras=("skills",)))
