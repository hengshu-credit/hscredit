# -*- coding: utf-8 -*-
"""数据透视表 OOXML 部件构建工具.

openpyxl 仅支持读取数据透视表而不支持创建，故本模块负责根据源数据与透视配置，
构建符合 ECMA-376 规范的透视表相关 XML 部件（pivotCacheDefinition / pivotCacheRecords /
pivotTable），由 :mod:`hscredit.excel.writer` 在 openpyxl 保存后注入到 xlsx（zip）中。

构建出的透视表会写入 ``refreshOnLoad="1"``，Excel 打开时会基于缓存记录自动刷新，
即便布局信息存在细微偏差也会被重新计算，从而提高兼容性。

设计要点:
- 行/列标签字段在缓存中以共享项（sharedItems）索引引用，值字段以字面量存储
- 行/列字段默认关闭分类汇总（defaultSubtotal=0），布局为「明细 + 总计」，便于精确枚举
- 多值字段时，值字段占位（fld=-2）作为列轴最内层
"""

from typing import List, Dict, Any, Optional, Tuple, Sequence

import numpy as np
import pandas as pd

# OOXML 命名空间 / 关系类型 / 内容类型
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

REL_CACHE_DEF = NS_REL + "/pivotCacheDefinition"
REL_CACHE_REC = NS_REL + "/pivotCacheRecords"
REL_PIVOT_TABLE = NS_REL + "/pivotTable"

CT_CACHE_DEF = "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"
CT_CACHE_REC = "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"
CT_PIVOT_TABLE = "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"

# 聚合方式 -> (OOXML subtotal 取值, 中文前缀)
# Excel 数据透视表原生仅支持以下 11 种汇总方式，自定义别名亦须映射到其一。
AGG_MAP: Dict[str, Tuple[str, str]] = {
    "sum": ("sum", "求和项"),
    "count": ("count", "计数项"),
    "counta": ("count", "计数项"),
    "average": ("average", "平均值项"),
    "mean": ("average", "平均值项"),
    "max": ("max", "最大值项"),
    "min": ("min", "最小值项"),
    "product": ("product", "乘积项"),
    "count_nums": ("countNums", "数值计数项"),
    "countnums": ("countNums", "数值计数项"),
    "std": ("stdDev", "标准偏差项"),
    "stddev": ("stdDev", "标准偏差项"),
    "stdp": ("stdDevp", "总体标准偏差项"),
    "stddevp": ("stdDevp", "总体标准偏差项"),
    "var": ("var", "方差项"),
    "varp": ("varp", "总体方差项"),
}

# 合法的 OOXML subtotal 取值（聚合扩展时校验）
_VALID_SUBTOTALS = {
    "sum", "count", "average", "max", "min", "product",
    "countNums", "stdDev", "stdDevp", "var", "varp",
}


def register_aggregation(key: str, subtotal: str, prefix: Optional[str] = None) -> None:
    """注册/扩展聚合方式别名（聚合函数可扩展）。

    Excel 原生汇总仅 11 种（sum/count/average/max/min/product/countNums/stdDev/stdDevp/var/varp），
    本函数用于为其登记自定义别名，使 ``values`` 参数可用更贴合业务的名称。

    :param key: 别名（大小写不敏感），如 ``'总和'``、``'均值'``
    :param subtotal: 对应的 OOXML 汇总取值，须为 11 种之一
    :param prefix: 中文标题前缀，默认为None（复用 ``subtotal`` 已有前缀，没有则用 key）

    **参考样例**

    >>> from hscredit.excel import _pivot
    >>> _pivot.register_aggregation('均值', 'average', '平均值项')
    >>> _pivot.register_aggregation('总和', 'sum')
    """
    if subtotal not in _VALID_SUBTOTALS:
        raise ValueError(
            "subtotal 须为 Excel 原生汇总之一: {}".format(", ".join(sorted(_VALID_SUBTOTALS)))
        )
    if prefix is None:
        # 复用已有同 subtotal 的中文前缀
        for _sub, _pre in AGG_MAP.values():
            if _sub == subtotal:
                prefix = _pre
                break
        prefix = prefix or key
    AGG_MAP[str(key).lower()] = (subtotal, prefix)


# 占比显示方式 -> (OOXML showDataAs 取值, 是否需 extLst 扩展, 默认数字格式)
SHOW_AS_MAP: Dict[str, Tuple[str, bool, Optional[str]]] = {
    "normal": ("normal", False, None),
    "占比": ("percentOfTotal", False, "0.00%"),
    "百分比": ("percentOfTotal", False, "0.00%"),
    "全局占比": ("percentOfTotal", False, "0.00%"),
    "percent_of_total": ("percentOfTotal", False, "0.00%"),
    "行占比": ("percentOfRow", False, "0.00%"),
    "percent_of_row": ("percentOfRow", False, "0.00%"),
    "列占比": ("percentOfCol", False, "0.00%"),
    "percent_of_col": ("percentOfCol", False, "0.00%"),
    "index": ("index", False, None),
    "差异": ("difference", False, None),
    "running_total": ("runTotal", False, None),
    # 组合占比（占父级分组比例，需 extLst，Excel 2010+）
    "组合占比": ("percentOfParentRow", True, "0.00%"),
    "行组合占比": ("percentOfParentRow", True, "0.00%"),
    "列组合占比": ("percentOfParentCol", True, "0.00%"),
    "percent_of_parent_row": ("percentOfParentRow", True, "0.00%"),
    "percent_of_parent_col": ("percentOfParentCol", True, "0.00%"),
}

# 常见数字格式 -> Excel 内置 numFmtId（无需注入 styles.xml）
BUILTIN_NUM_FMTS: Dict[str, int] = {
    "0": 1, "0.00": 2, "#,##0": 3, "#,##0.00": 4,
    "0%": 9, "0.00%": 10, "0.00e+00": 11,
    "#,##0;-#,##0": 37, "#,##0.00;-#,##0.00": 39,
}


def _esc(text: Any) -> str:
    """XML 文本/属性转义。"""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _as_list(value: Any) -> List[Any]:
    """将单值/None 规范化为列表。"""
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return list(value)
    return [value]


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float, np.integer, np.floating)) and not (
        isinstance(value, float) and (np.isnan(value) or np.isinf(value))
    )


def _parse_value_conf(field: Any, conf: Any) -> Dict[str, Any]:
    """解析单个值字段配置 -> {field, agg, show_as, name, number_format}。"""
    agg = show_as = name = number_format = None
    if conf is None:
        pass
    elif isinstance(conf, str):
        agg = conf
    elif isinstance(conf, dict):
        field = conf.get("field", field)
        agg = conf.get("agg")
        show_as = conf.get("show_as")
        name = conf.get("name")
        number_format = conf.get("number_format")
    elif isinstance(conf, (list, tuple)):
        if len(conf) >= 1:
            agg = conf[0]
        if len(conf) >= 2:
            show_as = conf[1]
        if len(conf) >= 3:
            name = conf[2]
    return {"field": field, "agg": agg, "show_as": show_as, "name": name, "number_format": number_format}


def normalize_values(values: Any, data: pd.DataFrame) -> List[Dict[str, Any]]:
    """将 ``values`` 参数规范化为值字段配置列表。

    支持的输入形式:
    - ``'col'`` 或 ``['c1', 'c2']``：默认聚合（数值列 sum，非数值列 count）
    - ``[('col', 'sum'), ('col2', 'mean')]``：显式聚合
    - ``[('col', 'sum', '全局占比')]``：聚合 + 占比显示
    - ``{'col': 'sum', 'col2': 'count'}``：字典（值为聚合名）
    - ``[{'field': 'col', 'agg': 'sum', 'show_as': '全局占比', 'name': '占比', 'number_format': '0.00%'}]``

    占比（``show_as``）支持：全局占比/行占比/列占比/组合占比（及 percent_of_* 英文别名）。

    :return: 形如 ``[{'col', 'agg', 'subtotal', 'name', 'show_as', 'show_as_ext', 'number_format'}]``
    """
    items: List[Dict[str, Any]] = []
    if isinstance(values, dict):
        items = [_parse_value_conf(k, v) for k, v in values.items()]
    else:
        for v in _as_list(values):
            if isinstance(v, dict):
                items.append(_parse_value_conf(v.get("field"), v))
            elif isinstance(v, (list, tuple)):
                items.append(_parse_value_conf(v[0], list(v[1:])))
            else:
                items.append(_parse_value_conf(v, None))

    result: List[Dict[str, Any]] = []
    for conf in items:
        col = conf["field"]
        if col not in data.columns:
            raise ValueError("值字段 '{}' 不存在于源数据列中".format(col))
        agg = conf["agg"]
        if agg is None:
            agg = "sum" if pd.api.types.is_numeric_dtype(data[col]) else "count"
        key = str(agg).lower()
        if key not in AGG_MAP:
            raise ValueError(
                "不支持的聚合方式 '{}'，可选: {}（可用 register_aggregation 扩展）".format(
                    agg, ", ".join(sorted(AGG_MAP))
                )
            )
        subtotal, prefix = AGG_MAP[key]

        show_as_raw = conf["show_as"]
        show_as = None
        show_as_ext = False
        number_format = conf["number_format"]
        if show_as_raw:
            skey = str(show_as_raw).lower() if not isinstance(show_as_raw, str) else show_as_raw
            skey = skey if skey in SHOW_AS_MAP else str(show_as_raw).lower()
            if skey not in SHOW_AS_MAP:
                raise ValueError(
                    "不支持的占比方式 '{}'，可选: {}".format(show_as_raw, ", ".join(sorted(SHOW_AS_MAP)))
                )
            show_as, show_as_ext, default_fmt = SHOW_AS_MAP[skey]
            if show_as == "normal":
                show_as = None
            if number_format is None:
                number_format = default_fmt

        result.append({
            "col": col,
            "agg": key,
            "subtotal": subtotal,
            "name": conf["name"] or "{}:{}".format(prefix, col),
            "show_as": show_as,
            "show_as_ext": show_as_ext,
            "number_format": number_format,
        })
    return result


def _shared_items(series: pd.Series) -> Tuple[List[Any], bool]:
    """计算某列的共享项（去重 + 排序）及是否含缺失值。

    :return: (共享项列表, 是否含缺失值)
    """
    has_missing = bool(series.isna().any())
    uniques = series.dropna().unique().tolist()
    try:
        uniques = sorted(uniques)
    except TypeError:
        uniques = sorted(uniques, key=lambda x: str(x))
    return uniques, has_missing


def compute_buckets(series: pd.Series, start: float, interval: float) -> Dict[str, Any]:
    """对数值列按起始值/步长分组，计算分桶标签、区间上界与「值 -> 桶索引」映射。

    分桶规则与 Excel「组合字段」一致：首尾各含一个溢出桶（``<start`` / ``>end``），
    中间为 ``[lo, hi)`` 左闭右开区间。

    :param series: 数值列
    :param start: 分组起始值
    :param interval: 分组步长（须 > 0）
    :return: ``{labels, start, end, interval, bucket_of}``，其中 ``bucket_of`` 为 value->索引函数
    """
    import math

    if interval <= 0:
        raise ValueError("分组步长 interval 必须大于 0")
    non_null = series.dropna().astype(float)
    vmax = float(non_null.max()) if len(non_null) else start + interval
    n_ranges = max(1, int(math.ceil((vmax - start) / interval + 1e-9)))
    end = start + n_ranges * interval
    while end <= vmax:  # 保证末区间上界严格大于最大值
        end += interval
        n_ranges += 1

    labels: List[str] = ["<{}".format(_fmt_num(start))]
    for j in range(n_ranges):
        lo = start + j * interval
        hi = lo + interval
        labels.append("{}-{}".format(_fmt_num(lo), _fmt_num(hi)))
    labels.append(">{}".format(_fmt_num(end)))

    n_labels = len(labels)

    def bucket_of(v: Any) -> int:
        fv = float(v)
        if fv < start:
            return 0
        if fv >= end:
            return n_labels - 1
        return 1 + int((fv - start) // interval)

    return {"labels": labels, "start": float(start), "end": float(end),
            "interval": float(interval), "bucket_of": bucket_of}


def build_pivot_spec(
    data: pd.DataFrame,
    source_sheet: str,
    source_ref: str,
    pivot_sheet: str,
    pivot_anchor: Tuple[int, int],
    rows: Sequence[Any],
    columns: Sequence[Any],
    values: List[Dict[str, Any]],
    filters: Sequence[Any],
    name: str,
    cache_id: int,
    show_row_totals: bool = True,
    show_col_totals: bool = True,
    groups: Optional[Dict[Any, Dict[str, float]]] = None,
    filter_items: Optional[Dict[Any, Sequence[Any]]] = None,
    subtotals: bool = False,
) -> Dict[str, Any]:
    """根据源数据与透视配置，计算构建 XML 所需的全部中间信息。

    :param groups: 数值字段分组，形如 ``{'放款金额': {'start': 0, 'interval': 1000}}``，
        对横轴/纵轴的数值特征按起始值与步长分桶统计
    :param filter_items: 字段筛选项，形如 ``{'MOB1': [1]}`` 或 ``{'区域': ['华东', '华南']}``，
        仅保留所列取值（可作用于行/列/筛选字段）
    :param subtotals: 是否对非最内层行/列字段显示分类汇总，默认为False
    :return: 透视表规格字典，供 :func:`render_*` 系列函数渲染 XML
    """
    rows = list(rows)
    columns = list(columns)
    filters = list(filters)
    groups = dict(groups or {})
    filter_items = dict(filter_items or {})
    all_columns = data.columns.tolist()

    label_cols = list(dict.fromkeys(rows + columns + filters))  # 去重保序
    for c in label_cols:
        if c not in all_columns:
            raise ValueError("字段 '{}' 不存在于源数据列中".format(c))
    for c in groups:
        if c not in all_columns:
            raise ValueError("分组字段 '{}' 不存在于源数据列中".format(c))

    grouped_set = set(groups)

    # 每个字段的共享项 / 分桶信息
    field_infos: List[Dict[str, Any]] = []
    shared_lookup: Dict[Any, Dict[Any, int]] = {}   # 离散标签字段：值 -> 共享项索引
    bucket_lookup: Dict[Any, Any] = {}              # 分组字段：值 -> 桶索引
    for col in all_columns:
        info: Dict[str, Any] = {"name": col, "is_label": col in label_cols}
        if col in label_cols and col in grouped_set:
            # 数值分组字段：保留数值 sharedItems + fieldGroup/分桶
            g = groups[col]
            bucket = compute_buckets(data[col], float(g["start"]), float(g["interval"]))
            non_null = data[col].dropna()
            info.update({
                "is_grouped": True,
                "shared": bucket["labels"],          # 分桶标签（pivotField items 用）
                "group_start": bucket["start"],
                "group_end": bucket["end"],
                "group_interval": bucket["interval"],
                "min": float(non_null.min()) if len(non_null) else 0.0,
                "max": float(non_null.max()) if len(non_null) else 0.0,
                "all_integer": bool(
                    len(non_null) and np.all(np.equal(np.mod(non_null.to_numpy(dtype=float), 1), 0))
                ),
                "has_missing": bool(data[col].isna().any()),
            })
            bucket_lookup[col] = bucket["bucket_of"]
        elif col in label_cols:
            shared, has_missing = _shared_items(data[col])
            info["shared"] = shared
            info["has_missing"] = has_missing
            info["all_number"] = all(_is_number(v) for v in shared) if shared else False
            shared_lookup[col] = {v: i for i, v in enumerate(shared)}
        else:
            ser = data[col]
            info["is_number"] = bool(pd.api.types.is_numeric_dtype(ser))
            if info["is_number"]:
                non_null = ser.dropna()
                info["min"] = float(non_null.min()) if len(non_null) else 0.0
                info["max"] = float(non_null.max()) if len(non_null) else 0.0
                info["all_integer"] = bool(
                    len(non_null) and np.all(np.equal(np.mod(non_null.to_numpy(dtype=float), 1), 0))
                )
            info["has_missing"] = bool(ser.isna().any())
        field_infos.append(info)

    col_index = {c: i for i, c in enumerate(all_columns)}

    # 字段 -> 隐藏的共享项索引集合（筛选）
    hidden_idx: Dict[Any, set] = {}
    for col, allowed in filter_items.items():
        if col not in label_cols:
            continue
        allowed_set = set(allowed if isinstance(allowed, (list, tuple, set)) else [allowed])
        if col in grouped_set:
            continue  # 分组字段暂不支持按桶筛选
        shared = field_infos[col_index[col]].get("shared", [])
        hidden_idx[col] = {i for i, v in enumerate(shared) if v not in allowed_set}

    # 缓存记录：离散标签用索引、分组字段用原始数值、值字段用字面量
    records: List[List[Tuple[str, Any]]] = []
    for _, row in data.iterrows():
        rec: List[Tuple[str, Any]] = []
        for col in all_columns:
            val = row[col]
            if pd.isna(val):
                rec.append(("m", None))
            elif col in grouped_set:
                rec.append(("n", float(val)))
            elif col in label_cols:
                rec.append(("x", shared_lookup[col][val]))
            elif _is_number(val):
                rec.append(("n", float(val)))
            else:
                rec.append(("s", val))
        records.append(rec)

    row_field_idx = [col_index[c] for c in rows]
    col_field_idx = [col_index[c] for c in columns]
    page_field_idx = [col_index[c] for c in filters]

    def _idx_of(field: Any, value: Any) -> int:
        if field in grouped_set:
            return bucket_lookup[field](value)
        return shared_lookup[field][value]

    # 行/列轴明细组合（应用筛选，按索引排序）
    def _combos(fields: List[Any]) -> List[Tuple[int, ...]]:
        if not fields:
            return []
        sub = data[fields].dropna(how="any").drop_duplicates()
        combos = set()
        for _, r in sub.iterrows():
            # 行级筛选：任一字段取值被隐藏则跳过
            skip = False
            for f in fields:
                if f in filter_items and f not in grouped_set:
                    allowed = filter_items[f]
                    allowed_set = set(allowed if isinstance(allowed, (list, tuple, set)) else [allowed])
                    if r[f] not in allowed_set:
                        skip = True
                        break
            if skip:
                continue
            combos.add(tuple(_idx_of(f, r[f]) for f in fields))
        return sorted(combos)

    row_combos = _combos(rows)
    col_combos = _combos(columns)

    n_data = len(values)

    return {
        "name": name,
        "cache_id": cache_id,
        "source_sheet": source_sheet,
        "source_ref": source_ref,
        "pivot_sheet": pivot_sheet,
        "pivot_anchor": pivot_anchor,
        "field_infos": field_infos,
        "records": records,
        "row_field_idx": row_field_idx,
        "col_field_idx": col_field_idx,
        "page_field_idx": page_field_idx,
        "values": values,
        "row_combos": row_combos,
        "col_combos": col_combos,
        "n_data": n_data,
        "show_row_totals": show_row_totals and bool(rows),
        "show_col_totals": show_col_totals and bool(columns),
        "hidden_idx": hidden_idx,
        "filter_items": filter_items,
        "grouped_set": grouped_set,
        "subtotals": subtotals,
        "shared_lookup": shared_lookup,
    }


def render_cache_definition_xml(spec: Dict[str, Any], rid_records: str) -> str:
    """渲染 pivotCacheDefinition XML。"""
    fields_xml = []
    for info in spec["field_infos"]:
        name = _esc(info["name"])
        if info.get("is_grouped"):
            # 数值分组字段：数值 sharedItems + fieldGroup/rangePr/groupItems
            num_attrs = (
                ' containsSemiMixedTypes="0" containsString="0" containsNumber="1"'
                ' minValue="{}" maxValue="{}"'.format(_fmt_num(info["min"]), _fmt_num(info["max"]))
            )
            if info.get("all_integer"):
                num_attrs += ' containsInteger="1"'
            labels = info["shared"]
            group_items = "".join('<s v="{}"/>'.format(_esc(lab)) for lab in labels)
            field_group = (
                '<fieldGroup>'
                '<rangePr autoStart="0" autoEnd="0" startNum="{s}" endNum="{e}" groupInterval="{i}"/>'
                '<groupItems count="{n}">{items}</groupItems>'
                '</fieldGroup>'
            ).format(
                s=_fmt_num(info["group_start"]), e=_fmt_num(info["group_end"]),
                i=_fmt_num(info["group_interval"]), n=len(labels), items=group_items,
            )
            shared_xml = "<sharedItems{}/>{}".format(num_attrs, field_group)
        elif info["is_label"]:
            shared = info["shared"]
            items = []
            for v in shared:
                if _is_number(v) and not isinstance(v, bool):
                    items.append('<n v="{}"/>'.format(_fmt_num(v)))
                else:
                    items.append('<s v="{}"/>'.format(_esc(v)))
            if info.get("has_missing"):
                items.append("<m/>")
            attrs = ' count="{}"'.format(len(items))
            if info.get("all_number"):
                attrs += ' containsNumber="1"'
            shared_xml = "<sharedItems{}>{}</sharedItems>".format(attrs, "".join(items))
        else:
            if info.get("is_number"):
                attrs = (
                    ' containsSemiMixedTypes="0" containsString="0" containsNumber="1"'
                    ' minValue="{}" maxValue="{}"'.format(_fmt_num(info["min"]), _fmt_num(info["max"]))
                )
                if info.get("all_integer"):
                    attrs += ' containsInteger="1"'
            else:
                attrs = ""
            if info.get("has_missing"):
                attrs = ' containsBlank="1"' + attrs
            shared_xml = "<sharedItems{}/>".format(attrs)
        fields_xml.append(
            '<cacheField name="{}" numFmtId="0">{}</cacheField>'.format(name, shared_xml)
        )

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        '<pivotCacheDefinition xmlns="{ns}" xmlns:r="{nsr}" r:id="{rid}" '
        'refreshOnLoad="1" refreshedBy="hscredit" createdVersion="6" refreshedVersion="6" '
        'minRefreshableVersion="3" recordCount="{rc}">'
        '<cacheSource type="worksheet">'
        '<worksheetSource ref="{ref}" sheet="{sheet}"/>'
        '</cacheSource>'
        '<cacheFields count="{fc}">{fields}</cacheFields>'
        '</pivotCacheDefinition>'
    ).format(
        ns=NS_MAIN, nsr=NS_REL, rid=rid_records, rc=len(spec["records"]),
        ref=_esc(spec["source_ref"]), sheet=_esc(spec["source_sheet"]),
        fc=len(spec["field_infos"]), fields="".join(fields_xml),
    )


def render_cache_records_xml(spec: Dict[str, Any]) -> str:
    """渲染 pivotCacheRecords XML。"""
    rows_xml = []
    for rec in spec["records"]:
        cells = []
        for tag, val in rec:
            if tag == "m":
                cells.append("<m/>")
            elif tag == "x":
                cells.append('<x v="{}"/>'.format(val))
            elif tag == "n":
                cells.append('<n v="{}"/>'.format(_fmt_num(val)))
            else:
                cells.append('<s v="{}"/>'.format(_esc(val)))
        rows_xml.append("<r>{}</r>".format("".join(cells)))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        '<pivotCacheRecords xmlns="{ns}" xmlns:r="{nsr}" count="{rc}">{rows}</pivotCacheRecords>'
    ).format(ns=NS_MAIN, nsr=NS_REL, rc=len(rows_xml), rows="".join(rows_xml))


def _fmt_num(value: float) -> str:
    """数字格式化：整数去掉 .0，其余保留原始浮点。"""
    f = float(value)
    if f.is_integer():
        return str(int(f))
    return repr(f)


# 适配主题色的自定义数据透视表样式名（全工作簿复用，仅注入一次）
THEME_PIVOT_STYLE_NAME = "HSCreditPivotStyle"


def _to_argb(color: str) -> str:
    """颜色规范化为 ARGB（8 位十六进制大写）。"""
    hex_color = str(color).lstrip("#").upper()
    return hex_color if len(hex_color) == 8 else "FF" + hex_color


def build_theme_dxfs(theme_color: str, stripe_color: str) -> List[str]:
    """构建主题样式所需的两个 dxf（表头：主题色底白字加粗；斑马纹：浅色调底）。"""
    header = _to_argb(theme_color)
    stripe = _to_argb(stripe_color)
    header_dxf = (
        '<dxf><font><b/><color rgb="FFFFFFFF"/></font>'
        '<fill><patternFill patternType="solid">'
        '<fgColor rgb="{c}"/><bgColor rgb="{c}"/></patternFill></fill></dxf>'
    ).format(c=header)
    stripe_dxf = (
        '<dxf><fill><patternFill patternType="solid">'
        '<fgColor rgb="{c}"/><bgColor rgb="{c}"/></patternFill></fill></dxf>'
    ).format(c=stripe)
    return [header_dxf, stripe_dxf]


def build_theme_tablestyle(name: str, header_dxf_id: int, stripe_dxf_id: int) -> str:
    """构建引用上述 dxf 的数据透视表 tableStyle（表头 + 首行斑马纹 + 总计行同表头色）。"""
    return (
        '<tableStyle name="{name}" table="0" pivot="1" count="3">'
        '<tableStyleElement type="headerRow" dxfId="{h}"/>'
        '<tableStyleElement type="firstRowStripe" dxfId="{s}"/>'
        '<tableStyleElement type="totalRow" dxfId="{h}"/>'
        '</tableStyle>'
    ).format(name=_esc(name), h=header_dxf_id, s=stripe_dxf_id)


def apply_pivot_styles(
    styles_xml: str,
    want_theme: bool,
    theme_color: str,
    stripe_color: str,
    custom_numfmts: Optional[List[Tuple[int, str]]] = None,
) -> str:
    """向 styles.xml 注入主题透视表样式（dxfs + tableStyle）与自定义数字格式（numFmts）。

    :param styles_xml: 原始 styles.xml 文本
    :param want_theme: 是否注入主题透视表样式
    :param theme_color: 主题色（表头底色）
    :param stripe_color: 斑马纹底色（一般为主题色浅色调）
    :param custom_numfmts: 待注入的自定义数字格式 [(numFmtId, formatCode)]
    :return: 修改后的 styles.xml 文本
    """
    import re as _re

    custom_numfmts = custom_numfmts or []

    # 1) 自定义 numFmts（须为 styleSheet 首个子元素）
    if custom_numfmts:
        add = "".join(
            '<numFmt numFmtId="{}" formatCode="{}"/>'.format(fid, _esc(code))
            for fid, code in custom_numfmts
        )
        m = _re.search(r'<numFmts count="(\d+)">', styles_xml)
        m_self = _re.search(r'<numFmts count="(\d+)"\s*/>', styles_xml)
        if m_self:
            cnt = int(m_self.group(1)) + len(custom_numfmts)
            styles_xml = styles_xml.replace(
                m_self.group(0), '<numFmts count="{}">{}</numFmts>'.format(cnt, add), 1
            )
        elif m:
            cnt = int(m.group(1)) + len(custom_numfmts)
            styles_xml = styles_xml.replace(m.group(0), '<numFmts count="{}">'.format(cnt), 1)
            styles_xml = styles_xml.replace("</numFmts>", add + "</numFmts>", 1)
        else:
            block = '<numFmts count="{}">{}</numFmts>'.format(len(custom_numfmts), add)
            styles_xml = _re.sub(r'(<styleSheet\b[^>]*>)', r'\1' + block, styles_xml, count=1)

    if not want_theme:
        return styles_xml

    # 2) dxfs：先取现有数量以确定新 dxf 索引
    dxfs = build_theme_dxfs(theme_color, stripe_color)
    m_self = _re.search(r'<dxfs count="(\d+)"\s*/>', styles_xml)
    m = _re.search(r'<dxfs count="(\d+)">', styles_xml)
    if m_self:
        base = int(m_self.group(1))
        styles_xml = styles_xml.replace(
            m_self.group(0),
            '<dxfs count="{}">{}</dxfs>'.format(base + len(dxfs), "".join(dxfs)), 1,
        )
    elif m:
        base = int(m.group(1))
        styles_xml = styles_xml.replace(m.group(0), '<dxfs count="{}">'.format(base + len(dxfs)), 1)
        styles_xml = styles_xml.replace("</dxfs>", "".join(dxfs) + "</dxfs>", 1)
    else:
        base = 0
        block = '<dxfs count="{}">{}</dxfs>'.format(len(dxfs), "".join(dxfs))
        # dxfs 位于 cellStyles 之后、tableStyles 之前；简单插入到 tableStyles 之前或末尾
        if "<tableStyles" in styles_xml:
            styles_xml = _re.sub(r'(<tableStyles)', block + r'\1', styles_xml, count=1)
        else:
            styles_xml = styles_xml.replace("</styleSheet>", block + "</styleSheet>", 1)
    header_id, stripe_id = base, base + 1

    # 3) tableStyles：追加自定义 pivot 样式
    tablestyle = build_theme_tablestyle(THEME_PIVOT_STYLE_NAME, header_id, stripe_id)
    m_self = _re.search(r'<tableStyles ([^>]*?)\s*/>', styles_xml)
    m = _re.search(r'<tableStyles ([^>]*?)>', styles_xml)
    if m_self:
        attrs = m_self.group(1)
        cnt_m = _re.search(r'count="(\d+)"', attrs)
        cnt = (int(cnt_m.group(1)) if cnt_m else 0) + 1
        attrs2 = _re.sub(r'count="\d+"', 'count="{}"'.format(cnt), attrs) if cnt_m else 'count="1" ' + attrs
        styles_xml = styles_xml.replace(
            m_self.group(0), '<tableStyles {}>{}</tableStyles>'.format(attrs2, tablestyle), 1,
        )
    elif m:
        attrs = m.group(1)
        cnt_m = _re.search(r'count="(\d+)"', attrs)
        cnt = (int(cnt_m.group(1)) if cnt_m else 0) + 1
        attrs2 = _re.sub(r'count="\d+"', 'count="{}"'.format(cnt), attrs) if cnt_m else 'count="1" ' + attrs
        styles_xml = styles_xml.replace(m.group(0), '<tableStyles {}>'.format(attrs2), 1)
        styles_xml = styles_xml.replace("</tableStyles>", tablestyle + "</tableStyles>", 1)
    else:
        block = '<tableStyles count="1" defaultTableStyle="TableStyleMedium2" defaultPivotStyle="PivotStyleLight16">{}</tableStyles>'.format(tablestyle)
        styles_xml = styles_xml.replace("</styleSheet>", block + "</styleSheet>", 1)

    return styles_xml


def _axis_items_xml(combos: List[Tuple[int, ...]], grand: bool) -> Tuple[str, int]:
    """根据明细组合枚举行/列轴 ``<i>`` 项（含游程压缩 ``r`` 属性）。

    :param combos: 已排序的索引组合列表，每个元素为各层级共享项索引构成的元组
    :param grand: 是否追加总计项
    :return: (拼接后的 XML, 项数)
    """
    out: List[str] = []
    prev: Tuple[int, ...] = ()
    for combo in combos:
        # 与上一行相同的前缀长度（游程）
        r = 0
        for a, b in zip(prev, combo):
            if a == b:
                r += 1
            else:
                break
        xs = "".join(
            "<x/>" if combo[k] == 0 else '<x v="{}"/>'.format(combo[k])
            for k in range(r, len(combo))
        )
        attr = ' r="{}"'.format(r) if r else ""
        out.append("<i{}>{}</i>".format(attr, xs))
        prev = combo
    if grand:
        out.append('<i t="grand"><x/></i>')
    return "".join(out), len(out)


def compute_pivot_layout(spec: Dict[str, Any]) -> Dict[str, Any]:
    """估算透视表在工作表中的落位（Excel 刷新时会重算，此处用于 location 与透视图引用）。

    :return: 含 ``first_header_row`` / ``first_data_row`` / ``first_data_col`` /
        ``n_row_leaf`` / ``n_col_leaf`` / ``n_row_label_cols`` / ``width`` / ``height`` / ``ref``
        （行列偏移均相对透视表左上角锚点，从 0 开始）
    """
    from openpyxl.utils import get_column_letter

    row_idx = spec["row_field_idx"]
    col_idx = spec["col_field_idx"]
    page_idx = spec["page_field_idx"]
    n_data = spec["n_data"]

    n_row_label_cols = max(1, len(row_idx))
    n_page_rows = len(page_idx) + (1 if page_idx else 0)
    col_header_rows = len(col_idx) + (1 if n_data > 1 else 0)
    first_header_row = n_page_rows
    first_data_row = first_header_row + max(1, col_header_rows)
    first_data_col = n_row_label_cols

    n_row_leaf = len(spec["row_combos"])
    base_combos = spec["col_combos"] if col_idx else [()]
    n_col_leaf = (len(base_combos) * n_data) if (col_idx or n_data > 1) else 1
    n_col_leaf = max(1, n_col_leaf)

    width = n_row_label_cols + n_col_leaf
    height = first_data_row + max(1, n_row_leaf) + (1 if spec["show_row_totals"] else 0)

    anchor_row, anchor_col = spec["pivot_anchor"]
    c0 = get_column_letter(anchor_col)
    c1 = get_column_letter(anchor_col + width - 1)
    ref = "{}{}:{}{}".format(c0, anchor_row, c1, anchor_row + height - 1)

    return {
        "first_header_row": first_header_row,
        "first_data_row": first_data_row,
        "first_data_col": first_data_col,
        "n_row_label_cols": n_row_label_cols,
        "n_row_leaf": n_row_leaf,
        "n_col_leaf": n_col_leaf,
        "width": width,
        "height": height,
        "ref": ref,
    }


def render_pivot_table_xml(spec: Dict[str, Any], fmt_id_map: Optional[Dict[str, int]] = None) -> str:
    """渲染 pivotTableDefinition XML。

    :param spec: :func:`build_pivot_spec` 产出的规格
    :param fmt_id_map: 自定义数字格式字符串 -> numFmtId 的映射（由注入端统一分配），默认为None
    """
    fmt_id_map = fmt_id_map or {}
    field_infos = spec["field_infos"]
    n_fields = len(field_infos)
    row_idx = spec["row_field_idx"]
    col_idx = spec["col_field_idx"]
    page_idx = spec["page_field_idx"]
    values = spec["values"]
    n_data = spec["n_data"]
    hidden_idx = spec.get("hidden_idx", {})
    filter_items = spec.get("filter_items", {})
    subtotals = spec.get("subtotals", False)

    axis_of = {}
    for i in row_idx:
        axis_of[i] = "axisRow"
    for i in col_idx:
        axis_of[i] = "axisCol"
    for i in page_idx:
        axis_of[i] = "axisPage"
    name_to_idx = {info["name"]: i for i, info in enumerate(field_infos)}
    data_field_idx = {name_to_idx[v["col"]] for v in values}

    # 各轴最内层字段索引（分类汇总不作用于最内层）
    innermost = set()
    if row_idx:
        innermost.add(row_idx[-1])
    if col_idx:
        innermost.add(col_idx[-1])

    # pivotFields
    pivot_fields_xml = []
    for i, info in enumerate(field_infos):
        attrs = ""
        is_axis = i in axis_of
        if is_axis:
            attrs += ' axis="{}"'.format(axis_of[i])
        if i in data_field_idx:
            attrs += ' dataField="1"'
        attrs += ' showAll="0"'
        # 多选筛选（页字段筛选到多个取值时）
        field_name = info["name"]
        field_hidden = hidden_idx.get(field_name, set())
        if axis_of.get(i) == "axisPage" and field_name in filter_items:
            allowed = filter_items[field_name]
            n_allowed = len(allowed if isinstance(allowed, (list, tuple, set)) else [allowed])
            if n_allowed > 1:
                attrs += ' multipleItemSelectionAllowed="1"'

        if is_axis:
            # 分类汇总：仅非最内层字段、且开启 subtotals 时显示
            field_has_subtotal = subtotals and (i not in innermost)
            if not field_has_subtotal:
                attrs += ' defaultSubtotal="0"'
            shared = info.get("shared", [])
            parts = []
            for k in range(len(shared)):
                h = ' h="1"' if k in field_hidden else ""
                parts.append('<item x="{}"{}/>'.format(k, h))
            # 缺失值项（分组字段不含）
            if info.get("has_missing") and not info.get("is_grouped"):
                parts.append('<item x="{}"/>'.format(len(shared)))
            if field_has_subtotal:
                parts.append('<item t="default"/>')
            pivot_fields_xml.append(
                '<pivotField{}><items count="{}">{}</items></pivotField>'.format(attrs, len(parts), "".join(parts))
            )
        else:
            pivot_fields_xml.append("<pivotField{}/>".format(attrs))

    # rowFields / rowItems
    if row_idx:
        row_fields_xml = '<rowFields count="{}">{}</rowFields>'.format(
            len(row_idx), "".join('<field x="{}"/>'.format(i) for i in row_idx)
        )
        items_xml, n_row_items = _axis_items_xml(spec["row_combos"], spec["show_row_totals"])
        row_items_xml = '<rowItems count="{}">{}</rowItems>'.format(n_row_items, items_xml)
    else:
        row_fields_xml = ""
        # 无行字段：仅一个总计行
        row_items_xml = '<rowItems count="1"><i><x/></i></rowItems>'

    # colFields / colItems（含多值字段占位 fld=-2）
    col_levels = list(col_idx)
    if n_data > 1:
        col_levels = col_levels + [-2]

    if not col_levels:
        col_fields_xml = ""
        col_items_xml = '<colItems count="1"><i/></colItems>'
    else:
        col_fields_xml = '<colFields count="{}">{}</colFields>'.format(
            len(col_levels), "".join('<field x="{}"/>'.format(i) for i in col_levels)
        )
        # 构造列轴组合：列标签组合 × 值字段索引
        base_combos = spec["col_combos"] if col_idx else [()]
        if n_data > 1:
            combos = [bc + (d,) for bc in base_combos for d in range(n_data)]
        else:
            combos = [bc for bc in base_combos]
        combos = sorted(combos)
        items_xml, n_col_items = _axis_items_xml(combos, spec["show_col_totals"])
        col_items_xml = '<colItems count="{}">{}</colItems>'.format(n_col_items, items_xml)

    # pageFields（含单选筛选项定位）
    if page_idx:
        page_parts = []
        for i in page_idx:
            field_name = field_infos[i]["name"]
            item_attr = ""
            if field_name in filter_items:
                allowed = filter_items[field_name]
                allowed = list(allowed) if isinstance(allowed, (list, tuple, set)) else [allowed]
                shared = field_infos[i].get("shared", [])
                sel = [shared.index(v) for v in allowed if v in shared]
                if len(sel) == 1:
                    item_attr = ' item="{}"'.format(sel[0])
            page_parts.append('<pageField fld="{}"{} hier="-1"/>'.format(i, item_attr))
        page_fields_xml = '<pageFields count="{}">{}</pageFields>'.format(len(page_idx), "".join(page_parts))
    else:
        page_fields_xml = ""

    # dataFields（聚合 + 占比显示 showDataAs + 数字格式）
    data_parts = []
    for v in values:
        fld = name_to_idx[v["col"]]
        attrs = ' fld="{}"'.format(fld)
        if v["subtotal"] != "sum":
            attrs += ' subtotal="{}"'.format(v["subtotal"])
        show_as = v.get("show_as")
        if show_as and not v.get("show_as_ext"):
            attrs += ' showDataAs="{}"'.format(show_as)
        attrs += ' baseField="0" baseItem="0"'
        # 数字格式
        num_fmt = v.get("number_format")
        if num_fmt:
            fmt_id = BUILTIN_NUM_FMTS.get(str(num_fmt).lower(), fmt_id_map.get(num_fmt))
            if fmt_id is not None:
                attrs += ' numFmtId="{}"'.format(fmt_id)
        # 组合占比等需 extLst（Excel 2010+）
        if show_as and v.get("show_as_ext"):
            ext = (
                '<extLst><ext uri="{{E15A36E0-9728-4e99-A89B-3F7291B0FE68}}" '
                'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
                '<x14:dataField pivotShowAs="{showas}"/></ext></extLst>'
            ).format(showas=show_as)
            data_parts.append('<dataField name="{}"{}>{}</dataField>'.format(_esc(v["name"]), attrs, ext))
        else:
            data_parts.append('<dataField name="{}"{}/>'.format(_esc(v["name"]), attrs))
    data_fields_xml = '<dataFields count="{}">{}</dataFields>'.format(n_data, "".join(data_parts))

    # location 估算（Excel 刷新时会重算）
    layout = compute_pivot_layout(spec)
    location_xml = (
        '<location ref="{ref}" firstHeaderRow="{fhr}" firstDataRow="{fdr}" firstDataCol="{fdc}"{page}/>'
    ).format(
        ref=layout["ref"], fhr=layout["first_header_row"], fdr=layout["first_data_row"],
        fdc=layout["first_data_col"],
        page=' rowPageCount="{}" colPageCount="1"'.format(len(page_idx)) if page_idx else "",
    )

    # 含自定义数字格式时需开启 applyNumberFormats
    apply_num = "1" if any(v.get("number_format") for v in values) else "0"
    show_stripes = "1" if spec.get("show_row_stripes") else "0"

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        '<pivotTableDefinition xmlns="{ns}" name="{name}" cacheId="{cid}" '
        'applyNumberFormats="{apply_num}" applyBorderFormats="0" applyFontFormats="0" '
        'applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1" '
        'dataCaption="值" updatedVersion="6" minRefreshableVersion="3" useAutoFormatting="1" '
        'itemPrintTitles="1" createdVersion="6" indent="0" compact="0" compactData="0" '
        'outline="1" outlineData="1" gridDropZones="0" multipleFieldFilters="0"{rgt}{cgt}>'
        '{location}'
        '<pivotFields count="{nf}">{pivot_fields}</pivotFields>'
        '{row_fields}{row_items}{col_fields}{col_items}{page_fields}{data_fields}'
        '<pivotTableStyleInfo name="{style}" showRowHeaders="1" showColHeaders="1" '
        'showRowStripes="{stripes}" showColStripes="0" showLastColumn="1"/>'
        '</pivotTableDefinition>'
    ).format(
        ns=NS_MAIN, name=_esc(spec["name"]), cid=spec["cache_id"],
        apply_num=apply_num,
        rgt="" if spec["show_row_totals"] else ' rowGrandTotals="0"',
        cgt="" if spec["show_col_totals"] else ' colGrandTotals="0"',
        location=location_xml,
        nf=n_fields, pivot_fields="".join(pivot_fields_xml),
        row_fields=row_fields_xml, row_items=row_items_xml,
        col_fields=col_fields_xml, col_items=col_items_xml,
        page_fields=page_fields_xml, data_fields=data_fields_xml,
        style=spec.get("style", "PivotStyleLight16"),
        stripes=show_stripes,
    )
