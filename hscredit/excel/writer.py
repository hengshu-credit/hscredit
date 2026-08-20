# -*- coding: utf-8 -*-
"""
Excel写入器

提供专业的Excel报告生成功能，支持丰富的样式和格式化选项。

核心功能:
- DataFrame数据写入，支持多层索引和多层列名
- 图片插入
- 超链接插入
- 条件格式设置
- 自定义样式

迁移自 scorecardpipeline.excel_writer
"""

import sys
import warnings
import re
import os
import copy
import math
import shutil
import tempfile
import zipfile
from typing import Optional, Union, List, Tuple, Dict, Any, Sequence

import numpy as np
import pandas as pd

from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import NamedStyle, Border, Side, Alignment, PatternFill, Font

from . import _pivot

# hscredit 可视化主题配色（与 core.viz 保持一致）：主题蓝 / 坏样本红 / 提升橙
HSCREDIT_CHART_COLORS = ["2639E9", "F76E6C", "FE7715"]


class ExcelWriter:
    """Excel写入器，提供专业的Excel报告生成功能。

    支持DataFrame数据写入、图片插入、超链接等功能。
    支持上下文管理器（with语句）自动保存。

    :param style_excel: 样式模板文件路径，默认使用包内的template.xlsx
    :param style_sheet_name: 模板文件内初始样式sheet名称，默认为"初始化"
    :param mode: 写入模式，可选'replace'或'append'，默认为'replace'
        - replace: 替换已有文件
        - append: 在已有文件基础上追加内容
    :param fontsize: 字体大小，默认为10
    :param font: 字体名称，默认为"阿里妈妈方圆体 VF Medium"；安装不可用时回退为"楷体"
    :param theme_color: 主题颜色（不包含#），默认为"2639E9"
    :param opacity: 颜色填充的透明度，默认为0.85
    :param system: 操作系统类型，可选'mac'、'windows'、'linux'，默认自动检测
    :param condition_color: 条件格式颜色（不包含#），默认为副主题色"F76E6C"

    **参考样例**

    >>> import pandas as pd
    >>> from hscredit.excel import ExcelWriter
    >>>
    >>> # 方法1：使用with语句（推荐，自动保存）
    >>> with ExcelWriter(theme_color='3f1dba').set_filename("report.xlsx") as writer:
    ...     worksheet = writer.get_sheet_by_name("模型报告")
    ...     writer.insert_value2sheet(worksheet, "B2", value="模型报告", style="header")
    ...     df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    ...     writer.insert_df2sheet(worksheet, df, "B4")
    >>> # 文件在退出with块时自动保存
    >>>
    >>> # 方法2：手动调用save（原有方式）
    >>> writer = ExcelWriter(theme_color='3f1dba')
    >>> worksheet = writer.get_sheet_by_name("模型报告")
    >>> writer.insert_value2sheet(worksheet, "B2", value="模型报告", style="header")
    >>> df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    >>> writer.insert_df2sheet(worksheet, df, "B4")
    >>> writer.save("report.xlsx")
    """

    AUTO_FAST_MIN_ROWS = 500
    AUTO_FAST_MIN_COLUMNS = 50
    AUTO_FAST_MIN_CELLS = 10_000

    def __init__(
        self,
        style_excel: Optional[str] = None,
        style_sheet_name: str = "初始化",
        mode: str = "replace",
        fontsize: int = 10,
        font: Optional[str] = None,
        theme_color: str = '2639E9',
        opacity: float = 0.85,
        system: Optional[str] = None,
        condition_color: str = "F76E6C"
    ):
        # 系统检测
        self.system = system
        if self.system is None:
            self.system = "mac" if sys.platform == "darwin" else "windows"

        if font is None:
            from ..utils.fonts import get_default_excel_font_name

            font = get_default_excel_font_name()

        # 样式参数
        self.english_width = 0.12
        self.chinese_width = 0.21
        self.mode = mode
        self.font = font
        self.opacity = opacity
        self.fontsize = fontsize
        self.theme_color = theme_color
        self.condition_color = condition_color

        # 加载模板
        if style_excel is None:
            # 使用resources目录下的template.xlsx
            package_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            style_excel = os.path.join(
                package_root,
                'resources', 'templates', 'template.xlsx'
            )

        if os.path.exists(style_excel):
            self.workbook = load_workbook(style_excel)
            if style_sheet_name in self.workbook.sheetnames:
                self.style_sheet = self.workbook[style_sheet_name]
            else:
                self.style_sheet = self.workbook.active
        else:
            # 模板文件不存在时，创建新的空 Workbook
            self.workbook = Workbook()
            self.style_sheet = self.workbook.active
            self.style_sheet.title = style_sheet_name

        # 初始化样式
        self.name_styles = []
        self.init_style(font, fontsize, theme_color)

        # 注册命名样式
        for style in self.name_styles:
            if style.name not in self.workbook.style_names:
                self.workbook.add_named_style(style)

        # 快速写入路径直接复用工作簿中已注册的样式数组，避免逐单元格按名称查找。
        self._style_cache = {
            style.name: copy.copy(style.as_tuple())
            for style in self.workbook._named_styles
        }

        # 用于上下文管理器的文件路径
        self._filename: Optional[str] = None

        # 迷你图（sparkline）规格缓存，保存时统一注入 worksheet XML
        # openpyxl 不支持写入 sparkline，故在 save() 后对 xlsx 进行 XML 注入
        self._sparkline_specs: List[Dict[str, Any]] = []

        # 数据透视表规格缓存，保存时注入 pivotCache/pivotTable 部件
        # openpyxl 不支持创建数据透视表，故在 save() 后对 xlsx 进行 XML 注入
        self._pivot_specs: List[Dict[str, Any]] = []
        # 数据透视图规格缓存：记录需在对应 chart XML 中注入 pivotSource 的透视图
        self._pivot_chart_specs: List[Dict[str, Any]] = []

    def __enter__(self) -> 'ExcelWriter':
        """进入上下文管理器。

        :return: ExcelWriter实例
        """
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """退出上下文管理器时自动保存文件。

        :param exc_type: 异常类型
        :param exc_val: 异常值
        :param exc_tb: 异常追踪信息
        """
        if self._filename is not None:
            self.save(self._filename)

    def set_filename(self, filename: str) -> 'ExcelWriter':
        """设置用于上下文管理器自动保存的文件路径。

        支持链式调用，可以与with语句配合使用。

        :param filename: 保存路径
        :return: self

        **参考样例**

        >>> # 方法1：使用set_filename设置路径
        >>> with ExcelWriter().set_filename("report.xlsx") as writer:
        ...     worksheet = writer.get_sheet_by_name("Sheet1")
        ...     writer.insert_value2sheet(worksheet, "B2", "Hello")
        >>> # 文件自动保存到report.xlsx

        >>> # 方法2：手动调用save（原有方式）
        >>> writer = ExcelWriter()
        >>> worksheet = writer.get_sheet_by_name("Sheet1")
        >>> writer.insert_value2sheet(worksheet, "B2", "Hello")
        >>> writer.save("report.xlsx")
        """
        self._filename = filename
        return self

    def add_conditional_formatting(
        self,
        worksheet: Worksheet,
        start_space: str,
        end_space: str,
        condition_color: Optional[str] = None
    ) -> None:
        """设置条件格式（数据条）。

        :param worksheet: 工作表对象
        :param start_space: 开始单元格位置，如'B2'
        :param end_space: 结束单元格位置，如'B10'
        :param condition_color: 条件格式颜色，默认使用ExcelWriter的condition_color
        """
        worksheet.conditional_formatting.add(
            f'{start_space}:{end_space}',
            DataBarRule(
                start_type='min',
                end_type='max',
                color=condition_color or self.condition_color
            )
        )

    @staticmethod
    def set_column_width(
        worksheet: Worksheet,
        column: Union[str, int],
        width: float
    ) -> None:
        """调整Excel列宽。

        :param worksheet: 工作表对象
        :param column: 列，可以是字母（如'B'）或索引（如2）
        :param width: 列宽
        """
        col_letter = column if isinstance(column, str) else get_column_letter(column)
        ExcelWriter._set_column_width_preserving_style(worksheet, col_letter, width)

    @staticmethod
    def _get_column_dimension_style(worksheet: Worksheet, col_letter: str) -> Any:
        """获取指定列在模板列维度中继承的样式。"""
        exact_dimension = worksheet.column_dimensions.get(col_letter)
        if exact_dimension is not None and exact_dimension.has_style:
            return copy.copy(exact_dimension._style)

        col_idx = column_index_from_string(col_letter)
        for dimension in worksheet.column_dimensions.values():
            start_idx = dimension.min or column_index_from_string(dimension.index)
            end_idx = dimension.max or start_idx
            if start_idx <= col_idx <= end_idx and dimension.has_style:
                return copy.copy(dimension._style)
        return None

    @staticmethod
    def _set_column_width_preserving_style(worksheet: Worksheet, col_letter: str, width: float) -> None:
        """设置列宽时保留模板列维度样式。"""
        style_snapshot = ExcelWriter._get_column_dimension_style(worksheet, col_letter)
        column_dimension = worksheet.column_dimensions[col_letter]
        if style_snapshot is not None:
            column_dimension._style = copy.copy(style_snapshot)
        column_dimension.width = width

    def _get_column_cells_data(self, worksheet: Worksheet, col_letter: str) -> List[Tuple[int, Any, Any]]:
        """获取指定列的单元格样式快照。

        :param worksheet: 工作表对象
        :param col_letter: 列字母
        :return: 列表，每项为 (row_idx, value, style_snapshot)
        """
        col_idx = column_index_from_string(col_letter)
        cells_data = []
        # 使用模板的max_row作为上限，确保处理所有行（包括空白行）
        max_row = max(worksheet.max_row, self.style_sheet.max_row)
        for row_idx in range(1, max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            style_snapshot = copy.copy(cell._style) if cell.has_style else None
            cells_data.append((row_idx, cell.value, style_snapshot))
        return cells_data

    def _reapply_styles_to_column(self, worksheet: Worksheet, col_letter: str, cells_data: List[Tuple[int, Any, Any]]) -> None:
        """重新应用样式到指定列的单元格。

        :param worksheet: 工作表对象
        :param col_letter: 列字母
        :param cells_data: 单元格数据列表
        """
        col_idx = column_index_from_string(col_letter)

        for row_idx, value, style_snapshot in cells_data:
            cell = worksheet.cell(row=row_idx, column=col_idx)
            # 检查是否需要保留文本格式（用于消除绿色感叹号）
            need_text_format = isinstance(value, str) and self.is_numeric_like_string(value)

            # 仅恢复调用前已经存在的样式；空白模板区域依赖列维度样式提供白底。
            cell._style = copy.copy(style_snapshot) if style_snapshot is not None else None

            # 如果值是类似数字的字符串，强制设置为文本格式以消除绿色感叹号
            # 这必须在应用样式之后执行，因为命名样式可能包含数字格式
            if need_text_format:
                cell.number_format = '@'

    def adjust_columns_width(
        self,
        worksheet: Worksheet,
        columns: Optional[Union[str, List[str], int, List[int]]] = None,
        start_col: Optional[Union[str, int]] = None,
        end_col: Optional[Union[str, int]] = None,
        max_width: float = 50,
        min_width: float = 8,
        extra_padding: float = 2.0
    ) -> None:
        """批量调整多列宽度，确保已有单元格和模板列样式不丢失。

        通过恢复调用前的单元格样式快照，并保留模板列维度样式，避免列宽调整影响空白区域白底。

        :param worksheet: 工作表对象
        :param columns: 需要调整宽度的列，可以是列字母或列索引列表，默认为None（自动检测所有有数据的列）
        :param start_col: 起始列，与 columns 参数互斥
        :param end_col: 结束列，与 columns 参数互斥
        :param max_width: 最大列宽，默认为50
        :param min_width: 最小列宽，默认为8
        :param extra_padding: 额外填充宽度，默认为2.0

        **参考样例**

        >>> # 调整指定列
        >>> writer.adjust_columns_width(worksheet, columns=['A', 'B', 'C'])
        >>>
        >>> # 调整列范围
        >>> writer.adjust_columns_width(worksheet, start_col='A', end_col='F')
        >>>
        >>> # 自动检测并调整所有有数据的列
        >>> writer.adjust_columns_width(worksheet)
        """
        # 确定要调整的列
        if columns is not None:
            if isinstance(columns, (str, int)):
                columns = [columns]
            col_letters = [c if isinstance(c, str) else get_column_letter(c) for c in columns]
        elif start_col is not None and end_col is not None:
            start_idx = column_index_from_string(start_col) if isinstance(start_col, str) else start_col
            end_idx = column_index_from_string(end_col) if isinstance(end_col, str) else end_col
            col_letters = [get_column_letter(i) for i in range(start_idx, end_idx + 1)]
        else:
            # 自动检测所有有数据的列
            max_col = worksheet.max_column
            col_letters = [get_column_letter(i) for i in range(1, max_col + 1)]

        # 保存每列的单元格数据
        columns_data = {}
        for col_letter in col_letters:
            columns_data[col_letter] = self._get_column_cells_data(worksheet, col_letter)

        # 计算并设置每列的宽度
        for col_letter in col_letters:
            col_idx = column_index_from_string(col_letter)
            max_content_width = min_width

            # 遍历该列所有单元格，计算最大内容宽度
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    value_str = str(cell.value)
                    _, eng_cnt, chi_cnt = self.check_contain_chinese(value_str)
                    content_width = (eng_cnt * self.english_width + chi_cnt * self.chinese_width) * self.fontsize + extra_padding
                    max_content_width = max(max_content_width, content_width)

            # 应用列宽限制
            final_width = min(max_content_width, max_width)
            self._set_column_width_preserving_style(worksheet, col_letter, final_width)

        # 重新应用样式到所有列
        for col_letter, cells_data in columns_data.items():
            self._reapply_styles_to_column(worksheet, col_letter, cells_data)

    def _calculate_content_width(
        self,
        value: Any,
        min_width: float = 8,
        max_width: float = 50,
        extra_padding: float = 2.0,
    ) -> float:
        """按现有中英文字符权重计算单个值需要的列宽。"""
        if value is None:
            return min_width
        _, eng_cnt, chi_cnt = self.check_contain_chinese(str(value))
        width = (eng_cnt * self.english_width + chi_cnt * self.chinese_width) * self.fontsize + extra_padding
        return min(max(width, min_width), max_width)

    def _initial_fast_widths(
        self,
        worksheet: Worksheet,
        start_col: int,
        end_col: int,
    ) -> Dict[int, float]:
        """读取写入前已有内容的列宽基线，不复制或修改单元格样式。"""
        widths = {col_idx: 8.0 for col_idx in range(start_col, end_col + 1)}
        for col_idx in widths:
            for row_idx in range(1, worksheet.max_row + 1):
                value = worksheet.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    widths[col_idx] = max(widths[col_idx], self._calculate_content_width(value))
        return widths

    def _apply_fast_widths(self, worksheet: Worksheet, widths: Dict[int, float]) -> None:
        """只更新列维度宽度，保留工作表和单元格已有样式。"""
        for col_idx, width in widths.items():
            self._set_column_width_preserving_style(worksheet, get_column_letter(col_idx), width)

    @staticmethod
    def set_number_format(
        worksheet: Worksheet,
        space: str,
        _format: str
    ) -> None:
        """设置数值显示格式。

        :param worksheet: 工作表对象
        :param space: 单元格范围，如'B2:B10'
        :param _format: 显示格式，如'0.00%'或'#,##0'
        """
        cells = worksheet[space]
        if isinstance(cells, Cell):
            cells = [cells]

        for cell in cells:
            if isinstance(cell, tuple):
                for c in cell:
                    c.number_format = _format
            else:
                cell.number_format = _format

    def set_freeze_panes(
        self,
        worksheet: Union[Worksheet, str],
        space: Union[str, Tuple[int, int]]
    ) -> None:
        """设置冻结窗格。

        :param worksheet: 工作表对象或名称
        :param space: 冻结位置，如'B2'或(2, 2)
        """
        if not isinstance(worksheet, Worksheet):
            worksheet = self.get_sheet_by_name(worksheet)

        if isinstance(space, (tuple, list)):
            space = self.get_cell_space(space)

        worksheet.freeze_panes = space

    def add_auto_filter(
        self,
        worksheet: Union[Worksheet, str],
        ref: Optional[str] = None,
    ) -> None:
        """给工作表添加自动筛选（auto_filter）。

        :param worksheet: 工作表对象或名称
        :param ref: 筛选区域，如"A1:E10"，默认为整个有数据的区域
        """
        if not isinstance(worksheet, Worksheet):
            worksheet = self.get_sheet_by_name(worksheet)

        if ref is None:
            worksheet.auto_filter.ref = worksheet.dimensions
        else:
            worksheet.auto_filter.ref = ref

    def get_sheet_by_name(self, name: str) -> Worksheet:
        """获取或创建指定名称的工作表。

        :param name: 工作表名称
        :return: 工作表对象
        """
        if name not in self.workbook.sheetnames:
            worksheet = self.workbook.copy_worksheet(self.style_sheet)
            worksheet.title = name
        else:
            worksheet = self.workbook[name]

        return worksheet

    def move_sheet(
        self,
        worksheet: Union[Worksheet, str],
        offset: int = 0,
        index: Optional[int] = None
    ) -> None:
        """移动工作表位置。

        :param worksheet: 工作表对象或名称
        :param offset: 相对移动位置，默认为0
        :param index: 移动到的目标绝对位置索引（从0开始），默认为None；传入时忽略 ``offset``
        """
        total_sheets = len(self.workbook.sheetnames)

        if index is not None:
            # 根据工作表当前位置换算到目标绝对位置所需的相对偏移
            # （openpyxl 的 move_sheet 采用「先删除再按 当前索引+offset 插入」语义，
            # 故 offset = 目标索引 - 当前索引 即可精确落位）
            sheet_name = worksheet.title if isinstance(worksheet, Worksheet) else worksheet
            current_index = self.workbook.sheetnames.index(sheet_name)
            # 将目标索引规范化到合法范围
            if index < 0:
                index = max(total_sheets + index, 0)
            else:
                index = min(index, total_sheets - 1)
            offset = index - current_index

        self.workbook.move_sheet(worksheet, offset=offset)

    def insert_hyperlink2sheet(
        self,
        worksheet: Worksheet,
        insert_space: Union[str, Tuple[int, int]],
        hyperlink: Optional[str] = None,
        file: Optional[str] = None,
        sheet: Optional[str] = None,
        target_space: Optional[Union[str, Tuple[int, int]]] = None
    ) -> None:
        """向单元格插入超链接。

        :param worksheet: 工作表对象
        :param insert_space: 插入位置，如'B2'或(2, 2)
        :param hyperlink: 超链接地址，与target_space互斥
        :param file: 超链接文件路径，默认当前文件
        :param sheet: 超链接sheet名称，默认当前sheet
        :param target_space: 超链接目标位置，如'B10'或(10, 2)

        **参考样例**

        >>> # 链接到当前sheet的其他位置
        >>> writer.insert_hyperlink2sheet(worksheet, "B2", target_space="B10")
        >>>
        >>> # 链接到其他sheet
        >>> writer.insert_hyperlink2sheet(worksheet, "B2", sheet="Sheet2", target_space="A1")
        >>>
        >>> # 链接到外部文件
        >>> writer.insert_hyperlink2sheet(worksheet, "B2", file="other.xlsx", sheet="Sheet1", target_space="A1")
        """
        # 解析插入位置
        if isinstance(insert_space, str):
            start_col = re.findall(r'\D+', insert_space)[0]
            start_row = int(re.findall(r"\d+", insert_space)[0])
        else:
            start_col = get_column_letter(insert_space[1])
            start_row = insert_space[0]

        cell = worksheet[f"{start_col}{start_row}"]

        # 构建超链接
        if hyperlink is None:
            if target_space is None:
                raise ValueError("hyperlink 和 target_space 必须传入一个")

            if sheet is None:
                sheet = worksheet.title

            # 解析目标位置
            if isinstance(target_space, str):
                target_col = re.findall(r'\D+', target_space)[0]
                target_row = int(re.findall(r"\d+", target_space)[0])
            else:
                target_col = get_column_letter(target_space[1])
                target_row = target_space[0]

            # 构建链接
            if file:
                hyperlink = f"file://{file} - #{sheet}!{target_col}{target_row}"
            else:
                hyperlink = f"#{sheet}!{target_col}{target_row}"

        cell.hyperlink = Hyperlink(
            ref=f"{start_col}{start_row}",
            location=hyperlink,
            display=f"{cell.value}"
        )

    def insert_value2sheet(
        self,
        worksheet: Worksheet,
        insert_space: Union[str, Tuple[int, int]],
        value: Any = "",
        style: str = "content",
        auto_width: bool = False,
        end_space: Optional[Union[str, Tuple[int, int]]] = None,
        align: Optional[Dict[str, str]] = None,
        max_col_width: int = 50,
        decimal: Optional[int] = 4,
    ) -> Tuple[int, int]:
        """向单元格插入内容。

        :param worksheet: 工作表对象
        :param insert_space: 插入位置，如'B2'或(2, 2)
        :param value: 插入的内容，默认为""
        :param style: 样式名称，默认为"content"
        :param auto_width: 是否自动调整列宽，默认为False（推荐在数据全部写入后使用 adjust_columns_width 批量调整）
        :param end_space: 合并单元格的结束位置，默认为None
        :param align: 文本对齐方式，默认为None，例如{'horizontal': 'left', 'vertical': 'center'}
        :param max_col_width: 最大列宽，默认为50
        :return: (下一行行号, 下一列列号)

        **参考样例**

        >>> # 插入普通内容
        >>> writer.insert_value2sheet(worksheet, "B2", value="模型报告", style="header")
        >>>
        >>> # 合并单元格
        >>> writer.insert_value2sheet(worksheet, "B2", value="标题", style="header", end_space="D2")
        >>>
        >>> # 批量调整列宽（推荐在所有数据写入完成后调用）
        >>> writer.adjust_columns_width(worksheet, columns=['A', 'B', 'C'])
        """
        # 解析位置
        if isinstance(insert_space, str):
            start_col = re.findall(r'\D+', insert_space)[0]
            start_row = int(re.findall(r"\d+", insert_space)[0])
        else:
            start_col = get_column_letter(insert_space[1])
            start_row = insert_space[0]

        # 设置单元格
        cell = worksheet[f"{start_col}{start_row}"]
        cell.style = style

        # 设置对齐方式
        if align:
            _align = {"horizontal": "center", "vertical": "center"}
            _align.update(align)
            cell.alignment = Alignment(**_align)

        # 合并单元格
        if end_space is not None:
            if isinstance(end_space, str):
                end_col = re.findall(r'\D+', end_space)[0]
                end_row = int(re.findall(r"\d+", end_space)[0])
            else:
                end_col = get_column_letter(end_space[1])
                end_row = end_space[0]

            if start_col != end_col or start_row != end_row:
                worksheet.merge_cells(f"{start_col}{start_row}:{end_col}{end_row}")

        # 格式化值
        formatted_value = self.astype_insertvalue(value, decimal=decimal)

        # 设置值
        worksheet[f"{start_col}{start_row}"] = formatted_value

        # 如果是类似数字的字符串，设置单元格为文本格式以避免绿色感叹号
        if self.is_numeric_like_string(value):
            cell.number_format = '@'

        # 自动调整列宽
        if auto_width:
            # 保存当前列的单元格数据
            cells_data = self._get_column_cells_data(worksheet, start_col)
            
            # 计算新宽度
            curr_width = worksheet.column_dimensions[start_col].width
            _, eng_cnt, chi_cnt = self.check_contain_chinese(str(formatted_value))
            calculated_width = min(
                max([
                    (eng_cnt * self.english_width + chi_cnt * self.chinese_width) * self.fontsize,
                    10,
                    curr_width
                ]),
                max_col_width
            )
            
            # 设置列宽
            self._set_column_width_preserving_style(worksheet, start_col, calculated_width)
            
            # 重新应用样式
            self._reapply_styles_to_column(worksheet, start_col, cells_data)

        # 返回下一个位置
        if end_space is not None:
            return end_row + 1, column_index_from_string(end_col) + 1
        else:
            return start_row + 1, column_index_from_string(start_col) + 1

    def insert_pic2sheet(
        self,
        worksheet: Worksheet,
        fig: str,
        insert_space: Union[str, Tuple[int, int]],
        figsize: Tuple[int, int] = (600, 250),
    ) -> Tuple[int, int]:
        """向Excel插入图片。

        :param worksheet: 工作表对象
        :param fig: 图片路径
        :param insert_space: 插入位置，如'B2'或(2, 2)
        :param figsize: 图片大小(宽, 高)，默认为(600, 250)
        :return: (下一行行号, 下一列列号)
        """
        # 解析位置
        if isinstance(insert_space, str):
            start_row = int(re.findall(r"\d+", insert_space)[0])
            start_col = re.findall(r'\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        # 插入图片
        image = Image(fig)
        image.width, image.height = figsize
        worksheet.add_image(image, f"{start_col}{start_row}")

        # 计算占用的行数
        row_height = 16.0 if self.system != 'mac' else 17.5
        occupied_rows = max(1, math.ceil(figsize[1] / row_height))

        return start_row + occupied_rows, column_index_from_string(start_col) + 8

    def _styled_chart_title(self, text: str):
        """构建带 hscredit 字体/主题色样式的图表标题对象。

        :param text: 标题文本
        :return: openpyxl Title 对象（构建失败时回退为纯文本字符串）
        """
        try:
            from openpyxl.chart.title import Title
            from openpyxl.chart.text import RichText
            from openpyxl.drawing.text import (
                Paragraph, ParagraphProperties, CharacterProperties,
                Font as DrawingFont, RegularTextRun,
            )

            char_props = CharacterProperties(
                latin=DrawingFont(typeface=self.font),
                sz=1200, b=True, solidFill=self.theme_color,
            )
            run = RegularTextRun(rPr=char_props, t=text)
            paragraph = Paragraph(
                pPr=ParagraphProperties(defRPr=char_props), r=[run]
            )
            return Title(tx=RichText(p=[paragraph]))
        except Exception:
            # 不同 openpyxl 版本富文本 API 可能存在差异，降级为纯文本标题
            return text

    def insert_chart2sheet(
        self,
        worksheet: Worksheet,
        insert_space: Union[str, Tuple[int, int]],
        chart: Any,
        width: float = 15.0,
        height: float = 7.5,
    ) -> Tuple[int, int]:
        """向Excel插入原生图表（openpyxl chart 对象）。

        与 :meth:`insert_pic2sheet` 不同，此方法插入的是Excel原生图表（基于单元格数据动态生成），
        而非静态图片，可在Excel中交互、随数据更新。

        :param worksheet: 工作表对象
        :param insert_space: 插入位置（图表左上角锚点），如'B2'或(2, 2)
        :param chart: openpyxl 图表对象（如 BarChart / LineChart）
        :param width: 图表宽度（厘米），默认为15.0
        :param height: 图表高度（厘米），默认为7.5
        :return: (下一行行号, 下一列列号)
        """
        # 解析锚点位置
        if isinstance(insert_space, str):
            start_row = int(re.findall(r"\d+", insert_space)[0])
            start_col = re.findall(r'\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        chart.width = width
        chart.height = height
        worksheet.add_chart(chart, f"{start_col}{start_row}")

        # 估算图表占用的行/列数（厘米 → 行高约0.5cm，列宽约1.8cm），用于返回下一可用位置
        occupied_rows = max(1, math.ceil(height / 0.5))
        occupied_cols = max(1, math.ceil(width / 1.8))

        return start_row + occupied_rows, column_index_from_string(start_col) + occupied_cols

    def insert_bin_chart2sheet(
        self,
        worksheet: Worksheet,
        data: pd.DataFrame,
        table_anchor: Union[str, Tuple[int, int]],
        chart_anchor: Optional[Union[str, Tuple[int, int]]] = None,
        bar_columns: Tuple[str, ...] = ("好样本数", "坏样本数"),
        line_columns: Tuple[str, ...] = ("坏样本率",),
        category_column: Optional[str] = None,
        title: Optional[str] = None,
        header: bool = True,
        index: bool = False,
        bar_stacked: bool = True,
        width: float = 18.0,
        height: float = 9.0,
    ) -> Tuple[int, int]:
        """基于已写入Excel的分箱统计表生成分箱图（柱状图 + 坏样本率折线，双坐标轴）。

        本方法复现 ``hscredit.core.viz.bin_plot`` 的样式：左轴堆叠柱状图展示各分箱好/坏样本数，
        右轴折线展示坏样本率，配色与字体均采用 hscredit 主题风格。图表数据**直接引用**
        worksheet 中已写入的单元格区域，故图表会随表格数据联动。

        使用前需先将 ``feature_bin_stats`` 输出的分箱表通过 :meth:`insert_df2sheet`
        （或 :func:`dataframe2excel`）写入到 ``table_anchor`` 位置（含表头）。

        :param worksheet: 工作表对象
        :param data: 分箱统计表（与写入Excel的DataFrame一致，用于定位列）
        :param table_anchor: 表格写入的左上角位置（含表头），如'B2'或(2, 2)
        :param chart_anchor: 图表插入位置，默认为None（自动置于表格右侧一列）
        :param bar_columns: 柱状图列名（左轴），默认为("好样本数", "坏样本数")
        :param line_columns: 折线图列名（右轴），默认为("坏样本率",)
        :param category_column: 分类轴列名（横轴），默认为None（优先取"分箱标签"，其次"分箱"）
        :param title: 图表标题，默认为None
        :param header: 表格写入时是否含表头，默认为True
        :param index: 表格写入时是否含索引，默认为False
        :param bar_stacked: 柱状图是否堆叠，默认为True
        :param width: 图表宽度（厘米），默认为18.0
        :param height: 图表高度（厘米），默认为9.0
        :return: (下一行行号, 下一列列号)

        **参考样例**

        >>> from hscredit.report import feature_bin_stats
        >>> from hscredit.excel import ExcelWriter
        >>> bin_table = feature_bin_stats(df, '某特征', target='target')
        >>> writer = ExcelWriter()
        >>> ws = writer.get_sheet_by_name('分箱图')
        >>> writer.insert_df2sheet(ws, bin_table, 'B2', fill=True)
        >>> writer.insert_bin_chart2sheet(ws, bin_table, 'B2', title='某特征分箱图')
        >>> writer.save('bin_chart.xlsx')
        """
        from openpyxl.chart import BarChart, LineChart, Reference, Series

        # 解析表格左上角位置
        if isinstance(table_anchor, str):
            start_row = int(re.findall(r"\d+", table_anchor)[0])
            start_col_idx = column_index_from_string(re.findall(r'\D+', table_anchor)[0])
        else:
            start_row, start_col_idx = table_anchor[0], table_anchor[1]

        # 列在Excel中的绝对列号
        idx_levels = data.index.nlevels if index else 0
        n_header_rows = data.columns.nlevels if header else 0

        def _col_letter_idx(col_name: str) -> Optional[int]:
            if col_name not in data.columns:
                return None
            return start_col_idx + idx_levels + data.columns.get_loc(col_name)

        # 数据行范围
        data_first_row = start_row + n_header_rows
        data_last_row = data_first_row + len(data) - 1
        header_row = data_first_row - 1  # 表头所在行（取值列标题）
        # 系列取数起始行：有表头时从表头行起（标题取自首格），无表头时从数据首行起
        series_min_row = header_row if header else data_first_row

        # 分类轴列
        if category_column is None:
            category_column = "分箱标签" if "分箱标签" in data.columns else "分箱"
        cat_col_idx = _col_letter_idx(category_column)

        # 柱状图：堆叠好/坏样本数
        bar = BarChart()
        bar.type = "col"
        bar.grouping = "stacked" if bar_stacked else "clustered"
        bar.overlap = 100 if bar_stacked else -27
        bar.gapWidth = 60
        bar.y_axis.title = "样本数"

        valid_bar_cols = [c for c in bar_columns if _col_letter_idx(c) is not None]
        for i, col_name in enumerate(valid_bar_cols):
            col_idx = _col_letter_idx(col_name)
            ref = Reference(worksheet, min_col=col_idx, min_row=series_min_row, max_row=data_last_row)
            series = Series(ref, title_from_data=True) if header else Series(ref, title=col_name)
            color = HSCREDIT_CHART_COLORS[i % len(HSCREDIT_CHART_COLORS)]
            try:
                series.graphicalProperties.solidFill = color
                series.graphicalProperties.line.solidFill = "FFFFFF"
            except Exception:
                pass
            bar.series.append(series)

        # 折线图：坏样本率（右轴）
        # 双坐标轴组合图必须为折线分配独立的分类轴（axId=500）与数值轴（axId=200），
        # 形成「分类轴↔数值轴」两两交叉引用的完整轴图；折线分类轴隐藏（delete），
        # 否则只有一条分类轴会导致次数值轴 crossAx 指向不一致，Excel 会判定损坏并删除图形。
        line = LineChart()
        line.x_axis.axId = 500
        line.x_axis.crossAx = 200       # 次分类轴 ↔ 次数值轴 互相交叉引用
        line.x_axis.delete = True
        line.x_axis.crosses = "max"     # 次数值轴显示在右侧
        line.y_axis.axId = 200
        line.y_axis.crossAx = 500
        line.y_axis.title = "坏样本率"
        line.y_axis.crosses = "max"

        valid_line_cols = [c for c in line_columns if _col_letter_idx(c) is not None]
        for j, col_name in enumerate(valid_line_cols):
            col_idx = _col_letter_idx(col_name)
            ref = Reference(worksheet, min_col=col_idx, min_row=series_min_row, max_row=data_last_row)
            series = Series(ref, title_from_data=True) if header else Series(ref, title=col_name)
            color = HSCREDIT_CHART_COLORS[1] if j == 0 else HSCREDIT_CHART_COLORS[2 % len(HSCREDIT_CHART_COLORS)]
            try:
                series.graphicalProperties.line.solidFill = color
                series.graphicalProperties.line.width = 20000  # EMU，约2pt
                series.smooth = False
                from openpyxl.chart.marker import Marker
                series.marker = Marker(symbol="circle", size=6)
                series.marker.graphicalProperties.solidFill = "FFFFFF"
                series.marker.graphicalProperties.line.solidFill = color
            except Exception:
                pass
            line.series.append(series)

        # 设置分类轴
        if cat_col_idx is not None:
            cats = Reference(worksheet, min_col=cat_col_idx, min_row=data_first_row, max_row=data_last_row)
            bar.set_categories(cats)
            line.set_categories(cats)

        # 组合双轴图
        if line.series:
            bar += line

        # 样式：标题、图例
        if title:
            bar.title = self._styled_chart_title(title)
        bar.legend.position = "b"
        bar.x_axis.delete = False
        bar.y_axis.delete = False

        # 图表位置：默认置于表格右侧
        if chart_anchor is None:
            table_end_col = start_col_idx + idx_levels + len(data.columns) + 1
            chart_anchor = (start_row, table_end_col)

        return self.insert_chart2sheet(worksheet, chart_anchor, bar, width=width, height=height)

    def insert_df2sheet(
        self,
        worksheet: Worksheet,
        data: pd.DataFrame,
        insert_space: Union[str, Tuple[int, int]],
        merge_column: Optional[Union[str, List[str]]] = None,
        header: bool = True,
        index: bool = False,
        auto_width: bool = False,
        fill: bool = False,
        merge: bool = False,
        merge_index: bool = True,
        merge_header: Union[bool, str, int, Sequence[int]] = True,
        decimal: Optional[int] = 4,
        speed: str = "auto",
    ) -> Tuple[int, int]:
        """向Excel插入DataFrame。

        :param worksheet: 工作表对象
        :param data: 需要插入的DataFrame
        :param insert_space: 插入位置，如'B2'或(2, 2)
        :param merge_column: 需要分组显示的列，默认为None
        :param header: 是否存储DataFrame的header，默认为True
        :param index: 是否存储DataFrame的index，默认为False
        :param auto_width: 是否自动调整列宽，默认为False（在所有数据写入完成后统一调整，避免边框样式丢失）
        :param fill: 是否使用颜色填充而非边框，默认为False
        :param merge: 是否合并单元格，默认为False
        :param merge_index: 当存储index时，是否合并连续相同的index值，默认为True
        :param merge_header: 多层列名是否横向合并相邻相同标题，默认True合并全部层级；
            可传False/``"none"``不合并，或传层级序号/序号列表仅合并指定层级
        :param decimal: 浮点值保留小数位数，默认4；None表示不主动舍入
        :param speed: 写入速度，可选``"auto"``、``"normal"``或``"fast"``，默认``"auto"``；
            自动模式根据数据行数、有效列数和单元格数选择保样式写入路径
        :return: (下一行行号, 下一列列号)

        **参考样例**

        >>> import pandas as pd
        >>> df = pd.DataFrame({'A': [1, 1, 2], 'B': [4, 5, 6], 'C': [7, 8, 9]})
        >>>
        >>> # 基本插入
        >>> writer.insert_df2sheet(worksheet, df, "B2")
        >>>
        >>> # 使用颜色填充
        >>> writer.insert_df2sheet(worksheet, df, "B10", fill=True)
        >>>
        >>> # 保存索引
        >>> writer.insert_df2sheet(worksheet, df.set_index('A'), "B20", index=True)
        >>>
        >>> # 分组显示
        >>> writer.insert_df2sheet(worksheet, df, "B30", merge_column='A', merge=True)
        """
        self._validate_decimal(decimal)
        resolved_speed = self._resolve_write_speed(data, speed, index=index)
        fast = resolved_speed == "fast"
        df = data if fast else data.copy()

        # 解析起始位置
        if isinstance(insert_space, str):
            start_row = int(re.findall(r"\d+", insert_space)[0])
            start_col = re.findall(r'\D+', insert_space)[0]
            start_col_idx = column_index_from_string(start_col)
        else:
            start_row, start_col = insert_space
            start_col_idx = start_col
            start_col = get_column_letter(start_col)

        table_end_col_idx = start_col_idx + len(data.columns) + (data.index.nlevels if index else 0)
        fast_widths = (
            self._initial_fast_widths(worksheet, start_col_idx, table_end_col_idx - 1)
            if fast and auto_width
            else None
        )
        fast_row_kwargs = {'widths': fast_widths} if fast else {}

        # 计算合并行
        def get_merge_rows(values, start_row):
            _rows = []
            item, start, length = self.calc_continuous_cnt(values)
            while start is not None:
                _rows.append(start + start_row)
                item, start, length = self.calc_continuous_cnt(values, start + length)
            _rows.append(len(values) + start_row)
            return _rows

        # 处理索引合并
        if index and merge_index:
            merge_index_cols = {
                i: get_column_letter(column_index_from_string(start_col) + i)
                for i in range(df.index.nlevels)
            }
            merge_index_rows = {
                i: get_merge_rows(
                    df.index.get_level_values(i).tolist(),
                    start_row + df.columns.nlevels if header else start_row
                )
                for i in range(df.index.nlevels)
            }
        else:
            merge_index_cols = None
            merge_index_rows = None

        # 处理列合并
        if merge_column:
            if not isinstance(merge_column, (list, np.ndarray)):
                merge_column = [merge_column]

            if isinstance(merge_column[0], (int, float)) and (merge_column[0] not in df.columns):
                merge_column = [
                    df.columns.tolist()[col] if col not in df.columns else col
                    for col in merge_column
                ]

            if index:
                merge_cols = {
                    col: get_column_letter(
                        df.columns.get_loc(col) + column_index_from_string(start_col) + df.index.nlevels
                    )
                    for col in merge_column
                }
            else:
                merge_cols = {
                    col: get_column_letter(
                        df.columns.get_loc(col) + column_index_from_string(start_col)
                    )
                    for col in merge_column
                }

            if header:
                merge_rows = {
                    col: get_merge_rows(df[col].tolist(), start_row + df.columns.nlevels)
                    for col in merge_column
                }
            else:
                merge_rows = {
                    col: get_merge_rows(df[col].tolist(), start_row)
                    for col in merge_column
                }
        else:
            merge_cols = None
            merge_rows = None

        def _resolve_merge_header_levels(value, nlevels: int) -> set:
            if nlevels <= 1:
                return set()

            def _normalize_level(level: int) -> int:
                level = int(level)
                if level < 0:
                    level += nlevels
                if level < 0 or level >= nlevels:
                    raise ValueError(f"merge_header 层级超出范围: {level}")
                return level

            if isinstance(value, bool):
                return set(range(nlevels)) if value else set()
            if value is None:
                return set(range(nlevels))
            if isinstance(value, str):
                option = value.strip().lower()
                if option in {"all", "true", "yes", "全部", "合并"}:
                    return set(range(nlevels))
                if option in {"none", "false", "no", "不合并"}:
                    return set()
                return {_normalize_level(int(option))}
            if isinstance(value, (int, np.integer)):
                return {_normalize_level(value)}
            if isinstance(value, (list, tuple, set, np.ndarray)):
                return {_normalize_level(level) for level in value}
            raise TypeError("merge_header 仅支持 bool、'all'/'none'、层级序号或层级序号列表")

        merge_header_levels = _resolve_merge_header_levels(merge_header, df.columns.nlevels)
        insert_row = self._insert_rows_fast if fast else self.insert_rows

        # 迭代行数据
        def _iter_rows(df, header=True, index=True):
            columns = df.columns.tolist()
            indexs = df.index.tolist()
            for i, row in enumerate(dataframe_to_rows(df, header=header, index=False)):
                if header:
                    if i < df.columns.nlevels:
                        if index:
                            if df.columns.nlevels > 1:
                                if i == df.columns.nlevels - 1:
                                    yield list(df.index.names) + [c[i] for c in columns]
                                else:
                                    yield [None] * df.index.nlevels + [c[i] for c in columns]
                            else:
                                yield list(df.index.names) + columns
                        else:
                            if df.columns.nlevels > 1 and i < df.columns.nlevels:
                                yield [c[i] for c in columns]
                            else:
                                yield columns
                    else:
                        if index:
                            if df.index.nlevels > 1:
                                yield list(indexs[int(i - df.columns.nlevels)]) + row
                            else:
                                yield [indexs[int(i - df.columns.nlevels)]] + row
                        else:
                            yield row
                else:
                    if index:
                        if df.index.nlevels > 1:
                            yield list(indexs[i]) + row
                        else:
                            yield [indexs[i]] + row
                    else:
                        yield row

        # 插入数据
        for i, row in enumerate(_iter_rows(df, header=header, index=index)):
            if fill:
                if header and i < df.columns.nlevels:
                    insert_row(
                        worksheet, row, start_row + i, start_col,
                        style="header",
                        auto_width=False,
                        multi_levels=df.columns.nlevels > 1 and i in merge_header_levels,
                        decimal=decimal,
                        **fast_row_kwargs,
                    )
                elif i == 0:
                    insert_row(
                        worksheet, row, start_row + i, start_col,
                        style="middle_even_first",
                        auto_width=False,
                        style_only=True,
                        decimal=decimal,
                        **fast_row_kwargs,
                    )
                else:
                    # 根据行数奇偶选择样式
                    if df.columns.nlevels % 2 == 1:
                        if i % 2 == 1:
                            style = "middle_odd_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_odd"
                        else:
                            style = "middle_even_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_even"
                    else:
                        if i % 2 == 1:
                            style = "middle_even_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_even"
                        else:
                            style = "middle_odd_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_odd"

                    insert_row(
                        worksheet, row, start_row + i, start_col,
                        style=style,
                        auto_width=False,
                        style_only=True,
                        decimal=decimal,
                        **fast_row_kwargs,
                    )
            else:
                if header and i < df.columns.nlevels:
                    insert_row(
                        worksheet, row, start_row + i, start_col,
                        style="header",
                        auto_width=False,
                        multi_levels=df.columns.nlevels > 1 and i in merge_header_levels,
                        decimal=decimal,
                        **fast_row_kwargs,
                    )
                elif i == 0:
                    insert_row(
                        worksheet, row, start_row + i, start_col,
                        style="first",
                        auto_width=False,
                        decimal=decimal,
                        **fast_row_kwargs,
                    )
                elif (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)):
                    insert_row(
                        worksheet, row, start_row + i, start_col,
                        style="last",
                        auto_width=False,
                        decimal=decimal,
                        **fast_row_kwargs,
                    )
                else:
                    if merge_rows and len(merge_rows) > 0:
                        insert_row(
                            worksheet, row, start_row + i, start_col,
                            auto_width=False,
                            merge_rows=sorted(set(_row for _rows in merge_rows.values() for _row in _rows)),
                            decimal=decimal,
                            **fast_row_kwargs,
                        )
                    else:
                        insert_row(
                            worksheet, row, start_row + i, start_col,
                            auto_width=False,
                            decimal=decimal,
                            **fast_row_kwargs,
                        )

        # 合并索引单元格
        if index and merge_index and merge_index_rows and len(merge_index_rows) > 0:
            for col in merge_index_cols.keys():
                merge_col = merge_index_cols[col]
                merge_row = merge_index_rows[col]

                for s, e in zip(merge_row[:-1], merge_row[1:]):
                    if e - s > 1:
                        self.merge_cells(worksheet, f"{merge_col}{s}", f"{merge_col}{e - 1}")

        # 合并列单元格
        if merge and merge_column and merge_cols and len(merge_cols) > 0:
            for col in merge_cols.keys():
                merge_col = merge_cols[col]
                merge_row = merge_rows[col]

                for s, e in zip(merge_row[:-1], merge_row[1:]):
                    if e - s > 1:
                        self.merge_cells(worksheet, f"{merge_col}{s}", f"{merge_col}{e - 1}")

        end_row = start_row + len(data) + df.columns.nlevels if header else start_row + len(data)
        end_col_idx = table_end_col_idx

        # 批量调整列宽（在所有数据写入完成后统一调整，避免边框样式丢失）
        if auto_width:
            if fast:
                self._apply_fast_widths(worksheet, fast_widths or {})
            else:
                self.adjust_columns_width(
                    worksheet,
                    start_col=start_col_idx,
                    end_col=end_col_idx - 1
                )

        return end_row, end_col_idx

    def _insert_value_fast(
        self,
        worksheet: Worksheet,
        row_index: int,
        col_index: int,
        value: Any,
        style: str,
        decimal: Optional[int],
        end_col: Optional[int] = None,
        widths: Optional[Dict[int, float]] = None,
    ) -> None:
        """使用整数坐标和缓存样式写入单元格。"""
        cell = worksheet.cell(row=row_index, column=col_index)
        style_array = self._style_cache.get(style)
        if style_array is None:
            cell.style = style
        else:
            cell._style = copy.copy(style_array)

        if end_col is not None and end_col != col_index:
            worksheet.merge_cells(
                start_row=row_index,
                start_column=col_index,
                end_row=row_index,
                end_column=end_col,
            )

        formatted_value = self.astype_insertvalue(value, decimal=decimal)
        cell.value = formatted_value
        if self.is_numeric_like_string(value):
            cell.number_format = '@'
        if widths is not None:
            widths[col_index] = max(
                widths.get(col_index, 8.0),
                self._calculate_content_width(formatted_value),
            )

    def _insert_rows_fast(
        self,
        worksheet: Worksheet,
        row: List,
        row_index: int,
        col_index: Union[str, int],
        merge_rows: Optional[List[int]] = None,
        style: str = "",
        auto_width: bool = False,
        style_only: bool = False,
        multi_levels: bool = False,
        decimal: Optional[int] = 4,
        widths: Optional[Dict[int, float]] = None,
    ) -> None:
        """按现有样式规则快速写入一行，不执行逐单元格列宽调整。"""
        curr_col = column_index_from_string(col_index) if isinstance(col_index, str) else col_index

        if multi_levels and style == "header":
            row = pd.Series(row).ffill().to_list()
            item, start, length = self.calc_continuous_cnt(row)
            while start is not None:
                if start + length < len(row):
                    cell_style = f"{style}_left" if start == 0 else f"{style}_middle"
                else:
                    cell_style = f"{style}_right"
                self._insert_value_fast(
                    worksheet,
                    row_index,
                    curr_col + start,
                    item,
                    cell_style,
                    decimal,
                    end_col=curr_col + start + length - 1,
                    widths=widths,
                )
                item, start, length = self.calc_continuous_cnt(row, start + length)
            return

        for j, value in enumerate(row):
            if merge_rows is not None and row_index + 1 not in merge_rows:
                if j == 0:
                    cell_style = "merge_left"
                elif j == len(row) - 1:
                    cell_style = "merge_right"
                else:
                    cell_style = "merge_middle"
            elif style_only or len(row) <= 1:
                cell_style = style or "middle"
            elif j == 0:
                cell_style = f"{style}_left" if style else "left"
            elif j == len(row) - 1:
                cell_style = f"{style}_right" if style else "right"
            else:
                cell_style = f"{style}_middle" if style else "middle"

            self._insert_value_fast(
                worksheet,
                row_index,
                curr_col + j,
                value,
                cell_style,
                decimal,
                widths=widths,
            )

    def insert_rows(
        self,
        worksheet: Worksheet,
        row: List,
        row_index: int,
        col_index: Union[str, int],
        merge_rows: Optional[List[int]] = None,
        style: str = "",
        auto_width: bool = False,
        style_only: bool = False,
        multi_levels: bool = False,
        decimal: Optional[int] = 4,
    ) -> None:
        """向Excel插入一行数据。

        :param worksheet: 工作表对象
        :param row: 行数据
        :param row_index: 行索引
        :param col_index: 起始列索引或字母
        :param merge_rows: 需要合并的行索引列表，默认为None
        :param style: 样式名称，默认为空
        :param auto_width: 是否自动调整列宽，默认为False（推荐在数据全部写入后使用 adjust_columns_width 批量调整）
        :param style_only: 是否仅应用样式，默认为False
        :param multi_levels: 是否多层索引，默认为False
        """
        curr_col = column_index_from_string(col_index) if isinstance(col_index, str) else col_index

        if multi_levels and style == "header":
            row = pd.Series(row).ffill().to_list()
            item, start, length = self.calc_continuous_cnt(row)

            while start is not None:
                if start + length < len(row):
                    if start == 0:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + start)}{row_index}',
                            item,
                            style=f"{style}_left" if style else "left",
                            auto_width=auto_width,
                            end_space=f'{get_column_letter(curr_col + start + length - 1)}{row_index}',
                            decimal=decimal,
                        )
                    else:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + start)}{row_index}',
                            item,
                            style=f"{style}_middle" if style else "middle",
                            auto_width=auto_width,
                            end_space=f'{get_column_letter(curr_col + start + length - 1)}{row_index}',
                            decimal=decimal,
                        )
                else:
                    self.insert_value2sheet(
                        worksheet,
                        f'{get_column_letter(curr_col + start)}{row_index}',
                        item,
                        style=f"{style}_right" if style else "right",
                        auto_width=auto_width,
                        end_space=f'{get_column_letter(curr_col + start + length - 1)}{row_index}',
                        decimal=decimal,
                    )

                item, start, length = self.calc_continuous_cnt(row, start + length)
        else:
            for j, v in enumerate(row):
                if merge_rows is not None and row_index + 1 not in merge_rows:
                    if j == 0:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            v,
                            style="merge_left",
                            auto_width=auto_width,
                            decimal=decimal,
                        )
                    elif j == len(row) - 1:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            v,
                            style="merge_right",
                            auto_width=auto_width,
                            decimal=decimal,
                        )
                    else:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            v,
                            style="merge_middle",
                            auto_width=auto_width,
                            decimal=decimal,
                        )
                elif style_only or len(row) <= 1:
                    self.insert_value2sheet(
                        worksheet,
                        f'{get_column_letter(curr_col + j)}{row_index}',
                        v,
                        style=style or "middle",
                        auto_width=auto_width,
                        decimal=decimal,
                    )
                else:
                    if j == 0:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            v,
                            style=f"{style}_left" if style else "left",
                            auto_width=auto_width,
                            decimal=decimal,
                        )
                    elif j == len(row) - 1:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            v,
                            style=f"{style}_right" if style else "right",
                            auto_width=auto_width,
                            decimal=decimal,
                        )
                    else:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            v,
                            style=f"{style}_middle" if style else "middle",
                            auto_width=auto_width,
                            decimal=decimal,
                        )

    def merge_cells(
        self,
        worksheet: Worksheet,
        start: Union[str, Tuple[int, int]],
        end: Union[str, Tuple[int, int]]
    ) -> None:
        """合并单元格并保证样式正确合并。

        :param worksheet: 工作表对象
        :param start: 开始位置，如'B2'或(2, 2)
        :param end: 结束位置，如'F10'或(10, 6)
        """
        # 解析位置
        if isinstance(start, str):
            start_col, start_row = self.get_cell_space(start)
        elif isinstance(start, (tuple, list)):
            start_col, start_row = start[0], start[1]
        else:
            raise TypeError("仅支持二元组或字符串")

        if isinstance(end, str):
            end_col, end_row = self.get_cell_space(end)
        elif isinstance(end, (tuple, list)):
            end_col, end_row = end[0], end[1]
        else:
            raise TypeError("仅支持二元组或字符串")

        # 确保起始列不大于结束列
        if start_col > end_col:
            start_col, end_col = end_col, start_col
        if start_row > end_row:
            start_row, end_row = end_row, start_row

        # 获取左上角单元格的样式
        top_left_cell = worksheet[f"{get_column_letter(start_col)}{start_row}"]
        cell_style = copy.deepcopy(top_left_cell.style)

        # 获取各边框样式
        top_border = top_left_cell.border.top
        left_cell = worksheet[f"{get_column_letter(start_col)}{start_row}"]
        left_border = left_cell.border.left
        right_cell = worksheet[f"{get_column_letter(end_col)}{start_row}"]
        right_border = right_cell.border.right
        bottom_cell = worksheet[f"{get_column_letter(start_col)}{end_row}"]
        bottom_border = bottom_cell.border.bottom

        # 创建合并后的边框样式
        border_style = Border(
            top=Side(style=top_border.style, color=top_border.color) if top_border else None,
            left=Side(style=left_border.style, color=left_border.color) if left_border else None,
            right=Side(style=right_border.style, color=right_border.color) if right_border else None,
            bottom=Side(style=bottom_border.style, color=bottom_border.color) if bottom_border else None,
        )

        # 将单元格样式应用到左上角单元格
        merged_cell = worksheet[f"{get_column_letter(start_col)}{start_row}"]
        merged_cell.style = cell_style
        merged_cell.border = border_style

        # 合并单元格
        start_cell = f"{get_column_letter(start_col)}{start_row}"
        end_cell = f"{get_column_letter(end_col)}{end_row}"
        if start_cell != end_cell:
            worksheet.merge_cells(f"{start_cell}:{end_cell}")

    @staticmethod
    def check_contain_chinese(check_str: str) -> Tuple[List[bool], int, int]:
        """检查字符串中是否包含中文。

        :param check_str: 需要检查的字符串
        :return: (每个字符是否是中文的列表, 英文字符个数, 中文字符个数)
        """
        out = []
        for ch in str(check_str).encode('utf-8').decode('utf-8'):
            if u'\u4e00' <= ch <= u'\u9fff':
                out.append(True)
            else:
                out.append(False)
        return out, len(out) - sum(out), sum(out)

    @staticmethod
    def _validate_decimal(decimal: Optional[int]) -> None:
        """校验写入 Excel 时的浮点数精度。"""
        if decimal is not None and (
            isinstance(decimal, (bool, np.bool_))
            or not isinstance(decimal, (int, np.integer))
            or int(decimal) < 0
        ):
            raise ValueError("decimal 必须是大于等于 0 的整数或 None")

    @classmethod
    def _resolve_write_speed(cls, data: pd.DataFrame, speed: str, index: bool = False) -> str:
        """解析 DataFrame 的 Excel 写入速度。"""
        if not isinstance(speed, str):
            raise ValueError("speed 必须是 'auto'、'normal' 或 'fast'")

        normalized = speed.strip().lower()
        if normalized not in {"auto", "normal", "fast"}:
            raise ValueError("speed 必须是 'auto'、'normal' 或 'fast'")
        if normalized != "auto":
            return normalized

        rows = len(data)
        columns = len(data.columns) + (data.index.nlevels if index else 0)
        if (
            rows >= cls.AUTO_FAST_MIN_ROWS
            or columns >= cls.AUTO_FAST_MIN_COLUMNS
            or rows * columns >= cls.AUTO_FAST_MIN_CELLS
        ):
            return "fast"
        return "normal"

    @staticmethod
    def astype_insertvalue(value: Any, decimal: Optional[int] = 4) -> Any:
        """格式化需要存储Excel的内容。

        :param value: 需要插入Excel的内容
        :param decimal: 如果是浮点型，需要保留的小数位数，默认为4；None 表示不主动舍入
        :return: 格式化后的内容
        """
        if re.search('tuple|list|set|numpy.ndarray|Categorical|numpy.dtype|Interval', str(type(value))):
            return str(value)
        elif re.search('float', str(type(value))):
            return float(value) if decimal is None else round(float(value), int(decimal))
        else:
            return value

    @staticmethod
    def is_numeric_like_string(value: Any) -> bool:
        """检查值是否为类似数字的字符串（如 "00123"、"123.45"、"12.34%"、"1,234.56"）。

        这类字符串在Excel中会显示绿色感叹号（以文本形式存储的数字）。

        :param value: 需要检查的值
        :return: 如果是类似数字的字符串返回True，否则返回False
        """
        if not isinstance(value, str):
            return False
        if not value:
            return False
        
        # 去除首尾空格
        value = value.strip()
        
        # 排除纯空格字符串
        if not value:
            return False
        
        # 排除常见的非数字字符串（如日期格式）
        if re.match(r'^\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2}$', value):
            return False
        
        # 检查是否为百分比格式（如 "12.34%"）
        if re.match(r'^[+-]?(\d{1,3}(,\d{3})*|\d+)(\.\d+)?%$', value):
            return True
        
        # 检查是否为带千位分隔符的数字（如 "1,234.56" 或 "1,234"）
        if re.match(r'^[+-]?(\d{1,3}(,\d{3})+|\d+)(\.\d+)?$', value):
            return True
        
        # 检查是否为纯数字字符串（包括前导零，如 "00123"）
        if re.match(r'^[+-]?\d+\.?\d*$', value):
            # 排除纯整数（非前导零情况）
            # 如果字符串长度大于1且以0开头且后面跟着数字，则是前导零数字
            if len(value) > 1 and value[0] == '0' and value[1].isdigit():
                return True
            # 检查是否包含小数点
            if '.' in value:
                return True
            # 检查是否带正负号
            if value[0] in '+-':
                return True
            return False
        
        # 检查是否为科学计数法（如 "1e5"、"1.2E-3"）
        if re.match(r'^[+-]?\d+(\.\d+)?[eE][+-]?\d+$', value):
            return True
        
        return False

    @staticmethod
    def _is_missing_merge_value(value: Any) -> bool:
        """Return True for scalar missing values used as blank merge placeholders."""
        try:
            missing = pd.isna(value)
        except (TypeError, ValueError):
            return False
        if isinstance(missing, (bool, np.bool_)):
            return bool(missing)
        return False

    @staticmethod
    def _merge_values_equal(left: Any, right: Any) -> bool:
        if ExcelWriter._is_missing_merge_value(left) and ExcelWriter._is_missing_merge_value(right):
            return True
        try:
            equal = left == right
        except (TypeError, ValueError):
            return False
        if isinstance(equal, (bool, np.bool_)):
            return bool(equal)
        return False

    @staticmethod
    def calc_continuous_cnt(list_: List, index_: int = 0) -> Tuple[Any, Optional[int], Optional[int]]:
        """计算列表中从某个索引开始连续出现某个元素的个数。

        :param list_: 需要检索的列表
        :param index_: 起始索引，默认为0
        :return: (元素值, 索引值, 连续出现的个数)

        **参考样例**

        >>> calc_continuous_cnt = ExcelWriter.calc_continuous_cnt
        >>> list_ = ['A','A','A','A','B','C','C','D','D','D']
        >>> calc_continuous_cnt(list_, 0)
        ('A', 0, 4)
        >>> calc_continuous_cnt(list_, 4)
        ('B', 4, 1)
        """
        if index_ >= len(list_):
            return None, None, None

        cnt, str_ = 0, list_[index_]
        for i in range(index_, len(list_), 1):
            if ExcelWriter._merge_values_equal(list_[i], str_):
                cnt = cnt + 1
            else:
                break
        return str_, index_, cnt

    @staticmethod
    def itlubber_border(border: List[str], color: List[str], white: bool = False) -> Border:
        """生成边框样式。

        :param border: 边框样式列表，长度为3或4。长度为3表示[左, 右, 下]，长度为4表示[左, 右, 下, 上]
        :param color: 边框颜色列表
        :param white: 是否显示白色边框，默认为False
        :return: 边框对象
        """
        if len(border) == 3:
            return Border(
                left=Side(border_style=None if not white and color[0] == "FFFFFF" else border[0], color=None if not white and color[0] == "FFFFFF" else color[0]),
                right=Side(border_style=None if not white and color[1] == "FFFFFF" else border[1], color=None if not white and color[1] == "FFFFFF" else color[1]),
                bottom=Side(border_style=border[2], color=color[2]),
            )
        else:
            return Border(
                left=Side(border_style=None if not white and color[0] == "FFFFFF" else border[0], color=None if not white and color[0] == "FFFFFF" else color[0]),
                right=Side(border_style=None if not white and color[1] == "FFFFFF" else border[1], color=None if not white and color[1] == "FFFFFF" else color[1]),
                bottom=Side(border_style=border[2], color=color[2]),
                top=Side(border_style=border[3], color=color[3]),
            )

    @staticmethod
    def get_cell_space(space: Union[str, Tuple[int, int]]) -> Union[Tuple[int, int], str]:
        """转换单元格位置格式。

        支持两种格式：
        - 字符串格式: 'B2'
        - 元组格式: (2, 2) 表示第2行第2列

        :param space: 单元格位置
        :return: 转换后的格式

        **参考样例**

        >>> get_cell_space = ExcelWriter.get_cell_space
        >>> get_cell_space("B3")
        (2, 3)
        >>> get_cell_space((2, 2))
        'B2'
        """
        if isinstance(space, str):
            start_row = int(re.findall(r"\d+", space)[0])
            start_col = re.findall(r'\D+', space)[0]
            return column_index_from_string(start_col), start_row
        else:
            start_row = space[0]
            if isinstance(space[1], int):
                start_col = get_column_letter(space[1])
            else:
                start_col = space[1]
            return f"{start_col}{start_row}"

    @staticmethod
    def calculate_rgba_color(hex_color: str, opacity: float, prefix: str = "#") -> str:
        """根据颜色和透明度计算对应的颜色值。

        :param hex_color: hex格式的颜色值
        :param opacity: 透明度，[0, 1]之间的数值
        :param prefix: 返回颜色的前缀，默认为"#"
        :return: 对应某个透明度的颜色
        """
        rgb_color = tuple(int(hex_color.lstrip('#')[i:i + 2], 16) for i in (0, 2, 4))
        rgba_color = tuple(int((1 - opacity) * c + opacity * 255) for c in rgb_color)
        return prefix + '{:02X}{:02X}{:02X}'.format(*rgba_color)

    def init_style(self, font: str, fontsize: int, theme_color: str) -> None:
        """初始化单元格样式。

        :param font: 字体名称
        :param fontsize: 字体大小
        :param theme_color: 主题颜色
        """
        # 创建所有命名样式
        header_style = NamedStyle(name="header")
        header_left_style = NamedStyle(name="header_left")
        header_middle_style = NamedStyle(name="header_middle")
        header_right_style = NamedStyle(name="header_right")

        last_style = NamedStyle(name="last")
        last_left_style = NamedStyle(name="last_left")
        last_middle_style = NamedStyle(name="last_middle")
        last_right_style = NamedStyle(name="last_right")

        content_style = NamedStyle(name="content")
        left_style = NamedStyle(name="left")
        middle_style = NamedStyle(name="middle")
        right_style = NamedStyle(name="right")

        merge_style = NamedStyle(name="merge")
        merge_left_style = NamedStyle(name="merge_left")
        merge_middle_style = NamedStyle(name="merge_middle")
        merge_right_style = NamedStyle(name="merge_right")

        first_style = NamedStyle(name="first")
        first_left_style = NamedStyle(name="first_left")
        first_middle_style = NamedStyle(name="first_middle")
        first_right_style = NamedStyle(name="first_right")

        # 字体和填充
        header_font = Font(size=fontsize, name=font, color="FFFFFF", bold=True)
        header_fill = PatternFill(fill_type="solid", start_color=theme_color)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        content_fill = PatternFill(fill_type="solid", start_color="FFFFFF")
        content_font = Font(size=fontsize, name=font, color="000000")
        even_fill = PatternFill(fill_type="solid", start_color=self.calculate_rgba_color(self.theme_color, self.opacity, prefix=""))

        # 设置header样式
        for style in [header_style, header_left_style, header_middle_style, header_right_style]:
            style.font = header_font
            style.fill = header_fill

        header_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_left_style.alignment = alignment
        header_middle_style.alignment = alignment
        header_right_style.alignment = alignment

        header_style.border = self.itlubber_border(["medium", "medium", "medium", "medium"], [theme_color, theme_color, theme_color, theme_color], white=True)
        header_left_style.border = self.itlubber_border(["medium", "thin", "medium", "medium"], [theme_color, "FFFFFF", theme_color, theme_color], white=True)
        header_middle_style.border = self.itlubber_border(["thin", "thin", "medium", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color], white=True)
        header_right_style.border = self.itlubber_border(["thin", "medium", "medium", "medium"], ["FFFFFF", theme_color, theme_color, theme_color], white=True)

        # 设置last样式
        for style in [last_style, last_left_style, last_middle_style, last_right_style]:
            style.font = content_font
            style.fill = content_fill
            style.alignment = alignment

        last_style.border = self.itlubber_border(["medium", "medium", "medium"], [theme_color, theme_color, theme_color])
        last_left_style.border = self.itlubber_border(["medium", "thin", "medium"], [theme_color, "FFFFFF", theme_color])
        last_middle_style.border = self.itlubber_border(["thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color])
        last_right_style.border = self.itlubber_border(["thin", "medium", "medium"], ["FFFFFF", theme_color, theme_color])

        # 设置content样式
        for style in [content_style, left_style, middle_style, right_style]:
            style.font = content_font
            style.fill = content_fill
            style.alignment = alignment

        content_style.border = self.itlubber_border(["medium", "medium", "thin"], [theme_color, theme_color, theme_color])
        left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", theme_color])
        middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", theme_color])
        right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, theme_color])

        # 设置merge样式
        for style in [merge_style, merge_left_style, merge_middle_style, merge_right_style]:
            style.font = content_font
            style.fill = content_fill
            style.alignment = alignment

        merge_style.border = self.itlubber_border(["medium", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", "FFFFFF"])
        merge_middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, "FFFFFF"])

        # 设置first样式
        for style in [first_style, first_left_style, first_middle_style, first_right_style]:
            style.font = content_font
            style.fill = content_fill
            style.alignment = alignment

        first_style.border = self.itlubber_border(["medium", "medium", "thin", "medium"], [theme_color, theme_color, theme_color, theme_color])
        first_left_style.border = self.itlubber_border(["medium", "thin", "thin", "medium"], [theme_color, "FFFFFF", theme_color, theme_color])
        first_middle_style.border = self.itlubber_border(["thin", "thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color])
        first_right_style.border = self.itlubber_border(["thin", "medium", "thin", "medium"], ["FFFFFF", theme_color, theme_color, theme_color])

        # 创建奇偶行样式
        middle_odd_style = NamedStyle(name="middle_odd")
        middle_odd_first_style = NamedStyle(name="middle_odd_first")
        middle_odd_last_style = NamedStyle(name="middle_odd_last")
        middle_even_style = NamedStyle(name="middle_even")
        middle_even_first_style = NamedStyle(name="middle_even_first")
        middle_even_last_style = NamedStyle(name="middle_even_last")

        for style in [middle_odd_style, middle_odd_first_style, middle_odd_last_style, middle_even_style, middle_even_first_style, middle_even_last_style]:
            style.font = content_font
            style.alignment = alignment

        middle_odd_style.fill = content_fill
        middle_odd_first_style.fill = content_fill
        middle_odd_last_style.fill = content_fill
        middle_even_style.fill = even_fill
        middle_even_first_style.fill = even_fill
        middle_even_last_style.fill = even_fill

        middle_odd_first_style.border = Border(top=Side(border_style="medium", color=self.theme_color))
        middle_odd_last_style.border = Border(bottom=Side(border_style="medium", color=self.theme_color))
        middle_even_first_style.border = Border(top=Side(border_style="medium", color=self.theme_color))
        middle_even_last_style.border = Border(bottom=Side(border_style="medium", color=self.theme_color))
        middle_even_style.border = Border(bottom=Side(border_style="thin", color="FFFFFF"))
        middle_odd_style.border = Border(bottom=Side(border_style="thin", color="FFFFFF"))

        # 收集所有样式
        self.name_styles.extend([
            header_style, header_left_style, header_middle_style, header_right_style,
            last_style, last_left_style, last_middle_style, last_right_style,
            content_style, left_style, middle_style, right_style,
            merge_style, merge_left_style, merge_middle_style, merge_right_style,
            first_style, first_left_style, first_middle_style, first_right_style,
            middle_odd_style, middle_even_first_style, middle_odd_last_style, middle_even_style, middle_odd_first_style, middle_even_last_style,
        ])

    @staticmethod
    def _to_argb(color: str) -> str:
        """将颜色规范化为 ARGB（8位十六进制，大写）。

        :param color: 颜色值，支持 ``#RRGGBB`` / ``RRGGBB`` / ``AARRGGBB``
        :return: ``AARRGGBB`` 格式颜色（默认不透明 FF）
        """
        # 复用 _pivot 中的实现，保持单一来源
        return _pivot._to_argb(color)

    @staticmethod
    def _quote_sheet_title(title: str) -> str:
        """为公式引用规范化 sheet 名称（含空格或特殊字符时用单引号包裹）。"""
        if re.match(r'^[A-Za-z一-鿿_][A-Za-z0-9一-鿿_]*$', title):
            return title
        return "'{}'".format(title.replace("'", "''"))

    def add_sparkline(
        self,
        worksheet: Union[Worksheet, str],
        location: Union[str, List[str]],
        data_range: Union[str, List[str]],
        type: str = "line",
        series_color: Optional[str] = None,
        negative_color: Optional[str] = None,
        markers: bool = False,
        marker_color: Optional[str] = None,
        high_point: bool = False,
        low_point: bool = False,
        first_point: bool = False,
        last_point: bool = False,
        negative_points: bool = False,
        high_color: Optional[str] = None,
        low_color: Optional[str] = None,
        first_color: Optional[str] = None,
        last_color: Optional[str] = None,
        display_x_axis: bool = False,
        show_empty_as: str = "gap",
        line_weight: Optional[float] = None,
    ) -> None:
        """向单元格插入迷你图（Sparkline）。

        在单个单元格内绘制折线图、柱状图或盈亏图，效果类似 xlsxwriter 的 ``add_sparkline``。
        由于 openpyxl 不支持写入迷你图，本方法仅记录配置，在 :meth:`save` 时将
        ``x14:sparklineGroups`` 注入到 worksheet XML 中。默认配色采用 hscredit 主题风格。

        .. note::
            迷你图通过直接修改 xlsx 内部 XML 实现，Excel 可正常显示。但 openpyxl 不识别该扩展，
            若用 openpyxl 重新打开并保存（含 ``mode='append'`` 追加模式），迷你图会丢失。
            建议迷你图在最终输出步骤添加。

        :param worksheet: 工作表对象或名称
        :param location: 迷你图所在单元格，如'H2'；可传列表与 ``data_range`` 一一对应批量生成同组迷你图
        :param data_range: 数据区域，如'B2:G2'或'Sheet1!B2:G2'（未含sheet名时默认当前sheet）；可传列表
        :param type: 迷你图类型，可选'line'(折线)、'column'(柱状)、'win_loss'(盈亏)，默认'line'
        :param series_color: 主体颜色，默认为主题色
        :param negative_color: 负值颜色（盈亏图/柱状图负值），默认为 hscredit 坏样本红
        :param markers: 是否显示数据点标记（仅折线图），默认为False
        :param marker_color: 标记颜色，默认同 ``series_color``
        :param high_point: 是否高亮最高点，默认为False
        :param low_point: 是否高亮最低点，默认为False
        :param first_point: 是否高亮首点，默认为False
        :param last_point: 是否高亮尾点，默认为False
        :param negative_points: 是否高亮负值点，默认为False
        :param high_color: 最高点颜色，默认同 ``series_color``
        :param low_color: 最低点颜色，默认同 ``negative_color``
        :param first_color: 首点颜色，默认为提升橙
        :param last_color: 尾点颜色，默认为提升橙
        :param display_x_axis: 是否显示横轴（数据含正负时分隔），默认为False
        :param show_empty_as: 空单元格显示方式，可选'gap'(留空)、'zero'(零)、'span'(连线)，默认'gap'
        :param line_weight: 折线粗细（磅），默认为None（使用Excel默认）

        **参考样例**

        >>> # 在 H2 单元格按 B2:G2 数据绘制折线迷你图
        >>> writer.add_sparkline(ws, "H2", "B2:G2", markers=True, high_point=True, low_point=True)
        >>>
        >>> # 柱状迷你图
        >>> writer.add_sparkline(ws, "H3", "B3:G3", type="column")
        >>>
        >>> # 盈亏迷你图
        >>> writer.add_sparkline(ws, "H4", "B4:G4", type="win_loss")
        """
        sheet_title = worksheet.title if isinstance(worksheet, Worksheet) else worksheet

        type_map = {"line": "line", "column": "column", "win_loss": "stacked"}
        if type not in type_map:
            raise ValueError("type 仅支持 'line'、'column' 或 'win_loss'")

        if show_empty_as not in ("gap", "zero", "span"):
            raise ValueError("show_empty_as 仅支持 'gap'、'zero' 或 'span'")

        # 规范化为并列列表
        locations = [location] if isinstance(location, str) else list(location)
        ranges = [data_range] if isinstance(data_range, str) else list(data_range)
        if len(locations) != len(ranges):
            raise ValueError("location 与 data_range 数量必须一致")

        # 为未含sheet名的数据区域补全当前sheet前缀
        quoted_title = self._quote_sheet_title(sheet_title)
        norm_ranges = []
        for rng in ranges:
            norm_ranges.append(rng if "!" in rng else f"{quoted_title}!{rng}")

        # 颜色（hscredit 默认配色）
        _series = self._to_argb(series_color or self.theme_color)
        _negative = self._to_argb(negative_color or HSCREDIT_CHART_COLORS[1])
        colors = {
            "colorSeries": _series,
            "colorNegative": _negative,
            "colorMarkers": self._to_argb(marker_color or series_color or self.theme_color),
            "colorFirst": self._to_argb(first_color or HSCREDIT_CHART_COLORS[2]),
            "colorLast": self._to_argb(last_color or HSCREDIT_CHART_COLORS[2]),
            "colorHigh": self._to_argb(high_color or series_color or self.theme_color),
            "colorLow": self._to_argb(low_color or negative_color or HSCREDIT_CHART_COLORS[1]),
        }

        # group 级属性
        attrs: Dict[str, str] = {"displayEmptyCellsAs": show_empty_as}
        if type_map[type] != "line":
            attrs["type"] = type_map[type]
        if markers:
            attrs["markers"] = "1"
        if high_point:
            attrs["high"] = "1"
        if low_point:
            attrs["low"] = "1"
        if first_point:
            attrs["first"] = "1"
        if last_point:
            attrs["last"] = "1"
        if negative_points:
            attrs["negative"] = "1"
        if display_x_axis:
            attrs["displayXAxis"] = "1"
        if line_weight is not None:
            attrs["lineWeight"] = str(line_weight)

        self._sparkline_specs.append({
            "sheet": sheet_title,
            "attrs": attrs,
            "colors": colors,
            "sparklines": list(zip(norm_ranges, locations)),
        })

    def _build_sparkline_groups_xml(self, specs: List[Dict[str, Any]]) -> str:
        """根据迷你图规格构建单个 sheet 的 ``x14:sparklineGroups`` 子元素 XML。"""
        # 颜色元素需按 schema 顺序输出
        color_order = [
            "colorSeries", "colorNegative", "colorAxis", "colorMarkers",
            "colorFirst", "colorLast", "colorHigh", "colorLow",
        ]
        groups = []
        for spec in specs:
            attr_str = "".join(f' {k}="{v}"' for k, v in spec["attrs"].items())
            color_str = "".join(
                f'<x14:{tag} rgb="{spec["colors"][tag]}"/>'
                for tag in color_order if tag in spec["colors"]
            )
            sparkline_str = "".join(
                f'<x14:sparkline><xm:f>{f_ref}</xm:f><xm:sqref>{sqref}</xm:sqref></x14:sparkline>'
                for f_ref, sqref in spec["sparklines"]
            )
            groups.append(
                f'<x14:sparklineGroup{attr_str}>{color_str}'
                f'<x14:sparklines>{sparkline_str}</x14:sparklines></x14:sparklineGroup>'
            )
        return "".join(groups)

    def _inject_sparklines(self, filename: str) -> None:
        """在 openpyxl 保存后，将迷你图 XML 注入到 xlsx 对应的 worksheet 部件中。

        openpyxl 不支持写入 sparkline，故通过直接修改 xlsx（zip）内的 worksheet XML 实现。

        :param filename: 已保存的 xlsx 文件路径
        """
        if not self._sparkline_specs:
            return

        # 按 sheet 名归集规格
        specs_by_sheet: Dict[str, List[Dict[str, Any]]] = {}
        for spec in self._sparkline_specs:
            specs_by_sheet.setdefault(spec["sheet"], []).append(spec)

        with zipfile.ZipFile(filename, "r") as zin:
            names = zin.namelist()
            workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")

            # sheet 名 -> r:id
            name_to_rid = {}
            for m in re.finditer(r"<sheet\b[^>]*/>", workbook_xml):
                tag = m.group(0)
                name_m = re.search(r'name="([^"]*)"', tag)
                rid_m = re.search(r'r:id="([^"]*)"', tag)
                if name_m and rid_m:
                    name_to_rid[name_m.group(1)] = rid_m.group(1)

            # r:id -> worksheet 部件路径
            rid_to_target = {}
            for m in re.finditer(r"<Relationship\b[^>]*/>", rels_xml):
                tag = m.group(0)
                if "worksheet" not in tag:
                    continue
                id_m = re.search(r'Id="([^"]*)"', tag)
                tgt_m = re.search(r'Target="([^"]*)"', tag)
                if id_m and tgt_m:
                    target = tgt_m.group(1)
                    if target.startswith("/"):
                        part = target.lstrip("/")
                    else:
                        part = "xl/" + target
                    rid_to_target[id_m.group(1)] = part

            # 修改目标 worksheet XML
            modified: Dict[str, bytes] = {}
            for sheet_name, specs in specs_by_sheet.items():
                rid = name_to_rid.get(sheet_name)
                part = rid_to_target.get(rid) if rid else None
                if not part or part not in names:
                    continue

                sheet_xml = zin.read(part).decode("utf-8")
                groups_xml = self._build_sparkline_groups_xml(specs)
                ext_xml = (
                    '<ext xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" '
                    'uri="{05C60535-1F16-4fd2-B633-F4F36F0B64E0}">'
                    '<x14:sparklineGroups xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
                    + groups_xml +
                    '</x14:sparklineGroups></ext>'
                )

                if "</extLst>" in sheet_xml:
                    # 已存在 extLst，将 ext 追加到最后一个 extLst 内
                    idx = sheet_xml.rfind("</extLst>")
                    sheet_xml = sheet_xml[:idx] + ext_xml + sheet_xml[idx:]
                else:
                    # 无 extLst，在 </worksheet> 前新增（extLst 必须为 worksheet 最后一个子元素）
                    idx = sheet_xml.rfind("</worksheet>")
                    sheet_xml = sheet_xml[:idx] + "<extLst>" + ext_xml + "</extLst>" + sheet_xml[idx:]

                modified[part] = sheet_xml.encode("utf-8")

            if not modified:
                return

            # 重写 zip
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(tmp_fd)
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = modified.get(item.filename, zin.read(item.filename))
                    zout.writestr(item, data)

        shutil.move(tmp_path, filename)

    # ------------------------------------------------------------------ #
    # 数据透视表 / 数据透视图
    # ------------------------------------------------------------------ #

    @staticmethod
    def _parse_anchor(space: Union[str, Tuple[int, int]]) -> Tuple[int, int]:
        """将 'B2' 或 (row, col) 统一解析为 (row, col) 整型元组。"""
        if isinstance(space, str):
            row = int(re.findall(r"\d+", space)[0])
            col = column_index_from_string(re.findall(r"\D+", space)[0])
            return row, col
        return int(space[0]), int(space[1])

    @staticmethod
    def _sanitize_sheet_title(title: str) -> str:
        """将字符串规范为合法的Excel工作表名（替换非法字符并截断到31字符）。

        Excel 工作表名不允许包含 ``: \\ / ? * [ ]``，且长度上限为31个字符。

        :param title: 原始名称
        :return: 合法的工作表名
        """
        cleaned = re.sub(r"[:\\/?*\[\]]", "_", str(title)).strip()
        return (cleaned[:31] or "Sheet").rstrip()

    def _unique_source_sheet_name(self, name: str) -> str:
        """为自动创建的源数据表生成合法且唯一的工作表名（≤31字符）。

        :param name: 透视表名称
        :return: 不与现有工作表重名的合法源数据表名
        """
        suffix = "_源数据"
        # 预留后缀长度，仅截断名称主体，保证 ``_源数据`` 标识不被裁掉且总长 ≤31
        stem = self._sanitize_sheet_title(name)[: 31 - len(suffix)].rstrip()
        base = stem + suffix
        if base not in self.workbook.sheetnames:
            return base
        i = 2
        while True:
            seq = str(i)
            candidate = (stem[: 31 - len(suffix) - len(seq)].rstrip() + suffix + seq)
            if candidate not in self.workbook.sheetnames:
                return candidate
            i += 1

    def insert_pivot_table2sheet(
        self,
        worksheet: Union[Worksheet, str],
        data: pd.DataFrame,
        pivot_anchor: Union[str, Tuple[int, int]],
        rows: Optional[Union[str, List[str]]] = None,
        columns: Optional[Union[str, List[str]]] = None,
        values: Optional[Any] = None,
        filters: Optional[Union[str, List[str], Dict[Any, Any]]] = None,
        filter_items: Optional[Dict[Any, Any]] = None,
        groups: Optional[Dict[Any, Any]] = None,
        subtotals: bool = False,
        source_sheet: Optional[Union[Worksheet, str]] = None,
        source_anchor: Union[str, Tuple[int, int]] = (1, 1),
        write_source: Optional[bool] = None,
        name: Optional[str] = None,
        show_row_totals: bool = True,
        show_col_totals: bool = True,
        theme_style: bool = True,
        style: Optional[str] = None,
        fill: bool = True,
    ) -> Tuple[int, int]:
        """向工作表插入Excel原生数据透视表。

        由于 openpyxl 不支持创建数据透视表，本方法仅记录配置，在 :meth:`save` 时将
        ``pivotCacheDefinition`` / ``pivotCacheRecords`` / ``pivotTable`` 等部件注入到
        xlsx 中，生成 Excel 可交互、可刷新的原生数据透视表。透视缓存写入 ``refreshOnLoad="1"``，
        Excel 打开时会基于源数据自动刷新。

        .. note::
            透视表通过在保存后直接注入 xlsx 内部部件实现。openpyxl 追加模式（``mode='append'``）
            会基于单元格值/样式重建工作簿，不保留原文件中的透视表、原生图、迷你图与图片。
            因此请勿用追加模式向已含透视表的文件追加内容，否则原透视表会丢失；建议透视表在最终输出步骤生成。

        :param worksheet: 透视表放置的工作表对象或名称
        :param data: 源数据 DataFrame（透视缓存基于此构建）
        :param pivot_anchor: 透视表左上角锚点，如'B2'或(2, 2)
        :param rows: 行字段（列名或列名列表），默认为None
        :param columns: 列字段（列名或列名列表），默认为None
        :param values: 值字段，支持多种形式：

            - ``'金额'`` 或 ``['金额', '数量']``：默认聚合（数值列求和，非数值列计数）
            - ``[('金额', 'sum'), ('数量', 'mean')]``：显式指定聚合
            - ``[('金额', 'sum', '全局占比')]``：聚合 + 占比显示（全局/行/列/组合占比）
            - ``{'金额': 'sum'}`` 或 ``[{'field': '金额', 'agg': 'sum', 'show_as': '全局占比',
              'name': '占比', 'number_format': '0.00%'}]``

            聚合：sum/count/average(mean)/max/min/product/count_nums/std/stdp/var/varp
            （可用 ``hscredit.excel._pivot.register_aggregation`` 扩展别名）
        :param filters: 页/筛选字段。列名/列名列表，或 ``{字段: [允许值]}`` 直接指定筛选项
        :param filter_items: 筛选项 ``{字段: [允许值]}``，可作用于行/列/筛选任一字段，仅保留所列取值
        :param groups: 数值字段分组 ``{字段: {'start': 起始值, 'interval': 步长}}``
            （亦支持 ``{字段: (起始, 步长)}``），对横轴/纵轴数值特征按步长分桶统计
        :param subtotals: 是否对非最内层行/列字段显示分类汇总，默认为False
        :param source_sheet: 源数据所在工作表对象或名称，默认为None（自动新建源数据表写入 ``data``）
        :param source_anchor: 源数据写入/定位的左上角（含表头），默认为(1, 1)
        :param write_source: 是否写入源数据，默认为None（``source_sheet`` 为None时自动写入）
        :param name: 透视表名称，默认为None（自动命名「数据透视表N」）
        :param show_row_totals: 是否显示行总计，默认为True
        :param show_col_totals: 是否显示列总计，默认为True
        :param theme_style: 是否套用适配 ``theme_color`` 的 hscredit 主题样式，默认为True
        :param style: 透视表样式名，默认为None（``theme_style`` 为True时用主题样式，否则用内置 PivotStyleLight16）
        :param fill: 自动写入源数据时是否使用颜色填充，默认为True
        :return: (透视表区域下一行行号, 下一列列号)

        **参考样例**

        >>> import pandas as pd
        >>> from hscredit.excel import ExcelWriter
        >>> df = pd.DataFrame({
        ...     '商品类别': ['数码', '服饰', '数码', '服饰'],
        ...     '区域': ['华东', '华东', '华南', '华南'],
        ...     '放款金额': [100, 200, 300, 400],
        ... })
        >>> with ExcelWriter(theme_color='2639E9').set_filename('pivot.xlsx') as writer:
        ...     ws = writer.get_sheet_by_name('透视表')
        ...     writer.insert_pivot_table2sheet(
        ...         ws, df, 'B2',
        ...         rows='商品类别', columns='区域',
        ...         values=[('放款金额', 'sum'), ('放款金额', 'sum', '全局占比')],
        ...         groups={'放款金额': {'start': 0, 'interval': 100}},
        ...         subtotals=True,
        ...     )
        """
        if values is None:
            raise ValueError("数据透视表至少需要指定一个值字段（values）")

        data = data.copy()
        rows = [] if rows is None else ([rows] if isinstance(rows, str) else list(rows))
        columns = [] if columns is None else ([columns] if isinstance(columns, str) else list(columns))

        # 解析筛选字段与筛选项（filters 可为列表或 {字段:[允许值]}）
        filter_items_all: Dict[Any, Any] = {}
        if isinstance(filters, dict):
            filter_fields = list(filters.keys())
            filter_items_all.update(filters)
        elif filters is None:
            filter_fields = []
        else:
            filter_fields = [filters] if isinstance(filters, str) else list(filters)
        if filter_items:
            filter_items_all.update(filter_items)

        # 规范化数值分组：{字段: {'start':, 'interval':}}
        groups_norm: Dict[Any, Dict[str, float]] = {}
        for f, g in (groups or {}).items():
            if isinstance(g, dict):
                groups_norm[f] = {"start": float(g["start"]), "interval": float(g["interval"])}
            elif isinstance(g, (list, tuple)) and len(g) >= 2:
                groups_norm[f] = {"start": float(g[0]), "interval": float(g[1])}
            else:
                raise ValueError("分组配置须为 {'start':.., 'interval':..} 或 (start, interval)")

        if not isinstance(worksheet, Worksheet):
            worksheet = self.get_sheet_by_name(worksheet)
        pivot_sheet = worksheet.title
        pivot_anchor = self._parse_anchor(pivot_anchor)

        name = name or "数据透视表{}".format(len(self._pivot_specs) + 1)
        value_fields = _pivot.normalize_values(values, data)

        # 解析/写入源数据
        src_row, src_col = self._parse_anchor(source_anchor)
        if source_sheet is None:
            source_ws = self.get_sheet_by_name(self._unique_source_sheet_name(name))
            do_write = True if write_source is None else write_source
        else:
            source_ws = source_sheet if isinstance(source_sheet, Worksheet) else self.get_sheet_by_name(source_sheet)
            do_write = bool(write_source)
        source_sheet_name = source_ws.title

        if do_write:
            self.insert_df2sheet(source_ws, data, (src_row, src_col), header=True, index=False, fill=fill)

        # 源数据区域引用（worksheetSource 的 sheet 单独存储，ref 不含 sheet 前缀/$）
        n_cols = len(data.columns)
        n_rows = len(data)
        source_ref = "{}{}:{}{}".format(
            get_column_letter(src_col), src_row,
            get_column_letter(src_col + n_cols - 1), src_row + n_rows,
        )

        cache_id = len(self._pivot_specs)
        spec = _pivot.build_pivot_spec(
            data=data,
            source_sheet=source_sheet_name,
            source_ref=source_ref,
            pivot_sheet=pivot_sheet,
            pivot_anchor=pivot_anchor,
            rows=rows, columns=columns, values=value_fields, filters=filter_fields,
            name=name, cache_id=cache_id,
            show_row_totals=show_row_totals, show_col_totals=show_col_totals,
            groups=groups_norm, filter_items=filter_items_all, subtotals=subtotals,
        )
        # 样式：默认套用适配主题色的 hscredit 自定义样式
        if style is not None:
            spec["style"] = style
            spec["_theme"] = False
        elif theme_style:
            spec["style"] = _pivot.THEME_PIVOT_STYLE_NAME
            spec["show_row_stripes"] = True
            spec["_theme"] = True
        else:
            spec["style"] = "PivotStyleLight16"
            spec["_theme"] = False
        self._pivot_specs.append(spec)

        layout = _pivot.compute_pivot_layout(spec)
        return pivot_anchor[0] + layout["height"], pivot_anchor[1] + layout["width"]

    def insert_pivot_chart2sheet(
        self,
        worksheet: Union[Worksheet, str],
        chart_anchor: Union[str, Tuple[int, int]],
        pivot_name: Optional[str] = None,
        chart_type: str = "bar",
        title: Optional[str] = None,
        width: float = 15.0,
        height: float = 7.5,
        bar_stacked: bool = False,
    ) -> Tuple[int, int]:
        """基于已创建的数据透视表插入Excel原生数据透视图。

        本方法先用 openpyxl 在透视表输出区域上构建普通图表（柱状/折线/饼图），
        并在 :meth:`save` 时向该图表 XML 注入 ``c:pivotSource``，使其成为绑定到透视表的
        原生数据透视图，可随透视表刷新联动。

        需先调用 :meth:`insert_pivot_table2sheet` 创建透视表。

        :param worksheet: 透视图放置的工作表对象或名称（通常与透视表同表）
        :param chart_anchor: 图表插入位置，如'H2'或(2, 8)
        :param pivot_name: 关联的透视表名称，默认为None（取最近创建的透视表）
        :param chart_type: 图表类型，可选'bar'(柱状)、'line'(折线)、'pie'(饼图)，默认'bar'
        :param title: 图表标题，默认为None
        :param width: 图表宽度（厘米），默认为15.0
        :param height: 图表高度（厘米），默认为7.5
        :param bar_stacked: 柱状图是否堆叠，默认为False
        :return: (下一行行号, 下一列列号)

        **参考样例**

        >>> writer.insert_pivot_table2sheet(ws, df, 'B2', rows='商品类别', values=[('放款金额', 'sum')])
        >>> writer.insert_pivot_chart2sheet(ws, 'H2', chart_type='bar', title='各类别放款金额')
        """
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        if not self._pivot_specs:
            raise ValueError("请先调用 insert_pivot_table2sheet 创建数据透视表")

        # 定位关联透视表
        spec = None
        if pivot_name is None:
            spec = self._pivot_specs[-1]
        else:
            for s in self._pivot_specs:
                if s["name"] == pivot_name:
                    spec = s
                    break
            if spec is None:
                raise ValueError("未找到名为 '{}' 的数据透视表".format(pivot_name))

        if not isinstance(worksheet, Worksheet):
            worksheet = self.get_sheet_by_name(worksheet)

        layout = _pivot.compute_pivot_layout(spec)
        anchor_row, anchor_col = spec["pivot_anchor"]
        # 透视表输出区域：类别在首个行标签列，数值在数据列
        cat_col = anchor_col
        header_row = anchor_row + layout["first_data_row"] - 1
        data_first_row = anchor_row + layout["first_data_row"]
        data_last_row = data_first_row + max(1, layout["n_row_leaf"]) - 1
        val_first_col = anchor_col + layout["first_data_col"]
        val_last_col = val_first_col + layout["n_col_leaf"] - 1
        pivot_ws = self.get_sheet_by_name(spec["pivot_sheet"])

        type_map = {"bar": BarChart, "line": LineChart, "pie": PieChart}
        if chart_type not in type_map:
            raise ValueError("chart_type 仅支持 'bar'、'line' 或 'pie'")
        chart = type_map[chart_type]()
        if chart_type == "bar":
            chart.type = "col"
            chart.grouping = "stacked" if bar_stacked else "clustered"
            if bar_stacked:
                chart.overlap = 100

        data_ref = Reference(
            pivot_ws, min_col=val_first_col, max_col=val_last_col,
            min_row=header_row, max_row=data_last_row,
        )
        cats_ref = Reference(pivot_ws, min_col=cat_col, min_row=data_first_row, max_row=data_last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        if title:
            chart.title = self._styled_chart_title(title)
        try:
            chart.legend.position = "b"
        except Exception:
            pass

        # 主题配色
        try:
            for i, series in enumerate(chart.series):
                color = HSCREDIT_CHART_COLORS[i % len(HSCREDIT_CHART_COLORS)]
                series.graphicalProperties.solidFill = color
        except Exception:
            pass

        next_pos = self.insert_chart2sheet(worksheet, chart_anchor, chart, width=width, height=height)

        # 记录待注入 pivotSource 的透视图（按透视表所在 sheet 名匹配 chart 部件）
        self._pivot_chart_specs.append({
            "pivot_name": spec["name"],
            "pivot_sheet": spec["pivot_sheet"],
            "cache_id": spec["cache_id"],
        })
        return next_pos

    def _inject_pivots(self, filename: str) -> None:
        """在 openpyxl 保存后，将数据透视表相关部件注入到 xlsx（zip）中。

        :param filename: 已保存的 xlsx 文件路径
        """
        if not self._pivot_specs:
            return

        with zipfile.ZipFile(filename, "r") as zin:
            names = set(zin.namelist())
            content_types = zin.read("[Content_Types].xml").decode("utf-8")
            workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
            wb_rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            styles_xml = zin.read("xl/styles.xml").decode("utf-8") if "xl/styles.xml" in names else None

            # 收集自定义数字格式并分配 numFmtId（避开内置与现有自定义 id）
            existing_fmt_ids = [int(x) for x in re.findall(r'numFmtId="(\d+)"', styles_xml or "")]
            next_fmt_id = max(existing_fmt_ids + [163]) + 1  # 内置 id 上限约 163
            fmt_id_map: Dict[str, int] = {}
            custom_numfmts: List[Tuple[int, str]] = []
            for spec in self._pivot_specs:
                for v in spec["values"]:
                    nf = v.get("number_format")
                    if nf and str(nf).lower() not in _pivot.BUILTIN_NUM_FMTS and nf not in fmt_id_map:
                        fmt_id_map[nf] = next_fmt_id
                        custom_numfmts.append((next_fmt_id, nf))
                        next_fmt_id += 1
            want_theme = any(spec.get("_theme") for spec in self._pivot_specs)

            # sheet 名 -> r:id -> worksheet 部件路径（复用 sparkline 注入同样的解析方式）
            name_to_rid = {}
            for m in re.finditer(r"<sheet\b[^>]*/>", workbook_xml):
                tag = m.group(0)
                name_m = re.search(r'name="([^"]*)"', tag)
                rid_m = re.search(r'r:id="([^"]*)"', tag)
                if name_m and rid_m:
                    name_to_rid[name_m.group(1)] = rid_m.group(1)
            rid_to_target = {}
            for m in re.finditer(r"<Relationship\b[^>]*/>", wb_rels):
                tag = m.group(0)
                if "worksheet" not in tag:
                    continue
                id_m = re.search(r'Id="([^"]*)"', tag)
                tgt_m = re.search(r'Target="([^"]*)"', tag)
                if id_m and tgt_m:
                    target = tgt_m.group(1)
                    part = target.lstrip("/") if target.startswith("/") else "xl/" + target
                    rid_to_target[id_m.group(1)] = part

            # 现有 rId 最大值，避免冲突
            existing_rids = [int(m) for m in re.findall(r'Id="rId(\d+)"', wb_rels)]
            next_rid = max(existing_rids) + 1 if existing_rids else 1

            new_parts: Dict[str, bytes] = {}
            sheet_pivot_rels: Dict[str, List[Tuple[str, str]]] = {}  # sheet part -> [(rid, pivotTable target)]
            wb_pivotcache_entries: List[Tuple[int, str]] = []       # (cacheId, rId)
            wb_new_rels: List[Tuple[str, str, str]] = []            # (rId, type, target)
            ct_overrides: List[Tuple[str, str]] = []                # (PartName, ContentType)

            for i, spec in enumerate(self._pivot_specs, start=1):
                cache_def_part = "xl/pivotCache/pivotCacheDefinition{}.xml".format(i)
                cache_rec_part = "xl/pivotCache/pivotCacheRecords{}.xml".format(i)
                pivot_part = "xl/pivotTables/pivotTable{}.xml".format(i)
                cache_def_rels = "xl/pivotCache/_rels/pivotCacheDefinition{}.xml.rels".format(i)
                pivot_rels = "xl/pivotTables/_rels/pivotTable{}.xml.rels".format(i)

                # cacheDefinition -> records 的关系（部件内固定 rId1）
                cache_def_xml = _pivot.render_cache_definition_xml(spec, "rId1")
                cache_rec_xml = _pivot.render_cache_records_xml(spec)
                pivot_xml = _pivot.render_pivot_table_xml(spec, fmt_id_map=fmt_id_map)

                new_parts[cache_def_part] = cache_def_xml.encode("utf-8")
                new_parts[cache_rec_part] = cache_rec_xml.encode("utf-8")
                new_parts[pivot_part] = pivot_xml.encode("utf-8")
                new_parts[cache_def_rels] = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    '<Relationship Id="rId1" Type="{}" Target="pivotCacheRecords{}.xml"/>'
                    '</Relationships>'.format(_pivot.REL_CACHE_REC, i)
                ).encode("utf-8")
                new_parts[pivot_rels] = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    '<Relationship Id="rId1" Type="{}" Target="../pivotCache/pivotCacheDefinition{}.xml"/>'
                    '</Relationships>'.format(_pivot.REL_CACHE_DEF, i)
                ).encode("utf-8")

                # workbook -> cacheDefinition 关系 + pivotCaches 条目
                rid = "rId{}".format(next_rid)
                next_rid += 1
                wb_new_rels.append((rid, _pivot.REL_CACHE_DEF, "pivotCache/pivotCacheDefinition{}.xml".format(i)))
                wb_pivotcache_entries.append((spec["cache_id"], rid))

                # sheet -> pivotTable 关系
                sheet_part = rid_to_target.get(name_to_rid.get(spec["pivot_sheet"]))
                if sheet_part:
                    sheet_rels_part = sheet_part.replace("xl/worksheets/", "xl/worksheets/_rels/") + ".rels"
                    rel_target = "../pivotTables/pivotTable{}.xml".format(i)
                    sheet_pivot_rels.setdefault(sheet_rels_part, []).append((_pivot.REL_PIVOT_TABLE, rel_target))

                ct_overrides.append(("/" + cache_def_part, _pivot.CT_CACHE_DEF))
                ct_overrides.append(("/" + cache_rec_part, _pivot.CT_CACHE_REC))
                ct_overrides.append(("/" + pivot_part, _pivot.CT_PIVOT_TABLE))

            # 1) [Content_Types].xml 追加 Override
            override_xml = "".join(
                '<Override PartName="{}" ContentType="{}"/>'.format(pn, ct) for pn, ct in ct_overrides
            )
            content_types = content_types.replace("</Types>", override_xml + "</Types>")

            # 1.5) styles.xml 注入 hscredit 主题透视表样式与自定义数字格式
            if styles_xml is not None:
                styles_xml = _pivot.apply_pivot_styles(
                    styles_xml=styles_xml,
                    want_theme=want_theme,
                    theme_color=self.theme_color,
                    stripe_color=self.calculate_rgba_color(self.theme_color, self.opacity, prefix=""),
                    custom_numfmts=custom_numfmts,
                )

            # 2) workbook.xml 追加 <pivotCaches>（位于 extLst / </workbook> 之前）
            pivotcaches_xml = "<pivotCaches>" + "".join(
                '<pivotCache cacheId="{}" r:id="{}"/>'.format(cid, rid)
                for cid, rid in wb_pivotcache_entries
            ) + "</pivotCaches>"
            if "<extLst" in workbook_xml:
                idx = workbook_xml.find("<extLst")
                workbook_xml = workbook_xml[:idx] + pivotcaches_xml + workbook_xml[idx:]
            else:
                workbook_xml = workbook_xml.replace("</workbook>", pivotcaches_xml + "</workbook>")
            # 确保 r 命名空间存在（openpyxl 默认会写入，稳妥起见兜底）
            if "xmlns:r=" not in workbook_xml.split(">", 1)[0]:
                workbook_xml = workbook_xml.replace(
                    "<workbook ",
                    '<workbook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                    1,
                )

            # 3) workbook.xml.rels 追加 workbook->cacheDefinition 关系
            wb_rel_add = "".join(
                '<Relationship Id="{}" Type="{}" Target="{}"/>'.format(rid, typ, tgt)
                for rid, typ, tgt in wb_new_rels
            )
            wb_rels = wb_rels.replace("</Relationships>", wb_rel_add + "</Relationships>")

            # 4) 各 sheet 的 _rels 追加 sheet->pivotTable 关系（无则新建）
            #    已存在的 rels 视为「修改」，不存在的视为「新增部件」（二者写入 zip 的方式不同）
            modified_sheet_rels: Dict[str, bytes] = {}
            for sheet_rels_part, rels in sheet_pivot_rels.items():
                if sheet_rels_part in names:
                    existing = zin.read(sheet_rels_part).decode("utf-8")
                    existing_ids = [int(m) for m in re.findall(r'Id="rId(\d+)"', existing)]
                    sid = max(existing_ids) + 1 if existing_ids else 1
                    add = ""
                    for typ, tgt in rels:
                        add += '<Relationship Id="rId{}" Type="{}" Target="{}"/>'.format(sid, typ, tgt)
                        sid += 1
                    existing = existing.replace("</Relationships>", add + "</Relationships>")
                    modified_sheet_rels[sheet_rels_part] = existing.encode("utf-8")
                else:
                    sid = 1
                    add = ""
                    for typ, tgt in rels:
                        add += '<Relationship Id="rId{}" Type="{}" Target="{}"/>'.format(sid, typ, tgt)
                        sid += 1
                    # 新建的 rels 部件须加入 new_parts，否则不会写入 zip
                    new_parts[sheet_rels_part] = (
                        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
                        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                        + add + '</Relationships>'
                    ).encode("utf-8")

            # 5) 透视图：向对应 chart XML 注入 pivotSource
            chart_modifications = self._build_pivot_chart_modifications(zin, names)

            # 汇总所有改动
            modified: Dict[str, bytes] = {
                "[Content_Types].xml": content_types.encode("utf-8"),
                "xl/workbook.xml": workbook_xml.encode("utf-8"),
                "xl/_rels/workbook.xml.rels": wb_rels.encode("utf-8"),
            }
            if styles_xml is not None:
                modified["xl/styles.xml"] = styles_xml.encode("utf-8")
            modified.update(modified_sheet_rels)
            modified.update(chart_modifications)

            # 重写 zip：保留原有部件（含改动），追加新部件
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(tmp_fd)
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    payload = modified.get(item.filename, zin.read(item.filename))
                    zout.writestr(item, payload)
                for part, payload in new_parts.items():
                    zout.writestr(part, payload)

        shutil.move(tmp_path, filename)

    def _build_pivot_chart_modifications(self, zin: 'zipfile.ZipFile', names: set) -> Dict[str, bytes]:
        """为数据透视图向对应 chart XML 注入 ``c:pivotSource``。

        按透视表所在 sheet 名匹配 chart 部件（chart 的系列引用透视表输出区域，
        其 ``<c:f>`` 含该 sheet 名）。

        :param zin: 打开的 xlsx ZipFile（只读）
        :param names: zip 内部件名集合
        :return: {chart 部件路径: 修改后的 XML 字节}
        """
        result: Dict[str, bytes] = {}
        if not self._pivot_chart_specs:
            return result

        chart_parts = sorted(n for n in names if re.match(r"xl/charts/chart\d+\.xml$", n))
        used = set()
        for pc in self._pivot_chart_specs:
            sheet = pc["pivot_sheet"]
            for part in chart_parts:
                if part in used:
                    continue
                xml = zin.read(part).decode("utf-8")
                if "pivotSource" in xml:
                    continue
                # chart 的系列公式引用透视表所在 sheet（openpyxl 可能以单引号包裹 sheet 名）
                if ("'{}'!".format(sheet) not in xml) and ("{}!".format(sheet) not in xml):
                    continue
                # 兼容前缀命名空间（c:）与默认命名空间（openpyxl 默认无前缀）
                m = re.search(r"<(c:)?chart(?:>|\s)", xml)
                if not m:
                    continue
                prefix = m.group(1) or ""
                name_ref = "{}!{}".format(sheet, pc["pivot_name"])
                pivot_source = (
                    "<{p}pivotSource><{p}name>{n}</{p}name>"
                    "<{p}fmtId val=\"0\"/></{p}pivotSource>".format(p=prefix, n=_pivot._esc(name_ref))
                )
                # pivotSource 是 chartSpace 的子元素，须位于 chart 元素之前
                idx = m.start()
                xml = xml[:idx] + pivot_source + xml[idx:]
                result[part] = xml.encode("utf-8")
                used.add(part)
                break
        return result

    def save(self, filename: str, close: bool = True) -> None:
        """保存Excel文件。

        .. note::
            迷你图、数据透视表/透视图均在 openpyxl 保存后通过注入 xlsx 内部 XML 实现。
            追加模式（``mode='append'``）会基于单元格值/样式重建工作簿，不保留原文件中的
            迷你图、透视表、原生图与图片，请在最终输出步骤再添加这些内容。

        :param filename: 保存路径
        :param close: 是否关闭workbook，默认为True
        """
        # 移除样式模板sheet
        if self.style_sheet.title in self.workbook.sheetnames:
            self.workbook.remove(self.style_sheet)

        # 处理append模式
        if os.path.exists(filename) and self.mode == "append":
            _workbook = load_workbook(filename)

            for _sheet_name in _workbook.sheetnames:
                if _sheet_name not in self.workbook.sheetnames:
                    _worksheet = self.get_sheet_by_name(_sheet_name)

                    for i, row in enumerate(_workbook[_sheet_name].iter_rows()):
                        for j, cell in enumerate(row):
                            _worksheet.cell(row=i + 1, column=j + 1).value = cell.value
                            _worksheet.cell(row=i + 1, column=j + 1).style = cell.style

                            if i == _workbook[_sheet_name].max_row - 1:
                                _worksheet.column_dimensions[get_column_letter(j + 1)].width = _workbook[_sheet_name].column_dimensions[get_column_letter(j + 1)].width

            _workbook.close()

        # 创建目录
        if os.path.dirname(filename) != "" and not os.path.exists(os.path.dirname(filename)):
            os.makedirs(os.path.dirname(filename), exist_ok=True)

        # 保存文件
        self.workbook.save(filename)

        # 注入迷你图（openpyxl 不支持写入 sparkline，需在保存后修改 xlsx XML）
        self._inject_sparklines(filename)

        # 注入数据透视表/透视图（openpyxl 不支持创建透视表，需在保存后修改 xlsx XML）
        self._inject_pivots(filename)

        if close:
            self.workbook.close()


def resolve_condition_color(
    condition_color: Optional[Union[str, List[str], Dict[Any, Union[str, List[str]]]]],
    column: Any,
    default_color: str,
) -> str:
    """解析某一列（或行）对应的条件格式单色（数据条等单色场景使用）。

    :param condition_color: 条件格式颜色配置，支持以下类型：

        - str：所有列/行统一使用该颜色
        - list/tuple：由 2 或 3 个颜色组成的色阶锚点（仅 ``color_cols``/``color_rows`` 颜色渐变完整生效；
          数据条等单色场景取末位颜色）
        - dict：以列名（或行索引值）为key分别指定颜色，值可为 str 或上述 list/tuple，匹配方式类似 ``condition_cols``：

          - 单层列名：key直接为列名
          - 多层级列名：key可以是完整的列名tuple，也可以是其中任意一层级的名称（优先匹配完整tuple，其次按由内到外的层级匹配）
          - dict中未匹配到的列/行回退使用 ``default_color``

    :param column: 当前列名或行索引值，多层级时为tuple
    :param default_color: ``condition_color`` 为None，或为dict且未匹配到时使用的颜色
    :return: 最终使用的单个颜色（list/tuple 锚点取末位），去除 ``#`` 前缀并大写

    **参考样例**

    >>> resolve_condition_color('F76E6C', 'iv', '2639E9')
    'F76E6C'
    >>> resolve_condition_color({'iv': 'F76E6C', 'ks': '5B8FF9'}, 'iv', '2639E9')
    'F76E6C'
    >>> resolve_condition_color({'iv': 'F76E6C'}, 'ks', '2639E9')
    '2639E9'
    >>> resolve_condition_color({('分组1', 'iv'): 'F76E6C'}, ('分组1', 'iv'), '2639E9')
    'F76E6C'
    >>> resolve_condition_color({'iv': 'F76E6C'}, ('分组1', 'iv'), '2639E9')
    'F76E6C'
    >>> resolve_condition_color(['#3B82F6', '#E879F9', '#EF4444'], 'iv', '2639E9')
    'EF4444'
    """
    value = _resolve_condition_value(condition_color, column, default_color)
    if isinstance(value, (list, tuple)):
        value = value[-1] if len(value) else default_color
    return _clean_hex(value)


def _clean_hex(color: str) -> str:
    """规范化颜色为openpyxl可用的十六进制：去除 ``#`` 前缀并大写。"""
    return str(color).lstrip("#").upper()


def _resolve_condition_value(
    condition_color: Optional[Union[str, List[str], Dict[Any, Union[str, List[str]]]]],
    column: Any,
    default_color: str,
) -> Union[str, List[str]]:
    """解析某一列（或行）配置的原始颜色值（单色 str 或多色锚点 list/tuple），未匹配回退 ``default_color``。

    与 :func:`resolve_condition_color` 共用匹配规则，但不折叠为单色，供颜色渐变构建多色阶使用。
    """
    if not isinstance(condition_color, dict):
        return condition_color if condition_color else default_color

    if column in condition_color:
        return condition_color[column]

    if isinstance(column, tuple):
        for level_value in reversed(column):
            if level_value in condition_color:
                return condition_color[level_value]

    return default_color


def _coerce_color_anchors(color_value: Union[str, List[str]]) -> List[str]:
    """将颜色配置规范化为色阶锚点列表（1~3 个，去 ``#``、大写）。

    多色锚点超过 3 个时自动取首/中/尾 3 个（``ColorScaleRule`` 最多支持 3 个色标）。
    """
    if isinstance(color_value, (list, tuple)):
        anchors = [_clean_hex(c) for c in color_value if c]
    else:
        anchors = [_clean_hex(color_value)] if color_value else []
    if len(anchors) > 3:
        anchors = [anchors[0], anchors[len(anchors) // 2], anchors[-1]]
    return anchors


def _build_color_scale_rule(values: pd.Series, color_value: Union[str, List[str]]) -> ColorScaleRule:
    """根据锚点颜色构建颜色渐变规则 ``ColorScaleRule``，支持单/双/三色阶。

    :param values: 当前列（或行）数据，用于确定最小/最大值
    :param color_value: 颜色配置，单色 str 或由 2/3 个颜色组成的异色锚点 list/tuple

        - 单色：沿用「两端同色 + 白心（0 为基准）」的发散写法，保持向后兼容
        - 双色 ``[低值色, 高值色]``：最小值 → 最大值 双色阶
        - 三色 ``[低值色, 中值色, 高值色]``（异色）：最小值 → 中位数 → 最大值 三色阶
    :return: 颜色渐变规则
    """
    anchors = _coerce_color_anchors(color_value)
    vmin, vmax = values.min(), values.max()

    if len(anchors) >= 3:
        return ColorScaleRule(
            start_type='num', start_value=vmin, start_color=anchors[0],
            mid_type='percentile', mid_value=50, mid_color=anchors[1],
            end_type='num', end_value=vmax, end_color=anchors[2],
        )
    if len(anchors) == 2:
        return ColorScaleRule(
            start_type='num', start_value=vmin, start_color=anchors[0],
            end_type='num', end_value=vmax, end_color=anchors[1],
        )
    _color = anchors[0]
    return ColorScaleRule(
        start_type='num', start_value=vmin, start_color=_color,
        mid_type='num', mid_value=0., mid_color='FFFFFF',
        end_type='num', end_value=vmax, end_color=_color,
    )


def dataframe2excel(
    data: pd.DataFrame,
    excel_writer: Union[str, ExcelWriter],
    sheet_name: Optional[str] = None,
    title: Optional[str] = None,
    header: bool = True,
    theme_color: str = "2639E9",
    condition_color: Optional[Union[str, List[str], Dict[Any, Union[str, List[str]]]]] = None,
    fill: bool = True,
    percent_cols: Optional[List] = None,
    condition_cols: Optional[List] = None,
    custom_cols: Optional[List] = None,
    custom_format: str = "#,##0",
    color_cols: Optional[List] = None,
    percent_rows: Optional[List] = None,
    condition_rows: Optional[List] = None,
    custom_rows: Optional[List] = None,
    color_rows: Optional[List] = None,
    left_cols: Optional[List] = None,
    right_cols: Optional[List] = None,
    start_col: int = 2,
    start_row: int = 2,
    mode: str = "replace",
    figures: Optional[Union[str, List[str]]] = None,
    figsize: Tuple[int, int] = (600, 350),
    image_bottom_padding_rows: int = 1,
    writer_params: Optional[Dict] = None,
    auto_filter: bool = False,
    decimal: Optional[int] = 4,
    speed: str = "auto",
    **kwargs
) -> Tuple[int, int]:
    """快速将DataFrame写入Excel。

    这是一个便捷函数，封装了ExcelWriter的常用操作。

    :param data: 需要保存的DataFrame
    :param excel_writer: 文件路径或ExcelWriter对象
    :param sheet_name: 工作表名称，默认为None
    :param title: 标题，默认为None
    :param header: 是否保存列名，默认为True
    :param theme_color: 主题颜色，默认为"2639E9"
    :param condition_color: 条件格式（数据条/颜色渐变）颜色，默认为None（使用ExcelWriter的condition_color，
        其默认值为副主题色"F76E6C"）。单次传入值优先于ExcelWriter配置。支持三种类型：

        - str：所有 condition_cols/color_cols/condition_rows/color_rows 统一使用该颜色
        - list/tuple：2 或 3 个颜色组成的异色色阶锚点，仅对 color_cols/color_rows 颜色渐变生效（2 色为
          ``[低值色, 高值色]`` 双色阶，3 色为 ``[低值色, 中值色, 高值色]`` 三色阶；超过 3 个自动取首/中/尾；
          数据条等单色场景取末位颜色）
        - dict：以列名（或行索引值）为key分别指定颜色，值可为 str 或上述 list/tuple，匹配方式类似 ``condition_cols`` —— 单层为列名，多层级可为完整列名tuple或其中任意层级名称；未匹配到的回退使用ExcelWriter的condition_color
    :param fill: 是否使用颜色填充，默认为True
    :param percent_cols: 需要显示为百分数的列，默认为None
    :param condition_cols: 需要显示数据条的列，默认为None
    :param custom_cols: 需要自定义格式的列，默认为None
    :param custom_format: 自定义格式，默认为"#,##0"
    :param color_cols: 需要显示颜色渐变的列，默认为None
    :param percent_rows: 需要显示为百分数的行，默认为None
    :param condition_rows: 需要显示数据条的行，默认为None
    :param custom_rows: 需要自定义格式的行，默认为None
    :param color_rows: 需要显示颜色渐变的行，默认为None
    :param left_cols: 需要左对齐的列名或列索引列表，默认为None（数据行，非表头）
    :param right_cols: 需要右对齐的列名或列索引列表，默认为None（数据行，非表头）
    :param start_col: 起始列，默认为2
    :param start_row: 起始行，默认为2
    :param mode: 写入模式，默认为"replace"
    :param figures: 需要插入的图片路径，默认为None
    :param figsize: 图片大小，默认为(600, 350)
    :param image_bottom_padding_rows: 图片区与下方表格之间的额外空行数，默认为1
    :param writer_params: ExcelWriter参数，默认为None
    :param decimal: 浮点值保留小数位数，默认4；None表示不主动舍入
    :param speed: 写入速度，可选``"auto"``、``"normal"``或``"fast"``，默认``"auto"``；
        自动模式在行数不少于500、有效列数不少于50或有效单元格数不少于10000时使用快速路径
    :param kwargs: 其他参数，传递给insert_df2sheet
    :return: (下一行行号, 下一列列号)

    **参考样例**

    >>> import pandas as pd
    >>> from hscredit.excel import dataframe2excel
    >>>
    >>> # 创建示例数据
    >>> df = pd.DataFrame({
    ...     'feature': ['A', 'B', 'C'],
    ...     'iv': [0.1, 0.2, 0.3],
    ...     'ks': [0.3, 0.4, 0.5],
    ...     'rate': [0.05, 0.10, 0.15]
    ... })
    >>>
    >>> # 快速写入Excel
    >>> dataframe2excel(
    ...     df,
    ...     "report.xlsx",
    ...     sheet_name="特征分析",
    ...     title="特征统计表",
    ...     percent_cols=['rate'],  # 百分比格式
    ...     condition_cols=['iv', 'ks'],  # 条件格式
    ...     auto_width=True
    ... )
    """
    writer_params = writer_params or {}
    ExcelWriter._validate_decimal(decimal)

    if isinstance(excel_writer, ExcelWriter):
        writer = excel_writer
    else:
        writer = ExcelWriter(theme_color=theme_color, mode=mode, **writer_params)
    default_condition_color = writer.condition_color

    if isinstance(sheet_name, Worksheet):
        worksheet = sheet_name
    else:
        worksheet = writer.get_sheet_by_name(sheet_name or "Sheet1")

    image_bottom_padding_rows = 0 if image_bottom_padding_rows is None else max(int(image_bottom_padding_rows), 0)

    # 插入标题
    if title:
        col_width = len(data.columns) + data.index.nlevels if kwargs.get("index", False) else len(data.columns)
        start_row, end_col = writer.insert_value2sheet(
            worksheet, (start_row, start_col),
            value=title,
            style="header",
            end_space=(start_row, start_col + col_width - 1)
        )
        start_row += 1

    # 插入图片
    if figures is not None:
        if isinstance(figures, str):
            figures = [figures]

        figures = [pic for pic in figures if pic]

        if figures:
            pic_row = start_row
            for i, pic in enumerate(figures):
                if i == 0:
                    start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, start_col), figsize=figsize)
                else:
                    start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, end_col - 1), figsize=figsize)

            start_row += image_bottom_padding_rows

    # 处理merge_column参数
    if "merge_column" in kwargs and kwargs["merge_column"]:
        if not isinstance(kwargs["merge_column"][0], (tuple, list)):
            kwargs["merge_column"] = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in kwargs["merge_column"]) or (not isinstance(c, tuple) and c in kwargs["merge_column"])]

    # 插入DataFrame
    end_row, end_col = writer.insert_df2sheet(
        worksheet, data, (start_row, start_col),
        fill=fill, header=header, decimal=decimal, speed=speed, **kwargs
    )

    # 设置百分比格式列
    if percent_cols and len(data) > 0:
        if not isinstance(percent_cols[0], (tuple, list)):
            percent_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in percent_cols) or (not isinstance(c, tuple) and c in percent_cols)]
        for c in [c for c in percent_cols if c in data.columns]:
            conditional_column = get_column_letter(
                start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c)
            )
            writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", "0.00%")

    # 设置自定义格式列
    if custom_cols and len(data) > 0:
        if not isinstance(custom_cols[0], (tuple, list)):
            custom_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in custom_cols) or (not isinstance(c, tuple) and c in custom_cols)]
        for c in [c for c in custom_cols if c in data.columns]:
            conditional_column = get_column_letter(
                start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c)
            )
            writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", custom_format)

    # 设置条件格式列
    if condition_cols and len(data) > 0:
        if not isinstance(condition_cols[0], (tuple, list)):
            condition_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in condition_cols) or (not isinstance(c, tuple) and c in condition_cols)]
        for c in [c for c in condition_cols if c in data.columns]:
            conditional_column = get_column_letter(
                start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c)
            )
            writer.add_conditional_formatting(
                worksheet,
                f'{conditional_column}{end_row - len(data)}',
                f'{conditional_column}{end_row - 1}',
                condition_color=resolve_condition_color(condition_color, c, default_condition_color)
            )

    # 设置颜色渐变列
    if color_cols and len(data) > 0:
        if not isinstance(color_cols[0], (tuple, list)):
            color_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in color_cols) or (not isinstance(c, tuple) and c in color_cols)]
        for c in [c for c in color_cols if c in data.columns]:
            try:
                rule = _build_color_scale_rule(
                    data[c],
                    _resolve_condition_value(condition_color, c, default_condition_color),
                )
                conditional_column = get_column_letter(
                    start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c)
                )
                worksheet.conditional_formatting.add(f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", rule)
            except Exception:
                import traceback
                traceback.print_exc()

    # 设置百分比格式行
    if percent_rows:
        if not isinstance(percent_rows[0], (tuple, list)):
            percent_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in percent_rows) or (not isinstance(c, tuple) and c in percent_rows)]
        for c in [c for c in percent_rows if c in data.index]:
            insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
            index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
            index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
            writer.set_number_format(worksheet, f"{get_column_letter(index_col)}{index_row}:{get_column_letter(index_col + len(data.columns) - 1)}{index_row}", "0.00%")

    # 设置自定义格式行
    if custom_rows:
        if not isinstance(custom_rows[0], (tuple, list)):
            custom_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in custom_rows) or (not isinstance(c, tuple) and c in custom_rows)]
        for c in [c for c in custom_rows if c in data.index]:
            insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
            index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
            index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
            writer.set_number_format(worksheet, f"{get_column_letter(index_col)}{index_row}:{get_column_letter(index_col + len(data.columns) - 1)}{index_row}", custom_format)

    # 设置条件格式行
    if condition_rows:
        if not isinstance(condition_rows[0], (tuple, list)):
            condition_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in condition_rows) or (not isinstance(c, tuple) and c in condition_rows)]
        for c in [c for c in condition_rows if c in data.index]:
            insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
            index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
            index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
            writer.add_conditional_formatting(
                worksheet,
                f'{get_column_letter(index_col)}{index_row}',
                f'{get_column_letter(index_col + len(data.columns) - 1)}{index_row}',
                condition_color=resolve_condition_color(condition_color, c, default_condition_color)
            )

    # 设置颜色渐变行
    if color_rows:
        if not isinstance(color_rows[0], (tuple, list)):
            color_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in color_rows) or (not isinstance(c, tuple) and c in color_rows)]
        for c in [c for c in color_rows if c in data.index]:
            try:
                insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
                rule = _build_color_scale_rule(
                    data.loc[c],
                    _resolve_condition_value(condition_color, c, default_condition_color),
                )
                index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
                index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
                worksheet.conditional_formatting.add(f"{get_column_letter(index_col)}{index_row}:{get_column_letter(index_col + len(data.columns) - 1)}{index_row}", rule)
            except Exception:
                import traceback
                traceback.print_exc()

    # 应用自定义列对齐（仅数据行，非表头）
    if left_cols or right_cols:
        from openpyxl.styles import Alignment

        # 计算表头行数（1行或 MultiIndex 层数）
        n_header_rows = data.columns.nlevels if header else 0
        data_start_row = start_row + n_header_rows
        data_end_row = end_row - 1

        # index 列的层数
        idx_levels = data.index.nlevels if kwargs.get("index", False) else 0

        # 解析 left_cols / right_cols → DataFrame 列索引集合
        def _resolve_col_items(items, df_cols):
            result = set()
            if not items:
                return result
            for c in items:
                if isinstance(c, int):
                    if 0 <= c < len(df_cols):
                        result.add(c)
                elif isinstance(c, str):
                    try:
                        loc = df_cols.get_loc(c)
                        if isinstance(loc, int):
                            result.add(loc)
                        else:
                            result.update(range(loc.start, loc.stop))
                    except Exception:
                        pass
            return result

        left_idx_set = _resolve_col_items(left_cols, data.columns)
        right_idx_set = _resolve_col_items(right_cols, data.columns)
        alignments = {
            "left": Alignment(horizontal="left", vertical="center"),
            "right": Alignment(horizontal="right", vertical="center"),
        }

        for col_idx in (left_idx_set | right_idx_set):
            horiz = "left" if col_idx in left_idx_set else "right"
            excel_col = start_col + col_idx + idx_levels
            for row in range(data_start_row, data_end_row + 1):
                worksheet.cell(row=row, column=excel_col).alignment = alignments[horiz]

    # 添加自动筛选（必须在保存之前，否则保存并关闭 workbook 后筛选不会写入文件）
    if auto_filter:
        last_data_row = end_row - 1
        last_data_col = end_col - 1
        writer.add_auto_filter(
            worksheet,
            f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_data_col)}{last_data_row}"
        )

    # 保存文件（如果不是传入的ExcelWriter对象）
    if not isinstance(excel_writer, ExcelWriter) and not isinstance(sheet_name, Worksheet):
        writer.save(excel_writer)

    return end_row, end_col
