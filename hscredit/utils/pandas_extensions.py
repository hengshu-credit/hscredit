from __future__ import annotations

"""Pandas DataFrame/Series 扩展模块.

为 pandas DataFrame 和 Series 提供额外的方法，包括：
- df.summary(): 综合特征描述统计
- df.eda_info(): EDA基础信息
- df.missing_analysis(): 缺失值分析
- df.show(): 美化展示分箱表
- df.save(): 保存到Excel

使用方式:
    >>> import pandas as pd
    >>> import hscredit  # 自动注册所有扩展
    >>> 
    >>> df = pd.DataFrame({...})
    >>> 
    >>> # 数据摘要
    >>> summary = df.summary()
    >>> 
    >>> # 保存到Excel
    >>> df.save("report.xlsx", sheet_name="数据", title="统计表")
    >>> 
    >>> # 美化展示分箱表
    >>> table.show(compact=True)
"""

import pandas as pd
import numpy as np
from typing import Optional, List, Dict, Any, Literal, Union, Tuple, TYPE_CHECKING

if TYPE_CHECKING:
    from ..excel import ExcelWriter


# =============================================================================
# 1. EDA相关扩展方法 (原pandas_ext.py)
# =============================================================================

def _summary_method(
    self,
    features: List[str] = None,
    y: Optional[Union[str, pd.Series]] = None,
    val_df: Optional[pd.DataFrame] = None,
    models: Optional[Dict[str, Any]] = None,
    model_type: Optional[Literal['xgboost', 'lightgbm', 'catboost', 'randomforest']] = None,
    model_params: Optional[Dict] = None,
    max_n_bins: int = 5,
    psi_method: Literal['random_split', 'group_col', 'date_col'] = 'random_split',
    psi_group_col: Optional[str] = None,
    psi_date_col: Optional[str] = None,
    psi_freq: str = 'M',
    psi_test_size: float = 0.3,
    percentiles: List[float] = None,
    random_state: int = 42,
    return_type: Literal['dataframe', 'dict'] = 'dataframe'
) -> Union[pd.DataFrame, Dict]:
    """DataFrame 数据分布摘要统计.
    
    快速获取数据集特征详情，包括基础统计、IV、KS、趋势、PSI和特征重要性。
    
    :param features: 特征列表，None则分析全部
    :param y: 目标变量，支持列名(str)或Series，不传则不计算IV/KS/趋势/特征重要性
    :param val_df: 验证集，用于计算PSI
    :param models: 已训练好的模型字典，用于获取特征重要性
    :param model_type: 模型类型，用于自动训练模型提取特征重要性
    :param model_params: 模型参数
    :param max_n_bins: IV计算分箱数，默认5
    :param psi_method: PSI计算方式
    :param psi_group_col: 分组列名（当psi_method='group_col'时使用）
    :param psi_date_col: 日期列名（当psi_method='date_col'时使用）
    :param psi_freq: 时间频率
    :param psi_test_size: 随机拆分比例
    :param percentiles: 分位数点，默认[0.01, 0.05, 0.25, 0.5, 0.75, 0.95, 0.99]
    :param random_state: 随机种子
    :param return_type: 返回类型，'dataframe' 或 'dict'
    :return: 综合特征描述DataFrame或字典
    
    Example:
        >>> # 基础统计
        >>> summary = df.summary()
        
        >>> # 包含IV、KS、趋势（传入目标变量）
        >>> summary = df.summary(y='target')
        
        >>> # 按日期分组计算PSI
        >>> summary = df.summary(y='target', psi_method='date_col', psi_date_col='apply_date')
    """
    from ..core.eda import feature_summary
    
    result = feature_summary(
        df=self,
        features=features,
        y=y,
        val_df=val_df,
        models=models,
        model_type=model_type,
        model_params=model_params,
        max_n_bins=max_n_bins,
        psi_method=psi_method,
        psi_group_col=psi_group_col,
        psi_date_col=psi_date_col,
        psi_freq=psi_freq,
        psi_test_size=psi_test_size,
        percentiles=percentiles,
        random_state=random_state
    )
    
    if return_type == 'dict':
        return result.to_dict(orient='records')
    return result


def _eda_info_method(self) -> Dict[str, Any]:
    """DataFrame EDA 基础信息.
    
    快速获取数据集基础信息，包括样本数、特征数、缺失值等。
    
    :return: 字典格式的数据集信息
    
    Example:
        >>> info = df.eda_info()
        >>> print(info['样本数'])
    """
    from ..core.eda import data_info
    
    result = data_info(self)
    return dict(zip(result['信息项'], result['值']))


def _missing_analysis_method(self, threshold: float = 0.0) -> pd.DataFrame:
    """DataFrame 缺失值分析.
    
    :param threshold: 缺失率阈值，仅返回缺失率>=该值的特征
    :return: 缺失值分析DataFrame
    
    Example:
        >>> missing = df.missing_analysis(threshold=0.05)
    """
    from ..core.eda import missing_analysis
    return missing_analysis(self, threshold=threshold)


# =============================================================================
# 2. Excel保存相关扩展方法 (原report/excel/pandas_extension.py)
# =============================================================================

def _dataframe_save(
    self,
    excel_writer: Union[str, ExcelWriter],
    worksheet: Optional[Any] = None,
    sheet_name: Optional[str] = None,
    title: Optional[str] = None,
    header: bool = True,
    theme_color: str = "2639E9",
    condition_color: Optional[str] = None,
    fill: bool = True,
    percent_cols: Optional[List] = None,
    condition_cols: Optional[List] = None,
    custom_cols: Optional[List] = None,
    custom_format: str = "#,##0",
    color_cols: Optional[List] = None,
    percent_rows: Optional[List] = None,
    condition_rows: Optional[List] = None,
    custom_rows: Optional[List] = None,
    color_rows: Optional[List] = None,
    start_col: int = 2,
    start_row: int = 2,
    mode: str = "replace",
    figures: Optional[Union[str, List[str]]] = None,
    figsize: Tuple[int, int] = (600, 350),
    image_bottom_padding_rows: int = 1,
    writer_params: Optional[Dict] = None,
    **kwargs
) -> Union[Tuple[int, int], ExcelWriter]:
    """
    将DataFrame保存到Excel文件或已有的ExcelWriter中。

    :param excel_writer: 文件路径或ExcelWriter对象
    :param worksheet: 工作表对象（如果提供，将写入该worksheet而不保存文件）
    :param sheet_name: 工作表名称，默认为None
    :param title: 标题，默认为None
    :param header: 是否保存列名，默认为True
    :param theme_color: 主题颜色，默认为"2639E9"
    :param condition_color: 条件格式颜色，默认为None
    :param fill: 是否使用颜色填充，默认为True
    :param percent_cols: 需要显示为百分数的列，默认为None
    :param condition_cols: 需要显示数据条的列，默认为None
    :param custom_cols: 需要自定义格式的列，默认为None
    :param custom_format: 自定义格式，默认为"#,##0"
    :param color_cols: 需要显示颜色渐变的列，默认为None
    :param percent_rows: 需要显示为百分数的行，默认为None
    :param condition_rows: 需要显示数据条的行，默认为None
    :param custom_rows: 需要自定义格式的行，默认为None
    :param color_rows: 需要显示颜色渐变的行，默认为None
    :param start_col: 起始列，默认为2
    :param start_row: 起始行，默认为2
    :param mode: 写入模式，默认为"replace"
    :param figures: 需要插入的图片路径，默认为None
    :param figsize: 图片大小，默认为(600, 350)
    :param image_bottom_padding_rows: 图片区与下方表格之间的额外空行数，默认为1
    :param writer_params: ExcelWriter参数，默认为None
    :param kwargs: 其他参数，传递给insert_df2sheet
    :return: 如果传入文件路径返回(end_row, end_col)；如果传入ExcelWriter且提供了worksheet，返回ExcelWriter

    **参考样例**

    >>> import pandas as pd
    >>> from hscredit.excel import ExcelWriter
    >>> 
    >>> df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    >>> 
    >>> # 方式1：直接保存到文件
    >>> df.save("report.xlsx", sheet_name="数据", title="统计表")
    >>> 
    >>> # 方式2：写入已有的ExcelWriter
    >>> writer = ExcelWriter()
    >>> worksheet = writer.get_sheet_by_name("Sheet1")
    >>> df.save(writer, worksheet=worksheet)
    >>> writer.save("report.xlsx")
    """
    from ..excel import ExcelWriter, dataframe2excel
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils import get_column_letter
    
    # 如果提供了worksheet，说明要使用已有的writer
    if worksheet is not None and isinstance(excel_writer, ExcelWriter):
        # 直接插入到指定的worksheet
        writer = excel_writer
        
        # 插入标题
        if title:
            col_width = len(self.columns) + self.index.nlevels if kwargs.get("index", False) else len(self.columns)
            _start_row, _end_col = writer.insert_value2sheet(
                worksheet, (start_row, start_col),
                value=title,
                style="header",
                end_space=(start_row, start_col + col_width - 1)
            )
            start_row += 1
        
        # 插入图片
        if figures is not None:
            if isinstance(figures, str):
                figures = [figures]
            figures = [pic for pic in figures if pic]
            if figures:
                pic_row = start_row
                for i, pic in enumerate(figures):
                    if i == 0:
                        start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, start_col), figsize=figsize)
                    else:
                        start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, end_col - 1), figsize=figsize)

                start_row += 0 if image_bottom_padding_rows is None else max(int(image_bottom_padding_rows), 0)
        
        # 处理merge_column参数
        if "merge_column" in kwargs and kwargs["merge_column"]:
            if not isinstance(kwargs["merge_column"][0], (tuple, list)):
                kwargs["merge_column"] = [c for c in self.columns if (isinstance(c, tuple) and c[-1] in kwargs["merge_column"]) or (not isinstance(c, tuple) and c in kwargs["merge_column"])]
        
        # 插入DataFrame
        end_row, end_col = writer.insert_df2sheet(
            worksheet, self, (start_row, start_col),
            fill=fill, header=header, **kwargs
        )
        
        # 设置百分比格式列
        if percent_cols:
            if not isinstance(percent_cols[0], (tuple, list)):
                percent_cols = [c for c in self.columns if (isinstance(c, tuple) and c[-1] in percent_cols) or (not isinstance(c, tuple) and c in percent_cols)]
            for c in [c for c in percent_cols if c in self.columns]:
                conditional_column = get_column_letter(
                    start_col + self.columns.get_loc(c) + self.index.nlevels if kwargs.get("index", False) else start_col + self.columns.get_loc(c)
                )
                writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(self)}:{conditional_column}{end_row - 1}", "0.00%")
        
        # 设置自定义格式列
        if custom_cols:
            if not isinstance(custom_cols[0], (tuple, list)):
                custom_cols = [c for c in self.columns if (isinstance(c, tuple) and c[-1] in custom_cols) or (not isinstance(c, tuple) and c in custom_cols)]
            for c in [c for c in custom_cols if c in self.columns]:
                conditional_column = get_column_letter(
                    start_col + self.columns.get_loc(c) + self.index.nlevels if kwargs.get("index", False) else start_col + self.columns.get_loc(c)
                )
                writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(self)}:{conditional_column}{end_row - 1}", custom_format)
        
        # 设置条件格式列
        if condition_cols:
            if not isinstance(condition_cols[0], (tuple, list)):
                condition_cols = [c for c in self.columns if (isinstance(c, tuple) and c[-1] in condition_cols) or (not isinstance(c, tuple) and c in condition_cols)]
            for c in [c for c in condition_cols if c in self.columns]:
                conditional_column = get_column_letter(
                    start_col + self.columns.get_loc(c) + self.index.nlevels if kwargs.get("index", False) else start_col + self.columns.get_loc(c)
                )
                writer.add_conditional_formatting(
                    worksheet,
                    f'{conditional_column}{end_row - len(self)}',
                    f'{conditional_column}{end_row - 1}',
                    condition_color=condition_color or theme_color
                )
        
        return writer
    
    # 使用dataframe2excel函数（传入文件路径或ExcelWriter但没有worksheet）
    return dataframe2excel(
        data=self,
        excel_writer=excel_writer,
        sheet_name=sheet_name,
        title=title,
        header=header,
        theme_color=theme_color,
        condition_color=condition_color,
        fill=fill,
        percent_cols=percent_cols,
        condition_cols=condition_cols,
        custom_cols=custom_cols,
        custom_format=custom_format,
        color_cols=color_cols,
        percent_rows=percent_rows,
        condition_rows=condition_rows,
        custom_rows=custom_rows,
        color_rows=color_rows,
        start_col=start_col,
        start_row=start_row,
        mode=mode,
        figures=figures,
        figsize=figsize,
        image_bottom_padding_rows=image_bottom_padding_rows,
        writer_params=writer_params,
        **kwargs
    )


def _series_save(
    self,
    excel_writer: Union[str, ExcelWriter],
    worksheet: Optional[Any] = None,
    sheet_name: Optional[str] = None,
    title: Optional[str] = None,
    header: bool = True,
    theme_color: str = "2639E9",
    condition_color: Optional[str] = None,
    fill: bool = True,
    percent_cols: Optional[List] = None,
    condition_cols: Optional[List] = None,
    custom_cols: Optional[List] = None,
    custom_format: str = "#,##0",
    color_cols: Optional[List] = None,
    percent_rows: Optional[List] = None,
    condition_rows: Optional[List] = None,
    custom_rows: Optional[List] = None,
    color_rows: Optional[List] = None,
    start_col: int = 2,
    start_row: int = 2,
    mode: str = "replace",
    figures: Optional[Union[str, List[str]]] = None,
    figsize: Tuple[int, int] = (600, 350),
    image_bottom_padding_rows: int = 1,
    writer_params: Optional[Dict] = None,
    **kwargs
) -> Union[Tuple[int, int], ExcelWriter]:
    """
    将Series保存到Excel文件或已有的ExcelWriter中。
    Series会被转换为单列DataFrame（保留name作为列名）。

    :param excel_writer: 文件路径或ExcelWriter对象
    :param worksheet: 工作表对象（如果提供，将写入该worksheet而不保存文件）
    :param sheet_name: 工作表名称，默认为None
    :param title: 标题，默认为None
    :param header: 是否保存列名，默认为True
    :param theme_color: 主题颜色，默认为"2639E9"
    :param condition_color: 条件格式颜色，默认为None
    :param fill: 是否使用颜色填充，默认为True
    :param percent_cols: 需要显示为百分数的列，默认为None
    :param condition_cols: 需要显示数据条的列，默认为None
    :param custom_cols: 需要自定义格式的列，默认为None
    :param custom_format: 自定义格式，默认为"#,##0"
    :param color_cols: 需要显示颜色渐变的列，默认为None
    :param percent_rows: 需要显示为百分数的行，默认为None
    :param condition_rows: 需要显示数据条的行，默认为None
    :param custom_rows: 需要自定义格式的行，默认为None
    :param color_rows: 需要显示颜色渐变的行，默认为None
    :param start_col: 起始列，默认为2
    :param start_row: 起始行，默认为2
    :param mode: 写入模式，默认为"replace"
    :param figures: 需要插入的图片路径，默认为None
    :param figsize: 图片大小，默认为(600, 350)
    :param image_bottom_padding_rows: 图片区与下方表格之间的额外空行数，默认为1
    :param writer_params: ExcelWriter参数，默认为None
    :param kwargs: 其他参数，传递给insert_df2sheet
    :return: 如果传入文件路径返回(end_row, end_col)；如果传入ExcelWriter且提供了worksheet，返回ExcelWriter

    **参考样例**

    >>> import pandas as pd
    >>> from hscredit.excel import ExcelWriter
    >>> 
    >>> s = pd.Series([1, 2, 3], name='数值', index=['a', 'b', 'c'])
    >>> 
    >>> # 方式1：直接保存到文件
    >>> s.save("report.xlsx", sheet_name="数据", title="序列数据")
    >>> 
    >>> # 方式2：写入已有的ExcelWriter
    >>> writer = ExcelWriter()
    >>> worksheet = writer.get_sheet_by_name("Sheet1")
    >>> s.save(writer, worksheet=worksheet)
    >>> writer.save("report.xlsx")
    """
    # 将Series转换为DataFrame
    df = self.to_frame()
    
    # 调用DataFrame的save方法
    return df.save(
        excel_writer=excel_writer,
        worksheet=worksheet,
        sheet_name=sheet_name,
        title=title,
        header=header,
        theme_color=theme_color,
        condition_color=condition_color,
        fill=fill,
        percent_cols=percent_cols,
        condition_cols=condition_cols,
        custom_cols=custom_cols,
        custom_format=custom_format,
        color_cols=color_cols,
        percent_rows=percent_rows,
        condition_rows=condition_rows,
        custom_rows=custom_rows,
        color_rows=color_rows,
        start_col=start_col,
        start_row=start_row,
        mode=mode,
        figures=figures,
        figsize=figsize,
        image_bottom_padding_rows=image_bottom_padding_rows,
        writer_params=writer_params,
        **kwargs
    )


# =============================================================================
# 3. 分箱表展示相关扩展方法 (原bin_table_display.py)
# =============================================================================

class BinTableDisplay:
    """分箱表展示器.
    
    提供链式调用接口，方便在 Jupyter 中展示美观的分箱表。
    
    **参考样例**
    
    >>> table = feature_bin_stats(data, 'score', target='target')
    >>> table.show()
    >>> table.show(compact=True)
    >>> table.show(highlight_iv=False)
    """
    
    def __init__(self, df: pd.DataFrame):
        """初始化展示器.
        
        :param df: 分箱统计表
        """
        self._df = df
        self._styler = None
    
    def show(
        self,
        max_rows: Optional[int] = None,
        highlight_iv: bool = True,
        highlight_bad_rate: bool = True,
        highlight_lift: bool = True,
        highlight_ks: bool = True,
        compact: bool = False,
        precision: Optional[Dict[str, int]] = None,
        index_as_bin: bool = False,
        percent_format: bool = True,
        high_tech_style: bool = False,
        **kwargs
    ) -> 'BinTableDisplay':
        """展示美化的分箱表.
        
        :param max_rows: 最大显示行数
        :param highlight_iv: 是否高亮 IV 值
        :param highlight_bad_rate: 是否高亮坏样本率
        :param highlight_lift: 是否高亮 LIFT 值
        :param highlight_ks: 是否高亮 KS 值
        :param compact: 是否使用紧凑模式
        :param precision: 自定义小数位数
        :param index_as_bin: 是否将分箱作为索引显示
        :param percent_format: 是否将百分比相关列显示为百分比格式（默认True）
        :param high_tech_style: 是否使用高科技/AI风格样式
        :param kwargs: 其他参数
        :return: self，支持链式调用
        """
        self._styler = _style_bin_table(
            self._df,
            max_rows=max_rows,
            highlight_iv=highlight_iv,
            highlight_bad_rate=highlight_bad_rate,
            highlight_lift=highlight_lift,
            highlight_ks=highlight_ks,
            compact=compact,
            precision=precision,
            index_as_bin=index_as_bin,
            percent_format=percent_format,
            high_tech_style=high_tech_style
        )
        try:
            from IPython.display import display
            display(self._styler)
        except ImportError:
            pass
        return self
    
    def highlight_bins(self, bins: Union[int, List[int]], color: str = '#e3f2fd') -> 'BinTableDisplay':
        """高亮指定的分箱行.
        
        :param bins: 要高亮的分箱索引或索引列表
        :param color: 高亮颜色
        :return: self，支持链式调用
        """
        if self._styler is None:
            self._styler = _style_bin_table(self._df)
        
        if isinstance(bins, int):
            bins = [bins]
        
        def highlight_row(row):
            if row.name in bins:
                return ['background-color: #e3f2fd; color: #1565C0'] * len(row)
            return [''] * len(row)
        
        self._styler = self._styler.apply(highlight_row, axis=1)
        try:
            from IPython.display import display
            display(self._styler)
        except ImportError:
            pass
        return self
    
    def export_html(self, filename: str) -> 'BinTableDisplay':
        """导出为 HTML 文件.
        
        :param filename: 文件名
        :return: self，支持链式调用
        """
        if self._styler is None:
            self._styler = _style_bin_table(self._df)
        
        html = self._styler.to_html()
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(f'''
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <title>分箱统计表</title>
                <style>
                    .dataframe {{ overflow-x: auto; }}
                    .dataframe th {{ white-space: nowrap; }}
                    .dataframe td {{ white-space: nowrap; }}
                </style>
            </head>
            <body>
                {html}
            </body>
            </html>
            ''')
        print(f"已导出到: {filename}")
        return self
    
    def to_excel(self, filename: str, sheet_name: str = '分箱统计') -> 'BinTableDisplay':
        """导出为 Excel 文件.
        
        :param filename: 文件名
        :param sheet_name: 工作表名称
        :return: self，支持链式调用
        """
        if self._styler is None:
            self._styler = _style_bin_table(self._df)
        
        self._styler.to_excel(filename, sheet_name=sheet_name, engine='openpyxl')
        print(f"已导出到: {filename}")
        return self


def _style_bin_table(
    df: pd.DataFrame,
    max_rows: Optional[int] = None,
    highlight_iv: bool = True,
    highlight_bad_rate: bool = True,
    highlight_lift: bool = True,
    highlight_ks: bool = True,
    compact: bool = False,
    precision: Optional[Dict[str, int]] = None,
    index_as_bin: bool = False,
    percent_format: bool = True,
    high_tech_style: bool = False,
) -> Any:
    """美化分箱表展示.

    使用 pandas Styler 对分箱表进行格式化和高亮，使其在 Jupyter 中更易读。

    :param df: 分箱统计表 DataFrame
    :param max_rows: 最大显示行数，None 表示显示全部
    :param highlight_iv: 是否高亮 IV 值
    :param highlight_bad_rate: 是否高亮坏样本率（进度条）
    :param highlight_lift: 是否高亮 LIFT 值（进度条）
    :param highlight_ks: 是否高亮 KS 值（进度条）
    :param compact: 是否使用紧凑模式（隐藏部分列）
    :param precision: 自定义小数位数，格式为 {'列名': 位数}
    :param index_as_bin: 是否将分箱作为索引显示
    :param percent_format: 是否将百分比相关列显示为百分比格式（默认True）
    :param high_tech_style: 是否使用高科技/AI风格样式
    :return: 格式化后的 Styler 对象
    """
    # 检查是否为多级表头
    is_multi_level = isinstance(df.columns, pd.MultiIndex)

    # 创建副本避免修改原数据
    df_display = df.copy()

    # index_as_bin: 将分箱列设置为索引，并合并指标名称/含义列
    bin_col = None
    if index_as_bin:
        # 查找分箱标签列
        bin_label_col = None
        indicator_name_col = None
        indicator_desc_col = None
        
        for col in df_display.columns:
            col_name = col[1] if isinstance(col, tuple) else col
            if col_name == '分箱标签':
                bin_label_col = col
            if col_name == '指标名称':
                indicator_name_col = col
            if col_name == '指标含义':
                indicator_desc_col = col
        
        # 先合并指标名称和指标含义
        if indicator_name_col is not None and indicator_desc_col is not None:
            # 合并为新列
            if is_multi_level:
                df_display['指标名称_合并'] = df_display[indicator_name_col].astype(str) + ' - ' + df_display[indicator_desc_col].astype(str)
                df_display = df_display.drop(columns=[indicator_name_col, indicator_desc_col])
                # 重建多级索引列名
                new_columns = []
                for col in df_display.columns:
                    if col == '指标名称_合并':
                        new_columns.append(('基本信息', '指标'))
                    elif isinstance(col, tuple):
                        new_columns.append(col)
                    else:
                        new_columns.append(col)
                df_display.columns = pd.MultiIndex.from_tuples(new_columns)
                is_multi_level = isinstance(df_display.columns, pd.MultiIndex)
        
        if bin_label_col is not None:
            # 将分箱标签设为索引名称
            df_display = df_display.set_index(bin_label_col)
            # 如果有分箱列，也删除它
            if bin_col is not None:
                try:
                    df_display = df_display.drop(columns=[bin_col])
                except KeyError:
                    pass
        
        # 重新检查是否为多级表头（因为可能改变了列结构）
        is_multi_level = isinstance(df_display.columns, pd.MultiIndex)

    # 限制行数
    if max_rows is not None and len(df_display) > max_rows:
        df_display = df_display.head(max_rows)

    # 紧凑模式下隐藏部分列
    if compact:
        # 确定要隐藏的列
        hide_cols = []
        if is_multi_level:
            # 多级表头
            all_cols = df_display.columns.tolist()
            # 保留核心列
            keep_patterns = ['指标', '分箱标签', '样本总数', 
                           '坏样本率', '分档WOE值', '分档IV值', '指标IV值', 'LIFT值', '分档KS值']
            for col in all_cols:
                col_name = col[1] if isinstance(col, tuple) else col
                if not any(p in col_name for p in keep_patterns):
                    hide_cols.append(col)
        else:
            # 单层表头
            all_cols = df_display.columns.tolist()
            keep_patterns = ['指标名称', '指标含义', '指标', '分箱标签', '样本总数', 
                           '坏样本率', '分档WOE值', '分档IV值', '指标IV值', 'LIFT值', '分档KS值']
            hide_cols = [c for c in all_cols if not any(p in str(c) for p in keep_patterns)]

        if hide_cols:
            df_display = df_display.drop(columns=hide_cols)

    # 创建 Styler
    styler = df_display.style

    # 定义默认小数位数
    default_precision = {
        '样本总数': 0,
        '好样本数': 0,
        '坏样本数': 0,
        '样本占比': 2,
        '好样本占比': 2,
        '坏样本占比': 2,
        '坏样本率': 2,
        '分档WOE值': 4,
        '分档IV值': 4,
        '指标IV值': 4,
        'LIFT值': 2,
        '坏账改善': 2,
        '累积LIFT值': 2,
        '累积坏账改善': 2,
        '分档KS值': 2,
    }

    # 定义需要显示为百分比的列
    percent_columns = {
        '样本占比': 2,
        '好样本占比': 2,
        '坏样本占比': 2,
        '坏样本率': 2,
        'LIFT值': 2,
        '坏账改善': 2,
        '累积LIFT值': 2,
        '累积坏账改善': 2,
        '分档KS值': 2,
    }

    # 预定义的百分比格式
    percent_formats = {
        2: '{:.2%}',
        3: '{:.3%}',
        1: '{:.1%}',
        0: '{:.0f}',
    }

    # 更新自定义精度
    if precision:
        default_precision.update(precision)

    # 格式化数字
    if is_multi_level:
        # 多级表头
        format_dict = {}
        for col in df_display.columns:
            col_name = col[1]
            if percent_format and col_name in percent_columns:
                precision_val = percent_columns[col_name]
                format_dict[col] = percent_formats.get(precision_val, '{:.2%}')
            elif col_name in default_precision:
                precision_val = default_precision[col_name]
                if precision_val == 0:
                    format_dict[col] = '{:.0f}'
                else:
                    format_dict[col] = f'{{:.{precision_val}f}}'
        if format_dict:
            styler = styler.format(format_dict, na_rep='-')
    else:
        # 单层表头
        format_dict = {}
        for col in df_display.columns:
            col_name = str(col)
            if percent_format and col_name in percent_columns:
                precision_val = percent_columns[col_name]
                format_dict[col] = percent_formats.get(precision_val, '{:.2%}')
            elif col_name in default_precision:
                precision_val = default_precision[col_name]
                if precision_val == 0:
                    format_dict[col] = '{:.0f}'
                else:
                    format_dict[col] = f'{{:.{precision_val}f}}'
        if format_dict:
            styler = styler.format(format_dict, na_rep='-')

    # ===== 条件格式：只对坏样本率、LIFT、KS使用进度条 =====
    # 进度条颜色
    bad_rate_color = '#5B8FF9'   # 蓝色
    lift_color = '#5AD8A6'       # 绿色
    ks_color = '#F6BD16'         # 金色

    # 高亮坏样本率（进度条）
    if highlight_bad_rate:
        bad_rate_cols = _find_columns(df_display, '坏样本率')
        for col in bad_rate_cols:
            values = df_display[col].dropna()
            vmin = 0
            if len(values) > 0:
                vmax = max(values.quantile(0.95), 0.1)
                vmax = min(vmax * 1.2, 1)
            else:
                vmax = 1
            styler = styler.bar(
                subset=[col],
                color=bad_rate_color,
                vmin=vmin,
                vmax=vmax,
                axis=0
            )

    # 高亮 LIFT 值（进度条）
    if highlight_lift:
        lift_cols = _find_columns(df_display, 'LIFT值')
        for col in lift_cols:
            lift_values = df_display[col].dropna()
            vmin = 0
            if len(lift_values) > 0:
                vmax = max(lift_values.quantile(0.95), 1.5)
                vmax = min(vmax * 1.2, 5)
            else:
                vmax = 2
            styler = styler.bar(
                subset=[col],
                color=lift_color,
                vmin=vmin,
                vmax=vmax,
                axis=0
            )

    # 高亮 KS 值（进度条）
    if highlight_ks:
        ks_cols = _find_columns(df_display, '分档KS值')
        for col in ks_cols:
            ks_values = df_display[col].dropna()
            vmin = 0
            if len(ks_values) > 0:
                vmax = max(ks_values.quantile(0.95), 0.1)
                vmax = min(vmax * 1.2, 1)
            else:
                vmax = 1
            styler = styler.bar(
                subset=[col],
                color=ks_color,
                vmin=vmin,
                vmax=vmax,
                axis=0
            )

    # 设置表格样式 - 列名不换行，横向滚动
    styler = styler.set_properties(**{
        'white-space': 'nowrap',
        'text-align': 'center',
        'font-size': '12px',
    })

    # 表头和行列样式 - 简洁清爽风格
    styler = styler.set_table_styles([
        {'selector': 'thead th', 'props': [
            ('background-color', '#f5f5f5'),
            ('color', '#333333'),
            ('font-weight', 'bold'),
            ('text-align', 'center'),
            ('font-size', '12px'),
            ('padding', '8px'),
            ('border', '1px solid #dddddd'),
            ('white-space', 'nowrap'),
        ]},
        {'selector': 'td', 'props': [
            ('padding', '6px 8px'),
            ('border', '1px solid #dddddd'),
        ]},
        # 奇偶行颜色不同
        {'selector': 'tr:nth-child(odd)', 'props': [
            ('background-color', '#ffffff'),
        ]},
        {'selector': 'tr:nth-child(even)', 'props': [
            ('background-color', '#fafafa'),
        ]},
        {'selector': 'tr:hover', 'props': [
            ('background-color', '#f0f0f0'),
        ]},
        # 进度条样式
        {'selector': '.pd-bar', 'props': [
            ('opacity', '0.7'),
        ]},
        # 选中样式 - 柔和的颜色对比
        {'selector': 'tr.selected, td.selected', 'props': [
            ('background-color', '#e3f2fd'),
            ('color', '#1565C0'),
        ]},
    ])

    return styler


def _find_columns(df: pd.DataFrame, pattern: str) -> List:
    """查找匹配的列名（支持多级表头）."""
    if isinstance(df.columns, pd.MultiIndex):
        return [col for col in df.columns if pattern in str(col[1])]
    else:
        return [col for col in df.columns if pattern in str(col)]


def _show_method(self, **kwargs):
    """展示美化的分箱表.
    
    :param kwargs: 传递给 BinTableDisplay.show() 的参数
    :return: BinTableDisplay 对象，支持链式调用
    
    **参考样例**
    
    >>> table = feature_bin_stats(data, 'score', target='target')
    >>> table.show()
    >>> table.show(compact=True)
    """
    display_obj = BinTableDisplay(self)
    return display_obj.show(**kwargs)


# =============================================================================
# 4. 注册所有扩展方法
# =============================================================================

# 标记是否已注册
_EXTENSIONS_REGISTERED = False


def register_extensions():
    """注册 pandas DataFrame/Series 扩展方法.
    
    在导入 hscredit 时自动调用，将以下方法添加到 pandas:
    - df.summary(): 综合特征描述统计
    - df.eda_info(): EDA基础信息
    - df.missing_analysis(): 缺失值分析
    - df.show(): 美化展示分箱表
    - df.save(): 保存到Excel
    - s.save(): Series保存到Excel
    
    Example:
        >>> import pandas as pd
        >>> import hscredit  # 自动注册扩展
        >>> 
        >>> df = pd.DataFrame({...})
        >>> summary = df.summary()
        >>> df.save("report.xlsx")
    """
    global _EXTENSIONS_REGISTERED
    
    if _EXTENSIONS_REGISTERED:
        return
    
    # EDA相关方法
    if not hasattr(pd.DataFrame, 'summary'):
        pd.DataFrame.summary = _summary_method
    
    if not hasattr(pd.DataFrame, 'eda_info'):
        pd.DataFrame.eda_info = _eda_info_method
    
    if not hasattr(pd.DataFrame, 'missing_analysis'):
        pd.DataFrame.missing_analysis = _missing_analysis_method
    
    # Excel保存方法
    if not hasattr(pd.DataFrame, 'save'):
        pd.DataFrame.save = _dataframe_save
    
    if not hasattr(pd.Series, 'save'):
        pd.Series.save = _series_save
    
    # 分箱表展示方法
    if not hasattr(pd.DataFrame, 'show'):
        pd.DataFrame.show = _show_method
    
    _EXTENSIONS_REGISTERED = True


# 自动注册扩展
register_extensions()
