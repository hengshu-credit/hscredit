"""
Pandas扩展 - 为DataFrame和Series添加save方法

使用方式：
    import pandas as pd
    from hscredit.report.excel.pandas_extension import install_save_method
    
    # 安装save方法（只需调用一次）
    install_save_method()
    
    # 之后可以直接使用
    df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    df.save("report.xlsx", sheet_name="Sheet1", title="数据表")
    
    # Series也会自动转为DataFrame保存
    s = pd.Series([1, 2, 3], name='数值')
    s.save("series_report.xlsx", title="序列数据")

或者直接导入hscredit时会自动安装：
    import hscredit  # 自动安装save方法
    import pandas as pd
    
    df = pd.DataFrame({'A': [1, 2, 3]})
    df.save("report.xlsx")
"""

from typing import Optional, Union, List, Dict, Any, Tuple
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from .writer import ExcelWriter, dataframe2excel


def _dataframe_save(
    self,
    excel_writer: Union[str, ExcelWriter],
    worksheet: Optional[Worksheet] = None,
    sheet_name: Optional[str] = None,
    title: Optional[str] = None,
    header: bool = True,
    theme_color: str = "2639E9",
    condition_color: Optional[str] = None,
    fill: bool = True,
    percent_cols: Optional[List] = None,
    condition_cols: Optional[List] = None,
    custom_cols: Optional[List] = None,
    custom_format: str = "#,##0",
    color_cols: Optional[List] = None,
    percent_rows: Optional[List] = None,
    condition_rows: Optional[List] = None,
    custom_rows: Optional[List] = None,
    color_rows: Optional[List] = None,
    start_col: int = 2,
    start_row: int = 2,
    mode: str = "replace",
    figures: Optional[Union[str, List[str]]] = None,
    figsize: Tuple[int, int] = (600, 350),
    writer_params: Optional[Dict] = None,
    **kwargs
) -> Union[Tuple[int, int], ExcelWriter]:
    """
    将DataFrame保存到Excel文件或已有的ExcelWriter中。

    :param excel_writer: 文件路径或ExcelWriter对象
    :param worksheet: 工作表对象（如果提供，将写入该worksheet而不保存文件）
    :param sheet_name: 工作表名称，默认为None
    :param title: 标题，默认为None
    :param header: 是否保存列名，默认为True
    :param theme_color: 主题颜色，默认为"2639E9"
    :param condition_color: 条件格式颜色，默认为None
    :param fill: 是否使用颜色填充，默认为True
    :param percent_cols: 需要显示为百分数的列，默认为None
    :param condition_cols: 需要显示数据条的列，默认为None
    :param custom_cols: 需要自定义格式的列，默认为None
    :param custom_format: 自定义格式，默认为"#,##0"
    :param color_cols: 需要显示颜色渐变的列，默认为None
    :param percent_rows: 需要显示为百分数的行，默认为None
    :param condition_rows: 需要显示数据条的行，默认为None
    :param custom_rows: 需要自定义格式的行，默认为None
    :param color_rows: 需要显示颜色渐变的行，默认为None
    :param start_col: 起始列，默认为2
    :param start_row: 起始行，默认为2
    :param mode: 写入模式，默认为"replace"
    :param figures: 需要插入的图片路径，默认为None
    :param figsize: 图片大小，默认为(600, 350)
    :param writer_params: ExcelWriter参数，默认为None
    :param kwargs: 其他参数，传递给insert_df2sheet
    :return: 如果传入文件路径返回(end_row, end_col)；如果传入ExcelWriter且提供了worksheet，返回ExcelWriter

    **参考样例**

    >>> import pandas as pd
    >>> from hscredit.report.excel import ExcelWriter
    >>> 
    >>> df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    >>> 
    >>> # 方式1：直接保存到文件
    >>> df.save("report.xlsx", sheet_name="数据", title="统计表")
    >>> 
    >>> # 方式2：写入已有的ExcelWriter
    >>> writer = ExcelWriter()
    >>> worksheet = writer.get_sheet_by_name("Sheet1")
    >>> df.save(writer, worksheet=worksheet)
    >>> writer.save("report.xlsx")
    """
    # 如果提供了worksheet，说明要使用已有的writer
    if worksheet is not None and isinstance(excel_writer, ExcelWriter):
        # 直接插入到指定的worksheet
        writer = excel_writer
        
        # 插入标题
        if title:
            col_width = len(self.columns) + self.index.nlevels if kwargs.get("index", False) else len(self.columns)
            _start_row, _end_col = writer.insert_value2sheet(
                worksheet, (start_row, start_col),
                value=title,
                style="header",
                end_space=(start_row, start_col + col_width - 1)
            )
            start_row += 1
        
        # 插入图片
        if figures is not None:
            if isinstance(figures, str):
                figures = [figures]
            pic_row = start_row
            for i, pic in enumerate(figures):
                if i == 0:
                    start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, start_col), figsize=figsize)
                else:
                    start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, end_col - 1), figsize=figsize)
        
        # 处理merge_column参数
        if "merge_column" in kwargs and kwargs["merge_column"]:
            if not isinstance(kwargs["merge_column"][0], (tuple, list)):
                kwargs["merge_column"] = [c for c in self.columns if (isinstance(c, tuple) and c[-1] in kwargs["merge_column"]) or (not isinstance(c, tuple) and c in kwargs["merge_column"])]
        
        # 插入DataFrame
        from openpyxl.utils import get_column_letter
        
        end_row, end_col = writer.insert_df2sheet(
            worksheet, self, (start_row, start_col),
            fill=fill, header=header, **kwargs
        )
        
        # 设置百分比格式列
        if percent_cols:
            if not isinstance(percent_cols[0], (tuple, list)):
                percent_cols = [c for c in self.columns if (isinstance(c, tuple) and c[-1] in percent_cols) or (not isinstance(c, tuple) and c in percent_cols)]
            for c in [c for c in percent_cols if c in self.columns]:
                conditional_column = get_column_letter(
                    start_col + self.columns.get_loc(c) + self.index.nlevels if kwargs.get("index", False) else start_col + self.columns.get_loc(c)
                )
                writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(self)}:{conditional_column}{end_row - 1}", "0.00%")
        
        # 设置自定义格式列
        if custom_cols:
            if not isinstance(custom_cols[0], (tuple, list)):
                custom_cols = [c for c in self.columns if (isinstance(c, tuple) and c[-1] in custom_cols) or (not isinstance(c, tuple) and c in custom_cols)]
            for c in [c for c in custom_cols if c in self.columns]:
                conditional_column = get_column_letter(
                    start_col + self.columns.get_loc(c) + self.index.nlevels if kwargs.get("index", False) else start_col + self.columns.get_loc(c)
                )
                writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(self)}:{conditional_column}{end_row - 1}", custom_format)
        
        # 设置条件格式列
        if condition_cols:
            if not isinstance(condition_cols[0], (tuple, list)):
                condition_cols = [c for c in self.columns if (isinstance(c, tuple) and c[-1] in condition_cols) or (not isinstance(c, tuple) and c in condition_cols)]
            for c in [c for c in condition_cols if c in self.columns]:
                conditional_column = get_column_letter(
                    start_col + self.columns.get_loc(c) + self.index.nlevels if kwargs.get("index", False) else start_col + self.columns.get_loc(c)
                )
                writer.add_conditional_formatting(
                    worksheet,
                    f'{conditional_column}{end_row - len(self)}',
                    f'{conditional_column}{end_row - 1}',
                    condition_color=condition_color or theme_color
                )
        
        return writer
    
    # 使用dataframe2excel函数（传入文件路径或ExcelWriter但没有worksheet）
    return dataframe2excel(
        data=self,
        excel_writer=excel_writer,
        sheet_name=sheet_name,
        title=title,
        header=header,
        theme_color=theme_color,
        condition_color=condition_color,
        fill=fill,
        percent_cols=percent_cols,
        condition_cols=condition_cols,
        custom_cols=custom_cols,
        custom_format=custom_format,
        color_cols=color_cols,
        percent_rows=percent_rows,
        condition_rows=condition_rows,
        custom_rows=custom_rows,
        color_rows=color_rows,
        start_col=start_col,
        start_row=start_row,
        mode=mode,
        figures=figures,
        figsize=figsize,
        writer_params=writer_params,
        **kwargs
    )


def _series_save(
    self,
    excel_writer: Union[str, ExcelWriter],
    worksheet: Optional[Worksheet] = None,
    sheet_name: Optional[str] = None,
    title: Optional[str] = None,
    header: bool = True,
    theme_color: str = "2639E9",
    condition_color: Optional[str] = None,
    fill: bool = True,
    percent_cols: Optional[List] = None,
    condition_cols: Optional[List] = None,
    custom_cols: Optional[List] = None,
    custom_format: str = "#,##0",
    color_cols: Optional[List] = None,
    percent_rows: Optional[List] = None,
    condition_rows: Optional[List] = None,
    custom_rows: Optional[List] = None,
    color_rows: Optional[List] = None,
    start_col: int = 2,
    start_row: int = 2,
    mode: str = "replace",
    figures: Optional[Union[str, List[str]]] = None,
    figsize: Tuple[int, int] = (600, 350),
    writer_params: Optional[Dict] = None,
    **kwargs
) -> Union[Tuple[int, int], ExcelWriter]:
    """
    将Series保存到Excel文件或已有的ExcelWriter中。
    Series会被转换为单列DataFrame（保留name作为列名）。

    :param excel_writer: 文件路径或ExcelWriter对象
    :param worksheet: 工作表对象（如果提供，将写入该worksheet而不保存文件）
    :param sheet_name: 工作表名称，默认为None
    :param title: 标题，默认为None
    :param header: 是否保存列名，默认为True
    :param theme_color: 主题颜色，默认为"2639E9"
    :param condition_color: 条件格式颜色，默认为None
    :param fill: 是否使用颜色填充，默认为True
    :param percent_cols: 需要显示为百分数的列，默认为None
    :param condition_cols: 需要显示数据条的列，默认为None
    :param custom_cols: 需要自定义格式的列，默认为None
    :param custom_format: 自定义格式，默认为"#,##0"
    :param color_cols: 需要显示颜色渐变的列，默认为None
    :param percent_rows: 需要显示为百分数的行，默认为None
    :param condition_rows: 需要显示数据条的行，默认为None
    :param custom_rows: 需要自定义格式的行，默认为None
    :param color_rows: 需要显示颜色渐变的行，默认为None
    :param start_col: 起始列，默认为2
    :param start_row: 起始行，默认为2
    :param mode: 写入模式，默认为"replace"
    :param figures: 需要插入的图片路径，默认为None
    :param figsize: 图片大小，默认为(600, 350)
    :param writer_params: ExcelWriter参数，默认为None
    :param kwargs: 其他参数，传递给insert_df2sheet
    :return: 如果传入文件路径返回(end_row, end_col)；如果传入ExcelWriter且提供了worksheet，返回ExcelWriter

    **参考样例**

    >>> import pandas as pd
    >>> from hscredit.report.excel import ExcelWriter
    >>> 
    >>> s = pd.Series([1, 2, 3], name='数值', index=['a', 'b', 'c'])
    >>> 
    >>> # 方式1：直接保存到文件
    >>> s.save("report.xlsx", sheet_name="数据", title="序列数据")
    >>> 
    >>> # 方式2：写入已有的ExcelWriter
    >>> writer = ExcelWriter()
    >>> worksheet = writer.get_sheet_by_name("Sheet1")
    >>> s.save(writer, worksheet=worksheet)
    >>> writer.save("report.xlsx")
    """
    # 将Series转换为DataFrame
    df = self.to_frame()
    
    # 调用DataFrame的save方法
    return df.save(
        excel_writer=excel_writer,
        worksheet=worksheet,
        sheet_name=sheet_name,
        title=title,
        header=header,
        theme_color=theme_color,
        condition_color=condition_color,
        fill=fill,
        percent_cols=percent_cols,
        condition_cols=condition_cols,
        custom_cols=custom_cols,
        custom_format=custom_format,
        color_cols=color_cols,
        percent_rows=percent_rows,
        condition_rows=condition_rows,
        custom_rows=custom_rows,
        color_rows=color_rows,
        start_col=start_col,
        start_row=start_row,
        mode=mode,
        figures=figures,
        figsize=figsize,
        writer_params=writer_params,
        **kwargs
    )


# 标记是否已安装
_SAVE_METHOD_INSTALLED = False


def install_save_method():
    """安装save方法到pandas的DataFrame和Series类。
    
    该方法可以安全地多次调用，但只有第一次会实际安装。
    """
    global _SAVE_METHOD_INSTALLED
    
    if _SAVE_METHOD_INSTALLED:
        return
    
    # 将save方法绑定到DataFrame和Series
    pd.DataFrame.save = _dataframe_save
    pd.Series.save = _series_save
    
    _SAVE_METHOD_INSTALLED = True


# 自动安装（当导入此模块时）
install_save_method()
