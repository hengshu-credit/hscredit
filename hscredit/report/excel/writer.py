# -*- coding: utf-8 -*-
"""
Excel写入器

提供专业的Excel报告生成功能，支持丰富的样式和格式化选项。

核心功能:
- DataFrame数据写入，支持多层索引和多层列名
- 图片插入
- 超链接插入
- 条件格式设置
- 自定义样式

迁移自 scorecardpipeline.excel_writer
"""

import sys
import warnings
import re
import os
import copy
from typing import Optional, Union, List, Tuple, Dict, Any

import numpy as np
import pandas as pd

from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import NamedStyle, Border, Side, Alignment, PatternFill, Font

warnings.filterwarnings("ignore")


class ExcelWriter:
    """Excel写入器，提供专业的Excel报告生成功能。

    支持DataFrame数据写入、图片插入、超链接等功能。
    支持上下文管理器（with语句）自动保存。

    :param style_excel: 样式模板文件路径，默认使用包内的template.xlsx
    :param style_sheet_name: 模板文件内初始样式sheet名称，默认为"初始化"
    :param mode: 写入模式，可选'replace'或'append'，默认为'replace'
        - replace: 替换已有文件
        - append: 在已有文件基础上追加内容
    :param fontsize: 字体大小，默认为10
    :param font: 字体名称，默认为"楷体"
    :param theme_color: 主题颜色（不包含#），默认为"2639E9"
    :param opacity: 颜色填充的透明度，默认为0.85
    :param system: 操作系统类型，可选'mac'、'windows'、'linux'，默认自动检测

    **参考样例**

    >>> import pandas as pd
    >>> from hscredit.report.excel import ExcelWriter
    >>>
    >>> # 方法1：使用with语句（推荐，自动保存）
    >>> with ExcelWriter(theme_color='3f1dba').set_filename("report.xlsx") as writer:
    ...     worksheet = writer.get_sheet_by_name("模型报告")
    ...     writer.insert_value2sheet(worksheet, "B2", value="模型报告", style="header")
    ...     df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    ...     writer.insert_df2sheet(worksheet, df, "B4")
    >>> # 文件在退出with块时自动保存
    >>>
    >>> # 方法2：手动调用save（原有方式）
    >>> writer = ExcelWriter(theme_color='3f1dba')
    >>> worksheet = writer.get_sheet_by_name("模型报告")
    >>> writer.insert_value2sheet(worksheet, "B2", value="模型报告", style="header")
    >>> df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    >>> writer.insert_df2sheet(worksheet, df, "B4")
    >>> writer.save("report.xlsx")
    """

    def __init__(
        self,
        style_excel: Optional[str] = None,
        style_sheet_name: str = "初始化",
        mode: str = "replace",
        fontsize: int = 10,
        font: str = '楷体',
        theme_color: str = '2639E9',
        opacity: float = 0.85,
        system: Optional[str] = None
    ):
        # 系统检测
        self.system = system
        if self.system is None:
            self.system = "mac" if sys.platform == "darwin" else "windows"

        # 样式参数
        self.english_width = 0.12
        self.chinese_width = 0.21
        self.mode = mode
        self.font = font
        self.opacity = opacity
        self.fontsize = fontsize
        self.theme_color = theme_color

        # 加载模板
        if style_excel is None:
            # 使用resources目录下的template.xlsx
            style_excel = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                'resources', 'templates', 'template.xlsx'
            )

        self.workbook = load_workbook(style_excel)
        self.style_sheet = self.workbook[style_sheet_name]

        # 初始化样式
        self.name_styles = []
        self.init_style(font, fontsize, theme_color)

        # 注册命名样式
        for style in self.name_styles:
            if style.name not in self.workbook.style_names:
                self.workbook.add_named_style(style)

        # 用于上下文管理器的文件路径
        self._filename: Optional[str] = None

    def __enter__(self) -> 'ExcelWriter':
        """进入上下文管理器。

        :return: ExcelWriter实例
        """
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """退出上下文管理器时自动保存文件。

        :param exc_type: 异常类型
        :param exc_val: 异常值
        :param exc_tb: 异常追踪信息
        """
        if self._filename is not None:
            self.save(self._filename)

    def set_filename(self, filename: str) -> 'ExcelWriter':
        """设置用于上下文管理器自动保存的文件路径。

        支持链式调用，可以与with语句配合使用。

        :param filename: 保存路径
        :return: self

        **参考样例**

        >>> # 方法1：使用set_filename设置路径
        >>> with ExcelWriter().set_filename("report.xlsx") as writer:
        ...     worksheet = writer.get_sheet_by_name("Sheet1")
        ...     writer.insert_value2sheet(worksheet, "B2", "Hello")
        >>> # 文件自动保存到report.xlsx

        >>> # 方法2：手动调用save（原有方式）
        >>> writer = ExcelWriter()
        >>> worksheet = writer.get_sheet_by_name("Sheet1")
        >>> writer.insert_value2sheet(worksheet, "B2", "Hello")
        >>> writer.save("report.xlsx")
        """
        self._filename = filename
        return self

    def add_conditional_formatting(
        self,
        worksheet: Worksheet,
        start_space: str,
        end_space: str,
        condition_color: Optional[str] = None
    ) -> None:
        """设置条件格式（数据条）。

        :param worksheet: 工作表对象
        :param start_space: 开始单元格位置，如'B2'
        :param end_space: 结束单元格位置，如'B10'
        :param condition_color: 条件格式颜色，默认使用主题色
        """
        worksheet.conditional_formatting.add(
            f'{start_space}:{end_space}',
            DataBarRule(
                start_type='min',
                end_type='max',
                color=condition_color or self.theme_color
            )
        )

    @staticmethod
    def set_column_width(
        worksheet: Worksheet,
        column: Union[str, int],
        width: float
    ) -> None:
        """调整Excel列宽。

        :param worksheet: 工作表对象
        :param column: 列，可以是字母（如'B'）或索引（如2）
        :param width: 列宽
        """
        col_letter = column if isinstance(column, str) else get_column_letter(column)
        worksheet.column_dimensions[col_letter].width = width

    @staticmethod
    def set_number_format(
        worksheet: Worksheet,
        space: str,
        _format: str
    ) -> None:
        """设置数值显示格式。

        :param worksheet: 工作表对象
        :param space: 单元格范围，如'B2:B10'
        :param _format: 显示格式，如'0.00%'或'#,##0'
        """
        cells = worksheet[space]
        if isinstance(cells, Cell):
            cells = [cells]

        for cell in cells:
            if isinstance(cell, tuple):
                for c in cell:
                    c.number_format = _format
            else:
                cell.number_format = _format

    def set_freeze_panes(
        self,
        worksheet: Union[Worksheet, str],
        space: Union[str, Tuple[int, int]]
    ) -> None:
        """设置冻结窗格。

        :param worksheet: 工作表对象或名称
        :param space: 冻结位置，如'B2'或(2, 2)
        """
        if not isinstance(worksheet, Worksheet):
            worksheet = self.get_sheet_by_name(worksheet)

        if isinstance(space, (tuple, list)):
            space = self.get_cell_space(space)

        worksheet.freeze_panes = space

    def get_sheet_by_name(self, name: str) -> Worksheet:
        """获取或创建指定名称的工作表。

        :param name: 工作表名称
        :return: 工作表对象
        """
        if name not in self.workbook.sheetnames:
            worksheet = self.workbook.copy_worksheet(self.style_sheet)
            worksheet.title = name
        else:
            worksheet = self.workbook[name]

        return worksheet

    def move_sheet(
        self,
        worksheet: Union[Worksheet, str],
        offset: int = 0,
        index: Optional[int] = None
    ) -> None:
        """移动工作表位置。

        :param worksheet: 工作表对象或名称
        :param offset: 相对移动位置，默认为0
        :param index: 移动到的位置索引，默认为None
        """
        total_sheets = len(self.workbook.sheetnames)

        if index is not None:
            offset = -(total_sheets - 1) + index
            if offset >= total_sheets:
                offset = total_sheets - 1

        self.workbook.move_sheet(worksheet, offset=offset)

    def insert_hyperlink2sheet(
        self,
        worksheet: Worksheet,
        insert_space: Union[str, Tuple[int, int]],
        hyperlink: Optional[str] = None,
        file: Optional[str] = None,
        sheet: Optional[str] = None,
        target_space: Optional[Union[str, Tuple[int, int]]] = None
    ) -> None:
        """向单元格插入超链接。

        :param worksheet: 工作表对象
        :param insert_space: 插入位置，如'B2'或(2, 2)
        :param hyperlink: 超链接地址，与target_space互斥
        :param file: 超链接文件路径，默认当前文件
        :param sheet: 超链接sheet名称，默认当前sheet
        :param target_space: 超链接目标位置，如'B10'或(10, 2)

        **参考样例**

        >>> # 链接到当前sheet的其他位置
        >>> writer.insert_hyperlink2sheet(worksheet, "B2", target_space="B10")
        >>>
        >>> # 链接到其他sheet
        >>> writer.insert_hyperlink2sheet(worksheet, "B2", sheet="Sheet2", target_space="A1")
        >>>
        >>> # 链接到外部文件
        >>> writer.insert_hyperlink2sheet(worksheet, "B2", file="other.xlsx", sheet="Sheet1", target_space="A1")
        """
        # 解析插入位置
        if isinstance(insert_space, str):
            start_col = re.findall(r'\D+', insert_space)[0]
            start_row = int(re.findall(r"\d+", insert_space)[0])
        else:
            start_col = get_column_letter(insert_space[1])
            start_row = insert_space[0]

        cell = worksheet[f"{start_col}{start_row}"]

        # 构建超链接
        if hyperlink is None:
            if target_space is None:
                raise ValueError("hyperlink 和 target_space 必须传入一个")

            if sheet is None:
                sheet = worksheet.title

            # 解析目标位置
            if isinstance(target_space, str):
                target_col = re.findall(r'\D+', target_space)[0]
                target_row = int(re.findall(r"\d+", target_space)[0])
            else:
                target_col = get_column_letter(target_space[1])
                target_row = target_space[0]

            # 构建链接
            if file:
                hyperlink = f"file://{file} - #{sheet}!{target_col}{target_row}"
            else:
                hyperlink = f"#{sheet}!{target_col}{target_row}"

        cell.hyperlink = Hyperlink(
            ref=f"{start_col}{start_row}",
            location=hyperlink,
            display=f"{cell.value}"
        )

    def insert_value2sheet(
        self,
        worksheet: Worksheet,
        insert_space: Union[str, Tuple[int, int]],
        value: Any = "",
        style: str = "content",
        auto_width: bool = False,
        end_space: Optional[Union[str, Tuple[int, int]]] = None,
        align: Optional[Dict[str, str]] = None,
        max_col_width: int = 50
    ) -> Tuple[int, int]:
        """向单元格插入内容。

        :param worksheet: 工作表对象
        :param insert_space: 插入位置，如'B2'或(2, 2)
        :param value: 插入的内容，默认为""
        :param style: 样式名称，默认为"content"
        :param auto_width: 是否自动调整列宽，默认为False
        :param end_space: 合并单元格的结束位置，默认为None
        :param align: 文本对齐方式，默认为None，例如{'horizontal': 'left', 'vertical': 'center'}
        :param max_col_width: 最大列宽，默认为50
        :return: (下一行行号, 下一列列号)

        **参考样例**

        >>> # 插入普通内容
        >>> writer.insert_value2sheet(worksheet, "B2", value="模型报告", style="header")
        >>>
        >>> # 合并单元格
        >>> writer.insert_value2sheet(worksheet, "B2", value="标题", style="header", end_space="D2")
        >>>
        >>> # 自动调整列宽
        >>> writer.insert_value2sheet(worksheet, "B2", value="内容", auto_width=True)
        """
        # 解析位置
        if isinstance(insert_space, str):
            start_col = re.findall(r'\D+', insert_space)[0]
            start_row = int(re.findall(r"\d+", insert_space)[0])
        else:
            start_col = get_column_letter(insert_space[1])
            start_row = insert_space[0]

        # 设置单元格
        cell = worksheet[f"{start_col}{start_row}"]
        cell.style = style

        # 设置对齐方式
        if align:
            _align = {"horizontal": "center", "vertical": "center"}
            _align.update(align)
            cell.alignment = Alignment(**_align)

        # 合并单元格
        if end_space is not None:
            if isinstance(end_space, str):
                end_col = re.findall(r'\D+', end_space)[0]
                end_row = int(re.findall(r"\d+", end_space)[0])
            else:
                end_col = get_column_letter(end_space[1])
                end_row = end_space[0]

            worksheet.merge_cells(f"{start_col}{start_row}:{end_col}{end_row}")

        # 设置值
        worksheet[f"{start_col}{start_row}"] = value

        # 自动调整列宽
        if auto_width:
            curr_width = worksheet.column_dimensions[start_col].width
            _, eng_cnt, chi_cnt = self.check_contain_chinese(value)
            auto_width = min(
                max([
                    (eng_cnt * self.english_width + chi_cnt * self.chinese_width) * self.fontsize,
                    10,
                    curr_width
                ]),
                max_col_width
            )
            worksheet.column_dimensions[start_col].width = auto_width

        # 返回下一个位置
        if end_space is not None:
            return end_row + 1, column_index_from_string(end_col) + 1
        else:
            return start_row + 1, column_index_from_string(start_col) + 1

    def insert_pic2sheet(
        self,
        worksheet: Worksheet,
        fig: str,
        insert_space: Union[str, Tuple[int, int]],
        figsize: Tuple[int, int] = (600, 250)
    ) -> Tuple[int, int]:
        """向Excel插入图片。

        :param worksheet: 工作表对象
        :param fig: 图片路径
        :param insert_space: 插入位置，如'B2'或(2, 2)
        :param figsize: 图片大小(宽, 高)，默认为(600, 250)
        :return: (下一行行号, 下一列列号)
        """
        # 解析位置
        if isinstance(insert_space, str):
            start_row = int(re.findall(r"\d+", insert_space)[0])
            start_col = re.findall(r'\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        # 插入图片
        image = Image(fig)
        image.width, image.height = figsize
        worksheet.add_image(image, f"{start_col}{start_row}")

        # 计算占用的行数
        row_height = 16.0 if self.system != 'mac' else 17.5
        occupied_rows = int(figsize[1] / row_height)

        return start_row + occupied_rows, column_index_from_string(start_col) + 8

    def insert_df2sheet(
        self,
        worksheet: Worksheet,
        data: pd.DataFrame,
        insert_space: Union[str, Tuple[int, int]],
        merge_column: Optional[Union[str, List[str]]] = None,
        header: bool = True,
        index: bool = False,
        auto_width: bool = False,
        fill: bool = False,
        merge: bool = False,
        merge_index: bool = True
    ) -> Tuple[int, int]:
        """向Excel插入DataFrame。

        :param worksheet: 工作表对象
        :param data: 需要插入的DataFrame
        :param insert_space: 插入位置，如'B2'或(2, 2)
        :param merge_column: 需要分组显示的列，默认为None
        :param header: 是否存储DataFrame的header，默认为True
        :param index: 是否存储DataFrame的index，默认为False
        :param auto_width: 是否自动调整列宽，默认为False
        :param fill: 是否使用颜色填充而非边框，默认为False
        :param merge: 是否合并单元格，默认为False
        :param merge_index: 当存储index时，是否合并连续相同的index值，默认为True
        :return: (下一行行号, 下一列列号)

        **参考样例**

        >>> import pandas as pd
        >>> df = pd.DataFrame({'A': [1, 1, 2], 'B': [4, 5, 6], 'C': [7, 8, 9]})
        >>>
        >>> # 基本插入
        >>> writer.insert_df2sheet(worksheet, df, "B2")
        >>>
        >>> # 使用颜色填充
        >>> writer.insert_df2sheet(worksheet, df, "B10", fill=True)
        >>>
        >>> # 保存索引
        >>> writer.insert_df2sheet(worksheet, df.set_index('A'), "B20", index=True)
        >>>
        >>> # 分组显示
        >>> writer.insert_df2sheet(worksheet, df, "B30", merge_column='A', merge=True)
        """
        df = data.copy()

        # 解析起始位置
        if isinstance(insert_space, str):
            start_row = int(re.findall(r"\d+", insert_space)[0])
            start_col = re.findall(r'\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        # 计算合并行
        def get_merge_rows(values, start_row):
            _rows = []
            item, start, length = self.calc_continuous_cnt(values)
            while start is not None:
                _rows.append(start + start_row)
                item, start, length = self.calc_continuous_cnt(values, start + length)
            _rows.append(len(values) + start_row)
            return _rows

        # 处理索引合并
        if index and merge_index:
            merge_index_cols = {
                i: get_column_letter(column_index_from_string(start_col) + i)
                for i in range(df.index.nlevels)
            }
            merge_index_rows = {
                i: get_merge_rows(
                    df.index.get_level_values(i).tolist(),
                    start_row + df.columns.nlevels if header else start_row
                )
                for i in range(df.index.nlevels)
            }
        else:
            merge_index_cols = None
            merge_index_rows = None

        # 处理列合并
        if merge_column:
            if not isinstance(merge_column, (list, np.ndarray)):
                merge_column = [merge_column]

            if isinstance(merge_column[0], (int, float)) and (merge_column[0] not in df.columns):
                merge_column = [
                    df.columns.tolist()[col] if col not in df.columns else col
                    for col in merge_column
                ]

            if index:
                merge_cols = {
                    col: get_column_letter(
                        df.columns.get_loc(col) + column_index_from_string(start_col) + df.index.nlevels
                    )
                    for col in merge_column
                }
            else:
                merge_cols = {
                    col: get_column_letter(
                        df.columns.get_loc(col) + column_index_from_string(start_col)
                    )
                    for col in merge_column
                }

            if header:
                merge_rows = {
                    col: get_merge_rows(df[col].tolist(), start_row + df.columns.nlevels)
                    for col in merge_column
                }
            else:
                merge_rows = {
                    col: get_merge_rows(df[col].tolist(), start_row)
                    for col in merge_column
                }
        else:
            merge_cols = None
            merge_rows = None

        # 迭代行数据
        def _iter_rows(df, header=True, index=True):
            columns = df.columns.tolist()
            indexs = df.index.tolist()
            for i, row in enumerate(dataframe_to_rows(df, header=header, index=False)):
                if header:
                    if i < df.columns.nlevels:
                        if index:
                            if df.columns.nlevels > 1:
                                if i == df.columns.nlevels - 1:
                                    yield list(df.index.names) + [c[i] for c in columns]
                                else:
                                    yield [None] * df.index.nlevels + [c[i] for c in columns]
                            else:
                                yield list(df.index.names) + columns
                        else:
                            if df.columns.nlevels > 1 and i < df.columns.nlevels:
                                yield [c[i] for c in columns]
                            else:
                                yield columns
                    else:
                        if index:
                            if df.index.nlevels > 1:
                                yield list(indexs[int(i - df.columns.nlevels)]) + row
                            else:
                                yield [indexs[int(i - df.columns.nlevels)]] + row
                        else:
                            yield row
                else:
                    if index:
                        if df.index.nlevels > 1:
                            yield list(indexs[i]) + row
                        else:
                            yield [indexs[i]] + row
                    else:
                        yield row

        # 插入数据
        for i, row in enumerate(_iter_rows(df, header=header, index=index)):
            if fill:
                if header and i < df.columns.nlevels:
                    self.insert_rows(
                        worksheet, row, start_row + i, start_col,
                        style="header",
                        auto_width=auto_width,
                        multi_levels=True if df.columns.nlevels > 1 else False
                    )
                elif i == 0:
                    self.insert_rows(
                        worksheet, row, start_row + i, start_col,
                        style="middle_even_first",
                        auto_width=auto_width,
                        style_only=True
                    )
                else:
                    # 根据行数奇偶选择样式
                    if df.columns.nlevels % 2 == 1:
                        if i % 2 == 1:
                            style = "middle_odd_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_odd"
                        else:
                            style = "middle_even_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_even"
                    else:
                        if i % 2 == 1:
                            style = "middle_even_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_even"
                        else:
                            style = "middle_odd_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_odd"

                    self.insert_rows(
                        worksheet, row, start_row + i, start_col,
                        style=style,
                        auto_width=auto_width,
                        style_only=True
                    )
            else:
                if header and i < df.columns.nlevels:
                    self.insert_rows(
                        worksheet, row, start_row + i, start_col,
                        style="header",
                        auto_width=auto_width,
                        multi_levels=True if df.columns.nlevels > 1 else False
                    )
                elif i == 0:
                    self.insert_rows(
                        worksheet, row, start_row + i, start_col,
                        style="first",
                        auto_width=auto_width
                    )
                elif (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)):
                    self.insert_rows(
                        worksheet, row, start_row + i, start_col,
                        style="last",
                        auto_width=auto_width
                    )
                else:
                    if merge_rows and len(merge_rows) > 0:
                        self.insert_rows(
                            worksheet, row, start_row + i, start_col,
                            auto_width=auto_width,
                            merge_rows=sorted(set(_row for _rows in merge_rows.values() for _row in _rows))
                        )
                    else:
                        self.insert_rows(
                            worksheet, row, start_row + i, start_col,
                            auto_width=auto_width
                        )

        # 合并索引单元格
        if index and merge_index and merge_index_rows and len(merge_index_rows) > 0:
            for col in merge_index_cols.keys():
                merge_col = merge_index_cols[col]
                merge_row = merge_index_rows[col]

                for s, e in zip(merge_row[:-1], merge_row[1:]):
                    if e - s > 1:
                        self.merge_cells(worksheet, f"{merge_col}{s}", f"{merge_col}{e - 1}")

        # 合并列单元格
        if merge and merge_column and merge_cols and len(merge_cols) > 0:
            for col in merge_cols.keys():
                merge_col = merge_cols[col]
                merge_row = merge_rows[col]

                for s, e in zip(merge_row[:-1], merge_row[1:]):
                    if e - s > 1:
                        self.merge_cells(worksheet, f"{merge_col}{s}", f"{merge_col}{e - 1}")

        end_row = start_row + len(data) + df.columns.nlevels if header else start_row + len(data)

        return end_row, column_index_from_string(start_col) + len(data.columns)

    def insert_rows(
        self,
        worksheet: Worksheet,
        row: List,
        row_index: int,
        col_index: Union[str, int],
        merge_rows: Optional[List[int]] = None,
        style: str = "",
        auto_width: bool = False,
        style_only: bool = False,
        multi_levels: bool = False
    ) -> None:
        """向Excel插入一行数据。

        :param worksheet: 工作表对象
        :param row: 行数据
        :param row_index: 行索引
        :param col_index: 起始列索引或字母
        :param merge_rows: 需要合并的行索引列表，默认为None
        :param style: 样式名称，默认为空
        :param auto_width: 是否自动调整列宽，默认为False
        :param style_only: 是否仅应用样式，默认为False
        :param multi_levels: 是否多层索引，默认为False
        """
        curr_col = column_index_from_string(col_index) if isinstance(col_index, str) else col_index

        if multi_levels and style == "header":
            row = pd.Series(row).ffill().to_list()
            item, start, length = self.calc_continuous_cnt(row)

            while start is not None:
                if start + length < len(row):
                    if start == 0:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + start)}{row_index}',
                            self.astype_insertvalue(item),
                            style=f"{style}_left" if style else "left",
                            auto_width=auto_width,
                            end_space=f'{get_column_letter(curr_col + start + length - 1)}{row_index}'
                        )
                    else:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + start)}{row_index}',
                            self.astype_insertvalue(item),
                            style=f"{style}_middle" if style else "middle",
                            auto_width=auto_width,
                            end_space=f'{get_column_letter(curr_col + start + length - 1)}{row_index}'
                        )
                else:
                    self.insert_value2sheet(
                        worksheet,
                        f'{get_column_letter(curr_col + start)}{row_index}',
                        self.astype_insertvalue(item),
                        style=f"{style}_right" if style else "right",
                        auto_width=auto_width,
                        end_space=f'{get_column_letter(curr_col + start + length - 1)}{row_index}'
                    )

                item, start, length = self.calc_continuous_cnt(row, start + length)
        else:
            for j, v in enumerate(row):
                if merge_rows is not None and row_index + 1 not in merge_rows:
                    if j == 0:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            self.astype_insertvalue(v),
                            style="merge_left",
                            auto_width=auto_width
                        )
                    elif j == len(row) - 1:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            self.astype_insertvalue(v),
                            style="merge_right",
                            auto_width=auto_width
                        )
                    else:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            self.astype_insertvalue(v),
                            style="merge_middle",
                            auto_width=auto_width
                        )
                elif style_only or len(row) <= 1:
                    self.insert_value2sheet(
                        worksheet,
                        f'{get_column_letter(curr_col + j)}{row_index}',
                        self.astype_insertvalue(v),
                        style=style or "middle",
                        auto_width=auto_width
                    )
                else:
                    if j == 0:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            self.astype_insertvalue(v),
                            style=f"{style}_left" if style else "left",
                            auto_width=auto_width
                        )
                    elif j == len(row) - 1:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            self.astype_insertvalue(v),
                            style=f"{style}_right" if style else "right",
                            auto_width=auto_width
                        )
                    else:
                        self.insert_value2sheet(
                            worksheet,
                            f'{get_column_letter(curr_col + j)}{row_index}',
                            self.astype_insertvalue(v),
                            style=f"{style}_middle" if style else "middle",
                            auto_width=auto_width
                        )

    def merge_cells(
        self,
        worksheet: Worksheet,
        start: Union[str, Tuple[int, int]],
        end: Union[str, Tuple[int, int]]
    ) -> None:
        """合并单元格并保证样式正确合并。

        :param worksheet: 工作表对象
        :param start: 开始位置，如'B2'或(2, 2)
        :param end: 结束位置，如'F10'或(10, 6)
        """
        # 解析位置
        if isinstance(start, str):
            start_col, start_row = self.get_cell_space(start)
        elif isinstance(start, (tuple, list)):
            start_col, start_row = start[0], start[1]
        else:
            raise TypeError("仅支持二元组或字符串")

        if isinstance(end, str):
            end_col, end_row = self.get_cell_space(end)
        elif isinstance(end, (tuple, list)):
            end_col, end_row = end[0], end[1]
        else:
            raise TypeError("仅支持二元组或字符串")

        # 确保起始列不大于结束列
        if start_col > end_col:
            start_col, end_col = end_col, start_col
        if start_row > end_row:
            start_row, end_row = end_row, start_row

        # 获取左上角单元格的样式
        top_left_cell = worksheet[f"{get_column_letter(start_col)}{start_row}"]
        cell_style = copy.deepcopy(top_left_cell.style)

        # 获取各边框样式
        top_border = top_left_cell.border.top
        left_cell = worksheet[f"{get_column_letter(start_col)}{start_row}"]
        left_border = left_cell.border.left
        right_cell = worksheet[f"{get_column_letter(end_col)}{start_row}"]
        right_border = right_cell.border.right
        bottom_cell = worksheet[f"{get_column_letter(start_col)}{end_row}"]
        bottom_border = bottom_cell.border.bottom

        # 创建合并后的边框样式
        border_style = Border(
            top=Side(style=top_border.style, color=top_border.color) if top_border else None,
            left=Side(style=left_border.style, color=left_border.color) if left_border else None,
            right=Side(style=right_border.style, color=right_border.color) if right_border else None,
            bottom=Side(style=bottom_border.style, color=bottom_border.color) if bottom_border else None,
        )

        # 将单元格样式应用到左上角单元格
        merged_cell = worksheet[f"{get_column_letter(start_col)}{start_row}"]
        merged_cell.style = cell_style
        merged_cell.border = border_style

        # 合并单元格
        start_cell = f"{get_column_letter(start_col)}{start_row}"
        end_cell = f"{get_column_letter(end_col)}{end_row}"
        worksheet.merge_cells(f"{start_cell}:{end_cell}")

    @staticmethod
    def check_contain_chinese(check_str: str) -> Tuple[List[bool], int, int]:
        """检查字符串中是否包含中文。

        :param check_str: 需要检查的字符串
        :return: (每个字符是否是中文的列表, 英文字符个数, 中文字符个数)
        """
        out = []
        for ch in str(check_str).encode('utf-8').decode('utf-8'):
            if u'\u4e00' <= ch <= u'\u9fff':
                out.append(True)
            else:
                out.append(False)
        return out, len(out) - sum(out), sum(out)

    @staticmethod
    def astype_insertvalue(value: Any, decimal_point: int = 4) -> Any:
        """格式化需要存储Excel的内容。

        :param value: 需要插入Excel的内容
        :param decimal_point: 如果是浮点型，需要保留的小数位数，默认为4
        :return: 格式化后的内容
        """
        if re.search('tuple|list|set|numpy.ndarray|Categorical|numpy.dtype|Interval', str(type(value))):
            return str(value)
        elif re.search('float', str(type(value))):
            return round(float(value), decimal_point)
        else:
            return value

    @staticmethod
    def calc_continuous_cnt(list_: List, index_: int = 0) -> Tuple[Any, Optional[int], Optional[int]]:
        """计算列表中从某个索引开始连续出现某个元素的个数。

        :param list_: 需要检索的列表
        :param index_: 起始索引，默认为0
        :return: (元素值, 索引值, 连续出现的个数)

        **参考样例**

        >>> calc_continuous_cnt = ExcelWriter.calc_continuous_cnt
        >>> list_ = ['A','A','A','A','B','C','C','D','D','D']
        >>> calc_continuous_cnt(list_, 0)
        ('A', 0, 4)
        >>> calc_continuous_cnt(list_, 4)
        ('B', 4, 1)
        """
        if index_ >= len(list_):
            return None, None, None

        cnt, str_ = 0, list_[index_]
        for i in range(index_, len(list_), 1):
            if list_[i] == str_:
                cnt = cnt + 1
            else:
                break
        return str_, index_, cnt

    @staticmethod
    def itlubber_border(border: List[str], color: List[str], white: bool = False) -> Border:
        """生成边框样式。

        :param border: 边框样式列表，长度为3或4。长度为3表示[左, 右, 下]，长度为4表示[左, 右, 下, 上]
        :param color: 边框颜色列表
        :param white: 是否显示白色边框，默认为False
        :return: 边框对象
        """
        if len(border) == 3:
            return Border(
                left=Side(border_style=None if not white and color[0] == "FFFFFF" else border[0], color=None if not white and color[0] == "FFFFFF" else color[0]),
                right=Side(border_style=None if not white and color[1] == "FFFFFF" else border[1], color=None if not white and color[1] == "FFFFFF" else color[1]),
                bottom=Side(border_style=border[2], color=color[2]),
            )
        else:
            return Border(
                left=Side(border_style=None if not white and color[0] == "FFFFFF" else border[0], color=None if not white and color[0] == "FFFFFF" else color[0]),
                right=Side(border_style=None if not white and color[1] == "FFFFFF" else border[1], color=None if not white and color[1] == "FFFFFF" else color[1]),
                bottom=Side(border_style=border[2], color=color[2]),
                top=Side(border_style=border[3], color=color[3]),
            )

    @staticmethod
    def get_cell_space(space: Union[str, Tuple[int, int]]) -> Union[Tuple[int, int], str]:
        """转换单元格位置格式。

        支持两种格式：
        - 字符串格式: 'B2'
        - 元组格式: (2, 2) 表示第2行第2列

        :param space: 单元格位置
        :return: 转换后的格式

        **参考样例**

        >>> get_cell_space = ExcelWriter.get_cell_space
        >>> get_cell_space("B3")
        (2, 3)
        >>> get_cell_space((2, 2))
        'B2'
        """
        if isinstance(space, str):
            start_row = int(re.findall(r"\d+", space)[0])
            start_col = re.findall(r'\D+', space)[0]
            return column_index_from_string(start_col), start_row
        else:
            start_row = space[0]
            if isinstance(space[1], int):
                start_col = get_column_letter(space[1])
            else:
                start_col = space[1]
            return f"{start_col}{start_row}"

    @staticmethod
    def calculate_rgba_color(hex_color: str, opacity: float, prefix: str = "#") -> str:
        """根据颜色和透明度计算对应的颜色值。

        :param hex_color: hex格式的颜色值
        :param opacity: 透明度，[0, 1]之间的数值
        :param prefix: 返回颜色的前缀，默认为"#"
        :return: 对应某个透明度的颜色
        """
        rgb_color = tuple(int(hex_color.lstrip('#')[i:i + 2], 16) for i in (0, 2, 4))
        rgba_color = tuple(int((1 - opacity) * c + opacity * 255) for c in rgb_color)
        return prefix + '{:02X}{:02X}{:02X}'.format(*rgba_color)

    def init_style(self, font: str, fontsize: int, theme_color: str) -> None:
        """初始化单元格样式。

        :param font: 字体名称
        :param fontsize: 字体大小
        :param theme_color: 主题颜色
        """
        # 创建所有命名样式
        header_style = NamedStyle(name="header")
        header_left_style = NamedStyle(name="header_left")
        header_middle_style = NamedStyle(name="header_middle")
        header_right_style = NamedStyle(name="header_right")

        last_style = NamedStyle(name="last")
        last_left_style = NamedStyle(name="last_left")
        last_middle_style = NamedStyle(name="last_middle")
        last_right_style = NamedStyle(name="last_right")

        content_style = NamedStyle(name="content")
        left_style = NamedStyle(name="left")
        middle_style = NamedStyle(name="middle")
        right_style = NamedStyle(name="right")

        merge_style = NamedStyle(name="merge")
        merge_left_style = NamedStyle(name="merge_left")
        merge_middle_style = NamedStyle(name="merge_middle")
        merge_right_style = NamedStyle(name="merge_right")

        first_style = NamedStyle(name="first")
        first_left_style = NamedStyle(name="first_left")
        first_middle_style = NamedStyle(name="first_middle")
        first_right_style = NamedStyle(name="first_right")

        # 字体和填充
        header_font = Font(size=fontsize, name=font, color="FFFFFF", bold=True)
        header_fill = PatternFill(fill_type="solid", start_color=theme_color)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        content_fill = PatternFill(fill_type="solid", start_color="FFFFFF")
        content_font = Font(size=fontsize, name=font, color="000000")
        even_fill = PatternFill(fill_type="solid", start_color=self.calculate_rgba_color(self.theme_color, self.opacity, prefix=""))

        # 设置header样式
        for style in [header_style, header_left_style, header_middle_style, header_right_style]:
            style.font = header_font
            style.fill = header_fill

        header_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_left_style.alignment = alignment
        header_middle_style.alignment = alignment
        header_right_style.alignment = alignment

        header_style.border = self.itlubber_border(["medium", "medium", "medium", "medium"], [theme_color, theme_color, theme_color, theme_color], white=True)
        header_left_style.border = self.itlubber_border(["medium", "thin", "medium", "medium"], [theme_color, "FFFFFF", theme_color, theme_color], white=True)
        header_middle_style.border = self.itlubber_border(["thin", "thin", "medium", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color], white=True)
        header_right_style.border = self.itlubber_border(["thin", "medium", "medium", "medium"], ["FFFFFF", theme_color, theme_color, theme_color], white=True)

        # 设置last样式
        for style in [last_style, last_left_style, last_middle_style, last_right_style]:
            style.font = content_font
            style.fill = content_fill
            style.alignment = alignment

        last_style.border = self.itlubber_border(["medium", "medium", "medium"], [theme_color, theme_color, theme_color])
        last_left_style.border = self.itlubber_border(["medium", "thin", "medium"], [theme_color, "FFFFFF", theme_color])
        last_middle_style.border = self.itlubber_border(["thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color])
        last_right_style.border = self.itlubber_border(["thin", "medium", "medium"], ["FFFFFF", theme_color, theme_color])

        # 设置content样式
        for style in [content_style, left_style, middle_style, right_style]:
            style.font = content_font
            style.fill = content_fill
            style.alignment = alignment

        content_style.border = self.itlubber_border(["medium", "medium", "thin"], [theme_color, theme_color, theme_color])
        left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", theme_color])
        middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", theme_color])
        right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, theme_color])

        # 设置merge样式
        for style in [merge_style, merge_left_style, merge_middle_style, merge_right_style]:
            style.font = content_font
            style.fill = content_fill
            style.alignment = alignment

        merge_style.border = self.itlubber_border(["medium", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", "FFFFFF"])
        merge_middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, "FFFFFF"])

        # 设置first样式
        for style in [first_style, first_left_style, first_middle_style, first_right_style]:
            style.font = content_font
            style.fill = content_fill
            style.alignment = alignment

        first_style.border = self.itlubber_border(["medium", "medium", "thin", "medium"], [theme_color, theme_color, theme_color, theme_color])
        first_left_style.border = self.itlubber_border(["medium", "thin", "thin", "medium"], [theme_color, "FFFFFF", theme_color, theme_color])
        first_middle_style.border = self.itlubber_border(["thin", "thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color])
        first_right_style.border = self.itlubber_border(["thin", "medium", "thin", "medium"], ["FFFFFF", theme_color, theme_color, theme_color])

        # 创建奇偶行样式
        middle_odd_style = NamedStyle(name="middle_odd")
        middle_odd_first_style = NamedStyle(name="middle_odd_first")
        middle_odd_last_style = NamedStyle(name="middle_odd_last")
        middle_even_style = NamedStyle(name="middle_even")
        middle_even_first_style = NamedStyle(name="middle_even_first")
        middle_even_last_style = NamedStyle(name="middle_even_last")

        for style in [middle_odd_style, middle_odd_first_style, middle_odd_last_style, middle_even_style, middle_even_first_style, middle_even_last_style]:
            style.font = content_font
            style.alignment = alignment

        middle_odd_style.fill = content_fill
        middle_odd_first_style.fill = content_fill
        middle_odd_last_style.fill = content_fill
        middle_even_style.fill = even_fill
        middle_even_first_style.fill = even_fill
        middle_even_last_style.fill = even_fill

        middle_odd_first_style.border = Border(top=Side(border_style="medium", color=self.theme_color))
        middle_odd_last_style.border = Border(bottom=Side(border_style="medium", color=self.theme_color))
        middle_even_first_style.border = Border(top=Side(border_style="medium", color=self.theme_color))
        middle_even_last_style.border = Border(bottom=Side(border_style="medium", color=self.theme_color))
        middle_even_style.border = Border(bottom=Side(border_style="thin", color="FFFFFF"))
        middle_odd_style.border = Border(bottom=Side(border_style="thin", color="FFFFFF"))

        # 收集所有样式
        self.name_styles.extend([
            header_style, header_left_style, header_middle_style, header_right_style,
            last_style, last_left_style, last_middle_style, last_right_style,
            content_style, left_style, middle_style, right_style,
            merge_style, merge_left_style, merge_middle_style, merge_right_style,
            first_style, first_left_style, first_middle_style, first_right_style,
            middle_odd_style, middle_even_first_style, middle_odd_last_style, middle_even_style, middle_odd_first_style, middle_even_last_style,
        ])

    def save(self, filename: str, close: bool = True) -> None:
        """保存Excel文件。

        :param filename: 保存路径
        :param close: 是否关闭workbook，默认为True
        """
        # 移除样式模板sheet
        if self.style_sheet.title in self.workbook.sheetnames:
            self.workbook.remove(self.style_sheet)

        # 处理append模式
        if os.path.exists(filename) and self.mode == "append":
            _workbook = load_workbook(filename)

            for _sheet_name in _workbook.sheetnames:
                if _sheet_name not in self.workbook.sheetnames:
                    _worksheet = self.get_sheet_by_name(_sheet_name)

                    for i, row in enumerate(_workbook[_sheet_name].iter_rows()):
                        for j, cell in enumerate(row):
                            _worksheet.cell(row=i + 1, column=j + 1).value = cell.value
                            _worksheet.cell(row=i + 1, column=j + 1).style = cell.style

                            if i == _workbook[_sheet_name].max_row - 1:
                                _worksheet.column_dimensions[get_column_letter(j + 1)].width = _workbook[_sheet_name].column_dimensions[get_column_letter(j + 1)].width

            _workbook.close()

        # 创建目录
        if os.path.dirname(filename) != "" and not os.path.exists(os.path.dirname(filename)):
            os.makedirs(os.path.dirname(filename), exist_ok=True)

        # 保存文件
        self.workbook.save(filename)

        if close:
            self.workbook.close()


def dataframe2excel(
    data: pd.DataFrame,
    excel_writer: Union[str, ExcelWriter],
    sheet_name: Optional[str] = None,
    title: Optional[str] = None,
    header: bool = True,
    theme_color: str = "2639E9",
    condition_color: Optional[str] = None,
    fill: bool = True,
    percent_cols: Optional[List] = None,
    condition_cols: Optional[List] = None,
    custom_cols: Optional[List] = None,
    custom_format: str = "#,##0",
    color_cols: Optional[List] = None,
    percent_rows: Optional[List] = None,
    condition_rows: Optional[List] = None,
    custom_rows: Optional[List] = None,
    color_rows: Optional[List] = None,
    start_col: int = 2,
    start_row: int = 2,
    mode: str = "replace",
    figures: Optional[Union[str, List[str]]] = None,
    figsize: Tuple[int, int] = (600, 350),
    writer_params: Optional[Dict] = None,
    **kwargs
) -> Tuple[int, int]:
    """快速将DataFrame写入Excel。

    这是一个便捷函数，封装了ExcelWriter的常用操作。

    :param data: 需要保存的DataFrame
    :param excel_writer: 文件路径或ExcelWriter对象
    :param sheet_name: 工作表名称，默认为None
    :param title: 标题，默认为None
    :param header: 是否保存列名，默认为True
    :param theme_color: 主题颜色，默认为"2639E9"
    :param condition_color: 条件格式颜色，默认为None
    :param fill: 是否使用颜色填充，默认为True
    :param percent_cols: 需要显示为百分数的列，默认为None
    :param condition_cols: 需要显示数据条的列，默认为None
    :param custom_cols: 需要自定义格式的列，默认为None
    :param custom_format: 自定义格式，默认为"#,##0"
    :param color_cols: 需要显示颜色渐变的列，默认为None
    :param percent_rows: 需要显示为百分数的行，默认为None
    :param condition_rows: 需要显示数据条的行，默认为None
    :param custom_rows: 需要自定义格式的行，默认为None
    :param color_rows: 需要显示颜色渐变的行，默认为None
    :param start_col: 起始列，默认为2
    :param start_row: 起始行，默认为2
    :param mode: 写入模式，默认为"replace"
    :param figures: 需要插入的图片路径，默认为None
    :param figsize: 图片大小，默认为(600, 350)
    :param writer_params: ExcelWriter参数，默认为None
    :param kwargs: 其他参数，传递给insert_df2sheet
    :return: (下一行行号, 下一列列号)

    **参考样例**

    >>> import pandas as pd
    >>> from hscredit.report.excel import dataframe2excel
    >>>
    >>> # 创建示例数据
    >>> df = pd.DataFrame({
    ...     'feature': ['A', 'B', 'C'],
    ...     'iv': [0.1, 0.2, 0.3],
    ...     'ks': [0.3, 0.4, 0.5],
    ...     'rate': [0.05, 0.10, 0.15]
    ... })
    >>>
    >>> # 快速写入Excel
    >>> dataframe2excel(
    ...     df,
    ...     "report.xlsx",
    ...     sheet_name="特征分析",
    ...     title="特征统计表",
    ...     percent_cols=['rate'],  # 百分比格式
    ...     condition_cols=['iv', 'ks'],  # 条件格式
    ...     auto_width=True
    ... )
    """
    writer_params = writer_params or {}

    if isinstance(excel_writer, ExcelWriter):
        writer = excel_writer
    else:
        writer = ExcelWriter(theme_color=theme_color, mode=mode, **writer_params)

    if isinstance(sheet_name, Worksheet):
        worksheet = sheet_name
    else:
        worksheet = writer.get_sheet_by_name(sheet_name or "Sheet1")

    # 插入标题
    if title:
        col_width = len(data.columns) + data.index.nlevels if kwargs.get("index", False) else len(data.columns)
        start_row, end_col = writer.insert_value2sheet(
            worksheet, (start_row, start_col),
            value=title,
            style="header",
            end_space=(start_row, start_col + col_width - 1)
        )
        start_row += 1

    # 插入图片
    if figures is not None:
        if isinstance(figures, str):
            figures = [figures]

        pic_row = start_row
        for i, pic in enumerate(figures):
            if i == 0:
                start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, start_col), figsize=figsize)
            else:
                start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, end_col - 1), figsize=figsize)

    # 处理merge_column参数
    if "merge_column" in kwargs and kwargs["merge_column"]:
        if not isinstance(kwargs["merge_column"][0], (tuple, list)):
            kwargs["merge_column"] = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in kwargs["merge_column"]) or (not isinstance(c, tuple) and c in kwargs["merge_column"])]

    # 插入DataFrame
    end_row, end_col = writer.insert_df2sheet(
        worksheet, data, (start_row, start_col),
        fill=fill, header=header, **kwargs
    )

    # 设置百分比格式列
    if percent_cols:
        if not isinstance(percent_cols[0], (tuple, list)):
            percent_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in percent_cols) or (not isinstance(c, tuple) and c in percent_cols)]
        for c in [c for c in percent_cols if c in data.columns]:
            conditional_column = get_column_letter(
                start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c)
            )
            writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", "0.00%")

    # 设置自定义格式列
    if custom_cols:
        if not isinstance(custom_cols[0], (tuple, list)):
            custom_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in custom_cols) or (not isinstance(c, tuple) and c in custom_cols)]
        for c in [c for c in custom_cols if c in data.columns]:
            conditional_column = get_column_letter(
                start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c)
            )
            writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", custom_format)

    # 设置条件格式列
    if condition_cols:
        if not isinstance(condition_cols[0], (tuple, list)):
            condition_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in condition_cols) or (not isinstance(c, tuple) and c in condition_cols)]
        for c in [c for c in condition_cols if c in data.columns]:
            conditional_column = get_column_letter(
                start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c)
            )
            writer.add_conditional_formatting(
                worksheet,
                f'{conditional_column}{end_row - len(data)}',
                f'{conditional_column}{end_row - 1}',
                condition_color=condition_color or theme_color
            )

    # 设置颜色渐变列
    if color_cols:
        if not isinstance(color_cols[0], (tuple, list)):
            color_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in color_cols) or (not isinstance(c, tuple) and c in color_cols)]
        for c in [c for c in color_cols if c in data.columns]:
            try:
                rule = ColorScaleRule(
                    start_type='num', start_value=data[c].min(), start_color=condition_color or theme_color,
                    mid_type='num', mid_value=0., mid_color='FFFFFF',
                    end_type='num', end_value=data[c].max(), end_color=condition_color or theme_color
                )
                conditional_column = get_column_letter(
                    start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c)
                )
                worksheet.conditional_formatting.add(f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", rule)
            except Exception:
                import traceback
                traceback.print_exc()

    # 设置百分比格式行
    if percent_rows:
        if not isinstance(percent_rows[0], (tuple, list)):
            percent_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in percent_rows) or (not isinstance(c, tuple) and c in percent_rows)]
        for c in [c for c in percent_rows if c in data.index]:
            insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
            index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
            index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
            writer.set_number_format(worksheet, f"{get_column_letter(index_col)}{index_row}:{get_column_letter(index_col + len(data.columns))}{index_row}", "0.00%")

    # 设置自定义格式行
    if custom_rows:
        if not isinstance(custom_rows[0], (tuple, list)):
            custom_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in custom_rows) or (not isinstance(c, tuple) and c in custom_rows)]
        for c in [c for c in custom_rows if c in data.index]:
            insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
            index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
            index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
            writer.set_number_format(worksheet, f"{get_column_letter(index_col)}{index_row}:{get_column_letter(index_col + len(data.columns))}{index_row}", custom_format)

    # 设置条件格式行
    if condition_rows:
        if not isinstance(condition_rows[0], (tuple, list)):
            condition_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in condition_rows) or (not isinstance(c, tuple) and c in condition_rows)]
        for c in [c for c in condition_rows if c in data.index]:
            insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
            index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
            index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
            writer.add_conditional_formatting(
                worksheet,
                f'{get_column_letter(index_col)}{index_row}',
                f'{get_column_letter(index_col + len(data.columns))}{index_row}',
                condition_color=condition_color or theme_color
            )

    # 设置颜色渐变行
    if color_rows:
        if not isinstance(color_rows[0], (tuple, list)):
            color_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in color_rows) or (not isinstance(c, tuple) and c in color_rows)]
        for c in [c for c in color_rows if c in data.index]:
            try:
                insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
                rule = ColorScaleRule(
                    start_type='num', start_value=data.loc[c].min(), start_color=condition_color or theme_color,
                    mid_type='num', mid_value=0., mid_color='FFFFFF',
                    end_type='num', end_value=data.loc[c].max(), end_color=condition_color or theme_color
                )
                index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
                index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
                worksheet.conditional_formatting.add(f"{get_column_letter(index_col)}{index_row}:{get_column_letter(index_col + len(data.columns))}{index_row}", rule)
            except Exception:
                import traceback
                traceback.print_exc()

    # 保存文件（如果不是传入的ExcelWriter对象）
    if not isinstance(excel_writer, ExcelWriter) and not isinstance(sheet_name, Worksheet):
        writer.save(excel_writer)

    return end_row, end_col
