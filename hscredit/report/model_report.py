# -*- coding: utf-8 -*-
"""模型评估报告快速输出.

参考风控建模标准报告模板，提供多 Sheet 结构的模型报告，包括：
- 目录（带超链接）
- 基本信息（项目目标、样本统计、分月分布）
- 模型性能（KS/AUC/PSI、TOP n% LIFT、分月PSI、评分分箱）
- 入模变量重要性 & 分布
- 入模变量有效性分析（逐特征分箱表 + 金额口径 + PSI）
- 模型参数（评分卡详情）
- 模型部署需求
"""

from __future__ import annotations

import logging
import inspect
import re
import warnings
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd
from joblib.externals import cloudpickle

from ._sample_stats import build_group_distribution_table, build_sample_stats_table
from .model_explanation import build_model_explanation, explanation_to_dict, normalize_explain_config
from ..exceptions import SerializationError, ValidationError
from ..utils.parallel import (
    _ACTIVE_BUDGET,
    ParallelBudget,
    ParallelWorkload,
    parallel_execute,
    resolve_n_jobs,
    split_parallel_budget,
    validate_parallel_config,
)

logger = logging.getLogger(__name__)

_SUMMARY_PERCENT_METRICS = {"KS", "AUC", "坏样本率"}


@dataclass(frozen=True)
class _ModelComparisonParallelPlan:
    """模型比较外层和同时运行的数据集子任务预算。"""

    total_workers: int
    outer_workers: int
    child_workers: int
    has_parallel_children: bool


def _plan_model_comparison_parallel(
    n_jobs,
    model_task_count: int,
    dataset_task_count: int,
) -> _ModelComparisonParallelPlan:
    """在当前活动预算内按真实任务数规划模型比较并发。"""
    total_workers = resolve_n_jobs(n_jobs) or 1
    active_budget = _ACTIVE_BUDGET.get()
    if active_budget is not None:
        total_workers = min(total_workers, active_budget.available)
    total_workers = max(1, total_workers)
    model_task_count = max(0, int(model_task_count))
    dataset_task_count = max(0, int(dataset_task_count))
    has_parallel_children = model_task_count > 1 and dataset_task_count > 1 and total_workers > 1

    if has_parallel_children:
        outer_workers, child_budget = split_parallel_budget(
            total_workers,
            model_task_count,
            True,
        )
        child_workers = min(child_budget, dataset_task_count)
    elif model_task_count <= 1:
        outer_workers = 1
        child_workers = min(total_workers, max(1, dataset_task_count))
    else:
        outer_workers = min(total_workers, model_task_count)
        child_workers = 1

    return _ModelComparisonParallelPlan(
        total_workers=total_workers,
        outer_workers=max(1, outer_workers),
        child_workers=max(1, child_workers),
        has_parallel_children=has_parallel_children,
    )


def _execute_model_comparison_plan(function, tasks, plan, **kwargs):
    """执行模型外层计划，并在根调用建立等于请求总量的活动预算。"""
    token = None
    if _ACTIVE_BUDGET.get() is None:
        token = _ACTIVE_BUDGET.set(ParallelBudget(plan.total_workers, 0))
    try:
        return parallel_execute(
            function,
            tasks,
            n_jobs=plan.outer_workers,
            has_parallel_children=plan.has_parallel_children,
            **kwargs,
        )
    finally:
        if token is not None:
            _ACTIVE_BUDGET.reset(token)


# ---------------------------------------------------------------------------
# 内部工具
# ---------------------------------------------------------------------------


def _normalize_feature_names(feature_names) -> Optional[List[str]]:
    """将特征名入参统一规整为 Python ``str`` 列表.

    兼容 numpy 数组、``np.str_`` 标量、pandas Index 等输入，
    避免对 numpy 数组直接做真值判断（``if``/``or``）时报错。
    """
    if feature_names is None:
        return None
    if isinstance(feature_names, (np.ndarray, pd.Index)):
        feature_names = feature_names.tolist()
    return [str(name) for name in list(feature_names)]


def _ensure_dataframe(X, feature_names: Optional[List[str]] = None) -> pd.DataFrame:
    if isinstance(X, pd.DataFrame):
        out = X.copy()
        # numpy 字符串标量（np.str_）列名统一为 Python str，保证下游比较/索引一致
        if any(type(col) is not str and isinstance(col, str) for col in out.columns):
            out.columns = [str(col) if isinstance(col, str) else col for col in out.columns]
        return out
    arr = np.asarray(X)
    if arr.ndim == 1:
        arr = arr.reshape(-1, 1)
    cols = _normalize_feature_names(feature_names)
    if cols is None:
        cols = [f"feature_{i}" for i in range(arr.shape[1])]
    return pd.DataFrame(arr, columns=cols)


def _ensure_series(y, name: str = "target") -> pd.Series:
    if isinstance(y, pd.Series):
        out = y.copy()
        if out.name is None:
            out.name = name
        return out
    return pd.Series(np.asarray(y), name=name)


def _proba_pos(model, X) -> np.ndarray:
    """获取正类概率."""
    proba = np.asarray(model.predict_proba(X), dtype=float)
    if proba.ndim == 2 and proba.shape[1] >= 2:
        classes = getattr(model, "classes_", None)
        if classes is not None:
            positive = np.flatnonzero(np.asarray(classes) == 1)
            if len(positive) == 1:
                return proba[:, positive[0]]
        return proba[:, 1]
    return proba.reshape(-1)


def _safe_binary_metric(metric, y_true, y_score) -> float:
    """计算二分类指标，单类别样本返回 NaN。"""
    y_arr = np.asarray(y_true)
    if len(y_arr) == 0 or len(np.unique(y_arr[~pd.isna(y_arr)])) < 2:
        return np.nan
    try:
        return float(metric(y_arr, y_score))
    except (TypeError, ValueError):
        return np.nan


def _summary_percent_cols(columns) -> List[Any]:
    """识别 summary 表中需要按百分比显示的列."""
    percent_cols: List[Any] = []
    for col in columns:
        metric = col[0] if isinstance(col, tuple) and col else col
        if metric in _SUMMARY_PERCENT_METRICS:
            percent_cols.append(col)
    return percent_cols


def _prepare_report_dataset(task):
    """校验并准备单个报告数据集；worker 不执行模型方法。"""
    model, key, label, X, y, y_dict, feature_names = task
    required_features: Optional[List[str]] = None
    if hasattr(model, "feature_names_") and model.feature_names_ is not None:
        required_features = _normalize_feature_names(model.feature_names_)
    elif hasattr(model, "feature_names_in_") and model.feature_names_in_ is not None:
        required_features = _normalize_feature_names(model.feature_names_in_)
    feature_names = _normalize_feature_names(feature_names)

    if required_features:
        missing = set(required_features).difference(X.columns)
        if missing:
            raise ValueError(f"数据集缺少以下模型特征: {missing}")
        X_for_pred = X[required_features]
    elif feature_names:
        missing = set(feature_names).difference(X.columns)
        if missing:
            raise ValueError(f"数据集缺少以下模型特征: {missing}")
        X_for_pred = X[feature_names]
    else:
        X_for_pred = X

    if len(X) != len(y):
        raise ValueError(f"特征与标签样本数不一致: X={len(X)}, y={len(y)}")

    return key, label, X, y, y_dict, X_for_pred


def _normalize_prediction(model, result, expected_length: int, probability: bool) -> np.ndarray:
    """把唯一方法结果规整为与数据集等长的一维有限数组。"""
    values = np.asarray(result, dtype=float)
    if probability and values.ndim == 2 and values.shape[1] >= 2:
        classes = getattr(model, "classes_", None)
        positive_index = 1
        if classes is not None:
            positive = np.flatnonzero(np.asarray(classes) == 1)
            if len(positive) == 1:
                positive_index = int(positive[0])
        values = values[:, positive_index]
    elif values.ndim == 2 and values.shape[1] == 1:
        values = values[:, 0]
    elif values.ndim != 1:
        raise ValidationError("method 返回值必须是一维数组或单列二维数组")
    if len(values) != expected_length:
        raise ValidationError(f"method 返回值长度 {len(values)} 与数据集长度 {expected_length} 不一致")
    if not np.isfinite(values).all():
        raise ValidationError("method 返回值必须全部为有限数值")
    return values


def _invoke_named_prediction(model, X, method: str) -> np.ndarray:
    """调用一次已校验的模型方法。"""
    method_name = "predict_proba" if method == "predict_prob" else method
    result = getattr(model, method_name)(X)
    return _normalize_prediction(model, result, len(X), probability=method_name == "predict_proba")


def _build_report_dataset(task):
    """准备数据，并在字符串 method 路径于共享 worker 中执行唯一预测。"""
    *base_task, method = task
    prepared = _prepare_report_dataset(tuple(base_task))
    if method is None:
        return prepared
    key, label, X, y, y_dict, X_for_pred = prepared
    prediction = _invoke_named_prediction(base_task[0], X_for_pred, method)
    return ReportDataset(name=key, label=label, X=X, y=y, prediction=prediction, y_dict=y_dict)


def _binary_metric_worker(task) -> Tuple[float, float, int, float]:
    """基于已缓存预测计算单个数据集/标签的确定性指标。"""
    from ..core.metrics import auc, ks

    y_true, y_proba = task
    y_arr = np.asarray(y_true, dtype=float)
    proba_arr = np.asarray(y_proba)
    valid = np.isfinite(y_arr)
    y_arr = y_arr[valid]
    proba_arr = proba_arr[valid]
    bad_rate = float(y_arr.mean()) if len(y_arr) else np.nan
    return (
        _safe_binary_metric(ks, y_arr, proba_arr),
        _safe_binary_metric(auc, y_arr, proba_arr),
        len(y_arr),
        bad_rate,
    )


def _psi_metric_worker(task) -> float:
    """计算一对评分分布 PSI。"""
    from ..core.metrics import psi

    expected, actual = task
    return float(psi(expected, actual))


def _feature_metric_worker(task) -> Tuple[float, float, float]:
    """计算单特征 IV/KS/PSI；不读取或修改报告实例状态。"""
    from ..core.metrics import iv, ks, psi

    y_true, train_values, test_values = task

    def optional_metric(metric, *values):
        try:
            return float(metric(*values))
        except (TypeError, ValueError, ZeroDivisionError):
            return np.nan

    iv_value = optional_metric(iv, y_true, train_values)
    ks_value = optional_metric(ks, y_true, train_values)
    psi_value = optional_metric(psi, train_values, test_values) if test_values is not None else np.nan
    return iv_value, ks_value, psi_value


def _compare_model_worker(task) -> Tuple[str, pd.DataFrame]:
    """构建单模型摘要；用于 ``compare_models`` 的模型外层并行。"""
    name, model, X_train, y_train, X_test, y_test, n_jobs, backend, config = task
    report = ModelReport(
        model=model,
        X_train=X_train,
        y_train=y_train,
        X_test=X_test,
        y_test=y_test,
        n_jobs=n_jobs,
        parallel_backend=backend,
        parallel_config=config,
    )
    return name, report.summary()


def _safe_close_figs():
    """安全关闭 matplotlib 图形以释放内存."""
    try:
        import matplotlib.pyplot as plt

        plt.close("all")
    except Exception:
        pass


def _safe_close_plot_result(result) -> None:
    """在线程内仅关闭当前绘图调用返回的画布，不误关其他线程图表."""
    try:
        import matplotlib.pyplot as plt

        candidate = result[0] if isinstance(result, tuple) and result else result
        if hasattr(candidate, "figure"):
            candidate = candidate.figure
        if hasattr(candidate, "savefig"):
            plt.close(candidate)
    except Exception:
        pass


@contextmanager
def _threaded_agg_rendering():
    """让当前报告工作线程使用独立 Agg 画布."""
    from ..core.viz.utils import _threaded_agg_rendering as rendering_context

    with rendering_context():
        yield


def _execute_plot_group(group):
    """执行一个完整图表分组；分组仅在线程后端中提交."""
    with _threaded_agg_rendering():
        return group()


def _merge_multi_label_bin_tables(tables: List[pd.DataFrame], labels: List[str]) -> pd.DataFrame:
    """将多个单标签分箱表合并为 MultiIndex 列的合并表。

    列结构：分箱详情列保持不变，其余列变为 (标签名, 列名) 的 MultiIndex。
    """
    if not tables:
        return pd.DataFrame()
    if len(tables) == 1:
        return tables[0]

    base = tables[0].copy()
    merge_cols = ["分箱标签", "指标含义", "指标名称", "指标明细"]
    available_merge = [c for c in merge_cols if c in base.columns]
    other_cols = [c for c in base.columns if c not in available_merge]

    # 重排列：合并列在前，明细列在后
    base = base[available_merge + other_cols]

    # 当没有可合并列时（cross-product 陷阱），改用 index 对齐拼接
    if not available_merge:
        result_parts = [base]
        for table, lbl in zip(tables[1:], labels[1:]):
            table_copy = table.copy()
            table_other = [c for c in table_copy.columns if c not in available_merge]
            table_multi_cols = [("分箱详情", c) for c in available_merge] + [(lbl, c) for c in table_other]
            table_copy.columns = pd.MultiIndex.from_tuples(table_multi_cols)
            result_parts.append(table_copy[table_other])
        result = pd.concat(result_parts, axis=1, keys=[labels[0]] + list(labels[1:]), names=["标签"])
        return result

    # 构建 MultiIndex 列
    multi_cols = []
    for col in base.columns:
        if col in available_merge:
            multi_cols.append(("分箱详情", col))
        else:
            multi_cols.append((labels[0], col))
    base.columns = pd.MultiIndex.from_tuples(multi_cols)

    for table, lbl in zip(tables[1:], labels[1:]):
        table_copy = table.copy()
        table_multi_cols = []
        for col in table_copy.columns:
            if col in available_merge:
                table_multi_cols.append(("分箱详情", col))
            else:
                table_multi_cols.append((lbl, col))
        table_copy.columns = pd.MultiIndex.from_tuples(table_multi_cols)
        merge_on = [("分箱详情", c) for c in available_merge]
        base = base.merge(table_copy, on=merge_on)

    return base


def _drop_bin_meta_cols(table: pd.DataFrame) -> pd.DataFrame:
    """删除分箱表中的 ``指标名称`` / ``指标含义`` 列（兼容扁平列与 MultiIndex 列）。"""
    drop_names = {"指标名称", "指标含义"}
    if isinstance(table.columns, pd.MultiIndex):
        keep = [c for c in table.columns if c[-1] not in drop_names]
    else:
        keep = [c for c in table.columns if c not in drop_names]
    return table[keep]


def _next_image_col(ws, start_col: int, width_px: int, default_width: float = 13.0) -> int:
    """根据图片像素宽度与各列实际列宽，计算紧邻图片右侧的下一列号。

    避免使用 ``insert_pic2sheet`` 固定 +8 列的返回值导致图片之间出现多余空列，
    使多张图片首尾紧挨着排列。
    """
    from openpyxl.utils import get_column_letter

    acc = 0.0
    col = start_col
    # 防御性上限，避免列宽异常时陷入死循环
    while acc < width_px and col < start_col + 60:
        dim = ws.column_dimensions.get(get_column_letter(col))
        width = dim.width if (dim is not None and dim.width) else default_width
        acc += width * 7.0 + 5.0
        col += 1
    return col


# ---------------------------------------------------------------------------
# 数据容器
# ---------------------------------------------------------------------------


@dataclass
class ReportDataset:
    name: str
    label: str  # 中文标签: "训练集" / "测试集" / "OOT"
    X: pd.DataFrame
    y: pd.Series
    prediction: np.ndarray
    y_dict: Optional[Dict[str, np.ndarray]] = None  # {label_name: y_array}，多标签场景下各标签的独立标签

    def __post_init__(self):
        self.y_proba = self.prediction
        self.score = self.prediction


# ---------------------------------------------------------------------------
# ModelReport
# ---------------------------------------------------------------------------


class ModelReport:
    """面向报表输出的快速模型报告封装.

    参考风控建模标准报告模板，对已训练模型一站式生成多 Sheet 结构的
    Excel / HTML 报告，并提供各分项结果的获取方法（指标、分箱表、特征重要性、
    描述统计、相关性等）。支持任意多个数据集（训练/测试/OOT…）的横向对比，
    以及 ``overdue`` + ``dpds`` 的多逾期标签构建。

    **参数**

    （完整说明见 :meth:`__init__`）

    :param model: 已训练好的模型，需实现 ``predict`` / ``predict_proba``
    :param datasets: 数据集字典或列表（推荐），如 ``{'train': df, 'test': df}``；
        DataFrame 需含目标列或配合 ``overdue``/``dpds`` 构建标签
    :param X_train/y_train/X_test/y_test: 兼容 sklearn 风格的数据传入方式
    :param target: 目标列名（sklearn/scorecardpipeline 风格）
    :param overdue: 逾期天数列名或列表，配合 ``dpds`` 自动构建 0/1 标签
    :param dpds: 逾期定义天数或列表（逾期天数 > dpds 记为坏样本）
    :param del_grey: 是否按每个逾期标签独立剔除 ``(0, dpd]`` 灰样本
    :param feature_names: 特征名称列表，可选；None 时自动从模型 feature_names_ / feature_names_in_ 获取

    **属性**

    - model: 传入的已训练模型
    - feature_names: 最终使用的特征名称列表
    - _datasets: 解析后的各数据集（key -> :class:`ReportDataset`）

    **参考样例**

    >>> from hscredit.report import ModelReport
    >>> report = ModelReport(model, datasets={'train': train_df, 'test': test_df})
    >>> report.get_metrics()            # 各数据集指标对比
    >>> report.get_feature_importance(top_n=20)
    >>> report.to_excel('模型报告.xlsx')  # 导出多 Sheet 报告
    """

    _PERCENT_COLS = [
        "样本占比",
        "好样本占比",
        "坏样本占比",
        "坏样本率",
        "LIFT值",
        "坏账改善",
        "累积LIFT值",
        "累积坏账改善",
        "分档KS值",
    ]
    _CONDITION_COLS = ["坏样本率", "LIFT值", "累积LIFT值"]

    def __init__(
        self,
        model,
        X_train=None,
        y_train=None,
        X_test=None,
        y_test=None,
        X_oot=None,
        y_oot=None,
        feature_names: Optional[List[str]] = None,
        target: Optional[Union[str, Dict]] = None,
        datasets: Optional[Union[List, Dict]] = None,
        overdue: Optional[Union[str, List[str]]] = None,
        dpds: Optional[Union[int, float, List[Union[int, float]]]] = None,
        del_grey: bool = False,
        method: Union[str, Callable] = "predict_proba",
        method_kwargs: Optional[Dict[str, Any]] = None,
        explain_config: Optional[Dict[str, Any]] = None,
        n_jobs=-1,
        parallel_backend: Optional[str] = None,
        parallel_config: Optional[Dict[str, Any]] = None,
        **kwargs,
    ):
        """初始化模型报告.

        数据集传入支持三种方式，内部统一规整为 ``{数据集名称: (X, y)}`` 的 dict 结构：

        1. datasets 为 dict：直接以 key 作为数据集名称，
           如 ``{'训练集': df, 'OOT': df}`` 或 ``{'train': (X, y), 'test': (X, y)}``
        2. datasets 为 list：依次命名为 数据集1、数据集2、...、数据集N
        3. X_train/X_test/X_oot 参数：依次命名为 训练集、测试集、跨时间验证集

        标签（y）解析遵循 hscredit 统一的两种传参风格：

        - sklearn 风格：显式传入 y（如 ``X_train=X, y_train=y`` 或
          ``datasets={'train': (X, y)}``），y 优先使用
        - scorecardpipeline 风格：数据全部在 X 中，通过 ``target='列名'`` 提取标签
        - overdue + dpds 组合：传入后直接忽略 target，按 逾期天数 > 阈值 构建标签

        示例::

            # 方式1: datasets dict（key 即数据集名称，X 中含目标列）
            report = ModelReport(model, datasets={'训练集': train_df, 'OOT': oot_df}, target='target')

            # 方式2: datasets list（自动命名为 数据集1、数据集2）
            report = ModelReport(model, datasets=[train_df, test_df], target='target')

            # overdue/dpds 自动构建标签（X 中不含目标列，忽略 target）
            report = ModelReport(
                model,
                datasets={'建模集': df},
                overdue='dpds',     # 逾期天数列名
                dpds=[15, 7, 0],    # 任一 MOB 下 DPD > threshold 则 y=1
            )

            # 方式3: 兼容 sklearn API（显式传入 y，y 优先于 target）
            report = ModelReport(model, X_train=X, y_train=y, X_test=X_val, y_test=y_val)

            # 方式3: 含跨时间验证集
            report = ModelReport(model, X_train=X, y_train=y, X_oot=X_oot, y_oot=y_oot)

        :param model: 训练好的模型（ScoreCard / XGBoost / LightGBM / sklearn 等）
        :param datasets: 数据集字典/列表（推荐方式），dict 的 key 直接作为数据集名称
        :param X_train: 训练集特征（命名为 训练集）
        :param y_train: 训练集标签，None 时从 X_train 中通过 target / overdue+dpds 构建
        :param X_test: 测试集特征（命名为 测试集）
        :param y_test: 测试集标签，None 时从 X_test 中通过 target / overdue+dpds 构建
        :param X_oot: 跨时间验证集特征（命名为 跨时间验证集）
        :param y_oot: 跨时间验证集标签，None 时从 X_oot 中通过 target / overdue+dpds 构建
        :param feature_names: 特征名称列表，可选；None 时自动从模型 feature_names_ / feature_names_in_ 获取
        :param target: 目标列配置
            - str: 列名，如 'target'，数据全部在 X 中时使用
            - dict: {'overdue': col, 'dpds': threshold} 或 {'overdue': col, 'dpds': [15, 7, 0]}
        :param overdue: 逾期列名（str）或多个列名（List[str]），传入后忽略 target
        :param dpds: 逾期天数阈值（int/float）或多个阈值（List），与 overdue 配合使用
        :param del_grey: overdue 模式下是否按各 DPD 独立剔除 ``(0, DPD]`` 灰样本
        :param method: 数据集唯一预测方法，支持 predict_proba/predict_prob/predict/predict_score/transform/callable
        :param method_kwargs: callable 同名参数的显式覆盖字典
        :param n_jobs: 并行工作数；-1 自动保留 CPU，None 使用兼容串行模式
        :param parallel_backend: joblib 后端，如 ``threading`` 或 ``loky``
        :param parallel_config: joblib 其他并行配置，保留调用者字典引用
        :param kwargs: 透传给 callable 的额外同名参数
        """
        self.model = model
        self.explain_config = normalize_explain_config(explain_config)
        self._feature_names = _normalize_feature_names(feature_names)
        if not isinstance(method, str) and not callable(method):
            raise ValidationError("method 必须是方法名字符串或 callable")
        if isinstance(method, str):
            normalized_method = "predict_proba" if method == "predict_prob" else method
            allowed_methods = {"predict_proba", "predict", "predict_score", "transform"}
            if normalized_method not in allowed_methods:
                raise ValidationError(f"不支持的 method: {method!r}，可选: {sorted(allowed_methods)}")
            if not callable(getattr(model, normalized_method, None)):
                raise ValidationError(f"模型 {type(model).__name__} 不支持方法 {normalized_method}")
        if method_kwargs is not None and not isinstance(method_kwargs, dict):
            raise ValidationError("method_kwargs 必须是字典或 None")
        self.method = method
        self.method_kwargs = dict(method_kwargs or {})
        self.kwargs = dict(kwargs)
        self.method_source_ = None
        if callable(method):
            try:
                self.method_source_ = inspect.getsource(method).strip()
            except (OSError, TypeError):
                pass
        self.n_jobs = n_jobs
        self.parallel_backend = parallel_backend
        self.parallel_config = parallel_config
        self.del_grey = bool(del_grey)
        self.init_params_ = {
            "model": model,
            "datasets": datasets,
            "X_train": X_train,
            "y_train": y_train,
            "X_test": X_test,
            "y_test": y_test,
            "X_oot": X_oot,
            "y_oot": y_oot,
            "feature_names": feature_names,
            "target": target,
            "overdue": overdue,
            "dpds": dpds,
            "del_grey": del_grey,
            "method": method,
            "method_kwargs": method_kwargs,
            "n_jobs": n_jobs,
            "parallel_backend": parallel_backend,
            "parallel_config": parallel_config,
            "explain_config": explain_config,
        }

        # 在任何预测任务启动前完成公共配置校验；保留调用者传入字典的对象身份。
        validate_parallel_config(parallel_backend, parallel_config)
        resolve_n_jobs(n_jobs, task_count=1)

        # 传入 overdue（配合 dpds）时直接忽略 target，按 逾期天数 > 阈值 构建标签
        if overdue is not None:
            self._target_cfg: Optional[Union[str, Dict]] = {
                "overdue": overdue,
                "dpds": dpds,
            }
        else:
            self._target_cfg = target

        # 构建数据集
        self._datasets: Dict[str, ReportDataset] = {}
        self._datasets_info: Dict[str, str] = {}  # key -> label

        # 多标签指标名列表（用于多标签报告），按构建顺序从第一个数据集获取
        # 当 overdue/dpds 产生多指标时填充，如 ['dpds_m1>15', 'dpds_m1>7', 'dpds_m1>0', ...]
        self._label_names: List[str] = []

        # 所有缓存仅在完整计算成功后提交。
        self._metrics_cache: Dict[Optional[str], pd.DataFrame] = {}
        self._summary_cache: Optional[pd.DataFrame] = None
        self._raw_importance_cache: Optional[pd.Series] = None
        self._importance_cache: Optional[pd.DataFrame] = None
        self._features_describe_cache: Optional[pd.DataFrame] = None
        self._corr_cache: Optional[pd.DataFrame] = None
        self._bin_table_cache: Dict[Tuple[Any, ...], pd.DataFrame] = {}
        self._feature_bin_table_cache: Dict[Tuple[Any, ...], pd.DataFrame] = {}
        self._lift_table_cache: Dict[Tuple[Any, ...], pd.DataFrame] = {}
        self._monthly_metrics_cache: Dict[str, pd.DataFrame] = {}
        self._monthly_psi_cache: Dict[str, pd.DataFrame] = {}
        self._features_summary_cache: Optional[pd.DataFrame] = None
        self._model_explanation_cache: Optional[Dict[str, Any]] = None

        # 确定目标列名
        self._target_name = self._resolve_target_name(self._target_cfg)

        # 将各种传入形式统一规整为 {数据集名称: (X, y)}，再基于该结构构建数据集
        normalized = self._normalize_datasets(
            datasets=datasets,
            X_train=X_train,
            y_train=y_train,
            X_test=X_test,
            y_test=y_test,
            X_oot=X_oot,
            y_oot=y_oot,
        )
        self._init_from_normalized(normalized)

        # 解析最终使用的特征名列表
        # 优先级：显式传入 feature_names > 模型自带 feature_names_ / feature_names_in_ > 数据集列名
        model_required: Optional[List[str]] = None
        if hasattr(self.model, "feature_names_") and self.model.feature_names_ is not None:
            model_required = _normalize_feature_names(self.model.feature_names_)
        elif hasattr(self.model, "feature_names_in_") and self.model.feature_names_in_ is not None:
            model_required = _normalize_feature_names(self.model.feature_names_in_)

        if self._feature_names:
            # 显式传入：按模型实际入模特征过滤（排除 MOB1、放款金额 等非入模字段），保留传入顺序
            self.feature_names = list(self._feature_names)
            if model_required:
                self.feature_names = [f for f in self.feature_names if f in model_required]
        elif model_required:
            # 未显式传入：直接从模型获取入模特征（顺序以模型为准）
            self.feature_names = list(model_required)
        elif self._datasets:
            first_ds = next(iter(self._datasets.values()))
            self.feature_names = list(first_ds.X.columns)
        else:
            self.feature_names = []

    def _invalidate_caches(self) -> None:
        """清除所有依赖数据集的派生结果。"""
        self._metrics_cache.clear()
        self._summary_cache = None
        self._raw_importance_cache = None
        self._importance_cache = None
        self._features_describe_cache = None
        self._corr_cache = None
        self._bin_table_cache.clear()
        self._feature_bin_table_cache.clear()
        self._lift_table_cache.clear()
        self._monthly_metrics_cache.clear()
        self._monthly_psi_cache.clear()
        self._features_summary_cache = None
        self._model_explanation_cache = None

    def _run_cache_transaction(self, function, *args, **kwargs):
        """隔离派生缓存写入；仅在整个公共输出操作成功后提交。"""
        cache_names = (
            "_metrics_cache",
            "_summary_cache",
            "_raw_importance_cache",
            "_importance_cache",
            "_features_describe_cache",
            "_corr_cache",
            "_bin_table_cache",
            "_feature_bin_table_cache",
            "_lift_table_cache",
            "_monthly_metrics_cache",
            "_monthly_psi_cache",
            "_features_summary_cache",
            "_model_explanation_cache",
        )
        original = {name: getattr(self, name) for name in cache_names}
        for name, value in original.items():
            # 映射缓存使用独立容器，避免失败尝试污染调用方持有的旧引用。
            setattr(self, name, dict(value) if isinstance(value, dict) else value)
        try:
            return function(*args, **kwargs)
        except Exception:
            for name, value in original.items():
                setattr(self, name, value)
            raise

    def __getstate__(self):
        """把 callable method 转为 cloudpickle 载荷，避免执行源码字符串。"""
        state = dict(self.__dict__)
        if callable(self.method):
            try:
                state["_method_payload"] = cloudpickle.dumps(self.method)
            except Exception as exc:
                raise SerializationError(f"method callable 无法序列化: {exc}") from exc
            state["method"] = None
            init_params = dict(state.get("init_params_", {}))
            init_params["method"] = None
            state["init_params_"] = init_params
        return state

    def __setstate__(self, state):
        """从可信制品恢复 callable method。"""
        payload = state.pop("_method_payload", None)
        self.__dict__.update(state)
        if payload is not None:
            try:
                self.method = cloudpickle.loads(payload)
            except Exception as exc:
                raise SerializationError(f"method callable 无法恢复: {exc}") from exc
            self.init_params_["method"] = self.method

    def _call_custom_method(self, X: pd.DataFrame):
        """按签名同名注入上下文并严格调用 callable 一次。"""
        signature = inspect.signature(self.method)
        available: Dict[str, Any] = dict(self.init_params_)
        for name, value in vars(self).items():
            available.setdefault(name, value)
        available.update(self.kwargs)
        available.update(self.method_kwargs)
        available.update({"self": self, "report": self, "x": X, "X": X})

        args: List[Any] = []
        call_kwargs: Dict[str, Any] = {}
        consumed = set()
        missing = []
        accepts_kwargs = False
        for parameter in signature.parameters.values():
            if parameter.kind == inspect.Parameter.VAR_POSITIONAL:
                continue
            if parameter.kind == inspect.Parameter.VAR_KEYWORD:
                accepts_kwargs = True
                continue
            if parameter.name in available:
                consumed.add(parameter.name)
                if parameter.kind == inspect.Parameter.POSITIONAL_ONLY:
                    args.append(available[parameter.name])
                else:
                    call_kwargs[parameter.name] = available[parameter.name]
            elif parameter.default is inspect.Parameter.empty:
                missing.append(parameter.name)

        if missing:
            raise ValidationError(
                f"method callable 缺少必填参数 {missing}，可用参数: {sorted(available)}"
            )
        if accepts_kwargs:
            call_kwargs.update({name: value for name, value in available.items() if name not in consumed})
        return self.method(*args, **call_kwargs)

    def _normalize_method_result(self, result, X: pd.DataFrame, probability: bool) -> np.ndarray:
        """把唯一方法结果规整为与数据集等长的一维有限数组。"""
        return _normalize_prediction(self.model, result, len(X), probability=probability)

    def _calculate_prediction(self, X: pd.DataFrame) -> np.ndarray:
        """调用一次指定 method 并返回唯一数据集结果。"""
        if callable(self.method):
            return self._normalize_method_result(self._call_custom_method(X), X, probability=False)

        return _invoke_named_prediction(self.model, X, self.method)

    def _dataset_from_prepared(self, prepared) -> ReportDataset:
        """在主进程调用 method，并在成功后构建数据集。"""
        key, label, X, y, y_dict, X_for_pred = prepared
        prediction = self._calculate_prediction(X_for_pred)
        return ReportDataset(
            name=key,
            label=label,
            X=X,
            y=y,
            prediction=prediction,
            y_dict=y_dict,
        )

    def _commit_dataset_specs(self, specs: List[Tuple[Any, ...]]) -> None:
        """并行计算一批数据集，并在全部成功后按输入顺序一次性提交。"""
        total_rows = sum(len(spec[3]) for spec in specs)
        max_columns = max((spec[3].shape[1] for spec in specs), default=1)
        data_bytes = sum(int(spec[3].memory_usage(deep=True).sum()) for spec in specs)
        worker_method = None if callable(self.method) else self.method
        tasks = [(*spec, worker_method) for spec in specs]
        prepared = parallel_execute(
            _build_report_dataset,
            tasks,
            n_jobs=self.n_jobs,
            parallel_backend=self.parallel_backend,
            parallel_config=self.parallel_config,
            task_labels=[spec[1] for spec in specs],
            default_backend="loky",
            has_parallel_children=True,
            workload=ParallelWorkload(
                task_count=len(specs),
                rows=total_rows,
                columns=max_columns,
                data_bytes=data_bytes,
                cost_per_item=5.0,
                capability="process_safe",
                has_parallel_children=True,
                operation="模型报告数据集预测",
            ),
        )
        results = (
            [self._dataset_from_prepared(item) for item in prepared]
            if callable(self.method)
            else prepared
        )
        self._datasets = {dataset.name: dataset for dataset in results}
        self._datasets_info = {dataset.name: dataset.label for dataset in results}

    def _resolve_target_name(self, target) -> str:
        """解析目标配置，返回标签列名.

        overdue/dpds 作为单独参数传入时，target 参数将被忽略，
        标签列名默认为 'target'。
        """
        if isinstance(target, str):
            return target
        if isinstance(target, dict) and "overdue" in target:
            return target.get("label", "target")
        return "target"

    def _build_y(self, X: pd.DataFrame, target_cfg) -> Tuple[pd.Series, Optional[Dict[str, np.ndarray]]]:
        """根据 target 配置从 X 构建 y 标签.

        支持三种配置：
        - None: 从 X 中查找 'target' 列
        - str: 直接取 X[target] 作为标签
        - dict: 联合构建标签
            - 单逾期列:  target={'overdue': col, 'dpds': threshold} 或
                        target={'overdue': col, 'dpds': [t1, t2, ...]}
            - 多逾期列:  target={'overdue': [col1, col2], 'dpds': [t1, t2, ...]}
                          每列 × 每阈值生成指标，任一为真则 label=1

        返回 (y_series, y_dict)：
        - y_series: 聚合标签（任一指标为真则 y=1）
        - y_dict: 多标签场景下各指标的独立标签数组（key 如 'dpds_m1>15'）
                  单标签或非多指标场景下返回 None
        """
        if target_cfg is None:
            for col in ("target", "label", "y", "flag", "overdue"):
                if col in X.columns:
                    return _ensure_series(X[col], name="target"), None
            raise ValueError("未找到目标列（target），请通过 target 参数指定标签列名，" "或传入 dict={'overdue': col, 'dpds': threshold} 联合构建")

        if isinstance(target_cfg, str):
            if target_cfg in X.columns:
                return _ensure_series(X[target_cfg], name=target_cfg), None
            raise ValueError(f"目标列 '{target_cfg}' 不存在于数据中")

        if isinstance(target_cfg, dict) and "overdue" in target_cfg:
            overdue_cols = target_cfg["overdue"]
            dpds_vals = target_cfg.get("dpds")
            threshold = target_cfg.get("threshold")
            label_name = target_cfg.get("label", "target")

            # 统一为列表
            if isinstance(overdue_cols, str):
                overdue_cols = [overdue_cols]

            # 支持旧格式 threshold 键，或新格式 dpds 作为阈值
            # 旧格式: {'overdue': col, 'dpds': col, 'threshold': 3}
            # 新格式: {'overdue': col, 'dpds': [15, 7, 0]}
            if threshold is not None:
                # 旧格式：dpds 为列名，threshold 为阈值
                dpds_col = dpds_vals if isinstance(dpds_vals, str) else None
                thresholds = [threshold]
            elif dpds_vals is not None:
                if isinstance(dpds_vals, (int, float)):
                    dpds_vals = [dpds_vals]
                thresholds = dpds_vals
                dpds_col = None
            else:
                # 只有 overdue，无 dpds/threshold：overdue 列值 > 0 → y=1
                thresholds = [0]
                dpds_col = None

            # 验证列名
            for col in overdue_cols:
                if col not in X.columns:
                    raise ValueError(f"逾期列 '{col}' 不存在，请检查列名")

            # 每列 × 每阈值，生成全指标
            indicators = pd.DataFrame(index=X.index, dtype=float)
            for col in overdue_cols:
                for t in thresholds:
                    if dpds_col is not None and dpds_col in X.columns:
                        # dpds 列 > threshold
                        overdue_values = X[dpds_col]
                    else:
                        # col 列 > threshold
                        overdue_values = X[col]
                    indicator = (overdue_values > t).astype(float)
                    if self.del_grey:
                        valid = (overdue_values == 0) | (overdue_values > t)
                        indicator = indicator.where(valid, np.nan)
                    indicators[f"{col}>{t}"] = indicator

            # 单逾期标签保留 NaN 灰样本标记，数据集初始化阶段会连同特征一起剔除；
            # 多标签则保留完整行并在各标签数组中分别标记，确保每个 DPD 可独立过滤。
            if self.del_grey and indicators.shape[1] == 1:
                y = indicators.iloc[:, 0].rename(label_name)
            else:
                y = indicators.fillna(0).astype(bool).any(axis=1).astype(int)
            # 多指标时返回各指标独立标签，供多标签报告使用
            y_dict: Optional[Dict[str, np.ndarray]] = None
            if len(overdue_cols) > 1 or (isinstance(dpds_vals, list) and len(dpds_vals) > 1):
                y_dict = {col: indicators[col].to_numpy(dtype=float) for col in indicators.columns}
            return _ensure_series(y, name=label_name), y_dict

        raise ValueError(f"target 参数格式错误：{target_cfg}")

    def _normalize_datasets(
        self,
        datasets=None,
        X_train=None,
        y_train=None,
        X_test=None,
        y_test=None,
        X_oot=None,
        y_oot=None,
    ) -> Dict[str, Tuple[Any, Any]]:
        """将各种传入形式统一规整为 ``{数据集名称: (X, y)}`` 的 dict 结构.

        命名规则：

        - datasets 为 dict：直接以 key 作为数据集名称；
          value 支持 ``DataFrame``（y 为 None，自动构建标签）或 ``(X, y)`` 元组
        - datasets 为 list/tuple：依次命名为 数据集1、数据集2、...、数据集N
        - X_train/X_test/X_oot 参数：依次命名为 训练集、测试集、跨时间验证集；
          y 为 None 时从 X 中通过 target / overdue+dpds 自动构建标签
        """
        normalized: Dict[str, Tuple[Any, Any]] = {}

        if datasets is not None:
            if isinstance(datasets, dict):
                entries = list(datasets.items())
            elif isinstance(datasets, (list, tuple)):
                entries = [(f"数据集{i + 1}", value) for i, value in enumerate(datasets)]
            else:
                raise ValueError("datasets 必须为字典或列表")
            for name, value in entries:
                if isinstance(value, (tuple, list)) and len(value) >= 2:
                    normalized[str(name)] = (value[0], value[1])
                else:
                    normalized[str(name)] = (value, None)
            return normalized

        if X_train is not None:
            normalized["训练集"] = (X_train, y_train)
        if X_test is not None:
            normalized["测试集"] = (X_test, y_test)
        if X_oot is not None:
            normalized["跨时间验证集"] = (X_oot, y_oot)
        if not normalized:
            raise ValueError("未提供任何数据集，请通过 datasets 或 X_train/y_train 等参数传入数据")
        return normalized

    def _init_from_normalized(self, normalized: Dict[str, Tuple[Any, Any]]):
        """基于统一的 ``{数据集名称: (X, y)}`` 结构构建各数据集.

        标签解析规则（y 优先，其次 target，最后 overdue+dpds）：

        - y 不为 None（sklearn 风格）：直接使用传入的 y
        - y 为 None（scorecardpipeline 风格）：从 X 中通过 target 列名提取，
          或通过 overdue+dpds 组合构建标签
        """
        specs: List[Tuple[Any, ...]] = []
        candidate_labels: List[str] = []
        for name, (X_raw, y_raw) in normalized.items():
            X_df = _ensure_dataframe(X_raw, feature_names=self._feature_names)
            if y_raw is None:
                y_s, y_dict = self._build_y(X_df, self._target_cfg)
                if self.del_grey and self._is_overdue_cfg() and not y_dict:
                    valid = y_s.notna()
                    X_df = X_df.loc[valid].copy()
                    y_s = y_s.loc[valid].astype(int)
            else:
                y_s = _ensure_series(y_raw, name=self._target_name)
                y_dict = None
            if y_dict and not candidate_labels:
                candidate_labels = list(y_dict)
            specs.append((self.model, name, name, X_df, y_s, y_dict, self._feature_names))

        self._commit_dataset_specs(specs)
        self._label_names = candidate_labels

    # ---------- 数据集管理 ----------

    def _add_dataset(
        self,
        key: str,
        label: str,
        X: pd.DataFrame,
        y: pd.Series,
        y_dict: Optional[Dict[str, np.ndarray]] = None,
    ):
        spec = (self.model, key, label, X, y, y_dict, self._feature_names)
        worker_method = None if callable(self.method) else self.method
        prepared = parallel_execute(
            _build_report_dataset,
            [(*spec, worker_method)],
            n_jobs=self.n_jobs,
            parallel_backend=self.parallel_backend,
            parallel_config=self.parallel_config,
            task_labels=[key],
            default_backend="loky",
            has_parallel_children=True,
            workload=ParallelWorkload(
                task_count=1,
                rows=len(X),
                columns=X.shape[1],
                data_bytes=int(X.memory_usage(deep=True).sum()),
                cost_per_item=5.0,
                capability="process_safe",
                has_parallel_children=True,
                operation="模型报告新增数据集预测",
            ),
        )[0]
        dataset = self._dataset_from_prepared(prepared) if callable(self.method) else prepared
        # worker 成功后才提交；已有数据集与缓存不会被失败任务破坏。
        self._datasets[key] = dataset
        self._datasets_info[key] = label
        self._invalidate_caches()

    def add_dataset(self, key: str, label: str, X, y=None, feature_names: Optional[List[str]] = None):
        """添加额外数据集（如 OOT）用于报告.

        :param key: 数据集标识
        :param label: 数据集标签
        :param X: DataFrame（含目标列时 y 可为 None，自动构建标签）
        :param y: 标签列，None 时从 X 中通过 target / overdue+dpds 自动构建
        :param feature_names: 特征名列表
        """
        X = _ensure_dataframe(X, feature_names=_normalize_feature_names(feature_names) or self.feature_names)
        # y=None 时从 X 中通过 overdue+dpds 自动构建标签（scorecardpipeline 风格）
        if y is None:
            y, y_dict = self._build_y(X, self._target_cfg)
            if self.del_grey and self._is_overdue_cfg() and not y_dict:
                valid = y.notna()
                X = X.loc[valid].copy()
                y = y.loc[valid].astype(int)
        else:
            y_dict = None
        y = _ensure_series(y, name=self._target_name)
        self._add_dataset(key, label, X, y, y_dict)
        if y_dict and not self._label_names:
            self._label_names = list(y_dict)

    def _is_multi_label(self) -> bool:
        """是否多标签模式."""
        return bool(self._label_names)

    def _get_y(self, dataset_key: str, label: Optional[str] = None) -> np.ndarray:
        """获取指定数据集的 y 数组.

        :param dataset_key: 数据集标识
        :param label: 标签名，None 时返回 combined y
        """
        ds = self._datasets[dataset_key]
        if label and ds.y_dict and label in ds.y_dict:
            return ds.y_dict[label]
        return ds.y.to_numpy()

    def _get_valid_target_arrays(
        self,
        dataset_key: str,
        label: Optional[str] = None,
    ) -> Tuple[np.ndarray, np.ndarray]:
        """返回同步剔除无效目标值后的标签与预测数组."""
        y_arr = np.asarray(self._get_y(dataset_key, label), dtype=float)
        prediction = np.asarray(self._datasets[dataset_key].y_proba)
        valid = self._get_target_valid_mask(dataset_key, label)
        return y_arr[valid], prediction[valid]

    def _get_target_valid_mask(self, dataset_key: str, label: Optional[str] = None) -> np.ndarray:
        """返回指定数据集和标签的有效样本掩码."""
        return np.isfinite(np.asarray(self._get_y(dataset_key, label), dtype=float))

    @property
    def _train_key(self) -> str:
        """训练集对应的内部 key：优先 ``'train'``，否则回退到第一个数据集."""
        if "train" in self._datasets:
            return "train"
        return next(iter(self._datasets))

    @property
    def _test_key(self) -> Optional[str]:
        """测试集对应的内部 key：优先 ``'test'``，否则回退到第二个数据集（不存在时为 None）."""
        if "test" in self._datasets:
            return "test"
        keys = [k for k in self._datasets if k != self._train_key]
        return keys[0] if keys else None

    def _resolve_dataset_key(self, dataset: str) -> str:
        """解析数据集标识，支持 ``'train'`` / ``'test'`` 在缺省时回退到第一 / 第二个数据集."""
        if dataset in self._datasets:
            return dataset
        if dataset == "train":
            return self._train_key
        if dataset == "test":
            test_key = self._test_key
            if test_key is not None:
                return test_key
        raise KeyError(f"数据集 '{dataset}' 不存在，可用数据集: {list(self._datasets)}")

    def _is_overdue_cfg(self) -> bool:
        """目标配置是否为 overdue + dpds 逾期联合标签模式."""
        return isinstance(self._target_cfg, dict) and "overdue" in self._target_cfg

    def _overdue_dpds(self) -> Tuple[List[str], List[Union[int, float]]]:
        """从 target 配置解析逾期列与逾期天数阈值列表（与 ``_build_y`` 保持一致）。

        返回 (overdue_cols, dpds_thresholds)，可直接透传给 ``feature_bin_stats``。
        """
        cfg = self._target_cfg or {}
        overdue = cfg.get("overdue")
        overdue_cols = [overdue] if isinstance(overdue, str) else list(overdue or [])
        dpds_vals = cfg.get("dpds")
        threshold = cfg.get("threshold")
        if threshold is not None:
            # 旧格式：dpds 为列名，threshold 为阈值，实际阈值列为 dpds 列
            col = dpds_vals if isinstance(dpds_vals, str) else None
            return ([col] if col else overdue_cols), [threshold]
        if dpds_vals is None:
            return overdue_cols, [0]
        if isinstance(dpds_vals, (int, float)):
            dpds_vals = [dpds_vals]
        return overdue_cols, list(dpds_vals)

    def _overdue_label_map(self, separator: str = ">") -> Dict[str, str]:
        """返回内部标签到报告展示标签的映射。"""
        overdue_cols, dpds_vals = self._overdue_dpds()
        display = [f"{col}{separator}{dpd}" for col in overdue_cols for dpd in dpds_vals]
        return dict(zip(self._label_names, display))

    def _normalize_overdue_bin_columns(self, table: pd.DataFrame) -> pd.DataFrame:
        """将 ``feature_bin_stats`` 的原生逾期标签统一为报告标签。"""
        if not isinstance(table.columns, pd.MultiIndex):
            return table
        overdue_cols, dpds_vals = self._overdue_dpds()
        native_labels = [f"{col}_{dpd}+" for col in overdue_cols for dpd in dpds_vals]
        rename_map = dict(zip(native_labels, self._label_names))
        renamed = table.copy()
        renamed.columns = pd.MultiIndex.from_tuples(
            [(rename_map.get(col[0], col[0]), *col[1:]) for col in table.columns],
            names=table.columns.names,
        )
        return renamed

    # ---------- 1. 模型性能指标 ----------

    def get_metrics(self, label: Optional[str] = None) -> pd.DataFrame:
        """KS / AUC / PSI 等核心指标.

        :param label: 多标签模式下指定标签名，None 时使用 combined y
        """
        if label in self._metrics_cache:
            return self._metrics_cache[label].copy()

        # 数据集按传入顺序展示（dict 的 key / list 顺序 / 训练集→测试集→跨时间验证集）
        ds_keys = list(self._datasets)
        labels_map = {k: self._datasets[k].label for k in ds_keys}
        tasks = [(self._get_y(key, label), self._datasets[key].y_proba) for key in ds_keys]
        metric_values = parallel_execute(
            _binary_metric_worker,
            tasks,
            n_jobs=self.n_jobs,
            parallel_backend=self.parallel_backend,
            parallel_config=self.parallel_config,
            task_labels=[f"指标:{label or 'combined'}:{key}" for key in ds_keys],
            default_backend="threading",
            workload=ParallelWorkload(
                task_count=len(tasks),
                rows=max((len(task[0]) for task in tasks), default=0),
                columns=1,
                data_bytes=sum(np.asarray(task[1]).nbytes for task in tasks),
                cost_per_item=4.0,
                capability="thread_safe",
                releases_gil=True,
                operation="模型报告二分类指标",
            ),
        )

        rows = [
            {"统计项": "KS", **{labels_map[k]: values[0] for k, values in zip(ds_keys, metric_values)}},
            {"统计项": "AUC", **{labels_map[k]: values[1] for k, values in zip(ds_keys, metric_values)}},
            {"统计项": "样本数", **{labels_map[k]: values[2] for k, values in zip(ds_keys, metric_values)}},
            {"统计项": "坏样本率", **{labels_map[k]: values[3] for k, values in zip(ds_keys, metric_values)}},
        ]

        # 多标签的单独标签视图保持原布局，不附加 PSI 行。
        if not (self._is_multi_label() and label) and len(ds_keys) >= 2:
            base_key = ds_keys[0]
            psi_tasks = [(self._datasets[base_key].score, self._datasets[key].score) for key in ds_keys[1:]]
            psi_values = parallel_execute(
                _psi_metric_worker,
                psi_tasks,
                n_jobs=self.n_jobs,
                parallel_backend=self.parallel_backend,
                parallel_config=self.parallel_config,
                task_labels=[f"PSI:{base_key}:{key}" for key in ds_keys[1:]],
                default_backend="threading",
                workload=ParallelWorkload(
                    task_count=len(psi_tasks),
                    rows=max((len(task[0]) + len(task[1]) for task in psi_tasks), default=0),
                    columns=1,
                    data_bytes=sum(np.asarray(task[0]).nbytes + np.asarray(task[1]).nbytes for task in psi_tasks),
                    cost_per_item=3.0,
                    capability="thread_safe",
                    releases_gil=True,
                    operation="模型报告评分PSI",
                ),
            )
            psi_row: Dict[str, Any] = {"统计项": "PSI", labels_map[base_key]: "\\"}
            psi_row.update({labels_map[key]: value for key, value in zip(ds_keys[1:], psi_values)})
            rows.append(psi_row)

        result = pd.DataFrame(rows)
        self._metrics_cache[label] = result.copy()
        return result

    # ---------- 2. 评分分箱效果表 ----------

    def get_bin_table(
        self,
        dataset: str = "train",
        method: str = "quantile",
        max_n_bins: int = 10,
        amount_col: Optional[str] = None,
        margins: bool = True,
        label: Optional[str] = None,
        labels: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """返回评分分箱表，并缓存同一报告调用中的确定性结果。"""
        cache_key = (dataset, method, max_n_bins, amount_col, margins, label, tuple(labels or ()))
        cached = self._bin_table_cache.get(cache_key)
        if cached is not None:
            return cached.copy()
        result = self._compute_bin_table(
            dataset=dataset,
            method=method,
            max_n_bins=max_n_bins,
            amount_col=amount_col,
            margins=margins,
            label=label,
            labels=labels,
        )
        self._bin_table_cache[cache_key] = result.copy()
        return result

    def _compute_bin_table(
        self,
        dataset: str = "train",
        method: str = "quantile",
        max_n_bins: int = 10,
        amount_col: Optional[str] = None,
        margins: bool = True,
        label: Optional[str] = None,
        labels: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """使用 feature_bin_stats 生成评分分箱效果表。

        :param label: 多标签模式下指定单个标签名
        :param labels: 多标签合并模式，传入标签名列表，返回 MultiIndex 列合并表
        """
        from .feature_analyzer import feature_bin_stats

        dataset = self._resolve_dataset_key(dataset)
        ds = self._datasets[dataset]
        score_col = "__score__"
        df = ds.X.copy()
        df[score_col] = ds.score

        score_return_cols = [
            "样本总数",
            "好样本数",
            "坏样本数",
            "样本占比",
            "好样本占比",
            "坏样本占比",
            "坏样本率",
            "LIFT值",
            "累积LIFT值",
            "坏账改善",
            "累积坏账改善",
            "分档KS值",
        ]

        # 多标签合并模式：直接通过 feature_bin_stats 的 overdue + dpds 计算多目标合并分箱表，
        # 由其原生输出多级表头（公共列归入「分箱详情」），并去除「指标名称 / 指标含义」列。
        if labels and self._is_multi_label() and self._is_overdue_cfg():
            overdue_cols, dpds_vals = self._overdue_dpds()
            kw: Dict[str, Any] = dict(
                feature=score_col,
                overdue=overdue_cols,
                dpds=dpds_vals,
                method=method,
                desc="模型评分",
                max_n_bins=max_n_bins,
                missing_separate=True,
                margins=margins,
                del_grey=self.del_grey,
                return_cols=score_return_cols,
                n_jobs=self.n_jobs,
                parallel_backend=self.parallel_backend,
                parallel_config=self.parallel_config,
            )
            if amount_col and amount_col in df.columns:
                kw["amount"] = amount_col
            table = feature_bin_stats(df, **kw)
            if isinstance(table, tuple):
                table = table[0]
            return _drop_bin_meta_cols(self._normalize_overdue_bin_columns(table))

        # 其他多标签合并模式（非 overdue 配置）：逐标签计算后合并
        if labels and self._is_multi_label():
            tables: List[pd.DataFrame] = []
            for lbl in labels:
                t = self.get_bin_table(
                    dataset=dataset,
                    method=method,
                    max_n_bins=max_n_bins,
                    amount_col=amount_col,
                    margins=margins,
                    label=lbl,
                )
                tables.append(t)
            return _drop_bin_meta_cols(_merge_multi_label_bin_tables(tables, labels))

        target_col = "__target__"
        df[target_col] = self._get_y(dataset, label)

        kw = dict(
            feature=score_col,
            target=target_col,
            method=method,
            desc="模型评分",
            max_n_bins=max_n_bins,
            missing_separate=True,
            margins=margins,
            return_cols=score_return_cols,
            n_jobs=self.n_jobs,
            parallel_backend=self.parallel_backend,
            parallel_config=self.parallel_config,
        )
        if amount_col and amount_col in df.columns:
            kw["amount"] = amount_col

        table = feature_bin_stats(df, **kw)
        if isinstance(table, tuple):
            table = table[0]
        return _drop_bin_meta_cols(table)

    # ---------- 3. 特征重要性 ----------

    def _get_raw_feature_importance(self) -> pd.Series:
        """按模型能力提取原始绝对重要性，并与最终入模字段对齐。"""
        if self._raw_importance_cache is not None:
            return self._raw_importance_cache.copy()

        candidates: List[Any] = []
        seen = set()
        for candidate in [
            self.model,
            getattr(self.model, "model_", None),
            getattr(self.model, "_model", None),
            getattr(self.model, "model", None),
            getattr(self.model, "lr_model_", None),
            getattr(self.model, "lr_model", None),
        ]:
            if candidate is None or id(candidate) in seen:
                continue
            seen.add(id(candidate))
            candidates.append(candidate)

        resolved: Optional[pd.Series] = None
        for candidate in candidates:
            values = None
            if callable(getattr(candidate, "get_feature_importances", None)):
                try:
                    values = candidate.get_feature_importances()
                except Exception:
                    values = None
            if values is None and hasattr(candidate, "feature_importances_"):
                try:
                    values = getattr(candidate, "feature_importances_")
                except Exception:
                    values = None
            if values is None and hasattr(candidate, "coef_"):
                try:
                    coefficients = np.asarray(getattr(candidate, "coef_"), dtype=float)
                    values = np.mean(np.abs(coefficients), axis=0) if coefficients.ndim > 1 else np.abs(coefficients)
                except Exception:
                    values = None
            if values is None:
                continue

            if isinstance(values, pd.Series):
                series = pd.to_numeric(values, errors="coerce")
            else:
                array = np.asarray(values, dtype=float).reshape(-1)
                if len(array) != len(self.feature_names):
                    continue
                series = pd.Series(array, index=self.feature_names)

            if not set(series.index).intersection(self.feature_names) and len(series) == len(self.feature_names):
                series.index = self.feature_names
            resolved = series.reindex(self.feature_names).abs()
            break

        if resolved is None:
            resolved = pd.Series(index=self.feature_names, dtype=float, name="特征重要性")
        else:
            resolved.name = "特征重要性"
        self._raw_importance_cache = resolved.copy()
        return resolved

    def _get_model_input_table(self, feature_map: Optional[Dict[str, str]] = None) -> pd.DataFrame:
        """构造入参字段、重要性占比及累计贡献表。"""
        importance = self._get_raw_feature_importance()
        table = pd.DataFrame(
            {
                "入参字段": self.feature_names,
                "特征重要性": importance.reindex(self.feature_names).to_numpy(dtype=float),
            }
        )
        if table["特征重要性"].notna().any():
            table = table.sort_values("特征重要性", ascending=False, kind="mergesort").reset_index(drop=True)
            total_importance = float(table["特征重要性"].sum())
            if total_importance > 0:
                table["特征重要性%"] = table["特征重要性"] / total_importance
                table["累积特征重要性%"] = table["特征重要性%"].cumsum()
            else:
                table["特征重要性%"] = np.nan
                table["累积特征重要性%"] = np.nan
        else:
            table["特征重要性%"] = np.nan
            table["累积特征重要性%"] = np.nan

        table.insert(0, "序号", np.arange(1, len(table) + 1))
        if feature_map:
            table.insert(2, "字段名称", table["入参字段"].map(feature_map).fillna(""))
        return table

    @staticmethod
    def _validate_feature_contribution_label_limit(value: Optional[int]) -> Optional[int]:
        """校验贡献图显示标签的最大特征数。"""
        if value is None:
            return None
        if isinstance(value, bool) or not isinstance(value, (int, np.integer)) or value < 0:
            raise ValueError("feature_contribution_label_max_features 必须是非负整数或 None")
        return int(value)

    def _create_feature_contribution_figure(
        self,
        table: pd.DataFrame,
        label_max_features: Optional[int] = 10,
    ):
        """绘制序号维度的单字段贡献柱和累计贡献折线。"""
        from matplotlib.ticker import PercentFormatter

        from ..core.viz.utils import (
            BAD_RATE_COLOR,
            DEFAULT_COLORS,
            _create_subplots,
            _layout_top_center_legend,
            setup_axis_style,
        )

        figure, primary = _create_subplots(figsize=(10, 5))
        secondary = primary.twinx()
        axis_theme = DEFAULT_COLORS[0]
        sequence = table["序号"].to_numpy(dtype=int)
        importance_ratio = table["特征重要性%"].to_numpy(dtype=float)
        cumulative_ratio = table["累积特征重要性%"].to_numpy(dtype=float)
        label_max_features = self._validate_feature_contribution_label_limit(label_max_features)
        show_labels = label_max_features is None or len(table) <= label_max_features

        secondary.bar(
            sequence,
            importance_ratio,
            width=0.66,
            color=axis_theme,
            edgecolor="white",
            linewidth=0.8,
            alpha=0.92,
            hatch="/",
            label="特征重要性%",
        )
        primary.plot(
            sequence,
            cumulative_ratio,
            color=BAD_RATE_COLOR,
            linestyle=(0, (4, 3)),
            marker="o",
            markerfacecolor="white",
            markeredgecolor=BAD_RATE_COLOR,
            markeredgewidth=1.4,
            markersize=5.5,
            linewidth=2.1,
            clip_on=False,
            label="累积特征重要性%",
        )
        primary.set(
            xlabel="序号",
            ylabel="累积特征重要性%",
            xticks=sequence,
            ylim=(0, 1),
        )
        secondary.set_ylabel("特征重要性%")
        secondary.set_ylim(0, 1)
        primary.yaxis.set_major_formatter(PercentFormatter(1, decimals=0, is_latex=True))
        secondary.yaxis.set_major_formatter(PercentFormatter(1, decimals=0, is_latex=True))
        if show_labels:
            for item, ratio in zip(sequence, cumulative_ratio):
                label_y = ratio - 0.035 if ratio >= 0.95 else ratio + 0.025
                primary.text(
                    item,
                    label_y,
                    f"{ratio:.2%}",
                    ha="center",
                    va="top" if ratio >= 0.95 else "bottom",
                    color=BAD_RATE_COLOR,
                    fontsize=10,
                    fontweight="semibold",
                    bbox={
                        "boxstyle": "round,pad=0.18",
                        "facecolor": "white",
                        "edgecolor": BAD_RATE_COLOR,
                        "linewidth": 0.6,
                        "alpha": 0.92,
                    },
                )
            for item, ratio in zip(sequence[1:], importance_ratio[1:]):
                secondary.text(
                    item,
                    ratio + 0.018,
                    f"{ratio:.2%}",
                    ha="center",
                    va="bottom",
                    color=axis_theme,
                    fontsize=10,
                    fontweight="semibold",
                    bbox={
                        "boxstyle": "round,pad=0.18",
                        "facecolor": "white",
                        "edgecolor": axis_theme,
                        "linewidth": 0.6,
                        "alpha": 0.92,
                    },
                )
        setup_axis_style(primary, [axis_theme], hide_top_right=False)
        setup_axis_style(secondary, [axis_theme], hide_top_right=False)
        primary.spines["top"].set_visible(False)
        secondary.spines["top"].set_visible(False)
        primary.tick_params(axis="both", colors=axis_theme)
        secondary.tick_params(axis="both", colors=axis_theme)
        primary.grid(False)
        secondary.grid(False)
        title_artist = figure.suptitle("入模特征贡献", fontsize=14, fontweight="bold")
        handles1, labels1 = primary.get_legend_handles_labels()
        handles2, labels2 = secondary.get_legend_handles_labels()
        legend = figure.legend(
            handles2 + handles1,
            labels2 + labels1,
            loc="upper center",
            bbox_to_anchor=(0.5, 0.90),
            ncol=2,
            frameon=False,
        )
        figure.tight_layout()
        _layout_top_center_legend(
            figure,
            legend,
            title=title_artist,
            axes=[primary, secondary],
        )
        return figure

    def get_feature_importance(self, top_n: Optional[int] = None) -> pd.DataFrame:
        if self._importance_cache is None:
            importances = self._get_raw_feature_importance().dropna()
            if importances.empty:
                self._importance_cache = pd.DataFrame(columns=["特征重要性", "IV", "KS", "PSI"])
            else:
                importances = importances.reindex([f for f in importances.index if f in self.feature_names])
                total = importances.abs().sum()
                importance_df = pd.DataFrame(index=importances.index)
                importance_df["特征重要性"] = importances.abs().values / total if total else importances.values

                train_ds = self._datasets[self._train_key]
                test_key = self._test_key
                y_arr = train_ds.y.to_numpy()
                metric_tasks = []
                for feat in importance_df.index:
                    train_values = train_ds.X[feat] if feat in train_ds.X.columns else pd.Series(dtype=float)
                    test_values = self._datasets[test_key].X[feat] if test_key is not None and feat in self._datasets[test_key].X.columns else None
                    metric_tasks.append((y_arr, train_values, test_values))
                metric_values = parallel_execute(
                    _feature_metric_worker,
                    metric_tasks,
                    n_jobs=self.n_jobs,
                    parallel_backend=self.parallel_backend,
                    parallel_config=self.parallel_config,
                    task_labels=[f"特征指标:{feat}" for feat in importance_df.index],
                    default_backend="threading",
                    workload=ParallelWorkload(
                        task_count=len(metric_tasks),
                        rows=len(train_ds.X),
                        columns=len(metric_tasks),
                        data_bytes=int(train_ds.X.memory_usage(deep=True).sum()),
                        cost_per_item=6.0,
                        capability="thread_safe",
                        releases_gil=True,
                        operation="模型报告字段IV KS PSI",
                    ),
                )
                importance_df["IV"] = [values[0] for values in metric_values]
                importance_df["KS"] = [values[1] for values in metric_values]
                importance_df["PSI"] = [values[2] for values in metric_values]
                self._importance_cache = importance_df.sort_values("特征重要性", ascending=False)

        df = self._importance_cache.copy()
        if top_n is not None:
            df = df.head(top_n)
        return df

    def _is_scorecard_model(self) -> bool:
        """按评分卡能力识别 hscredit、toad 与 scorecardpipeline 对象。"""
        model_name = type(self.model).__name__.lower()
        has_points = callable(getattr(self.model, "scorecard_points", None))
        has_rules = isinstance(getattr(self.model, "rules", None), dict) or isinstance(
            getattr(self.model, "rules_", None), dict
        )
        has_export = callable(getattr(self.model, "export", None))
        return "scorecard" in model_name and (has_points or (has_rules and has_export))

    def _scorecard_lr_model(self):
        """解析三类评分卡实际持有的逻辑回归模型。"""
        for attribute in ("lr_model_", "lr_model", "pretrain_lr", "model"):
            candidate = getattr(self.model, attribute, None)
            if candidate is not None and candidate is not self.model:
                return candidate
        return None

    @staticmethod
    def _mapping_table(mapping: Dict[str, Any], key_name: str = "配置项", value_name: str = "配置值") -> pd.DataFrame:
        """把有序配置映射转为 Excel 友好的两列表。"""
        return pd.DataFrame([{key_name: key, value_name: value} for key, value in mapping.items()])

    def _get_score_conversion_sections(self) -> Optional[Dict[str, pd.DataFrame]]:
        """返回普通模型已拟合概率评分转换器的选型、参数与公式。"""
        if self.method != "predict_score" or self._is_scorecard_model():
            return None

        owner = getattr(self.model, "scorecard_", None)
        if owner is None and hasattr(self.model, "transformer_"):
            owner = self.model
        converter = getattr(self.model, "score_transformer_", None)
        if converter is None and owner is not None:
            converter = getattr(owner, "transformer_", None)
        if converter is None:
            return None

        actual_converter = getattr(converter, "transformer_", converter)
        is_fitted = any(
            bool(getattr(candidate, "_is_fitted", False))
            for candidate in (owner, converter, actual_converter)
            if candidate is not None
        )
        if not is_fitted:
            return None

        selection = {
            "转换入口": type(converter).__name__,
            "转换方法": getattr(converter, "method", getattr(owner, "method", None)),
            "实际转换器": type(actual_converter).__name__,
        }

        params: Dict[str, Any] = {}
        if owner is not None and callable(getattr(owner, "get_params_info", None)):
            try:
                params.update(owner.get_params_info())
            except Exception:
                pass
        if not params:
            for name in (
                "method",
                "direction_",
                "direction",
                "lower",
                "upper",
                "decimal",
                "clip",
                "base_odds",
                "base_score",
                "pdo",
                "rate",
                "A_",
                "B_",
            ):
                for candidate in (converter, actual_converter):
                    if hasattr(candidate, name):
                        params[name.rstrip("_")] = getattr(candidate, name)
                        break

        formula: Dict[str, Any] = {}
        for candidate in (owner, converter, actual_converter):
            if candidate is None or not callable(getattr(candidate, "score_formula", None)):
                continue
            try:
                result = candidate.score_formula()
                formula = result if isinstance(result, dict) else {"公式": str(result)}
                break
            except Exception:
                continue

        return {
            "selection": self._mapping_table(selection),
            "params": self._mapping_table(params),
            "formula": self._mapping_table(formula, key_name="公式项", value_name="公式内容"),
        }

    def _scorecard_lr_summary(self) -> pd.DataFrame:
        """优先读取 LR summary，不支持时按系数构造统一表。"""
        lr_model = self._scorecard_lr_model()
        if lr_model is None:
            return pd.DataFrame()
        if callable(getattr(lr_model, "summary", None)):
            try:
                summary = lr_model.summary()
                if isinstance(summary, pd.DataFrame):
                    return summary
                tables = getattr(summary, "tables", None)
                if tables and len(tables) > 1 and hasattr(tables[1], "data"):
                    rows = tables[1].data
                    return pd.DataFrame(rows[1:], columns=rows[0])
            except Exception:
                pass

        try:
            coefficients = np.asarray(lr_model.coef_, dtype=float)
        except Exception:
            return pd.DataFrame()
        coefficients = coefficients.mean(axis=0) if coefficients.ndim > 1 else coefficients
        feature_names = getattr(lr_model, "feature_names_in_", self.feature_names)
        feature_names = list(feature_names) if len(feature_names) == len(coefficients) else self.feature_names
        rows = [{"变量": name, "系数": float(value)} for name, value in zip(feature_names, coefficients)]
        intercept = np.asarray(getattr(lr_model, "intercept_", []), dtype=float).reshape(-1)
        if len(intercept):
            rows.insert(0, {"变量": "截距", "系数": float(intercept[0])})
        return pd.DataFrame(rows)

    def _direct_lr_summary(self, feature_map: Optional[Dict[str, str]] = None) -> pd.DataFrame:
        """读取直接传入的 LR 统计摘要，并补充字段含义。"""
        if self._is_scorecard_model() or not callable(getattr(self.model, "summary", None)):
            return pd.DataFrame()

        class_names = {base.__name__.lower().replace("_", "") for base in type(self.model).__mro__}
        if not any("logisticregression" in name or name == "lr" for name in class_names):
            return pd.DataFrame()

        try:
            summary = self.model.summary()
        except Exception as exc:
            logger.warning("读取逻辑回归统计摘要失败 [模型=%s]: %s", type(self.model).__name__, exc)
            return pd.DataFrame()
        if not isinstance(summary, pd.DataFrame):
            logger.warning("逻辑回归 summary() 未返回 DataFrame [模型=%s]", type(self.model).__name__)
            return pd.DataFrame()

        table = summary.copy()
        feature_column = next(
            (column for column in ("入参字段", "Features", "变量", "feature", "特征") if column in table.columns),
            None,
        )
        if feature_column is None and not isinstance(table.index, pd.RangeIndex):
            table = table.rename_axis("入参字段").reset_index()
            feature_column = "入参字段"
        if feature_map and feature_column is not None:
            table.insert(
                table.columns.get_loc(feature_column) + 1,
                "字段名称",
                table[feature_column].map(feature_map).fillna(""),
            )
        return table

    def _scorecard_scale_table(self) -> pd.DataFrame:
        """统一评分卡基础刻度参数。"""
        if callable(getattr(self.model, "scorecard_scale", None)):
            try:
                scale = self.model.scorecard_scale()
                if isinstance(scale, pd.DataFrame):
                    return scale
            except Exception:
                pass
        rows = []
        descriptions = {
            "base_odds": "基础 Odds",
            "base_score": "基础 Odds 对应分数",
            "rate": "Odds 倍率",
            "pdo": "Odds 增长 rate 倍时的分数变化",
            "factor": "评分转换系数",
            "offset": "评分转换截距",
        }
        for name, description in descriptions.items():
            if hasattr(self.model, name):
                rows.append({"刻度项": name, "刻度值": getattr(self.model, name), "备注": description})
        return pd.DataFrame(rows)

    def _scorecard_formula_table(self) -> pd.DataFrame:
        """统一评分卡概率转评分公式。"""
        if callable(getattr(self.model, "score_formula", None)):
            try:
                formula = self.model.score_formula()
                if isinstance(formula, dict):
                    return self._mapping_table(formula, key_name="公式项", value_name="公式内容")
                return self._mapping_table({"公式": str(formula)}, key_name="公式项", value_name="公式内容")
            except Exception:
                pass
        factor = getattr(self.model, "factor", getattr(self.model, "B_", None))
        offset = getattr(self.model, "offset", getattr(self.model, "A_", None))
        if factor is None or offset is None:
            return pd.DataFrame()
        formula = {
            "公式": f"Score = {float(offset):.4f} - {float(factor):.4f} × ln(P / (1 - P))",
            "offset": float(offset),
            "factor": float(factor),
        }
        return self._mapping_table(formula, key_name="公式项", value_name="公式内容")

    def _scorecard_points_table(self, feature_map: Optional[Dict[str, str]] = None) -> pd.DataFrame:
        """统一评分卡分箱分值表。"""
        if callable(getattr(self.model, "scorecard_points", None)):
            try:
                points = self.model.scorecard_points(feature_map=feature_map)
            except TypeError:
                points = self.model.scorecard_points()
            except Exception:
                points = None
            if isinstance(points, pd.DataFrame):
                return points

        if callable(getattr(self.model, "export", None)):
            try:
                points = self.model.export(to_frame=True)
                if isinstance(points, pd.DataFrame):
                    points = points.rename(
                        columns={"name": "变量名称", "value": "变量分箱", "score": "对应分数"}
                    )
                    if feature_map:
                        points.insert(
                            points.columns.get_loc("变量名称") + 1,
                            "变量含义",
                            points["变量名称"].map(feature_map).fillna(""),
                        )
                    return points
            except Exception:
                pass
        return pd.DataFrame()

    def _scorecard_odds_reference_table(self) -> pd.DataFrame:
        """统一评分、Odds 与理论逾期率参考表。"""
        try:
            reference = getattr(self.model, "score_odds_reference")
        except Exception:
            reference = None
        if isinstance(reference, pd.DataFrame):
            return reference

        factor = getattr(self.model, "factor", getattr(self.model, "B_", None))
        offset = getattr(self.model, "offset", getattr(self.model, "A_", None))
        base_score = getattr(self.model, "base_score", None)
        pdo = getattr(self.model, "pdo", None)
        if factor is None or offset is None or base_score is None or pdo is None:
            return pd.DataFrame()
        scores = np.arange(float(base_score) + 5 * float(pdo), float(base_score) - 5 * float(pdo) - 1, -float(pdo))
        odds = np.exp((float(offset) - scores) / float(factor))
        bad_rate = odds / (1 + odds)
        return pd.DataFrame(
            {
                "评分": scores,
                "理论Odds(坏好比)": odds,
                "理论逾期率": bad_rate,
            }
        )

    def _score_psi_matrix(self) -> pd.DataFrame:
        """计算全部数据集两两评分 PSI。"""
        if len(self._datasets) < 2:
            return pd.DataFrame()
        from ..core.metrics import psi

        keys = list(self._datasets)
        labels = [self._datasets[key].label for key in keys]
        matrix = pd.DataFrame(np.nan, index=labels, columns=labels)
        for row_index, left_key in enumerate(keys):
            for column_index, right_key in enumerate(keys):
                if row_index == column_index:
                    matrix.iloc[row_index, column_index] = 0.0
                else:
                    try:
                        matrix.iloc[row_index, column_index] = psi(
                            self._datasets[left_key].score,
                            self._datasets[right_key].score,
                        )
                    except Exception:
                        pass
        matrix.index.name = "基准数据集"
        return matrix

    # ---------- 4. 特征描述 ----------

    def get_features_describe(self) -> pd.DataFrame:
        """入模变量重要性及描述性统计."""
        if self._features_describe_cache is not None:
            return self._features_describe_cache.copy()

        importance = self.get_feature_importance()
        features = importance.index.tolist()
        train_ds = self._datasets[self._train_key]
        train_X = train_ds.X[features] if features else train_ds.X[self.feature_names]
        desc_stats = train_X.describe(percentiles=[0.01, 0.1, 0.5, 0.75, 0.9, 0.99]).T
        desc_stats = desc_stats.rename(
            columns={
                "count": "样本数",
                "mean": "平均值",
                "std": "标准差",
                "min": "最小值",
                "max": "最大值",
                "1%": "1%",
                "10%": "10%",
                "50%": "50%",
                "75%": "75%",
                "90%": "90%",
                "99%": "99%",
            }
        )
        desc_stats["缺失率"] = train_X.isnull().mean()
        desc_stats["字段类型"] = train_X.dtypes.astype(str)
        desc_stats["枚举数"] = train_X.nunique()

        keep_cols = [
            "字段类型",
            "缺失率",
            "枚举数",
            "平均值",
            "标准差",
            "最小值",
            "1%",
            "10%",
            "50%",
            "75%",
            "90%",
            "99%",
            "最大值",
        ]
        keep_cols = [c for c in keep_cols if c in desc_stats.columns]
        result = importance.join(desc_stats[keep_cols], how="left")
        result = result.drop(columns=["样本数"], errors="ignore")
        self._features_describe_cache = result
        return self._features_describe_cache.copy()

    # ---------- 5. 特征相关性 ----------

    def get_features_corr(self) -> pd.DataFrame:
        if self._corr_cache is not None:
            return self._corr_cache.copy()
        importance = self.get_feature_importance()
        features = importance.index.tolist()
        if not features:
            features = self.feature_names
        result = self._datasets[self._train_key].X[features].corr()
        self._corr_cache = result.copy()
        return result

    # ---------- 6. 特征分箱分析 ----------

    def get_feature_bin_table(
        self,
        feature: str,
        dataset: str = "train",
        max_n_bins: int = 10,
        method: str = "quantile",
        margins: bool = True,
        amount_col: Optional[str] = None,
        label: Optional[str] = None,
        labels: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """返回特征分箱表，并缓存同一报告调用中的确定性结果。"""
        cache_key = (feature, dataset, max_n_bins, method, margins, amount_col, label, tuple(labels or ()))
        cached = self._feature_bin_table_cache.get(cache_key)
        if cached is not None:
            return cached.copy()
        result = self._compute_feature_bin_table(
            feature=feature,
            dataset=dataset,
            max_n_bins=max_n_bins,
            method=method,
            margins=margins,
            amount_col=amount_col,
            label=label,
            labels=labels,
        )
        self._feature_bin_table_cache[cache_key] = result.copy()
        return result

    def _compute_feature_bin_table(
        self,
        feature: str,
        dataset: str = "train",
        max_n_bins: int = 10,
        method: str = "quantile",
        margins: bool = True,
        amount_col: Optional[str] = None,
        label: Optional[str] = None,
        labels: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """单特征分箱效果表，优先使用模型 binner。

        :param label: 多标签模式下指定单个标签名
        :param labels: 多标签合并模式，传入标签名列表，返回 MultiIndex 列合并表
        """
        from .feature_analyzer import feature_bin_stats

        dataset = self._resolve_dataset_key(dataset)
        ds = self._datasets[dataset]
        df = ds.X.copy()
        binner = getattr(self.model, "binner", None)

        # 多标签合并模式：通过 feature_bin_stats 的 overdue + dpds 计算多目标合并分箱表，
        # 多级表头由其原生输出（公共列归入「分箱详情」），并去除「指标名称 / 指标含义」列。
        if labels and self._is_multi_label() and self._is_overdue_cfg():
            overdue_cols, dpds_vals = self._overdue_dpds()
            kw: Dict[str, Any] = dict(
                feature=feature,
                overdue=overdue_cols,
                dpds=dpds_vals,
                method=method,
                max_n_bins=max_n_bins,
                margins=margins,
                del_grey=self.del_grey,
                missing_separate=True,
                n_jobs=self.n_jobs,
                parallel_backend=self.parallel_backend,
                parallel_config=self.parallel_config,
            )
            if binner is not None:
                kw["binner"] = binner
            if amount_col and amount_col in df.columns:
                kw["amount"] = amount_col
            table = feature_bin_stats(df, **kw)
            if isinstance(table, tuple):
                table = table[0]
            return _drop_bin_meta_cols(self._normalize_overdue_bin_columns(table))

        # 其他多标签合并模式（非 overdue 配置）：逐标签计算后合并
        if labels and self._is_multi_label():
            tables: List[pd.DataFrame] = []
            for lbl in labels:
                t = self.get_feature_bin_table(
                    feature=feature,
                    dataset=dataset,
                    max_n_bins=max_n_bins,
                    method=method,
                    margins=margins,
                    amount_col=amount_col,
                    label=lbl,
                )
                tables.append(t)
            return _drop_bin_meta_cols(_merge_multi_label_bin_tables(tables, labels))

        target_col = "__target__"
        df[target_col] = self._get_y(dataset, label)

        kw = dict(
            feature=feature,
            target=target_col,
            method=method,
            max_n_bins=max_n_bins,
            margins=margins,
            missing_separate=True,
            n_jobs=self.n_jobs,
            parallel_backend=self.parallel_backend,
            parallel_config=self.parallel_config,
        )
        if binner is not None:
            kw["binner"] = binner

        if amount_col and amount_col in df.columns:
            kw["amount"] = amount_col

        table = feature_bin_stats(df, **kw)
        if isinstance(table, tuple):
            table = table[0]
        return _drop_bin_meta_cols(table)

    # ---------- 8. 图表导出 ----------

    def _get_top_n_lift_table(
        self,
        percentiles: Tuple[float, ...] = (0.01, 0.03, 0.05, 0.10),
        amount_col: Optional[str] = None,
        label: Optional[str] = None,
        labels: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        cache_key = (tuple(percentiles), amount_col, label, tuple(labels or ()))
        cached = self._lift_table_cache.get(cache_key)
        if cached is not None:
            return cached.copy()
        result = self._compute_top_n_lift_table(
            percentiles=percentiles,
            amount_col=amount_col,
            label=label,
            labels=labels,
        )
        self._lift_table_cache[cache_key] = result.copy()
        return result

    def _compute_top_n_lift_table(
        self,
        percentiles: Tuple[float, ...] = (0.01, 0.03, 0.05, 0.10),
        amount_col: Optional[str] = None,
        label: Optional[str] = None,
        labels: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """构建 TOP n% 尾部区分能力表。

        :param percentiles: TOP n% 的百分位列表
        :param amount_col: 金额字段名（可选），指定时输出金额口径
        :param label: 单标签模式下指定标签名，None 时使用 combined y
        :param labels: 多标签合并模式，返回 MultiIndex 列合并表
        """
        # 多标签合并模式：labels 列表优先，返回 MultiIndex 列结构
        if labels and self._is_multi_label():
            return self._get_top_n_lift_table_multi(
                percentiles=percentiles,
                amount_col=amount_col,
                labels=labels,
            )

        rows: List[Dict[str, Any]] = []
        for ds_key, ds in self._datasets.items():
            tag = ds.label
            valid = self._get_target_valid_mask(ds_key, label)
            y_arr, prediction = self._get_valid_target_arrays(ds_key, label)
            n = len(y_arr)
            overall_bad_rate = float(y_arr.mean())

            sorted_idx = np.argsort(-prediction)
            sorted_y = y_arr[sorted_idx]

            bad_rates: Dict[str, float] = {}
            lifts: Dict[str, float] = {}
            improvements: Dict[str, float] = {}

            for pct in percentiles:
                top_n = max(1, int(n * pct))
                top_bad_rate = float(sorted_y[:top_n].mean())
                lift = top_bad_rate / overall_bad_rate if overall_bad_rate > 0 else 0.0
                improvement = (top_bad_rate - overall_bad_rate) / overall_bad_rate if overall_bad_rate > 0 else 0.0
                key = f"TOP {int(pct * 100)}%"
                bad_rates[key] = top_bad_rate
                lifts[key] = lift
                improvements[key] = improvement

            bad_rates["TOTAL"] = overall_bad_rate
            lifts["TOTAL"] = 1.0
            improvements["TOTAL"] = 0.0

            rows.append({"数据集": tag, "统计项": "坏样本率", **bad_rates})
            rows.append({"数据集": tag, "统计项": "LIFT值", **lifts})
            rows.append({"数据集": tag, "统计项": "坏账改善", **improvements})

            # 金额口径替代订单口径；调用方会将两个结果并排展示。
            if amount_col and amount_col in ds.X.columns:
                amounts = pd.to_numeric(ds.X[amount_col], errors="coerce").fillna(0).clip(lower=0).to_numpy(dtype=float)[valid]
                amounts_sorted = amounts[sorted_idx]
                overall_bad_amount = float((sorted_y * amounts_sorted).sum() / amounts_sorted.sum()) if amounts_sorted.sum() > 0 else overall_bad_rate

                amt_bad_rates: Dict[str, float] = {}
                amt_lifts: Dict[str, float] = {}
                amt_improvements: Dict[str, float] = {}

                for pct in percentiles:
                    top_n = max(1, int(n * pct))
                    top_amt = amounts_sorted[:top_n]
                    top_y_sorted = sorted_y[:top_n]
                    top_bad_amt = float((top_y_sorted * top_amt).sum() / top_amt.sum()) if top_amt.sum() > 0 else 0.0
                    lift_amt = top_bad_amt / overall_bad_amount if overall_bad_amount > 0 else 0.0
                    imp_amt = (top_bad_amt - overall_bad_amount) / overall_bad_amount if overall_bad_amount > 0 else 0.0
                    key = f"TOP {int(pct * 100)}%"
                    amt_bad_rates[key] = top_bad_amt
                    amt_lifts[key] = lift_amt
                    amt_improvements[key] = imp_amt

                amt_bad_rates["TOTAL"] = overall_bad_amount
                amt_lifts["TOTAL"] = 1.0
                amt_improvements["TOTAL"] = 0.0

                rows[-3:] = [
                    {"数据集": tag, "统计项": "坏样本率", **amt_bad_rates},
                    {"数据集": tag, "统计项": "LIFT值", **amt_lifts},
                    {"数据集": tag, "统计项": "坏账改善", **amt_improvements},
                ]

        return pd.DataFrame(rows)

    def _get_top_n_lift_table_multi(
        self,
        percentiles: Tuple[float, ...] = (0.01, 0.03, 0.05, 0.10),
        amount_col: Optional[str] = None,
        labels: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """多标签合并模式的 LIFT 表。

        行: TOP 1% / TOP 3% / TOP 5% / TOP 10% / TOTAL
        列: (数据集, 标签, 指标) 三层 MultiIndex
        """
        if labels is None:
            labels = self._label_names

        # 逐标签复用单标签 TOP n% 表，再沿列方向以「标签」为外层拼接，
        # 形成 行=(数据集, 统计项)、列=(标签, TOP n%) 的二级表头结构。
        per_label: Dict[str, pd.DataFrame] = {}
        for lbl in labels:
            sub = self._get_top_n_lift_table(
                percentiles=percentiles,
                amount_col=amount_col,
                label=lbl,
            )
            per_label[lbl] = sub.set_index(["数据集", "统计项"])

        result_df = pd.concat(per_label, axis=1)
        result_df.columns = pd.MultiIndex.from_tuples(list(result_df.columns), names=["标签", "统计指标"])
        result_df.index = result_df.index.set_names(["数据集", "统计项"])
        return result_df

    def _get_features_summary(self) -> pd.DataFrame:
        """使用 pd.DataFrame.summary() 获取入模变量综合统计."""
        if self._features_summary_cache is not None:
            return self._features_summary_cache.copy()
        importance = self.get_feature_importance()
        features = importance.index.tolist() if not importance.empty else self.feature_names

        target_col = self._target_name or "target"
        train_ds = self._datasets[self._train_key]
        train_df = train_ds.X[features].copy()
        train_df[target_col] = train_ds.y.values

        test_df = None
        test_key = self._test_key
        if test_key is not None:
            test_ds = self._datasets[test_key]
            test_df = test_ds.X[features].copy()
            test_df[target_col] = test_ds.y.values

        try:
            summary_result = train_df.summary(
                features=features,
                y=target_col,
                val_df=test_df,
                n_jobs=self.n_jobs,
                parallel_backend=self.parallel_backend,
                parallel_config=self.parallel_config,
            )
            self._features_summary_cache = summary_result.copy()
            return summary_result
        except Exception:
            result = self.get_features_describe()
            self._features_summary_cache = result.copy()
            return result

    # 分月评分分布分位数（与特征效率分析保持一致的分位数口径）
    _SCORE_DIST_QUANTILES = [0.01, 0.03, 0.05, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 0.95, 0.97, 0.99]

    def _get_monthly_metrics(self, date_col: str) -> pd.DataFrame:
        cached = self._monthly_metrics_cache.get(date_col)
        if cached is not None:
            return cached.copy()
        result = self._compute_monthly_metrics(date_col)
        self._monthly_metrics_cache[date_col] = result.copy()
        return result

    def _compute_monthly_metrics(self, date_col: str) -> pd.DataFrame:
        """分月计算 KS/AUC 及评分分布（均值/标准差/极值/分位数）."""
        from ..core.metrics import ks, auc

        q_cols = [f"{int(round(q * 100))}%" for q in self._SCORE_DIST_QUANTILES]

        rows: List[Dict[str, Any]] = []
        for ds_key, ds in self._datasets.items():
            if date_col not in ds.X.columns:
                continue
            dates = pd.to_datetime(ds.X[date_col])
            months = dates.dt.to_period("M")
            for month in sorted(months.unique()):
                mask = months == month
                y_m = ds.y[mask.values]
                proba_m = ds.y_proba[mask.values]
                score_m = np.asarray(ds.score)[mask.values]
                if len(y_m) == 0:
                    continue
                try:
                    row: Dict[str, Any] = {
                        "数据集": ds.label,
                        "月份": str(month),
                        "样本数": len(y_m),
                        "坏样本率": float(y_m.mean()),
                        "KS": _safe_binary_metric(ks, y_m, proba_m),
                        "AUC": _safe_binary_metric(auc, y_m, proba_m),
                        "均值": float(np.nanmean(score_m)),
                        "标准差": float(np.nanstd(score_m)),
                        "最小值": float(np.nanmin(score_m)),
                        "最大值": float(np.nanmax(score_m)),
                    }
                    q_values = np.nanpercentile(score_m, [q * 100 for q in self._SCORE_DIST_QUANTILES])
                    for q_col, q_val in zip(q_cols, q_values):
                        row[q_col] = float(q_val)
                    rows.append(row)
                except Exception as exc:
                    logger.warning("生成 %s %s 的分月模型效果失败: %s", ds.label, month, exc)
        return pd.DataFrame(rows) if rows else pd.DataFrame()

    def _get_monthly_psi_matrix(self, date_col: str) -> pd.DataFrame:
        cached = self._monthly_psi_cache.get(date_col)
        if cached is not None:
            return cached.copy()
        result = self._compute_monthly_psi_matrix(date_col)
        self._monthly_psi_cache[date_col] = result.copy()
        return result

    def _compute_monthly_psi_matrix(self, date_col: str) -> pd.DataFrame:
        """分月 PSI 交叉矩阵."""
        from ..core.metrics import psi

        month_scores: Dict[str, np.ndarray] = {}
        for ds in self._datasets.values():
            if date_col not in ds.X.columns:
                continue
            dates = pd.to_datetime(ds.X[date_col])
            months = dates.dt.to_period("M")
            for month in sorted(months.unique()):
                mask = months == month
                key = str(month)
                if key in month_scores:
                    month_scores[key] = np.concatenate([month_scores[key], ds.score[mask.values]])
                else:
                    month_scores[key] = ds.score[mask.values]

        if len(month_scores) < 2:
            return pd.DataFrame()

        labels = sorted(month_scores.keys())
        matrix = pd.DataFrame(np.nan, index=labels, columns=labels)
        for i, m1 in enumerate(labels):
            for j, m2 in enumerate(labels):
                try:
                    matrix.loc[m1, m2] = psi(month_scores[m1], month_scores[m2])
                except Exception:
                    pass
        return matrix

    # ---------- 8. 图表导出 ----------

    def _export_plots_serial(
        self,
        output_dir: Path,
        n_bins: int = 10,
        bin_method: str = "quantile",
        amount_col: Optional[str] = None,
        show_lift: bool = True,
        feature_contribution_label_max_features: Optional[int] = 10,
    ) -> Tuple[Dict[str, List[str]], Dict[str, pd.DataFrame]]:
        """串行导出全部图表，作为线程实现的结果与性能参考."""
        from ..core.viz import ks_plot, bin_plot, corr_plot, psi_plot, lift_plot, hist_plot

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        paths: Dict[str, List[str]] = {}
        tables: Dict[str, pd.DataFrame] = {}

        # --- 模型级图表（用于模型性能 Sheet） ---
        for ds_key, ds in self._datasets.items():
            tag = ds.label
            model_figs: List[str] = []

            p = str(output_dir / f"bin_{ds_key}.png")
            try:
                bt = self.get_bin_table(ds_key, method=bin_method, max_n_bins=n_bins, margins=True)
                bd = bt.iloc[:-1].reset_index(drop=True) if len(bt) > 1 else bt
                bin_plot(bd, desc="模型评分", ending=f" {tag}", save=p, figsize=(12, 7))
                _safe_close_figs()
                model_figs.append(p)
            except Exception as exc:
                logger.warning("生成模型评分分箱图失败 [数据集=%s, 文件=%s]: %s", tag, p, exc)

            p = str(output_dir / f"ks_{ds_key}.png")
            try:
                ks_plot(ds.score, ds.y, title=f"{tag} KS曲线", save=p, figsize=(12, 7))
                _safe_close_figs()
                model_figs.append(p)
            except Exception as exc:
                logger.warning("生成模型 KS 图失败 [数据集=%s, 文件=%s]: %s", tag, p, exc)

            if show_lift:
                p = str(output_dir / f"lift_{ds_key}.png")
                try:
                    lift_plot(ds.y, ds.y_proba, n_bins=20, title=f"{tag} LIFT曲线", save=p, figsize=(12, 7))
                    _safe_close_figs()
                    model_figs.append(p)
                except Exception as exc:
                    logger.warning("生成模型 LIFT 图失败 [数据集=%s, 文件=%s]: %s", tag, p, exc)

            p = str(output_dir / f"hist_{ds_key}.png")
            try:
                hist_plot(ds.score, y_true=ds.y, title=f"{tag} 评分分布", save=p, figsize=(12, 7))
                _safe_close_figs()
                model_figs.append(p)
            except Exception as exc:
                logger.warning("生成模型评分分布图失败 [数据集=%s, 文件=%s]: %s", tag, p, exc)

            if model_figs:
                paths[f"model_{ds_key}"] = model_figs

        contribution_table = self._get_model_input_table()
        if contribution_table["特征重要性%"].notna().any():
            p = str(output_dir / "feature_contribution.png")
            try:
                from ..core.viz.utils import save_figure

                figure = self._create_feature_contribution_figure(
                    contribution_table,
                    label_max_features=feature_contribution_label_max_features,
                )
                save_figure(figure, p)
                _safe_close_plot_result(figure)
                paths["feature_contribution"] = [p]
            except Exception as exc:
                logger.warning("生成入模特征贡献图失败 [文件=%s]: %s", p, exc)

        # --- 特征相关性图 ---
        importance = self.get_feature_importance()
        top_features = importance.index.tolist()
        if len(top_features) >= 2:
            p = str(output_dir / "feature_corr.png")
            try:
                corr_plot(self._datasets[self._train_key].X[top_features], annot=False, save=p)
                _safe_close_figs()
                paths["feature_corr"] = [p]
            except Exception as exc:
                logger.warning("生成特征相关性图失败 [数据集=训练集, 文件=%s]: %s", p, exc)

        # --- 逐特征图表（分箱图、分布图、PSI图） ---
        ds_keys = list(self._datasets.keys())
        for feat in top_features or self.feature_names:
            # 分箱图：按 train/test 顺序分组
            bin_figs: List[str] = []
            for ds_key, ds in self._datasets.items():
                p = str(output_dir / f"bin_{feat}_{ds_key}.png")
                try:
                    ft = self.get_feature_bin_table(feat, ds_key, max_n_bins=n_bins, method=bin_method, margins=True)
                    fd = ft.iloc[:-1].reset_index(drop=True) if len(ft) > 1 else ft
                    bin_plot(fd, desc=feat, ending=f" {ds.label}", save=p, figsize=(12, 7))
                    _safe_close_figs()
                    bin_figs.append(p)
                except Exception as exc:
                    logger.warning(
                        "生成特征分箱图失败 [特征=%s, 数据集=%s, 文件=%s]: %s",
                        feat,
                        ds.label,
                        p,
                        exc,
                    )
            if bin_figs:
                paths[f"feat_bin_{feat}"] = bin_figs

            # 特征KS分布图（替换直方图，显示特征对好坏样本的区分能力）
            # 处理缺失值和类别特征
            ks_figs: List[str] = []
            for ds_key, ds in self._datasets.items():
                p = str(output_dir / f"ks_{feat}_{ds_key}.png")
                try:
                    col_raw = ds.X[feat]
                    col = col_raw.dropna()
                    # 检查是否为类别特征或低基数的数值特征
                    is_categorical = col.dtype == "object" or (col.dtype in ["int64", "float64"] and col.nunique() <= 10)
                    if is_categorical:
                        # 类别特征跳过KS图
                        continue
                    y_f = ds.y.loc[col.index]
                    # 确保标签是二分类
                    if y_f.nunique() < 2:
                        continue
                    ks_plot(col, y_f, title=f"{ds.label} {feat}", save=p, figsize=(12, 7))
                    _safe_close_figs()
                    ks_figs.append(p)
                except Exception as exc:
                    logger.warning(
                        "生成特征 KS 图失败 [特征=%s, 数据集=%s, 文件=%s]: %s",
                        feat,
                        ds.label,
                        p,
                        exc,
                    )
            if ks_figs:
                paths[f"feat_hist_{feat}"] = ks_figs

            # PSI 图（训练集 vs 第一个非训练集），传入 y 以便图与表均包含坏样本率信息
            if len(ds_keys) >= 2:
                p = str(output_dir / f"psi_{feat}.png")
                try:
                    train_ds = self._datasets[ds_keys[0]]
                    test_ds = self._datasets[ds_keys[1]]
                    train_mask = train_ds.X[feat].notna()
                    test_mask = test_ds.X[feat].notna()
                    train_vals = train_ds.X[feat][train_mask]
                    test_vals = test_ds.X[feat][test_mask]
                    psi_y = np.concatenate(
                        [
                            train_ds.y.to_numpy()[train_mask.to_numpy()],
                            test_ds.y.to_numpy()[test_mask.to_numpy()],
                        ]
                    )
                    psi_result = psi_plot(
                        train_vals,
                        test_vals,
                        y=psi_y,
                        desc=feat,
                        save=p,
                        result=True,
                        plot=True,
                        figsize=(15, 8),
                    )
                    _safe_close_figs()
                    paths[f"feat_psi_{feat}"] = [p]
                    if isinstance(psi_result, pd.DataFrame):
                        tables[f"feat_psi_{feat}"] = psi_result
                except Exception as exc:
                    logger.warning("生成特征 PSI 图表失败 [特征=%s, 文件=%s]: %s", feat, p, exc)

        # --- 评分卡专属图表 ---
        if self._is_scorecard_model():
            lr_model = self._scorecard_lr_model()
            if lr_model is not None:
                p = str(output_dir / "plot_weights.png")
                try:
                    from ..core.viz import plot_weights as _pw

                    figure = _pw(lr_model, save=p)
                    _safe_close_plot_result(figure)
                    paths["model_weights"] = [p]
                except Exception as exc:
                    logger.warning("生成评分卡权重图失败 [文件=%s]: %s", p, exc)

            if len(ds_keys) >= 2:
                p = str(output_dir / "score_psi.png")
                try:
                    train_ds = self._datasets[ds_keys[0]]
                    test_ds = self._datasets[ds_keys[1]]
                    score_train = pd.Series(train_ds.score)
                    score_test = pd.Series(test_ds.score)
                    train_mask = score_train.notna()
                    test_mask = score_test.notna()
                    score_train = score_train[train_mask]
                    score_test = score_test[test_mask]
                    score_y = np.concatenate(
                        [
                            train_ds.y.to_numpy()[train_mask.to_numpy()],
                            test_ds.y.to_numpy()[test_mask.to_numpy()],
                        ]
                    )
                    score_psi_df = psi_plot(
                        score_train,
                        score_test,
                        y=score_y,
                        desc="模型评分",
                        save=p,
                        result=True,
                        plot=True,
                        figsize=(15, 8),
                    )
                    _safe_close_figs()
                    paths["score_psi"] = [p]
                    if isinstance(score_psi_df, pd.DataFrame):
                        tables["score_psi"] = score_psi_df
                except Exception as exc:
                    logger.warning("生成模型评分 PSI 图表失败 [文件=%s]: %s", p, exc)

        return paths, tables

    def _export_plots(
        self,
        output_dir: Path,
        n_bins: int = 10,
        bin_method: str = "quantile",
        amount_col: Optional[str] = None,
        show_lift: bool = True,
        feature_contribution_label_max_features: Optional[int] = 10,
    ) -> Tuple[Dict[str, List[str]], Dict[str, pd.DataFrame]]:
        """使用线程并行导出全部图表，并按原报告顺序汇总结果."""
        from ..core.viz import bin_plot, corr_plot, hist_plot, ks_plot, lift_plot, psi_plot

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        plot_groups = []
        task_labels = []

        def add_plot_group(label: str, group) -> None:
            plot_groups.append(group)
            task_labels.append(label)

        for ds_key, ds in self._datasets.items():
            def render_model_group(ds_key=ds_key, ds=ds):
                group_paths: Dict[str, List[str]] = {}
                model_figs: List[str] = []
                tag = ds.label

                path = str(output_dir / f"bin_{ds_key}.png")
                try:
                    table = self.get_bin_table(ds_key, method=bin_method, max_n_bins=n_bins, margins=True)
                    plot_data = table.iloc[:-1].reset_index(drop=True) if len(table) > 1 else table
                    figure = bin_plot(
                        plot_data,
                        desc="模型评分",
                        ending=f" {tag}",
                        save=path,
                        figsize=(12, 7),
                    )
                    _safe_close_plot_result(figure)
                    model_figs.append(path)
                except Exception as exc:
                    logger.warning("生成模型评分分箱图失败 [数据集=%s, 文件=%s]: %s", tag, path, exc)

                path = str(output_dir / f"ks_{ds_key}.png")
                try:
                    figure = ks_plot(ds.score, ds.y, title=f"{tag} KS曲线", save=path, figsize=(12, 7))
                    _safe_close_plot_result(figure)
                    model_figs.append(path)
                except Exception as exc:
                    logger.warning("生成模型 KS 图失败 [数据集=%s, 文件=%s]: %s", tag, path, exc)

                if show_lift:
                    path = str(output_dir / f"lift_{ds_key}.png")
                    try:
                        figure = lift_plot(
                            ds.y,
                            ds.y_proba,
                            n_bins=20,
                            title=f"{tag} LIFT曲线",
                            save=path,
                            figsize=(12, 7),
                        )
                        _safe_close_plot_result(figure)
                        model_figs.append(path)
                    except Exception as exc:
                        logger.warning("生成模型 LIFT 图失败 [数据集=%s, 文件=%s]: %s", tag, path, exc)

                path = str(output_dir / f"hist_{ds_key}.png")
                try:
                    figure = hist_plot(
                        ds.score,
                        y_true=ds.y,
                        title=f"{tag} 评分分布",
                        save=path,
                        figsize=(12, 7),
                    )
                    _safe_close_plot_result(figure)
                    model_figs.append(path)
                except Exception as exc:
                    logger.warning("生成模型评分分布图失败 [数据集=%s, 文件=%s]: %s", tag, path, exc)

                if model_figs:
                    group_paths[f"model_{ds_key}"] = model_figs
                return group_paths, {}

            add_plot_group(f"模型图表:{ds.label}", render_model_group)

        contribution_table = self._get_model_input_table()
        if contribution_table["特征重要性%"].notna().any():
            def render_contribution_group():
                group_paths: Dict[str, List[str]] = {}
                path = str(output_dir / "feature_contribution.png")
                try:
                    from ..core.viz.utils import save_figure

                    figure = self._create_feature_contribution_figure(
                        contribution_table,
                        label_max_features=feature_contribution_label_max_features,
                    )
                    save_figure(figure, path)
                    _safe_close_plot_result(figure)
                    group_paths["feature_contribution"] = [path]
                except Exception as exc:
                    logger.warning("生成入模特征贡献图失败 [文件=%s]: %s", path, exc)
                return group_paths, {}

            add_plot_group("入模特征贡献图", render_contribution_group)

        importance = self.get_feature_importance()
        top_features = importance.index.tolist()
        if len(top_features) >= 2:
            def render_corr_group():
                group_paths: Dict[str, List[str]] = {}
                path = str(output_dir / "feature_corr.png")
                try:
                    figure = corr_plot(self._datasets[self._train_key].X[top_features], annot=False, save=path)
                    _safe_close_plot_result(figure)
                    group_paths["feature_corr"] = [path]
                except Exception as exc:
                    logger.warning("生成特征相关性图失败 [数据集=训练集, 文件=%s]: %s", path, exc)
                return group_paths, {}

            add_plot_group("特征相关性图", render_corr_group)

        ds_keys = list(self._datasets.keys())
        for feature in top_features or self.feature_names:
            def render_feature_group(feature=feature):
                group_paths: Dict[str, List[str]] = {}
                group_tables: Dict[str, pd.DataFrame] = {}

                bin_figs: List[str] = []
                for ds_key, ds in self._datasets.items():
                    path = str(output_dir / f"bin_{feature}_{ds_key}.png")
                    try:
                        table = self.get_feature_bin_table(
                            feature,
                            ds_key,
                            max_n_bins=n_bins,
                            method=bin_method,
                            margins=True,
                        )
                        plot_data = table.iloc[:-1].reset_index(drop=True) if len(table) > 1 else table
                        figure = bin_plot(
                            plot_data,
                            desc=feature,
                            ending=f" {ds.label}",
                            save=path,
                            figsize=(12, 7),
                        )
                        _safe_close_plot_result(figure)
                        bin_figs.append(path)
                    except Exception as exc:
                        logger.warning(
                            "生成特征分箱图失败 [特征=%s, 数据集=%s, 文件=%s]: %s",
                            feature,
                            ds.label,
                            path,
                            exc,
                        )
                if bin_figs:
                    group_paths[f"feat_bin_{feature}"] = bin_figs

                ks_figs: List[str] = []
                for ds_key, ds in self._datasets.items():
                    path = str(output_dir / f"ks_{feature}_{ds_key}.png")
                    try:
                        column = ds.X[feature].dropna()
                        is_categorical = column.dtype == "object" or (
                            column.dtype in ["int64", "float64"] and column.nunique() <= 10
                        )
                        if is_categorical:
                            continue
                        target = ds.y.loc[column.index]
                        if target.nunique() < 2:
                            continue
                        figure = ks_plot(
                            column,
                            target,
                            title=f"{ds.label} {feature}",
                            save=path,
                            figsize=(12, 7),
                        )
                        _safe_close_plot_result(figure)
                        ks_figs.append(path)
                    except Exception as exc:
                        logger.warning(
                            "生成特征 KS 图失败 [特征=%s, 数据集=%s, 文件=%s]: %s",
                            feature,
                            ds.label,
                            path,
                            exc,
                        )
                if ks_figs:
                    group_paths[f"feat_hist_{feature}"] = ks_figs

                if len(ds_keys) >= 2:
                    path = str(output_dir / f"psi_{feature}.png")
                    try:
                        train_ds = self._datasets[ds_keys[0]]
                        test_ds = self._datasets[ds_keys[1]]
                        train_mask = train_ds.X[feature].notna()
                        test_mask = test_ds.X[feature].notna()
                        train_values = train_ds.X[feature][train_mask]
                        test_values = test_ds.X[feature][test_mask]
                        psi_target = np.concatenate(
                            [
                                train_ds.y.to_numpy()[train_mask.to_numpy()],
                                test_ds.y.to_numpy()[test_mask.to_numpy()],
                            ]
                        )
                        psi_result = psi_plot(
                            train_values,
                            test_values,
                            y=psi_target,
                            desc=feature,
                            save=path,
                            result=True,
                            plot=True,
                            figsize=(15, 8),
                        )
                        group_paths[f"feat_psi_{feature}"] = [path]
                        if isinstance(psi_result, pd.DataFrame):
                            group_tables[f"feat_psi_{feature}"] = psi_result
                    except Exception as exc:
                        logger.warning("生成特征 PSI 图表失败 [特征=%s, 文件=%s]: %s", feature, path, exc)

                return group_paths, group_tables

            add_plot_group(f"特征图表:{feature}", render_feature_group)

        if self._is_scorecard_model():
            def render_scorecard_group():
                group_paths: Dict[str, List[str]] = {}
                group_tables: Dict[str, pd.DataFrame] = {}
                lr_model = self._scorecard_lr_model()
                if lr_model is not None:
                    path = str(output_dir / "plot_weights.png")
                    try:
                        from ..core.viz import plot_weights

                        figure = plot_weights(lr_model, save=path)
                        _safe_close_plot_result(figure)
                        group_paths["model_weights"] = [path]
                    except Exception as exc:
                        logger.warning("生成评分卡权重图失败 [文件=%s]: %s", path, exc)

                if len(ds_keys) >= 2:
                    path = str(output_dir / "score_psi.png")
                    try:
                        train_ds = self._datasets[ds_keys[0]]
                        test_ds = self._datasets[ds_keys[1]]
                        score_train = pd.Series(train_ds.score)
                        score_test = pd.Series(test_ds.score)
                        train_mask = score_train.notna()
                        test_mask = score_test.notna()
                        score_train = score_train[train_mask]
                        score_test = score_test[test_mask]
                        score_target = np.concatenate(
                            [
                                train_ds.y.to_numpy()[train_mask.to_numpy()],
                                test_ds.y.to_numpy()[test_mask.to_numpy()],
                            ]
                        )
                        score_psi = psi_plot(
                            score_train,
                            score_test,
                            y=score_target,
                            desc="模型评分",
                            save=path,
                            result=True,
                            plot=True,
                            figsize=(15, 8),
                        )
                        group_paths["score_psi"] = [path]
                        if isinstance(score_psi, pd.DataFrame):
                            group_tables["score_psi"] = score_psi
                    except Exception as exc:
                        logger.warning("生成模型评分 PSI 图表失败 [文件=%s]: %s", path, exc)

                return group_paths, group_tables

            add_plot_group("评分卡图表", render_scorecard_group)

        rendered_groups = parallel_execute(
            _execute_plot_group,
            plot_groups,
            n_jobs=self.n_jobs,
            parallel_backend="threading",
            parallel_config={"batch_size": 1},
            task_labels=task_labels,
            default_backend="threading",
            preserve_exceptions=True,
        )

        paths: Dict[str, List[str]] = {}
        tables: Dict[str, pd.DataFrame] = {}
        for group_paths, group_tables in rendered_groups:
            paths.update(group_paths)
            tables.update(group_tables)
        return paths, tables

    # ---------- 9. 模型摘要 ----------

    # 模型摘要默认展示的指标（每个指标横向展开到各数据集）
    _SUMMARY_METRICS = ["KS", "AUC", "样本数", "坏样本率"]

    def summary(self) -> pd.DataFrame:
        """模型核心指标摘要表（多层列：统计指标 × 数据集；行：逾期指标）.

        参考 :mod:`hscredit.report.rule_strategy` 中拒绝规则策略表的展示方式：
        列为「统计指标 × 数据集」两层表头，便于同一指标在各数据集上横向对比；
        行为不同逾期指标。``overdue`` + ``dpds`` 多标签模式下每个逾期标签独占一行，
        单标签模式下仅一行（行名为目标列名）。

        :return: 行索引为 ``逾期指标``、列为 ``(统计指标, 数据集)`` 两层表头的 DataFrame

        **参考样例**

        >>> report = ModelReport(model, datasets={'train': tr, 'test': te},
        ...                           overdue=['MOB1'], dpds=[7, 3, 0])
        >>> report.summary()  # 行: MOB1@7 / MOB1@3 / MOB1@0；列: (KS, 训练集) ...
        """
        if self._summary_cache is not None:
            return self._summary_cache.copy()

        # 数据集按传入顺序展示（dict 的 key / list 顺序 / 训练集→测试集→跨时间验证集）
        ds_keys = list(self._datasets)
        ds_labels = [self._datasets[k].label for k in ds_keys]

        if self._is_multi_label():
            targets = list(self._label_names)
            display_map = self._overdue_label_map(separator="@") if self._is_overdue_cfg() else {}
            index_labels = [display_map.get(t, t) for t in targets]
        else:
            targets = [None]
            index_labels = [self._target_name or "target"]

        rows: List[Dict[tuple, Any]] = []
        for target_name in targets:
            metrics = self.get_metrics(target_name).set_index("统计项")
            row: Dict[tuple, Any] = {}
            for ds_label in ds_labels:
                for metric_name in self._SUMMARY_METRICS:
                    row[(metric_name, ds_label)] = metrics.at[metric_name, ds_label]
            rows.append(row)

        columns = pd.MultiIndex.from_tuples(
            [(metric, ds_label) for metric in self._SUMMARY_METRICS for ds_label in ds_labels],
            names=["统计指标", "数据集"],
        )
        index = pd.Index(index_labels, name="逾期指标")
        result = pd.DataFrame(rows, index=index, columns=columns)
        self._summary_cache = result.copy()
        return result

    # ---------- 10. 控制台输出 ----------

    def print_report(self, n_bins: int = 10, **kwargs) -> None:
        is_multi = self._is_multi_label()
        labels_arg = self._label_names if is_multi else None

        print("=" * 72)
        print("模型评估快速报告")
        print("=" * 72)
        print("\n【模型性能指标】")
        if is_multi:
            # 多标签模式：逐逾期标签展示「统计指标 × 数据集」摘要表
            print(self.summary().to_string())
        else:
            print(self.get_metrics().to_string(index=False))

        importance = self.get_feature_importance(top_n=10)
        if not importance.empty:
            print("\n【Top 10 特征重要性】")
            print(importance.to_string())

        for ds_key, ds in self._datasets.items():
            print(f"\n【{ds.label}评分分箱效果】")
            print(self.get_bin_table(ds_key, max_n_bins=n_bins, labels=labels_arg).to_string(index=False))
        print("\n" + "=" * 72)

    # ---------- 11. to_excel ----------

    def _precompute_excel_tables(
        self,
        *,
        n_bins: int,
        bin_method: str,
        amount_col: Optional[str],
        date_col: Optional[str],
        show_importance: bool,
    ) -> None:
        """在创建绘图与工作簿对象前完成报告核心数据计算。"""
        self.summary()
        if self._is_multi_label():
            for label in self._label_names:
                self.get_metrics(label)
        else:
            self.get_metrics()
        importance = self.get_feature_importance()
        self.get_features_corr()
        if show_importance:
            self.get_features_describe()

        labels_arg = self._label_names if self._is_multi_label() else None
        for dataset, ds in self._datasets.items():
            try:
                self.get_bin_table(
                    dataset,
                    method=bin_method,
                    max_n_bins=n_bins,
                    margins=True,
                    labels=labels_arg,
                )
            except Exception as exc:
                raise RuntimeError(f"生成评分分箱失败 [数据集={ds.label}]") from exc
            if amount_col:
                try:
                    self.get_bin_table(
                        dataset,
                        method=bin_method,
                        max_n_bins=n_bins,
                        amount_col=amount_col,
                        margins=True,
                        labels=labels_arg,
                    )
                except Exception as exc:
                    raise RuntimeError(f"生成金额口径评分分箱失败 [数据集={ds.label}, 金额字段={amount_col}]") from exc
        self._get_top_n_lift_table(labels=labels_arg)
        if amount_col:
            self._get_top_n_lift_table(amount_col=amount_col, labels=labels_arg)
        if date_col:
            self._get_monthly_metrics(date_col)
            self._get_monthly_psi_matrix(date_col)

        feature_list = importance.index.tolist() if not importance.empty else self.feature_names
        for feature in feature_list:
            for dataset, ds in self._datasets.items():
                try:
                    self.get_feature_bin_table(
                        feature,
                        dataset,
                        max_n_bins=n_bins,
                        method=bin_method,
                        margins=True,
                        labels=labels_arg,
                    )
                except Exception as exc:
                    raise RuntimeError(f"生成特征有效性分箱失败 [特征={feature}, 数据集={ds.label}]") from exc
                if amount_col:
                    try:
                        self.get_feature_bin_table(
                            feature,
                            dataset,
                            max_n_bins=n_bins,
                            method=bin_method,
                            margins=True,
                            amount_col=amount_col,
                            labels=labels_arg,
                        )
                    except Exception as exc:
                        raise RuntimeError(f"生成金额口径特征分箱失败 [特征={feature}, 数据集={ds.label}, 金额字段={amount_col}]") from exc

    def get_model_explanation(self) -> Dict[str, Any]:
        """返回显式配置的结构化模型解释结果。"""
        if not self.explain_config["enabled"]:
            raise ValidationError("模型解释未启用，请设置 explain_config={'enabled': True}")
        if self._model_explanation_cache is not None:
            return self._model_explanation_cache
        config = dict(self.explain_config)
        train = self._datasets[self._train_key].X
        feature_names = self.feature_names or list(train.columns)
        if config["data"] is None:
            config["data"] = train[feature_names]
        elif isinstance(config["data"], pd.DataFrame):
            config["data"] = config["data"][feature_names]
        if config["background_data"] is None:
            config["background_data"] = train[feature_names]
        elif isinstance(config["background_data"], pd.DataFrame):
            config["background_data"] = config["background_data"][feature_names]
        if config["stability_mode"] == "refit":
            config["X_train"] = train[feature_names] if config["X_train"] is None else config["X_train"]
            config["y_train"] = self._datasets[self._train_key].y if config["y_train"] is None else config["y_train"]
            config["X_validation"] = config["data"] if config["X_validation"] is None else config["X_validation"]
        explanation = build_model_explanation(self.model, config)
        self._model_explanation_cache = explanation
        return explanation

    def to_excel(
        self,
        filepath: str,
        *,
        n_bins: int = 10,
        bin_method: str = "quantile",
        amount_col: Optional[str] = None,
        date_col: Optional[str] = None,
        date_freq: Optional[str] = None,
        group_col: Optional[str] = None,
        with_plots: bool = True,
        model_name: Optional[str] = None,
        project_desc: Optional[str] = None,
        feature_map: Optional[Dict[str, str]] = None,
        feature_info: Optional[pd.DataFrame] = None,
        show_lift: bool = True,
        show_importance: bool = True,
        feature_contribution_label_max_features: Optional[int] = 10,
        data_source: Optional[str] = None,
        loc_cols: Optional[Union[str, List[str]]] = None,
    ) -> str:
        """事务性生成 Excel；失败时恢复进入调用前的全部派生缓存。"""
        return self._run_cache_transaction(
            self._to_excel_impl,
            filepath,
            n_bins=n_bins,
            bin_method=bin_method,
            amount_col=amount_col,
            date_col=date_col,
            date_freq=date_freq,
            group_col=group_col,
            with_plots=with_plots,
            model_name=model_name,
            project_desc=project_desc,
            feature_map=feature_map,
            feature_info=feature_info,
            show_lift=show_lift,
            show_importance=show_importance,
            feature_contribution_label_max_features=feature_contribution_label_max_features,
            data_source=data_source,
            loc_cols=loc_cols,
        )

    def _to_excel_impl(
        self,
        filepath: str,
        *,
        n_bins: int = 10,
        bin_method: str = "quantile",
        amount_col: Optional[str] = None,
        date_col: Optional[str] = None,
        date_freq: Optional[str] = None,
        group_col: Optional[str] = None,
        with_plots: bool = True,
        model_name: Optional[str] = None,
        project_desc: Optional[str] = None,
        feature_map: Optional[Dict[str, str]] = None,
        feature_info: Optional[pd.DataFrame] = None,
        show_lift: bool = True,
        show_importance: bool = True,
        feature_contribution_label_max_features: Optional[int] = 10,
        data_source: Optional[str] = None,
        loc_cols: Optional[Union[str, List[str]]] = None,
    ) -> str:
        """生成多 Sheet 结构的 Excel 模型报告.

        Sheet 结构：
        - 目录
        - 1-基本信息（项目目标、样本统计、分月/分组分布）
        - 2-模型性能（指标、TOP n%、PSI矩阵、分箱效果）
        - 3-入模变量分析（重要性、相关性、逐特征分箱/KS/PSI）

        :param show_lift: 是否生成并插入 LIFT 曲线；LIFT 数值表始终保留
        :param show_importance: 是否显示入模变量重要性及分布汇总章节
        :param feature_contribution_label_max_features: 贡献图显示数据标签的最大特征数；默认 10，
            ``None`` 始终显示，0 始终隐藏
        :param loc_cols: 定位字段（订单号等），支持 str 或 List[str]，仅用于生产订单测试用例
        """
        from ..excel import ExcelWriter, dataframe2excel as _dataframe2excel

        def dataframe2excel(*args, **kwargs):
            """模型报告由大量小表组成，统一使用保样式快速写入。"""
            data = args[0] if args else kwargs.get("data")
            if isinstance(data, pd.DataFrame) and data.empty:
                for option in ("percent_cols", "custom_cols", "condition_cols", "color_cols"):
                    kwargs[option] = None
            worksheet = kwargs.get("sheet_name")
            worksheet_title = getattr(worksheet, "title", worksheet if isinstance(worksheet, str) else None)
            if worksheet_title in {"2-模型性能", "3-入模变量分析"}:
                kwargs.setdefault("auto_width", True)
            kwargs["speed"] = "fast"
            return _dataframe2excel(*args, **kwargs)

        self._precompute_excel_tables(
            n_bins=n_bins,
            bin_method=bin_method,
            amount_col=amount_col,
            date_col=date_col,
            show_importance=show_importance,
        )
        feature_contribution_label_max_features = self._validate_feature_contribution_label_limit(
            feature_contribution_label_max_features
        )
        model_name = model_name or self.model.__class__.__name__

        plot_paths: Dict[str, List[str]] = {}
        psi_tables: Dict[str, pd.DataFrame] = {}
        if with_plots:
            plot_dir = Path(filepath).parent / f"{Path(filepath).stem}_assets"
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                plot_paths, psi_tables = self._export_plots(
                    plot_dir,
                    n_bins=n_bins,
                    bin_method=bin_method,
                    amount_col=amount_col,
                    show_lift=show_lift,
                    feature_contribution_label_max_features=feature_contribution_label_max_features,
                )

        writer = ExcelWriter()
        title_pattern = re.compile(r"^(?:[一二三四五六七八九十]+|\d+(?:\.\d+)*)、")

        def _report_title_level(row_idx: int, value: Any) -> Optional[int]:
            if not isinstance(value, str) or not value:
                return None
            if row_idx == 2:
                return 0
            if not title_pattern.match(value):
                return None
            prefix = value.split("、", 1)[0]
            if re.match(r"^\d+(?:\.\d+)*$", prefix):
                return prefix.count(".") + 1
            return 1

        def _content_max_col(worksheet, start_row: int, end_row: int, start_col: int = 2) -> int:
            if end_row < start_row:
                return start_col
            max_content_col = start_col
            for row in worksheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col,
                max_col=worksheet.max_column,
            ):
                for cell in row:
                    if cell.value is not None:
                        max_content_col = max(max_content_col, cell.column)
            for cell_range in worksheet.merged_cells.ranges:
                if cell_range.max_row >= start_row and cell_range.min_row <= end_row and cell_range.max_col >= start_col:
                    max_content_col = max(max_content_col, cell_range.max_col)
            return max_content_col

        def _merge_report_title(worksheet, row_idx: int, end_col: int, start_col: int = 2) -> None:
            for cell_range in list(worksheet.merged_cells.ranges):
                if cell_range.min_row == row_idx and cell_range.max_row == row_idx and cell_range.min_col == start_col:
                    worksheet.unmerge_cells(str(cell_range))
            if end_col > start_col:
                worksheet.merge_cells(
                    start_row=row_idx,
                    start_column=start_col,
                    end_row=row_idx,
                    end_column=end_col,
                )

        def _adjust_report_title_merges(worksheet) -> None:
            title_rows = [(row_cell.row, level) for row_cells in worksheet.iter_rows(min_col=2, max_col=2) for row_cell in row_cells for level in [_report_title_level(row_cell.row, row_cell.value)] if level is not None]
            for idx, (row_idx, level) in enumerate(title_rows):
                if level == 0:
                    end_row_for_title = worksheet.max_row
                else:
                    next_boundary = next(
                        (next_row for next_row, next_level in title_rows[idx + 1 :] if next_level <= level),
                        None,
                    )
                    end_row_for_title = next_boundary - 1 if next_boundary is not None else worksheet.max_row
                max_content_col = _content_max_col(worksheet, row_idx + 1, end_row_for_title)
                _merge_report_title(worksheet, row_idx, max_content_col)

        def _insert_required_hyperlink(worksheet, cell, hyperlink: str, purpose: str) -> None:
            """插入报告导航链接；失败时保留明确上下文并终止不完整报告。"""
            try:
                writer.insert_hyperlink2sheet(worksheet, cell, hyperlink=hyperlink)
            except Exception as exc:
                coordinate = writer.get_cell_space(cell)
                raise RuntimeError(f"生成{purpose}失败 [工作表={worksheet.title}, 单元格={coordinate}, 目标={hyperlink}]") from exc

        # ============================================================
        # 目录 Sheet
        # ============================================================
        contents = pd.DataFrame(
            [
                {"序号": 1, "内容": "1-基本信息", "备注": "项目目标、样本选取、样本坏率分布"},
                {"序号": 2, "内容": "2-模型性能", "备注": "模型效果、区分度、稳定性等内容"},
                {"序号": 3, "内容": "3-入模变量分析", "备注": "模型变量有效性及不同数据集分箱情况"},
                {"序号": 4, "内容": "4-稳定性分析", "备注": "评分分布、PSI、CSI等稳定性分析"},
                {"序号": 5, "内容": "5-模型参数", "备注": "模型选型、参数及评分卡配置"},
                {"序号": 6, "内容": "6-模型部署需求", "备注": "入模变量信息及测试用例"},
            ]
        )
        if self.explain_config["enabled"]:
            contents.loc[len(contents)] = {
                "序号": 7,
                "内容": "7-模型解释",
                "备注": "SHAP贡献、解释稳定性、代表样本及原因码",
            }

        ws = writer.get_sheet_by_name("目录")
        end_row, _ = writer.insert_value2sheet(ws, (2, 2), value="模型评估报告", style="header_middle")
        end_row, _ = dataframe2excel(contents, writer, sheet_name=ws, start_row=end_row + 1, left_cols=["内容", "备注"])

        for i, row in contents.iterrows():
            target_cell = writer.get_cell_space((2, 2))
            _insert_required_hyperlink(
                ws,
                (end_row - len(contents) + i, 3),
                hyperlink=f"#'{row['内容']}'!{target_cell}",
                purpose="目录链接",
            )

        _, _ = writer.insert_value2sheet(ws, (end_row + 1, 2), value="版本号:", style="middle", end_space=(end_row + 1, 2))
        end_row, _ = writer.insert_value2sheet(ws, (end_row + 1, 3), value="V1.0", style="middle", end_space=(end_row + 1, 4))
        _, _ = writer.insert_value2sheet(ws, (end_row, 2), value="创建日期:", style="middle", end_space=(end_row, 2))
        end_row, _ = writer.insert_value2sheet(ws, (end_row, 3), value=date.today().strftime("%Y-%m-%d"), style="middle", end_space=(end_row, 4))
        _, _ = writer.insert_value2sheet(ws, (end_row, 2), value="模型名称:", style="middle", end_space=(end_row, 2))
        end_row, _ = writer.insert_value2sheet(ws, (end_row, 3), value=model_name, style="middle", end_space=(end_row, 4))
        writer.adjust_columns_width(ws, start_col=2, end_col=4)

        # ============================================================
        # 1-基本信息 Sheet
        # ============================================================
        ws = writer.get_sheet_by_name("1-基本信息")
        end_row, _ = writer.insert_value2sheet(ws, (2, 2), value="一、基本信息", style="header_middle")
        _insert_required_hyperlink(ws, (2, 2), hyperlink="#'目录'!B2", purpose="返回目录链接")

        # 1.1 项目目标
        end_row, _ = writer.insert_value2sheet(ws, (end_row + 2, 2), value="1、项目目标", style="header_middle", align={"horizontal": "left"})
        desc_text = project_desc or f"使用 {model_name} 模型进行信用风险评估"
        end_row, _ = writer.insert_value2sheet(
            ws,
            (end_row, 2),
            value=desc_text,
            style="middle",
            align={"horizontal": "left"},
        )

        # 1.2 数据样本描述
        end_row, _ = writer.insert_value2sheet(ws, (end_row + 2, 2), value="2、数据样本描述", style="header_middle", align={"horizontal": "left"})

        label_text = self._target_name or "TARGET"

        # ---- Step 2: 合并所有数据集，计算整体各逾期标签的逾期数据 ----
        # 如果有日期字段，先提取全局日期范围
        global_date_prefix = ""
        if date_col:
            all_dates = []
            for ds in self._datasets.values():
                if ds is not None and date_col in ds.X.columns:
                    all_dates.append(pd.to_datetime(ds.X[date_col]))
            if all_dates:
                all_dates_combined = pd.concat(all_dates, ignore_index=True).dropna()
                if not all_dates_combined.empty:
                    global_date_prefix = f"{all_dates_combined.min().strftime('%Y-%m-%d')} ~ " f"{all_dates_combined.max().strftime('%Y-%m-%d')}  "
        sample_interval = global_date_prefix if global_date_prefix else ""

        is_multi = self._is_multi_label()
        ds_keys_list = list(self._datasets.keys())
        dataset_labels = [self._datasets[k].label for k in ds_keys_list]

        # ---- Step 2: 整体样本描述（多标签时逐标签展示坏样本率） ----
        target_specific_totals = self.del_grey and self._is_overdue_cfg()
        overall_n = sum(len(self._datasets[k].y) for k in ds_keys_list)
        if overall_n > 0:
            if is_multi:
                display_labels = self._overdue_label_map(separator="@")
                label_parts = []
                for lbl in self._label_names:
                    all_y = np.concatenate([self._get_y(k, lbl) for k in ds_keys_list])
                    valid_y = all_y[np.isfinite(all_y)]
                    bad_rate = round(float(valid_y.mean()) * 100, 2) if len(valid_y) else 0.0
                    if target_specific_totals:
                        label_parts.append(
                            f"{display_labels.get(lbl, lbl)}: 样本数 {len(valid_y)}, 坏样本率 {bad_rate}%"
                        )
                    else:
                        label_parts.append(f"{display_labels.get(lbl, lbl)}: {bad_rate}%")
                overall_desc = ", ".join(label_parts)
                if not target_specific_totals:
                    overall_desc = f"样本数: {overall_n}, " + overall_desc
            else:
                overall_bad = int(sum(int(self._datasets[k].y.sum()) for k in ds_keys_list))
                overall_bad_rate = overall_bad / overall_n * 100
                overall_desc = f"样本数: {overall_n}, " f"{label_text}: {round(overall_bad_rate, 2)}%"
        else:
            overall_desc = "N/A"

        # ---- Step 3: 固定描述行 ----
        data_source_str = data_source if data_source else "N/A"
        fixed_rows: List[Dict[str, Any]] = [
            {"统计项": "样本区间", "统计内容": sample_interval or "N/A"},
            {"统计项": "整体样本", "统计内容": overall_desc},
            {"统计项": "模型名称", "统计内容": model_name or "N/A"},
            {"统计项": "取样逻辑", "统计内容": project_desc or "N/A"},
            {"统计项": "数据源", "统计内容": data_source_str},
        ]

        # ---- Step 4: 各数据集描述行（与整体样本同一张表，多标签逐标签展示坏样本率） ----
        ds_rows: List[Dict[str, Any]] = []
        for ds_key, ds_label in zip(ds_keys_list, dataset_labels):
            n_samples = len(self._datasets[ds_key].y)
            if is_multi:
                display_labels = self._overdue_label_map(separator="@")
                label_parts = []
                for lbl in self._label_names:
                    y_values = np.asarray(self._get_y(ds_key, lbl), dtype=float)
                    valid_y = y_values[np.isfinite(y_values)]
                    bad_rate = round(float(valid_y.mean()) * 100, 2) if len(valid_y) else 0.0
                    if target_specific_totals:
                        label_parts.append(
                            f"{display_labels.get(lbl, lbl)}: 样本数 {len(valid_y)}, 坏样本率 {bad_rate}%"
                        )
                    else:
                        label_parts.append(f"{display_labels.get(lbl, lbl)}: {bad_rate}%")
                content = ", ".join(label_parts)
                if not target_specific_totals:
                    content = f"样本数: {n_samples}, " + content
            else:
                bad_rate = round(float(self._datasets[ds_key].y.mean()) * 100, 2)
                content = f"样本数: {n_samples}, {label_text}: {bad_rate}%"
            ds_rows.append({"统计项": ds_label, "统计内容": content})

        desc_df = pd.DataFrame(fixed_rows + ds_rows)
        end_row, _ = dataframe2excel(desc_df, writer, sheet_name=ws, start_row=end_row + 1, left_cols=["统计项", "统计内容"])

        # 1.3 数据样本统计
        end_row, _ = writer.insert_value2sheet(ws, (end_row + 2, 2), value="3、数据样本统计", style="header_middle", align={"horizontal": "left"})
        ds_keys_list = list(self._datasets.keys())
        dataset_labels = [self._datasets[k].label for k in ds_keys_list]

        y_maps = [{label: self._get_y(ds_key, label) for label in self._label_names} if is_multi else {label_text: self._get_y(ds_key)} for ds_key in ds_keys_list]
        stat_df, stat_pct_cols = build_sample_stats_table(
            dataset_labels,
            y_maps,
            self._label_names if is_multi else [label_text],
            display_labels=self._overdue_label_map(separator="@") if is_multi else None,
            flat_total_col="样本数",
            n_jobs=self.n_jobs,
            parallel_backend=self.parallel_backend,
            parallel_config=self.parallel_config,
            target_specific_totals=target_specific_totals,
        )
        if is_multi:
            stat_start_row = end_row + 1
            end_row, _ = dataframe2excel(
                stat_df,
                writer,
                sheet_name=ws,
                start_row=stat_start_row,
                index=True,
                percent_cols=stat_pct_cols,
            )
        else:
            end_row, _ = dataframe2excel(stat_df, writer, sheet_name=ws, start_row=end_row + 1, percent_cols=stat_pct_cols)

        # 1.4 样本分布情况
        freq_label_map = {"D": "日", "W": "周", "M": "月", "Q": "季度", "Y": "年"}
        if date_col or group_col:
            end_row, _ = writer.insert_value2sheet(ws, (end_row + 2, 2), value="4、样本分布情况", style="header_middle", align={"horizontal": "left"})

            def _write_distribution(group_of_ds, title: str):
                """构建并写入一张「数据集 × 数据分组」的分布表（数据集逐段堆叠）。

                :param group_of_ds: 函数 (ds_key, ds) -> 分组值序列(与 ds.y 等长) 或 None
                """
                distribution_dataset_labels: List[str] = []
                distribution_y_maps: List[Dict[str, Any]] = []
                group_values: List[Any] = []
                for ds_key, ds_label in zip(ds_keys_list, dataset_labels):
                    ds = self._datasets[ds_key]
                    gvals = group_of_ds(ds_key, ds)
                    if gvals is None:
                        continue
                    distribution_dataset_labels.append(ds_label)
                    group_values.append(gvals)
                    distribution_y_maps.append({label: self._get_y(ds_key, label) for label in self._label_names} if is_multi else {label_text: self._get_y(ds_key)})
                if not distribution_y_maps:
                    return
                dist_df, pct = build_group_distribution_table(
                    distribution_dataset_labels,
                    distribution_y_maps,
                    group_values,
                    self._label_names if is_multi else [label_text],
                    display_labels=self._overdue_label_map(separator="@") if is_multi else None,
                    n_jobs=self.n_jobs,
                    parallel_backend=self.parallel_backend,
                    parallel_config=self.parallel_config,
                    target_specific_totals=target_specific_totals,
                )
                result = dataframe2excel(
                    dist_df,
                    writer,
                    sheet_name=ws,
                    title=title,
                    start_row=end_row + 1,
                    index=not isinstance(dist_df.columns, pd.MultiIndex),
                    percent_cols=pct,
                )
                return result

            # 时间分布
            if date_col:
                freq = date_freq or "M"
                period_col_name = freq_label_map.get(freq, freq)

                def _period_of(ds_key, ds, _freq=freq):
                    if date_col not in ds.X.columns:
                        return None
                    dates = pd.to_datetime(ds.X[date_col])
                    try:
                        return dates.dt.to_period(_freq).astype(str).values
                    except Exception:
                        return dates.dt.to_period("M").astype(str).values

                ret = _write_distribution(_period_of, title=f"{period_col_name}度分布")
                if ret is not None:
                    end_row, _ = ret

            # 分组分布
            if group_col:

                def _group_of(ds_key, ds):
                    if group_col not in ds.X.columns:
                        return None
                    return ds.X[group_col].values

                ret = _write_distribution(_group_of, title=f"{group_col}分布")
                if ret is not None:
                    end_row, _ = ret

        # ============================================================
        # 2-模型性能 Sheet
        # ============================================================
        ws = writer.get_sheet_by_name("2-模型性能")
        end_row, _ = writer.insert_value2sheet(ws, (2, 2), value="二、模型性能评估", style="header_middle")
        _insert_required_hyperlink(ws, (2, 2), hyperlink="#'目录'!B2", purpose="返回目录链接")

        section_idx = 1

        # 2.1 性能指标
        end_row, _ = writer.insert_value2sheet(
            ws,
            (end_row + 2, 2),
            value=f"{section_idx}、模型性能验证指标",
            style="header_middle",
            align={"horizontal": "left"},
        )
        if is_multi:
            # 多标签模式：列为「标签 × 数据集」二级表头，行为 KS/AUC/样本数/坏样本率。
            import itertools
            from ..core.metrics import ks, auc

            metric_items = ["KS", "AUC", "样本总数", "坏样本率"]
            col_tuples = list(itertools.product(self._label_names, dataset_labels))
            multi_cols = pd.MultiIndex.from_tuples(col_tuples, names=["统计项", "统计指标"])
            rows_list: List[Dict[tuple, Any]] = []
            for metric in metric_items:
                row: Dict[tuple, Any] = {}
                for lbl in self._label_names:
                    for ds_key, ds_label in zip(ds_keys_list, dataset_labels):
                        y_arr, proba = self._get_valid_target_arrays(ds_key, lbl)
                        if metric == "样本总数":
                            val = len(y_arr)
                        elif metric == "坏样本率":
                            val = float(y_arr.mean())
                        elif metric == "KS":
                            val = _safe_binary_metric(ks, y_arr, proba)
                        else:
                            val = _safe_binary_metric(auc, y_arr, proba)
                        row[(lbl, ds_label)] = val
                rows_list.append(row)
            metrics_df = pd.DataFrame(rows_list, index=metric_items, columns=multi_cols)
            metrics_df.index.name = "统计指标"
            metrics_start_row = end_row + 1
            end_row, _ = dataframe2excel(
                metrics_df,
                writer,
                sheet_name=ws,
                start_row=metrics_start_row,
                index=True,
                percent_rows=["KS", "AUC", "坏样本率"],
                custom_rows=["样本总数"],
                custom_format="#,##0",
            )
        else:
            metrics = self.get_metrics().replace({"统计项": {"样本数": "样本总数"}}).set_index("统计项")
            end_row, _ = dataframe2excel(
                metrics,
                writer,
                sheet_name=ws,
                start_row=end_row + 1,
                index=True,
                percent_rows=["KS", "AUC", "坏样本率"],
                custom_rows=["样本总数"],
                custom_format="#,##0",
            )
        section_idx += 1

        # 2.2 分月模型效果
        if date_col:
            monthly_metrics = self._get_monthly_metrics(date_col)
            if not monthly_metrics.empty:
                end_row, _ = writer.insert_value2sheet(
                    ws,
                    (end_row + 2, 2),
                    value=f"{section_idx}、分月模型效果",
                    style="header_middle",
                    align={"horizontal": "left"},
                )
                end_row, _ = dataframe2excel(
                    monthly_metrics,
                    writer,
                    sheet_name=ws,
                    start_row=end_row + 1,
                    percent_cols=["坏样本率", "KS", "AUC"],
                )
                section_idx += 1

        # 2.3 模型尾部区分能力
        end_row, _ = writer.insert_value2sheet(
            ws,
            (end_row + 2, 2),
            value=f"{section_idx}、模型尾部区分能力（TOP n%）",
            style="header_middle",
            align={"horizontal": "left"},
        )
        pct_keys = ["TOP 1%", "TOP 3%", "TOP 5%", "TOP 10%", "TOTAL"]
        if is_multi:
            # 多标签合并模式：一次性生成 MultiIndex 列的 LIFT 表
            if amount_col:
                # 订单口径 + 金额口径左右并排
                lift_table = self._get_top_n_lift_table(
                    percentiles=(0.01, 0.03, 0.05, 0.10),
                    labels=self._label_names,
                )
                lift_amt = self._get_top_n_lift_table(
                    percentiles=(0.01, 0.03, 0.05, 0.10),
                    amount_col=amount_col,
                    labels=self._label_names,
                )
                table_start = end_row + 1
                end_row1, end_col1 = dataframe2excel(
                    lift_table,
                    writer,
                    sheet_name=ws,
                    title="订单口径",
                    start_row=table_start,
                    start_col=2,
                    percent_cols=list(lift_table.columns),
                    index=True,
                )
                end_row2, _ = dataframe2excel(
                    lift_amt,
                    writer,
                    sheet_name=ws,
                    title="金额口径",
                    start_row=table_start,
                    start_col=end_col1 + 1,
                    percent_cols=list(lift_amt.columns),
                    index=True,
                )
                end_row = max(end_row1, end_row2)
                writer.insert_value2sheet(ws, (table_start + 2, 2), value="统计指标", style="header_left")
                writer.insert_value2sheet(
                    ws,
                    (table_start + 2, end_col1 + 1),
                    value="统计指标",
                    style="header_left",
                )
                from openpyxl.utils import get_column_letter

                # 金额口径表起始列为 end_col1 + 1，占据 索引层级 + 数据列 共 nlevels+ncols
                # 列，故其最后一列为 (end_col1 + 1) + nlevels + ncols - 1。
                filter_end_col = end_col1 + 1 + lift_amt.index.nlevels + len(lift_amt.columns) - 1
                header_row = table_start + lift_table.columns.nlevels + 1
                writer.add_auto_filter(
                    ws,
                    f"B{header_row}:{get_column_letter(filter_end_col)}{end_row - 1}",
                )
            else:
                lift_table = self._get_top_n_lift_table(
                    percentiles=(0.01, 0.03, 0.05, 0.10),
                    labels=self._label_names,
                )
                table_start = end_row + 1
                end_row, _ = dataframe2excel(
                    lift_table,
                    writer,
                    sheet_name=ws,
                    start_row=table_start,
                    percent_cols=list(lift_table.columns),
                    index=True,
                )
                writer.insert_value2sheet(ws, (table_start, 2), value="统计指标", style="header_middle")
                from openpyxl.utils import get_column_letter

                filter_end_col = 2 + lift_table.index.nlevels + len(lift_table.columns) - 1
                writer.add_auto_filter(
                    ws,
                    f"B{table_start + lift_table.columns.nlevels}:" f"{get_column_letter(filter_end_col)}{end_row - 1}",
                )
            section_idx += 1
        elif amount_col:
            lift_table = self._get_top_n_lift_table(percentiles=(0.01, 0.03, 0.05, 0.10), amount_col=None)
            lift_amt = self._get_top_n_lift_table(percentiles=(0.01, 0.03, 0.05, 0.10), amount_col=amount_col)
            table_start = end_row + 1
            end_row1, end_col1 = dataframe2excel(
                lift_table,
                writer,
                sheet_name=ws,
                title="订单口径",
                start_row=table_start,
                start_col=2,
                percent_cols=pct_keys,
            )
            end_row2, _ = dataframe2excel(
                lift_amt,
                writer,
                sheet_name=ws,
                title="金额口径",
                start_row=table_start,
                start_col=end_col1 + 2,
                percent_cols=pct_keys,
            )
            end_row = max(end_row1, end_row2)
            n_lift_cols = len(lift_table.columns)
            filter_end_col = end_col1 + 2 + n_lift_cols - 1
            from openpyxl.utils import get_column_letter

            writer.add_auto_filter(ws, f"B{table_start + 2}:{get_column_letter(filter_end_col)}{end_row - 1}")
        else:
            lift_table = self._get_top_n_lift_table()
            table_start = end_row + 3
            end_row, _ = dataframe2excel(
                lift_table,
                writer,
                sheet_name=ws,
                start_row=table_start,
                percent_cols=pct_keys,
            )
            from openpyxl.utils import get_column_letter

            writer.add_auto_filter(ws, f"B{table_start}:{get_column_letter(len(lift_table.columns) + 1)}{end_row - 1}")
        if not is_multi:
            section_idx += 1

        # 2.4 分月PSI矩阵
        if date_col:
            psi_matrix = self._get_monthly_psi_matrix(date_col)
            if not psi_matrix.empty:
                end_row, _ = writer.insert_value2sheet(
                    ws,
                    (end_row + 2, 2),
                    value=f"{section_idx}、分月对比PSI",
                    style="header_middle",
                    align={"horizontal": "left"},
                )
                end_row, _ = dataframe2excel(psi_matrix, writer, sheet_name=ws, start_row=end_row + 1, index=True)
                section_idx += 1

        # 2.5 各数据集评分排序性
        for ds_key, ds in self._datasets.items():
            tag = ds.label
            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{section_idx}、{tag}评分排序性",
                style="header_middle",
                align={"horizontal": "left"},
            )

            figs = plot_paths.get(f"model_{ds_key}", [])

            # 先插入图表（同一行、左右排列，避免 figures 参数导致标题与分箱表之间出现图）
            img_start_row = end_row + 1
            current_col = 2
            max_img_end_row = img_start_row
            for fig in figs:
                try:
                    img_end_row, _ = writer.insert_pic2sheet(ws, fig, (img_start_row, current_col), figsize=(500, 300))
                    current_col = _next_image_col(ws, current_col, 500)
                    max_img_end_row = max(max_img_end_row, img_end_row)
                except Exception as exc:
                    logger.warning("插入可选图表失败 [工作表=%s, 文件=%s]: %s", ws.title, fig, exc)
            if figs:
                end_row = max_img_end_row

            if is_multi:
                # 多标签合并模式：所有标签合并为一个 MultiIndex 列的分箱表
                if amount_col:
                    order_table = self.get_bin_table(
                        ds_key,
                        method=bin_method,
                        max_n_bins=n_bins,
                        margins=True,
                        labels=self._label_names,
                    )
                    amount_table = self.get_bin_table(
                        ds_key,
                        method=bin_method,
                        max_n_bins=n_bins,
                        amount_col=amount_col,
                        margins=True,
                        labels=self._label_names,
                    )
                    # 从 MultiIndex 列中提取百分位列
                    pct_cols = [c for c in order_table.columns if (c[1] if isinstance(c, tuple) else c) in self._PERCENT_COLS]
                    cond_cols = [c for c in order_table.columns if (c[1] if isinstance(c, tuple) else c) in self._CONDITION_COLS]
                    amt_pct = [c for c in amount_table.columns if (c[1] if isinstance(c, tuple) else c) in self._PERCENT_COLS]
                    amt_cond = [c for c in amount_table.columns if (c[1] if isinstance(c, tuple) else c) in self._CONDITION_COLS]
                    order_start_row = end_row + 1
                    order_end_row, order_end_col = dataframe2excel(
                        order_table,
                        writer,
                        sheet_name=ws,
                        title=f"{tag} 订单口径",
                        start_row=order_start_row,
                        start_col=2,
                        percent_cols=pct_cols,
                        condition_cols=cond_cols,
                        condition_color="F76E6C",
                    )
                    amount_end_row, _ = dataframe2excel(
                        amount_table,
                        writer,
                        sheet_name=ws,
                        title=f"{tag} 金额口径",
                        start_row=order_start_row,
                        start_col=order_end_col + 1,
                        percent_cols=amt_pct,
                        condition_cols=amt_cond,
                        condition_color="F76E6C",
                    )
                    end_row = max(order_end_row, amount_end_row)
                else:
                    order_table = self.get_bin_table(
                        ds_key,
                        method=bin_method,
                        max_n_bins=n_bins,
                        margins=True,
                        labels=self._label_names,
                    )
                    pct_cols = [c for c in order_table.columns if (c[1] if isinstance(c, tuple) else c) in self._PERCENT_COLS]
                    cond_cols = [c for c in order_table.columns if (c[1] if isinstance(c, tuple) else c) in self._CONDITION_COLS]
                    end_row, _ = dataframe2excel(
                        order_table,
                        writer,
                        sheet_name=ws,
                        title=f"{tag} 评分有效性",
                        start_row=end_row + 1,
                        percent_cols=pct_cols,
                        condition_cols=cond_cols,
                        condition_color="F76E6C",
                    )
            elif amount_col:
                # 订单口径和金额口径左右并排（参考3-入模变量分析的分箱表布局）
                order_table = self.get_bin_table(ds_key, method=bin_method, max_n_bins=n_bins, margins=True)
                order_start_row = end_row + 1
                pct_cols = [c for c in self._PERCENT_COLS if c in order_table.columns]
                cond_cols = [c for c in self._CONDITION_COLS if c in order_table.columns]
                order_end_row, order_end_col = dataframe2excel(
                    order_table,
                    writer,
                    sheet_name=ws,
                    title=f"{tag} 订单口径",
                    start_row=order_start_row,
                    start_col=2,
                    percent_cols=pct_cols,
                    condition_cols=cond_cols,
                    condition_color="F76E6C",
                )
                try:
                    amount_table = self.get_bin_table(ds_key, method=bin_method, max_n_bins=n_bins, amount_col=amount_col, margins=True)
                    amt_pct = [c for c in self._PERCENT_COLS if c in amount_table.columns]
                    amt_cond = [c for c in self._CONDITION_COLS if c in amount_table.columns]
                    amount_end_row, _ = dataframe2excel(
                        amount_table,
                        writer,
                        sheet_name=ws,
                        title=f"{tag} 金额口径",
                        start_row=order_start_row,
                        start_col=order_end_col + 1,
                        percent_cols=amt_pct,
                        condition_cols=amt_cond,
                        condition_color="F76E6C",
                    )
                    end_row = max(order_end_row, amount_end_row)
                except Exception as exc:
                    raise RuntimeError(f"生成金额口径评分分箱失败 [数据集={tag}, 金额字段={amount_col}]") from exc
            else:
                order_table = self.get_bin_table(ds_key, method=bin_method, max_n_bins=n_bins, margins=True)
                pct_cols = [c for c in self._PERCENT_COLS if c in order_table.columns]
                cond_cols = [c for c in self._CONDITION_COLS if c in order_table.columns]
                end_row, _ = dataframe2excel(
                    order_table,
                    writer,
                    sheet_name=ws,
                    title=f"{tag} 评分有效性",
                    start_row=end_row + 1,
                    percent_cols=pct_cols,
                    condition_cols=cond_cols,
                    condition_color="F76E6C",
                )
            section_idx += 1

        # ============================================================
        # ============================================================
        # 3-入模变量分析 Sheet
        # ============================================================
        ws = writer.get_sheet_by_name("3-入模变量分析")
        end_row, _ = writer.insert_value2sheet(ws, (2, 2), value="三、入模变量分析", style="header_middle")
        _insert_required_hyperlink(ws, (2, 2), hyperlink="#'目录'!B2", purpose="返回目录链接")

        feature_section = 1
        feature_name_col: Optional[int] = None
        features_summary_rows: Dict[str, int] = {}
        if show_importance:
            # 3.1 入模变量重要性及分布情况
            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{feature_section}、入模变量重要性及分布情况",
                style="header_middle",
                align={"horizontal": "left"},
            )
            features_summary = self._get_features_summary()
            if "特征名" not in features_summary.columns:
                index_name = features_summary.index.name or "index"
                features_summary = features_summary.reset_index().rename(columns={index_name: "特征名"})
            features_summary_start_row = end_row + 1
            end_row, _ = dataframe2excel(
                features_summary,
                writer,
                sheet_name=ws,
                start_row=features_summary_start_row,
                right_cols=[0],
                condition_color='F76E6C',
                percent_cols=['缺失率', 'KS', 'PSI', '众数占比', '零值率', '负值率', '重复率'],
                condition_cols=['IV', 'KS'],
            )
            feature_name_col = 2 + features_summary.columns.get_loc("特征名")
            features_summary_rows = {str(feat): features_summary_start_row + features_summary.columns.nlevels + position for position, feat in enumerate(features_summary["特征名"])}
            feature_section += 1

        # 入模变量相关性
        end_row, _ = writer.insert_value2sheet(
            ws,
            (end_row + 2, 2),
            value=f"{feature_section}、入模变量相关性",
            style="header_middle",
            align={"horizontal": "left"},
        )
        corr_df = self.get_features_corr()
        corr_figs = plot_paths.get("feature_corr", [])
        end_row, _ = dataframe2excel(
            corr_df,
            writer,
            sheet_name=ws,
            start_row=end_row + 1,
            percent_cols=corr_df.columns.tolist(),
            index=True,
            figures=corr_figs,
            right_cols=[0],
        )
        feature_section += 1

        # 入模变量有效性分析
        effectiveness_section = feature_section
        end_row, _ = writer.insert_value2sheet(
            ws,
            (end_row + 2, 2),
            value=f"{effectiveness_section}、入模变量有效性分析",
            style="header_middle",
            align={"horizontal": "left"},
        )

        importance = self.get_feature_importance()
        feature_list = importance.index.tolist() if not importance.empty else self.feature_names
        ds_keys_list = list(self._datasets.keys())

        for i, feat in enumerate(feature_list):
            feature_title_row = end_row + 2
            end_row, _ = writer.insert_value2sheet(
                ws,
                (feature_title_row, 2),
                value=f"{effectiveness_section}.{i + 1}、{feat} 有效性分析",
                style="header_middle",
                align={"horizontal": "left"},
            )

            summary_row = features_summary_rows.get(str(feat))
            if summary_row is not None and feature_name_col is not None:
                _insert_required_hyperlink(
                    ws,
                    (summary_row, feature_name_col),
                    hyperlink=f"#'{ws.title}'!B{feature_title_row}",
                    purpose=f"特征导航链接 [特征={feat}]",
                )
                _insert_required_hyperlink(
                    ws,
                    (feature_title_row, 2),
                    hyperlink=f"#'{ws.title}'!{writer.get_cell_space((summary_row, feature_name_col))}",
                    purpose=f"特征返回链接 [特征={feat}]",
                )

            # 插入图表（同一行、左右排列，避免 figures 参数导致标题与分箱表之间出现图）
            bin_figs = plot_paths.get(f"feat_bin_{feat}", [])
            hist_figs = plot_paths.get(f"feat_hist_{feat}", [])
            all_figs = bin_figs + hist_figs
            img_start_row = end_row + 1
            current_col = 2
            max_img_end_row = img_start_row
            for fig in all_figs:
                try:
                    img_end_row, _ = writer.insert_pic2sheet(ws, fig, (img_start_row, current_col), figsize=(500, 300))
                    current_col = _next_image_col(ws, current_col, 500)
                    max_img_end_row = max(max_img_end_row, img_end_row)
                except Exception as exc:
                    logger.warning("插入可选图表失败 [工作表=%s, 文件=%s]: %s", ws.title, fig, exc)
            if all_figs:
                end_row = max_img_end_row  # 跳过图片占用的所有行，避免重叠

            # 每个数据集分别输出；多标签仅改变表头，不改变数据集粒度。
            for ds_key, ds in self._datasets.items():
                try:
                    labels_arg = self._label_names if is_multi else None
                    ft = self.get_feature_bin_table(
                        feat,
                        ds_key,
                        max_n_bins=n_bins,
                        method=bin_method,
                        margins=True,
                        labels=labels_arg,
                    )
                    ft_pct = [c for c in ft.columns if c[-1] in self._PERCENT_COLS] if isinstance(ft.columns, pd.MultiIndex) else [c for c in self._PERCENT_COLS if c in ft.columns]
                    ft_cond = [c for c in ft.columns if c[-1] in self._CONDITION_COLS] if isinstance(ft.columns, pd.MultiIndex) else [c for c in self._CONDITION_COLS if c in ft.columns]

                    if amount_col:
                        table_start = end_row + 1
                        order_end_row, order_end_col = dataframe2excel(
                            ft,
                            writer,
                            sheet_name=ws,
                            title=f"{ds.label} 订单口径",
                            start_row=table_start,
                            start_col=2,
                            percent_cols=ft_pct,
                            condition_cols=ft_cond,
                            condition_color="F76E6C",
                        )
                        ft_amt = self.get_feature_bin_table(
                            feat,
                            ds_key,
                            max_n_bins=n_bins,
                            method=bin_method,
                            margins=True,
                            amount_col=amount_col,
                            labels=labels_arg,
                        )
                        amt_pct = [c for c in ft_amt.columns if c[-1] in self._PERCENT_COLS] if isinstance(ft_amt.columns, pd.MultiIndex) else [c for c in self._PERCENT_COLS if c in ft_amt.columns]
                        amt_cond = [c for c in ft_amt.columns if c[-1] in self._CONDITION_COLS] if isinstance(ft_amt.columns, pd.MultiIndex) else [c for c in self._CONDITION_COLS if c in ft_amt.columns]
                        amount_end_row, _ = dataframe2excel(
                            ft_amt,
                            writer,
                            sheet_name=ws,
                            title=f"{ds.label} 金额口径",
                            start_row=table_start,
                            start_col=order_end_col + 1,
                            percent_cols=amt_pct,
                            condition_cols=amt_cond,
                            condition_color="F76E6C",
                        )
                        end_row = max(order_end_row, amount_end_row)
                    else:
                        end_row, _ = dataframe2excel(
                            ft,
                            writer,
                            sheet_name=ws,
                            title=ds.label,
                            start_row=end_row + 1,
                            percent_cols=ft_pct,
                            condition_cols=ft_cond,
                            condition_color="F76E6C",
                        )
                except Exception as exc:
                    raise RuntimeError(f"生成特征有效性分箱表失败 [特征={feat}, 数据集={ds.label}]") from exc

            # PSI 图表和数据表
            psi_fig_paths = plot_paths.get(f"feat_psi_{feat}", [])
            psi_df = psi_tables.get(f"feat_psi_{feat}")
            if psi_fig_paths:
                for fig_path in psi_fig_paths:
                    try:
                        end_row, _ = writer.insert_pic2sheet(ws, fig_path, (end_row + 1, 2), figsize=(500, 300))
                    except Exception as exc:
                        logger.warning(
                            "插入可选图表失败 [工作表=%s, 文件=%s]: %s",
                            ws.title,
                            fig_path,
                            exc,
                        )
            if isinstance(psi_df, pd.DataFrame) and not psi_df.empty:
                psi_percent_cols = [
                    column
                    for column in psi_df.columns
                    if "占比" in str(column)
                    or "样本率" in str(column)
                    or str(column) in {"实际% - 预期%", "分档PSI值", "总体PSI值"}
                ]
                psi_condition_cols = [
                    column
                    for column in psi_df.columns
                    if str(column) in {"实际% - 预期%", "分档PSI值"}
                ]
                end_row, _ = dataframe2excel(
                    psi_df,
                    writer,
                    sheet_name=ws,
                    title="PSI稳定性分析",
                    start_row=end_row + 1,
                    percent_cols=psi_percent_cols,
                    condition_cols=psi_condition_cols,
                )

        # ============================================================
        # 4-稳定性分析 Sheet
        # ============================================================
        ws = writer.get_sheet_by_name("4-稳定性分析")
        end_row, _ = writer.insert_value2sheet(ws, (2, 2), value="四、模型稳定性分析", style="header_middle")
        _insert_required_hyperlink(ws, (2, 2), hyperlink="#'目录'!B2", purpose="返回目录链接")

        stab_section = 1

        # 4.1 评分分布统计
        end_row, _ = writer.insert_value2sheet(
            ws,
            (end_row + 2, 2),
            value=f"{stab_section}、评分分布统计",
            style="header_middle",
            align={"horizontal": "left"},
        )
        score_dist_rows: List[Dict[str, Any]] = []
        for ds_key, ds in self._datasets.items():
            sc = ds.score
            row: Dict[str, Any] = {"数据集": ds.label}
            row["样本数"] = len(sc)
            row["均值"] = float(np.nanmean(sc))
            row["标准差"] = float(np.nanstd(sc))
            row["最小值"] = float(np.nanmin(sc))
            row["25%分位"] = float(np.nanpercentile(sc, 25))
            row["中位数"] = float(np.nanpercentile(sc, 50))
            row["75%分位"] = float(np.nanpercentile(sc, 75))
            row["最大值"] = float(np.nanmax(sc))
            score_dist_rows.append(row)
        score_dist_df = pd.DataFrame(score_dist_rows)
        end_row, _ = dataframe2excel(
            score_dist_df,
            writer,
            sheet_name=ws,
            start_row=end_row + 1,
        )
        stab_section += 1

        # 4.2 评分PSI矩阵（数据集两两对比）
        if len(self._datasets) >= 2:
            from ..core.metrics import psi as _psi

            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{stab_section}、评分PSI对比矩阵",
                style="header_middle",
                align={"horizontal": "left"},
            )
            ds_keys_list = list(self._datasets.keys())
            labels = [self._datasets[k].label for k in ds_keys_list]
            psi_matrix = pd.DataFrame(np.nan, index=labels, columns=labels)
            for i, k1 in enumerate(ds_keys_list):
                for j, k2 in enumerate(ds_keys_list):
                    if i == j:
                        psi_matrix.iloc[i, j] = 0.0
                    else:
                        try:
                            psi_matrix.iloc[i, j] = _psi(self._datasets[k1].score, self._datasets[k2].score)
                        except Exception as exc:
                            logger.warning(
                                "计算评分 PSI 失败 [基准数据集=%s, 对比数据集=%s]: %s",
                                self._datasets[k1].label,
                                self._datasets[k2].label,
                                exc,
                            )
            end_row, psi_matrix_end_col = dataframe2excel(
                psi_matrix,
                writer,
                sheet_name=ws,
                start_row=end_row + 1,
                index=True,
            )

            # 评分PSI参考阈值说明
            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 1, 2),
                value="PSI参考标准：<0.1 稳定 | 0.1~0.25 略变 | >0.25 不稳定",
                style="middle",
                end_space=(end_row + 1, psi_matrix_end_col - 1),
                align={"horizontal": "left"},
            )
            stab_section += 1

        # 4.3 评分漂移分析（以训练集为基准）
        if len(self._datasets) >= 2:
            base_key = self._train_key
            base_label = self._datasets[base_key].label
            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{stab_section}、评分漂移分析（vs {base_label}）",
                style="header_middle",
                align={"horizontal": "left"},
            )
            drift_rows: List[Dict[str, Any]] = []
            base_scores = self._datasets[base_key].score
            for ds_key, ds in self._datasets.items():
                if ds_key == base_key:
                    continue
                sc = ds.score
                drift = {
                    "数据集": ds.label,
                    "vs": base_label,
                    "均值偏移": float(np.nanmean(sc) - np.nanmean(base_scores)),
                    "均值偏移%": float((np.nanmean(sc) - np.nanmean(base_scores)) / (np.nanstd(base_scores) + 1e-9)),
                    "中位数偏移": float(np.nanmedian(sc) - np.nanmedian(base_scores)),
                    "好样本(评分>600)占比": float((sc > 600).sum() / len(sc)),
                    "坏样本(评分<500)占比": float((sc < 500).sum() / len(sc)),
                }
                drift_rows.append(drift)
            if drift_rows:
                drift_df = pd.DataFrame(drift_rows)
                pct_cols = [c for c in drift_df.columns if "%" in c or "占比" in c]
                end_row, _ = dataframe2excel(
                    drift_df,
                    writer,
                    sheet_name=ws,
                    start_row=end_row + 1,
                    percent_cols=pct_cols,
                )
            stab_section += 1

        # 4.4 逐特征PSI稳定性表
        if len(self._datasets) >= 2:
            from ..core.metrics import psi as _psi_feat

            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{stab_section}、入模特征PSI稳定性",
                style="header_middle",
                align={"horizontal": "left"},
            )
            importance = self.get_feature_importance()
            feat_list = importance.index.tolist() if not importance.empty else self.feature_names
            psi_rows: List[Dict[str, Any]] = []
            base_key = self._train_key
            base_ds = self._datasets[base_key]
            other_ds_keys = [k for k in self._datasets if k != base_key]

            for feat in feat_list:
                row: Dict[str, Any] = {"特征": feat}
                has_psi = False
                for dk in other_ds_keys:
                    if dk in self._datasets and feat in self._datasets[dk].X.columns:
                        try:
                            psi_val = _psi_feat(base_ds.X[feat], self._datasets[dk].X[feat])
                            row[f"PSI({self._datasets[dk].label})"] = psi_val
                            has_psi = True
                        except Exception:
                            row[f"PSI({self._datasets[dk].label})"] = np.nan
                if has_psi:
                    psi_rows.append(row)
            if psi_rows:
                psi_feat_df = pd.DataFrame(psi_rows)
                end_row, _ = dataframe2excel(
                    psi_feat_df,
                    writer,
                    sheet_name=ws,
                    start_row=end_row + 1,
                )
            stab_section += 1

        # ============================================================
        # 5-模型参数 Sheet
        # ============================================================
        ws = writer.get_sheet_by_name("5-模型参数")
        end_row, _ = writer.insert_value2sheet(ws, (2, 2), value="五、模型选型及参数", style="header_middle")
        _insert_required_hyperlink(ws, (2, 2), hyperlink="#'目录'!B2", purpose="返回目录链接")

        param_section = 1

        # 5.1 模型选型
        end_row, _ = writer.insert_value2sheet(
            ws,
            (end_row + 2, 2),
            value=f"{param_section}、模型选型",
            style="header_middle",
            align={"horizontal": "left"},
        )
        end_row, _ = writer.insert_value2sheet(ws, (end_row, 2), value=model_name, style="middle", align={"horizontal": "left"})
        param_section += 1

        # 5.2 模型参数
        end_row, _ = writer.insert_value2sheet(
            ws,
            (end_row + 2, 2),
            value=f"{param_section}、模型参数",
            style="header_middle",
            align={"horizontal": "left"},
        )
        params_str = ""
        if hasattr(self.model, "get_params"):
            try:
                params_str = str(self.model.get_params())
            except Exception as exc:
                logger.warning("读取模型参数失败 [模型=%s]: %s", model_name, exc)
        if not params_str and hasattr(self.model, "__dict__"):
            params_str = str({k: v for k, v in self.model.__dict__.items() if not k.startswith("_") and not callable(v)})
        end_row, _ = writer.insert_value2sheet(ws, (end_row, 2), value=params_str or "N/A", style="middle", align={"horizontal": "left"})
        param_section += 1

        # 5.3 入模特征列表
        end_row, _ = writer.insert_value2sheet(
            ws,
            (end_row + 2, 2),
            value=f"{param_section}、入模特征列表",
            style="header_middle",
            align={"horizontal": "left"},
        )
        features_df = self._get_model_input_table(feature_map)
        features_start_row = end_row + 1
        left_columns = [column for column in ("入参字段", "字段名称") if column in features_df.columns]
        features_end_row, features_end_col = dataframe2excel(
            features_df,
            writer,
            sheet_name=ws,
            start_row=features_start_row,
            percent_cols=["特征重要性%", "累积特征重要性%"],
            left_cols=left_columns,
        )
        end_row = features_end_row
        contribution_figures = plot_paths.get("feature_contribution", [])
        if contribution_figures:
            try:
                contribution_end_row, _ = writer.insert_pic2sheet(
                    ws,
                    contribution_figures[0],
                    (features_start_row, features_end_col + 1),
                    figsize=(500, 300),
                )
                end_row = max(end_row, contribution_end_row)
            except Exception as exc:
                logger.warning(
                    "插入入模特征贡献图失败 [工作表=%s, 文件=%s]: %s",
                    ws.title,
                    contribution_figures[0],
                    exc,
                )
        param_section += 1

        direct_lr_summary = self._direct_lr_summary(feature_map)
        if not direct_lr_summary.empty:
            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{param_section}、逻辑回归拟合结果",
                style="header_middle",
                align={"horizontal": "left"},
            )
            end_row, _ = dataframe2excel(
                direct_lr_summary,
                writer,
                sheet_name=ws,
                start_row=end_row + 1,
                left_cols=[column for column in ("入参字段", "Features", "变量", "feature", "特征", "字段名称") if column in direct_lr_summary.columns],
            )
            param_section += 1

        score_conversion_sections = self._get_score_conversion_sections()
        if score_conversion_sections is not None:
            for title, key in (
                ("评分转换器选型", "selection"),
                ("评分转换基础参数配置", "params"),
                ("概率转评分公式", "formula"),
            ):
                end_row, _ = writer.insert_value2sheet(
                    ws,
                    (end_row + 2, 2),
                    value=f"{param_section}、{title}",
                    style="header_middle",
                    align={"horizontal": "left"},
                )
                section_table = score_conversion_sections[key]
                if not section_table.empty:
                    end_row, _ = dataframe2excel(
                        section_table,
                        writer,
                        sheet_name=ws,
                        start_row=end_row + 1,
                        left_cols=list(section_table.columns),
                    )
                param_section += 1

        # 5.4+ 评分卡专属内容（hscredit / toad / scorecardpipeline）
        if self._is_scorecard_model():
            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{param_section}、逻辑回归拟合结果",
                style="header_middle",
                align={"horizontal": "left"},
            )
            weights_figs = plot_paths.get("model_weights", [])
            if weights_figs:
                for fig_path in weights_figs:
                    try:
                        end_row, _ = writer.insert_pic2sheet(ws, fig_path, (end_row + 1, 2), figsize=(500, 300))
                    except Exception as exc:
                        logger.warning(
                            "插入可选图表失败 [工作表=%s, 文件=%s]: %s",
                            ws.title,
                            fig_path,
                            exc,
                        )
            lr_summary = self._scorecard_lr_summary()
            if not lr_summary.empty:
                end_row, _ = dataframe2excel(
                    lr_summary,
                    writer,
                    sheet_name=ws,
                    start_row=end_row + 1,
                    title="逻辑回归系数",
                    left_cols=[column for column in ("变量",) if column in lr_summary.columns],
                )
            param_section += 1

            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{param_section}、评分卡基础参数配置",
                style="header_middle",
                align={"horizontal": "left"},
            )
            scale_df = self._scorecard_scale_table()
            if not scale_df.empty:
                end_row, _ = dataframe2excel(
                    scale_df,
                    writer,
                    sheet_name=ws,
                    start_row=end_row + 1,
                    right_cols=["刻度项"],
                    left_cols=["备注"],
                )
            param_section += 1

            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{param_section}、评分卡转换公式",
                style="header_middle",
                align={"horizontal": "left"},
            )
            scorecard_formula = self._scorecard_formula_table()
            if not scorecard_formula.empty:
                end_row, _ = dataframe2excel(
                    scorecard_formula,
                    writer,
                    sheet_name=ws,
                    start_row=end_row + 1,
                    left_cols=list(scorecard_formula.columns),
                )
            param_section += 1

            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{param_section}、评分卡分值表",
                style="header_middle",
                align={"horizontal": "left"},
            )
            sc_points = self._scorecard_points_table(feature_map)
            if not sc_points.empty:
                end_row, _ = dataframe2excel(
                    sc_points,
                    writer,
                    sheet_name=ws,
                    start_row=end_row + 1,
                    right_cols=["对应分数", "变量分箱", "变量名称"],
                )
            param_section += 1

            end_row, _ = writer.insert_value2sheet(
                ws,
                (end_row + 2, 2),
                value=f"{param_section}、评分、ODDS与逾期率参考表",
                style="header_middle",
                align={"horizontal": "left"},
            )
            odds_ref = self._scorecard_odds_reference_table()
            if not odds_ref.empty:
                percent_columns = [column for column in odds_ref.columns if "逾期率" in str(column)]
                end_row, _ = dataframe2excel(
                    odds_ref,
                    writer,
                    sheet_name=ws,
                    start_row=end_row + 1,
                    percent_cols=percent_columns,
                )
            param_section += 1

            if len(self._datasets) >= 2:
                end_row, _ = writer.insert_value2sheet(
                    ws,
                    (end_row + 2, 2),
                    value=f"{param_section}、评分稳定性分析",
                    style="header_middle",
                    align={"horizontal": "left"},
                )
                score_psi_matrix = self._score_psi_matrix()
                if not score_psi_matrix.empty:
                    end_row, _ = dataframe2excel(
                        score_psi_matrix,
                        writer,
                        sheet_name=ws,
                        start_row=end_row + 1,
                        index=True,
                    )
                score_psi_figs = plot_paths.get("score_psi", [])
                if score_psi_figs:
                    for fig_path in score_psi_figs:
                        try:
                            end_row, _ = writer.insert_pic2sheet(ws, fig_path, (end_row + 1, 2), figsize=(500, 300))
                        except Exception as exc:
                            logger.warning(
                                "插入可选图表失败 [工作表=%s, 文件=%s]: %s",
                                ws.title,
                                fig_path,
                                exc,
                            )
                score_psi_df = psi_tables.get("score_psi")
                if isinstance(score_psi_df, pd.DataFrame) and not score_psi_df.empty:
                    end_row, _ = dataframe2excel(score_psi_df, writer, sheet_name=ws, start_row=end_row + 1, title="评分PSI")
                param_section += 1

        # ============================================================
        # 6-模型部署需求 Sheet
        # ============================================================
        ws = writer.get_sheet_by_name("6-模型部署需求")
        end_row, _ = writer.insert_value2sheet(ws, (2, 2), value="六、模型部署需求", style="header_middle")
        _insert_required_hyperlink(ws, (2, 2), hyperlink="#'目录'!B2", purpose="返回目录链接")

        # 6.1 入模变量信息
        end_row, _ = writer.insert_value2sheet(ws, (end_row + 2, 2), value="1、入模变量信息", style="header_middle", align={"horizontal": "left"})
        if feature_info is not None and isinstance(feature_info, pd.DataFrame) and not feature_info.empty:
            end_row, _ = dataframe2excel(feature_info, writer, sheet_name=ws, start_row=end_row + 1)
        else:
            fi_rows: List[Dict[str, Any]] = []
            for idx, feat in enumerate(self.feature_names):
                fi_rows.append(
                    {
                        "序号": idx + 1,
                        "特征名称": feat,
                        "特征含义": (feature_map or {}).get(feat, ""),
                        "字段类型": str(self._datasets[self._train_key].X[feat].dtype),
                        "缺失值处理": "默认处理",
                    }
                )
            end_row, _ = dataframe2excel(pd.DataFrame(fi_rows), writer, sheet_name=ws, start_row=end_row + 1)

        # 6.2 生产订单测试用例
        end_row, _ = writer.insert_value2sheet(ws, (end_row + 2, 2), value="2、生产订单测试用例", style="header_middle", align={"horizontal": "left"})
        try:
            train_ds = self._datasets[self._train_key]
            sample_n = min(5, len(train_ds.X))
            sample_X = train_ds.X[self.feature_names].iloc[:sample_n].copy()

            # 支持定位字段（订单号等）显示在最前方
            if loc_cols:
                if isinstance(loc_cols, str):
                    loc_cols = [loc_cols]
                loc_cols = [c for c in loc_cols if c in train_ds.X.columns]

            test_cases = sample_X.reset_index(drop=True)
            if loc_cols:
                loc_df = train_ds.X[loc_cols].iloc[:sample_n].reset_index(drop=True)
                for i, col in enumerate(loc_cols):
                    test_cases.insert(i, col, loc_df[col])
            test_cases.insert(0, "序号", range(1, sample_n + 1))
            test_cases["模型分数"] = train_ds.score[:sample_n]
            end_row, _ = dataframe2excel(
                test_cases,
                writer,
                sheet_name=ws,
                start_row=end_row + 1,
                auto_filter=True,
            )
        except Exception as exc:
            raise RuntimeError("生成生产订单测试用例失败") from exc

        # ============================================================
        # 7-模型解释 Sheet（显式启用）
        # ============================================================
        if self.explain_config["enabled"]:
            explanation = self.get_model_explanation()
            ws = writer.get_sheet_by_name("7-模型解释")
            end_row, _ = writer.insert_value2sheet(ws, (2, 2), value="七、模型解释", style="header_middle")
            _insert_required_hyperlink(ws, (2, 2), hyperlink="#'目录'!B2", purpose="返回目录链接")
            if "失败原因" in explanation:
                end_row, _ = dataframe2excel(
                    pd.DataFrame([{"失败原因": explanation["失败原因"]}]),
                    writer,
                    sheet_name=ws,
                    start_row=end_row + 2,
                    title="1、解释计算状态",
                )
            else:
                metadata_table = pd.DataFrame(
                    [{"项目": key, "值": str(value)} for key, value in explanation["元信息"].items()]
                )
                end_row, _ = dataframe2excel(
                    metadata_table, writer, sheet_name=ws, start_row=end_row + 2, title="1、解释范围与元信息"
                )
                for title, key in [
                    ("2、全局SHAP重要性", "全局解释"),
                    ("3、SHAP贡献相关性", "相关性"),
                    ("4、特征聚类", "特征聚类"),
                    ("5、主要交互", "交互"),
                    ("6、解释稳定性", "稳定性"),
                    ("7、代表样本", "代表样本"),
                    ("8、业务原因码", "原因码"),
                ]:
                    table = explanation[key]
                    end_row, _ = dataframe2excel(
                        table, writer, sheet_name=ws, start_row=end_row + 2, title=title, index=key == "相关性"
                    )
                local_tables = []
                for sample_id, table in explanation["样本解释"].items():
                    local = table.copy()
                    local.insert(0, "代表样本", sample_id)
                    local_tables.append(local)
                if local_tables:
                    end_row, _ = dataframe2excel(
                        pd.concat(local_tables, ignore_index=True),
                        writer,
                        sheet_name=ws,
                        start_row=end_row + 2,
                        title="9、代表样本局部贡献",
                    )
                end_row, _ = writer.insert_value2sheet(
                    ws,
                    (end_row + 2, 2),
                    value="说明：模型贡献不等于因果关系，也不应单独作为审批或授信依据。",
                    style="middle",
                )
                if with_plots:
                    figure = explanation["解释器"].plot_explanation_overview(
                        explanation["解释结果"], show=False
                    )
                    try:
                        end_row, _ = writer.insert_pic2sheet(ws, figure, (end_row + 2, 2), figsize=(800, 550))
                    finally:
                        import matplotlib.pyplot as plt

                        plt.close(figure)

        # ============================================================
        # 保存
        # ============================================================
        basic_info_sheet = writer.workbook["1-基本信息"]
        if basic_info_sheet.max_column >= 4:
            writer.adjust_columns_width(
                basic_info_sheet,
                start_col=4,
                end_col=basic_info_sheet.max_column,
            )
        for sheet_name in [
            "目录",
            "1-基本信息",
            "2-模型性能",
            "3-入模变量分析",
            "4-稳定性分析",
            "5-模型参数",
            "6-模型部署需求",
            "7-模型解释",
        ]:
            if sheet_name in writer.workbook.sheetnames:
                _adjust_report_title_merges(writer.workbook[sheet_name])
        writer.save(filepath)
        return filepath

    # ---------- 12. to_dict ----------

    def to_dict(self) -> Dict[str, Any]:
        """事务性返回报告字典；任一表失败时恢复调用前缓存。"""
        return self._run_cache_transaction(self._to_dict_impl)

    def _to_dict_impl(self) -> Dict[str, Any]:
        labels_arg = self._label_names if self._is_multi_label() else None
        result: Dict[str, Any] = {
            "summary": self.summary().reset_index().to_dict(orient="records"),
            "metrics": self.get_metrics().to_dict(orient="records"),
            "feature_importance": self.get_feature_importance().reset_index().to_dict(orient="records"),
        }
        for ds_key in self._datasets:
            result[f"bin_table_{ds_key}"] = self.get_bin_table(ds_key, labels=labels_arg).to_dict(orient="records")
        if self.explain_config["enabled"]:
            result["模型解释"] = explanation_to_dict(self.get_model_explanation())
        return result


# ---------------------------------------------------------------------------
# 快捷函数
# ---------------------------------------------------------------------------


def auto_model_report(
    model,
    datasets: Optional[Union[List, Dict]] = None,
    X_train=None,
    y_train=None,
    X_test=None,
    y_test=None,
    X_oot=None,
    y_oot=None,
    feature_names: Optional[List[str]] = None,
    target: Optional[Union[str, Dict]] = None,
    overdue: Optional[Union[str, List[str]]] = None,
    dpds: Optional[Union[int, float, List[Union[int, float]]]] = None,
    del_grey: bool = False,
    excel_path: Optional[str] = None,
    verbose: bool = True,
    n_bins: int = 10,
    bin_method: str = "quantile",
    amount_col: Optional[str] = None,
    date_col: Optional[str] = None,
    date_freq: Optional[str] = None,
    group_col: Optional[str] = None,
    with_plots: bool = True,
    model_name: Optional[str] = None,
    project_desc: Optional[str] = None,
    feature_map: Optional[Dict[str, str]] = None,
    feature_info: Optional[pd.DataFrame] = None,
    show_lift: bool = True,
    show_importance: bool = True,
    feature_contribution_label_max_features: Optional[int] = 10,
    data_source: Optional[str] = None,
    loc_cols: Optional[Union[str, List[str]]] = None,
    method: Union[str, Callable] = "predict_proba",
    method_kwargs: Optional[Dict[str, Any]] = None,
    n_jobs=-1,
    parallel_backend: Optional[str] = None,
    parallel_config: Optional[Dict[str, Any]] = None,
    **kwargs,
) -> ModelReport:
    """一键生成模型报告.

    数据集传入支持三种方式，内部统一规整为 ``{数据集名称: (X, y)}`` 结构：

    1. datasets 为 dict：直接以 key 作为数据集名称，
       如 ``{'建模集': df, 'OOT': df}``，DataFrame 需包含目标列，或通过 overdue/dpds 自动构建标签
    2. datasets 为 list：依次命名为 数据集1、数据集2、...、数据集N
    3. X_train/X_test/X_oot 参数：依次命名为 训练集、测试集、跨时间验证集

    标签解析遵循 hscredit 统一传参风格：显式传入 y 优先（sklearn 风格）；
    否则通过 target 列名从 X 中提取（scorecardpipeline 风格）；
    传入 overdue+dpds 组合时直接忽略 target。

    overdue/dpds 用法（自动从 X 构建二分类标签）::

        # 单阈值：MOB 任一期间 DPD > 5 则 y=1
        auto_model_report(model, X_train=df, overdue='dpds', dpds=5)

        # 多阈值：MOB1 DPD>15 或 MOB3 DPD>7 任一触发则 y=1
        auto_model_report(
            model,
            X_train=df,
            overdue=['dpds_m1', 'dpds_m3'],
            dpds=[15, 7, 0],
        )

    示例::

        # 方式1: datasets dict（DataFrame 直接传入，X 中含目标列）
        auto_model_report(model, datasets={'train': train_df, 'test': test_df}, excel_path='report.xlsx')

        # 方式1: datasets list（自动命名）
        auto_model_report(model, datasets=[train_df, test_df], excel_path='report.xlsx')

        # 方式1: overdue/dpds 自动构建标签
        auto_model_report(
            model,
            datasets={'train': df},
            overdue='dpds',
            dpds=[15, 7, 0],
            excel_path='report.xlsx',
        )

        # 方式2: 兼容 sklearn API（分离 X/y）
        auto_model_report(
            model,
            X_train=X, y_train=y,
            X_test=X_val, y_test=y_val,
            excel_path='report.xlsx',
        )

    :param model: 训练好的模型（ScoreCard / XGBoost / LightGBM / sklearn 等）
    :param datasets: 数据集字典/列表，字典键直接作为数据集名称（推荐）
    :param X_train: 训练集特征（命名为 训练集）
    :param y_train: 训练集标签，None 时从 X_train 中自动构建
    :param X_test: 测试集特征（命名为 测试集）
    :param y_test: 测试集标签，None 时从 X_test 中自动构建
    :param X_oot: 跨时间验证集特征（命名为 跨时间验证集）
    :param y_oot: 跨时间验证集标签，None 时从 X_oot 中自动构建
    :param feature_names: 特征名称列表，可选；None 时自动从模型 feature_names_ / feature_names_in_ 获取
    :param target: 目标列配置，str 为列名，dict 为 {'overdue': col, 'dpds': threshold}
    :param overdue: 逾期列名（str）或多个列名（List[str]），与 dpds 配合自动构建标签
    :param dpds: 逾期天数阈值（int/float）或多个阈值（List），与 overdue 配合使用
    :param del_grey: overdue 模式下是否按每个 DPD 独立剔除 ``(0, DPD]`` 灰样本
    :param excel_path: Excel 报告输出路径
    :param verbose: 是否打印控制台报告
    :param n_bins: 分箱数
    :param bin_method: 分箱方法
    :param amount_col: 金额字段（用于金额口径分析）
    :param date_col: 日期字段（用于分月分析）
    :param date_freq: 日期频率，支持 'D', 'W', 'M', 'Q' 等（默认自动推断）
    :param group_col: 分组字段（用于分组坏样本率分析）
    :param with_plots: 是否生成图表
    :param model_name: 模型名称
    :param project_desc: 项目描述
    :param feature_map: 特征名称到含义的映射
    :param feature_info: 特征部署信息表
    :param show_lift: 是否在报告中显示 LIFT 曲线
    :param show_importance: 是否在报告中显示特征重要性
    :param feature_contribution_label_max_features: 贡献图显示数据标签的最大特征数；默认 10，
        ``None`` 始终显示，0 始终隐藏
    :param data_source: 数据源描述
    :param loc_cols: 定位字段（订单号等），支持 str 或 List[str]，用于生产测试用例列
    :param method: 数据集唯一预测方法，默认 ``predict_proba``，也支持 callable
    :param method_kwargs: callable 同名参数的显式覆盖字典
    :param n_jobs: 并行工作数；-1 自动保留 CPU，None 使用兼容串行模式
    :param parallel_backend: joblib 后端，如 ``threading`` 或 ``loky``
    :param parallel_config: joblib 其他并行配置
    :param kwargs: 透传给 callable 的额外同名参数
    :return: ModelReport 实例
    """
    report = ModelReport(
        model=model,
        datasets=datasets,
        X_train=X_train,
        y_train=y_train,
        X_test=X_test,
        y_test=y_test,
        X_oot=X_oot,
        y_oot=y_oot,
        feature_names=feature_names,
        target=target,
        overdue=overdue,
        dpds=dpds,
        del_grey=del_grey,
        method=method,
        method_kwargs=method_kwargs,
        n_jobs=n_jobs,
        parallel_backend=parallel_backend,
        parallel_config=parallel_config,
        **kwargs,
    )

    if verbose:
        report.print_report(n_bins=n_bins)

    if excel_path:
        report.to_excel(
            excel_path,
            n_bins=n_bins,
            bin_method=bin_method,
            amount_col=amount_col,
            date_col=date_col,
            date_freq=date_freq,
            group_col=group_col,
            with_plots=with_plots,
            model_name=model_name,
            project_desc=project_desc,
            feature_map=feature_map,
            feature_info=feature_info,
            show_lift=show_lift,
            show_importance=show_importance,
            feature_contribution_label_max_features=feature_contribution_label_max_features,
            data_source=data_source,
            loc_cols=loc_cols,
        )
        if verbose:
            logger.info(f"\nExcel 报告已保存: {excel_path}")

    return report


def compare_models(
    models: Dict[str, object],
    X_train,
    y_train,
    X_test=None,
    y_test=None,
    excel_path: Optional[str] = None,
    n_jobs=-1,
    parallel_backend: Optional[str] = None,
    parallel_config: Optional[Dict[str, Any]] = None,
) -> pd.DataFrame:
    """横向对比多个模型的评估指标.

    对每个模型分别构建 :class:`ModelReport` 并取其 :meth:`~ModelReport.summary`，
    按输入映射顺序纵向拼接为一张对比表；任一模型失败时立即抛出并保留原始异常链。

    :param models: 模型名称到模型对象的映射，如 ``{'XGB': xgb_model, 'LR': lr_model}``
    :param X_train: 训练集特征
    :param y_train: 训练集标签（0/1）
    :param X_test: 测试集特征，可选
    :param y_test: 测试集标签，可选
    :param excel_path: 可选，若提供则将对比表导出到该 Excel 路径
    :param n_jobs: 模型外层并行工作数；-1 自动保留 CPU
    :param parallel_backend: joblib 后端，如 ``threading`` 或 ``loky``
    :param parallel_config: joblib 其他并行配置
    :return: 含 ``模型名称`` 列的指标对比 ``DataFrame``

    **参考样例**

    >>> from hscredit.report import compare_models
    >>> result = compare_models(
    ...     {'XGBoost': xgb_model, '逻辑回归': lr_model},
    ...     X_train, y_train, X_test, y_test,
    ...     excel_path='模型对比.xlsx',
    ... )
    >>> print(result)
    """
    validate_parallel_config(parallel_backend, parallel_config)
    dataset_task_count = 1 + int(X_test is not None)
    plan = _plan_model_comparison_parallel(
        n_jobs,
        model_task_count=len(models),
        dataset_task_count=dataset_task_count,
    )
    tasks = [
        (
            name,
            model,
            X_train,
            y_train,
            X_test,
            y_test,
            plan.child_workers,
            parallel_backend,
            parallel_config,
        )
        for name, model in models.items()
    ]
    computed = _execute_model_comparison_plan(
        _compare_model_worker,
        tasks,
        plan,
        parallel_backend=parallel_backend,
        parallel_config=parallel_config,
        task_labels=list(models),
        default_backend="loky",
    )
    parts: Dict[str, pd.DataFrame] = {name: summary for name, summary in computed}

    # 以「模型名称」作为最外层行索引纵向拼接，保留 summary 的「统计指标 × 数据集」多层列
    result = pd.concat(parts, names=["模型名称"]) if parts else pd.DataFrame()
    if excel_path and not result.empty:
        from ..excel import dataframe2excel

        dataframe2excel(
            result,
            excel_path,
            index=True,
            percent_cols=_summary_percent_cols(result.columns),
        )
    return result


# 向后兼容别名：旧名称 QuickModelReport 等价于 ModelReport
QuickModelReport = ModelReport

__all__ = ["ModelReport", "QuickModelReport", "auto_model_report", "compare_models"]
