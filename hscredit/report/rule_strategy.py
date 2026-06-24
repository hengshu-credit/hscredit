"""拒绝规则策略文档表格转换工具."""

from collections import OrderedDict
from typing import Any, Callable, Dict, List, Mapping, Optional, Sequence, Tuple, Union

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font

from ..core.rules import Rule


_DETAIL_GROUPS = ("分箱详情", "规则详情")
_DETAIL_COLUMNS = ("规则分类", "指标名称", "指标含义", "分箱")
_BIN_ORDER = ("命中", "未命中", "合计")
_DEFAULT_METRICS = ("样本总数", "样本占比", "坏样本率", "LIFT值")
_DEFAULT_TARGET_METRICS = ("样本总数", "样本占比", "坏样本率", "LIFT值", "坏账改善", "风险拒绝比")
_DEFAULT_GROUP_METRICS = OrderedDict((("坏样本率", "坏样本率"), ("样本占比", "样本占比"), ("LIFT指标", "LIFT值"), ("样本总数", "样本总数")))


def _ordered_unique(values: Sequence[str]) -> list:
    return list(dict.fromkeys(values))


def _resolve_detail_group(report: pd.DataFrame) -> str:
    groups = _ordered_unique(report.columns.get_level_values(0).tolist())
    for group in _DETAIL_GROUPS:
        if group in groups and (group, "分箱") in report.columns:
            return group
    raise ValueError("rule.report 结果缺少分箱详情列")


def _display_target(target: str, target_names: Optional[Mapping[str, str]]) -> str:
    return target_names.get(target, target) if target_names else target


def _extract_rule_name(report: pd.DataFrame, rule_name: Optional[str], detail_group: Optional[str]) -> str:
    if rule_name:
        return rule_name

    column = (detail_group, "指标名称") if detail_group else "指标名称"
    if column in report.columns:
        values = report[column].dropna().astype(str)
        values = values[values != "合计"]
        if not values.empty:
            return values.iloc[0]
    return "规则"


def _value_from_row(row: pd.Series, target: str, metric: str, detail_group: Optional[str]):
    if detail_group is None:
        return row.get(metric, np.nan)

    target_column = (target, metric)
    detail_column = (detail_group, metric)
    if target_column in row.index:
        return row[target_column]
    if detail_column in row.index:
        return row[detail_column]
    return np.nan


def _total_metric(rows: pd.DataFrame, metric: str):
    if metric in {"样本总数", "好样本数", "坏样本数"}:
        return rows[metric].sum(min_count=1)
    if metric in {"样本占比", "好样本占比", "坏样本占比"}:
        value = rows[metric].sum(min_count=1)
        return min(float(value), 1.0) if pd.notna(value) else np.nan
    if metric == "坏样本率":
        if {"坏样本数", "样本总数"}.issubset(rows.columns):
            total = rows["样本总数"].sum(min_count=1)
            bad = rows["坏样本数"].sum(min_count=1)
            if pd.notna(total) and total != 0 and pd.notna(bad):
                return bad / total
        if "样本总数" in rows.columns:
            weights = pd.to_numeric(rows["样本总数"], errors="coerce")
            values = pd.to_numeric(rows[metric], errors="coerce")
            valid = weights.notna() & values.notna()
            if valid.any() and weights[valid].sum() != 0:
                return np.average(values[valid], weights=weights[valid])
    if metric == "LIFT值":
        return 1.0
    if metric in {"坏账改善", "风险拒绝比"}:
        return 0.0
    return np.nan


def _normalize_report(
    report: pd.DataFrame,
    rule_name: Optional[str] = None,
    target_names: Optional[Mapping[str, str]] = None,
    target_name: str = "target",
) -> Tuple[pd.DataFrame, str]:
    """将单层或多层 ``Rule.report`` 结果转换为统一长表."""
    if not isinstance(report, pd.DataFrame) or report.empty:
        raise ValueError("report 必须是非空的 DataFrame")

    detail_group = _resolve_detail_group(report) if isinstance(report.columns, pd.MultiIndex) else None
    bin_column = (detail_group, "分箱") if detail_group else "分箱"
    if bin_column not in report.columns:
        raise ValueError("rule.report 结果缺少分箱列")

    if detail_group:
        groups = _ordered_unique(report.columns.get_level_values(0).tolist())
        targets = [group for group in groups if group != detail_group]
        detail_metrics = [column[1] for column in report.columns if column[0] == detail_group]
        target_metrics = [column[1] for column in report.columns if column[0] in targets]
        metrics = _ordered_unique(detail_metrics + target_metrics)
    else:
        targets = [target_name]
        metrics = report.columns.tolist()

    metrics = [metric for metric in metrics if metric not in _DETAIL_COLUMNS]
    resolved_rule_name = _extract_rule_name(report, rule_name, detail_group)
    records = []
    for target in targets:
        display_target = _display_target(str(target), target_names)
        for bin_name in _BIN_ORDER:
            matched = report.loc[report[bin_column] == bin_name]
            if matched.empty:
                continue
            row = matched.iloc[0]
            record = {"规则名称": resolved_rule_name, "逾期指标": display_target, "命中情况": bin_name}
            record.update({metric: _value_from_row(row, target, metric, detail_group) for metric in metrics})
            records.append(record)

        target_rows = [record for record in records if record["逾期指标"] == display_target]
        if not any(record["命中情况"] == "合计" for record in target_rows):
            detail = pd.DataFrame([record for record in target_rows if record["命中情况"] in {"命中", "未命中"}])
            total = {"规则名称": resolved_rule_name, "逾期指标": display_target, "命中情况": "合计"}
            total.update({metric: _total_metric(detail, metric) for metric in metrics})
            records.append(total)

    normalized = pd.DataFrame(records)
    if normalized.empty:
        raise ValueError("rule.report 结果中没有可用的命中、未命中或合计数据")
    target_order = _ordered_unique(normalized["逾期指标"].tolist())
    normalized["逾期指标"] = pd.Categorical(normalized["逾期指标"], categories=target_order, ordered=True)
    normalized["命中情况"] = pd.Categorical(normalized["命中情况"], categories=_BIN_ORDER, ordered=True)
    normalized = normalized.sort_values(["逾期指标", "命中情况"], kind="stable").reset_index(drop=True)
    normalized["逾期指标"] = normalized["逾期指标"].astype(object)
    normalized["命中情况"] = normalized["命中情况"].astype(object)
    return normalized, resolved_rule_name


def rule_report_table(
    report: pd.DataFrame,
    rule_name: Optional[str] = None,
    target_names: Optional[Mapping[str, str]] = None,
    metrics: Optional[Sequence[str]] = None,
    target_name: str = "target",
) -> pd.DataFrame:
    """生成按逾期指标横向展开的规则详情表.

    :param report: ``Rule.report`` 返回的 DataFrame
    :param rule_name: 展示用规则名称，默认使用报告中的指标名称
    :param target_names: 逾期指标名称映射，如 ``{'MOB1 1+': 'fpd1'}``
    :param metrics: 每个逾期指标需要展示的字段，默认使用内置 ``_DEFAULT_METRICS``
    :param target_name: 单标签报告的逾期指标名称，默认 ``"target"``
    :return: 两层列头的规则详情表

    **参考样例**

    >>> from hscredit.core.rules import Rule
    >>> from hscredit.report import rule_report_table
    >>> rep = Rule("score < 600").report(data, target='FPD')
    >>> rule_report_table(rep, rule_name='低分拒绝')
    """
    normalized, resolved_rule_name = _normalize_report(report, rule_name, target_names, target_name)
    metrics = list(metrics or _DEFAULT_METRICS)
    _validate_metrics(normalized, metrics)

    targets = _ordered_unique(normalized["逾期指标"].tolist())
    rows = []
    for bin_name in _BIN_ORDER:
        row = [resolved_rule_name, bin_name]
        for target in targets:
            matched = normalized[(normalized["逾期指标"] == target) & (normalized["命中情况"] == bin_name)]
            row.extend([matched.iloc[0][metric] if not matched.empty else np.nan for metric in metrics])
        rows.append(row)

    columns = [("规则详情", "规则名称"), ("规则详情", "分箱")]
    columns.extend((target, metric) for target in targets for metric in metrics)
    return pd.DataFrame(rows, columns=pd.MultiIndex.from_tuples(columns))


def rule_target_analysis(
    report: pd.DataFrame,
    current_pass_rate: Optional[float] = 1.0,
    rule_name: Optional[str] = None,
    target_names: Optional[Mapping[str, str]] = None,
    target_name: str = "target",
) -> pd.DataFrame:
    """生成拒绝规则目标分析表.

    绝对逾期改善为合计坏样本率减未命中坏样本率，绝对通过率为规则样本内的
    未命中占比；相对逾期改善以合计坏样本率为分母，相对通过率则在现有策略
    通过率基础上计算，即 ``current_pass_rate * 绝对通过率``。

    :param report: ``Rule.report`` 返回的 DataFrame
    :param current_pass_rate: 规则执行前的当前通过率，取值范围为[0, 1]
    :param rule_name: 展示用规则名称
    :param target_names: 逾期指标名称映射
    :param target_name: 单标签报告的逾期指标名称，默认 ``"target"``
    :return: 两层列头的目标分析表

    **参考样例**

    >>> from hscredit.core.rules import Rule
    >>> from hscredit.report import rule_target_analysis
    >>> rep = Rule("score < 600").report(data, target='FPD')
    >>> # 当前通过率 0.8 时，评估该拒绝规则带来的逾期改善与通过率变化
    >>> rule_target_analysis(rep, current_pass_rate=0.8, rule_name='低分拒绝')
    """
    if not isinstance(current_pass_rate, (int, float, np.integer, np.floating)) or not 0 <= current_pass_rate <= 1:
        raise ValueError("current_pass_rate 必须是[0, 1]范围内的数值")

    normalized, resolved_rule_name = _normalize_report(report, rule_name, target_names, target_name)
    required = ["坏样本率", "样本总数", "样本占比", "LIFT值", "风险拒绝比"]
    _validate_metrics(normalized, required)

    rows = []
    for target in _ordered_unique(normalized["逾期指标"].tolist()):
        target_rows = normalized[normalized["逾期指标"] == target].set_index("命中情况")
        hit, miss, total = target_rows.loc["命中"], target_rows.loc["未命中"], target_rows.loc["合计"]
        absolute_overdue = total["坏样本率"] - miss["坏样本率"]
        absolute_pass = miss["样本占比"]
        relative_overdue = absolute_overdue / total["坏样本率"] if total["坏样本率"] else 0.0
        relative_pass = current_pass_rate * absolute_pass
        rows.append(
            [
                target,
                total["坏样本率"],
                miss["坏样本率"],
                hit["坏样本率"],
                total["样本总数"],
                miss["样本总数"],
                hit["样本总数"],
                hit["样本占比"],
                hit["LIFT值"],
                hit["风险拒绝比"],
                absolute_overdue,
                absolute_pass,
                relative_overdue,
                relative_pass,
            ]
        )

    columns = [
        (resolved_rule_name, "逾期指标"),
        ("坏样本率", "合计"),
        ("坏样本率", "未命中"),
        ("坏样本率", "命中"),
        ("样本总数", "合计"),
        ("样本总数", "未命中"),
        ("样本总数", "命中"),
        ("样本占比", "命中"),
        ("规则指标", "拒绝LIFT"),
        ("规则指标", "风险拒绝比"),
        ("绝对比例", "逾期改善"),
        ("绝对比例", "通过率"),
        ("相对比例", "逾期改善"),
        ("相对比例", "通过率"),
    ]
    return pd.DataFrame(rows, columns=pd.MultiIndex.from_tuples(columns))


def rule_target_table(
    report: pd.DataFrame,
    rule_name: Optional[str] = None,
    target_names: Optional[Mapping[str, str]] = None,
    metrics: Optional[Sequence[str]] = None,
    target_name: str = "target",
) -> pd.DataFrame:
    """生成规则、逾期指标和命中情况组成的纵向明细表.

    与 :func:`rule_report_table` 的横向展开不同，本函数按
    ``规则 × 逾期指标 × 命中情况`` 逐行纵向罗列各项指标，便于直接落库或透视。

    :param report: ``Rule.report`` 返回的 DataFrame
    :param rule_name: 展示用规则名称，默认使用报告中的指标名称
    :param target_names: 逾期指标名称映射，如 ``{'MOB1 1+': 'fpd1'}``
    :param metrics: 需要展示的字段列表，默认使用内置 ``_DEFAULT_TARGET_METRICS``
    :param target_name: 单标签报告的逾期指标名称，默认 ``"target"``
    :return: 含 ``规则详情`` / ``逾期指标`` / ``命中情况`` 及各指标列的纵向明细表

    **参考样例**

    >>> from hscredit.core.rules import Rule
    >>> from hscredit.report import rule_target_table
    >>> rep = Rule("score < 600").report(data, target='FPD')
    >>> rule_target_table(rep, rule_name='低分拒绝')
    """
    normalized, _ = _normalize_report(report, rule_name, target_names, target_name)
    metrics = list(metrics or _DEFAULT_TARGET_METRICS)
    _validate_metrics(normalized, metrics)
    return normalized[["规则名称", "逾期指标", "命中情况"] + metrics].rename(columns={"规则名称": "规则详情"})


def rule_group_hit_table(
    group_reports: Mapping[str, pd.DataFrame],
    rule_name: Optional[str] = None,
    target_names: Optional[Mapping[str, str]] = None,
    metrics: Optional[Mapping[str, str]] = None,
    target_name: str = "target",
) -> pd.DataFrame:
    """生成多个样本分组下的规则命中效果对比表.

    :param group_reports: 分组名称到 ``Rule.report`` 结果的映射
    :param rule_name: 展示用规则名称
    :param target_names: 逾期指标名称映射
    :param metrics: 顶层展示名称到 ``Rule.report`` 字段名的映射
    :param target_name: 单标签报告的逾期指标名称，默认 ``"target"``
    :return: 两层列头的分组命中对比表，不包含合计行

    **参考样例**

    >>> from hscredit.core.rules import Rule
    >>> from hscredit.report import rule_group_hit_table
    >>> r = Rule("score < 600")
    >>> # 对比训练集 / 测试集 / OOT 三个分组下同一规则的命中效果
    >>> reports = {
    ...     '训练集': r.report(train, target='FPD'),
    ...     '测试集': r.report(test, target='FPD'),
    ...     'OOT': r.report(oot, target='FPD'),
    ... }
    >>> rule_group_hit_table(reports, rule_name='低分拒绝')
    """
    if not isinstance(group_reports, Mapping) or not group_reports:
        raise ValueError("group_reports 必须是非空的分组名称到 DataFrame 的映射")

    metrics = OrderedDict(metrics or _DEFAULT_GROUP_METRICS)
    normalized_groups: Dict[str, pd.DataFrame] = {}
    resolved_rule_name = rule_name
    for group, report in group_reports.items():
        normalized, extracted_name = _normalize_report(report, resolved_rule_name, target_names, target_name)
        _validate_metrics(normalized, list(metrics.values()))
        normalized_groups[str(group)] = normalized
        resolved_rule_name = resolved_rule_name or extracted_name

    targets = _ordered_unique(
        [target for normalized in normalized_groups.values() for target in normalized["逾期指标"].tolist()]
    )
    rows = []
    for target in targets:
        for bin_name in ("命中", "未命中"):
            row = [resolved_rule_name, target, bin_name]
            for source_metric in metrics.values():
                for normalized in normalized_groups.values():
                    matched = normalized[(normalized["逾期指标"] == target) & (normalized["命中情况"] == bin_name)]
                    row.append(matched.iloc[0][source_metric] if not matched.empty else np.nan)
            rows.append(row)

    columns = [("规则详情", "规则名称"), ("规则详情", "逾期指标"), ("规则详情", "是否命中")]
    columns.extend((display_metric, group) for display_metric in metrics for group in normalized_groups.keys())
    return pd.DataFrame(rows, columns=pd.MultiIndex.from_tuples(columns))


_FREQ_LABELS = OrderedDict((("D", "日"), ("W", "周"), ("M", "月"), ("Q", "季度")))
_VALID_GROUP_ORDER = ("asc", "desc", "appearance")

# group_order 接受：``None``/``"asc"``（升序，默认）、``"desc"``（降序）、
# ``"appearance"``（数据出现顺序）、可调用排序键、或显式分组名称序列
GroupOrder = Union[None, str, Callable[[Any], Any], Sequence[Any]]


def _order_groups(present: Sequence[Any], group_order: GroupOrder) -> list:
    """按 ``group_order`` 规则对已出现的分组标签排序，缺省升序排列."""
    present = _ordered_unique(list(present))  # 去重并保留出现顺序

    if group_order is None or (isinstance(group_order, str) and group_order == "asc"):
        try:
            return sorted(present)
        except TypeError:
            return present
    if isinstance(group_order, str):
        if group_order == "desc":
            try:
                return sorted(present, reverse=True)
            except TypeError:
                return list(reversed(present))
        if group_order == "appearance":
            return present
        raise ValueError(f"group_order 字符串仅支持 {_VALID_GROUP_ORDER} 之一")
    if callable(group_order):
        return sorted(present, key=group_order)
    if isinstance(group_order, (list, tuple, pd.Index, np.ndarray)):
        specified = list(group_order)
        specified_set = set(specified)
        ordered = [group for group in specified if group in present]
        remaining = [group for group in present if group not in specified_set]
        return ordered + remaining
    raise ValueError("group_order 必须是 'asc'/'desc'/'appearance'、可调用对象或分组名称序列")


def _resolve_group_labels(
    data: pd.DataFrame,
    date_col: Optional[str],
    freq: str,
    group_col: Optional[str],
    dropna: bool,
    group_order: GroupOrder,
) -> Tuple[pd.Series, list]:
    """根据日期+频率或分组字段，解析每条样本所属的分组标签与有序分组列表."""
    if (date_col is None) == (group_col is None):
        raise ValueError("date_col 与 group_col 必须且只能指定其中一个")

    if group_col is not None:
        if group_col not in data.columns:
            raise ValueError(f"数据集缺少分组字段列: {group_col}")
        labels = data[group_col]
    else:
        if date_col not in data.columns:
            raise ValueError(f"数据集缺少日期列: {date_col}")
        if freq not in _FREQ_LABELS:
            raise ValueError("freq必须是'D'/'W'/'M'/'Q'之一")
        parsed = pd.to_datetime(data[date_col], errors="coerce")
        labels = parsed.dt.to_period(freq).astype(str)
        labels = labels.where(parsed.notna(), other=np.nan)

    labels = pd.Series(labels, index=data.index)
    valid = labels.notna()
    if not dropna and not valid.all():
        labels = labels.where(valid, other="缺失")
        valid = pd.Series(True, index=data.index)

    unique_values = labels[valid].dropna().unique().tolist()
    return labels, _order_groups(unique_values, group_order)


def rule_group_compare(
    data: pd.DataFrame,
    rule: Union[str, Rule],
    date_col: Optional[str] = None,
    freq: str = "M",
    group_col: Optional[str] = None,
    target: str = "target",
    overdue: Optional[Union[str, List[str]]] = None,
    dpds: Optional[Union[int, List[int]]] = None,
    rule_name: Optional[str] = None,
    target_names: Optional[Mapping[str, str]] = None,
    metrics: Optional[Mapping[str, str]] = None,
    prior_rules: Optional[Rule] = None,
    amount: Optional[str] = None,
    del_grey: bool = False,
    dropna: bool = True,
    group_order: GroupOrder = "asc",
    **kwargs: Any,
) -> pd.DataFrame:
    """直接从原始数据生成分组下的规则命中效果对比表.

    相比 :func:`rule_group_hit_table` 需要在函数外手工切分样本并逐组调用
    ``Rule.report``，本函数接收原始明细数据，按 ``日期列 + 频率`` 或 ``分组字段``
    自动切分样本，对每个分组调用同一规则的 ``Rule.report``（支持
    ``target`` 单标签或 ``overdue + dpds`` 多标签口径），再汇总为分组对比表。

    :param data: 原始明细数据 DataFrame，需包含规则所需字段、目标/逾期字段及分组依据列
    :param rule: 规则表达式字符串或 :class:`~hscredit.core.rules.Rule` 实例
    :param date_col: 日期列名，与 ``freq`` 配合按时间周期分组（与 ``group_col`` 二选一）
    :param freq: 时间频率，``'D'`` 日 / ``'W'`` 周 / ``'M'`` 月 / ``'Q'`` 季度，默认 ``'M'``
    :param group_col: 分组字段列名，按其取值分组（与 ``date_col`` 二选一）
    :param target: 目标变量列名，默认 ``"target"``，0=好样本，1=坏样本
    :param overdue: 逾期天数字段名（可选，传入时以逾期天数>DPD定义坏样本，支持多标签）
    :param dpds: 逾期定义方式，逾期天数 > DPD 为坏样本，可传入列表支持多DPD联合分析
    :param rule_name: 展示用规则名称，默认使用规则自身名称或报告中的指标名称
    :param target_names: 逾期指标名称映射，如 ``{'MOB1 1+': 'fpd1'}``
    :param metrics: 顶层展示名称到 ``Rule.report`` 字段名的映射，默认 ``_DEFAULT_GROUP_METRICS``
    :param prior_rules: 先验规则（可选），每个分组内先排除命中先验规则的样本再评估
    :param amount: 金额字段名（可选），传入时以金额口径而非样本数口径统计
    :param del_grey: 是否删除逾期天数在(0, DPD]区间内的灰度样本，默认为False
    :param dropna: 是否丢弃分组依据缺失的样本，默认为True；为False时缺失样本归入“缺失”分组
    :param group_order: 分组排列方式，默认 ``"asc"`` 升序。支持：

        * ``"asc"`` / ``"desc"`` — 按分组标签升序 / 降序
        * ``"appearance"`` — 按分组在数据中首次出现的顺序
        * 可调用对象 — 作为 ``sorted`` 的 ``key`` 排序键
        * 分组名称序列 — 按给定顺序排列，未列出的分组按出现顺序追加在末尾

    :param kwargs: 透传给 ``Rule.report`` 的其他参数（如 ``desc``、``filter_cols``、``margins`` 等）
    :return: 两层列头的分组命中对比表，不包含合计行；列头第二层为各分组名称
    :raises ValueError: ``date_col`` 与 ``group_col`` 未二选一，或所需列缺失时

    **参考样例**

    >>> from hscredit.report import rule_group_compare
    >>> # 按放款月份对比同一拒绝规则在各月样本上的命中效果（多标签口径）
    >>> rule_group_compare(
    ...     data, "score < 600", date_col='放款时间', freq='M',
    ...     overdue=['MOB1'], dpds=[7, 0], rule_name='低分拒绝',
    ... )
    >>> # 按商品类别分组、金额口径，并自定义分组展示顺序
    >>> rule_group_compare(
    ...     data, "score < 600", group_col='商品类别', target='FPD',
    ...     amount='放款金额', group_order=['手机通讯', '电脑数码', '家用电器'],
    ... )
    """
    if not isinstance(data, pd.DataFrame) or data.empty:
        raise ValueError("data 必须是非空的 DataFrame")

    rule = rule if isinstance(rule, Rule) else Rule(rule)
    labels, ordered_groups = _resolve_group_labels(data, date_col, freq, group_col, dropna, group_order)
    if not ordered_groups:
        raise ValueError("根据 date_col/group_col 未能切分出任何有效分组")

    group_reports: "OrderedDict[str, pd.DataFrame]" = OrderedDict()
    for group in ordered_groups:
        subset = data.loc[labels == group]
        if subset.empty:
            continue
        group_reports[str(group)] = rule.report(
            subset,
            target=target,
            overdue=overdue,
            dpds=dpds,
            del_grey=del_grey,
            prior_rules=prior_rules,
            amount=amount,
            **kwargs,
        )

    return rule_group_hit_table(
        group_reports,
        rule_name=rule_name or rule.name,
        target_names=target_names,
        metrics=metrics,
        target_name=target,
    )


def _validate_metrics(normalized: pd.DataFrame, metrics: Sequence[str]) -> None:
    missing = [metric for metric in metrics if metric not in normalized.columns]
    if missing:
        raise ValueError(f"rule.report 结果缺少以下指标列: {missing}")


# ======================================================================================
# 策略迭代（拒绝规则置换）报告 —— swap_out_report
# ======================================================================================

# 业务影响/规则效果报告中需展示的完整指标（与 Rule.report 字段一致）
_SWAP_EFFECT_METRICS = (
    "样本总数", "样本占比", "好样本数", "好样本占比", "坏样本数", "坏样本占比",
    "坏样本率", "LIFT值", "坏账改善", "风险拒绝比",
)
# Excel 中按百分比格式展示的列（按多层列名的最末层匹配）
_SWAP_PERCENT_NAMES = [
    "样本占比", "好样本占比", "坏样本占比", "坏样本率", "金额占比",
    "通过率", "逾期改善", "坏账改善",
]
# Excel 中按千分位整数格式展示的列
_SWAP_COUNT_NAMES = ["样本总数", "好样本数", "坏样本数", "金额总数", "累积好样本数", "累积坏样本数"]

# 各表条件格式：仅对关键列着色（数据条/色阶），保持与模板一致的克制风格，避免满屏色块。
# 业务/规则口径同时列出样本与金额两种列名，便于订单/金额口径表各自匹配到对应列。
_CF_IMPACT_BAR = ["样本占比", "金额占比", "拒绝LIFT", "风险拒绝比"]   # 业务影响：占比/拒绝LIFT/风险拒绝比 → 数据条
_CF_EFFECT_BAR = ["样本占比", "金额占比", "坏样本率"]                  # 规则效果：占比/坏样本率 → 数据条
_CF_STABILITY_BAR = ["坏样本率", "样本占比"]                          # 稳定性：坏样本率/样本占比 → 数据条
_CF_BIN_BAR = ["LIFT值"]                                            # 分箱详情：LIFT → 数据条
_CF_BIN_SCALE = ["分档KS值"]                                        # 分箱详情：KS → 色阶
_CF_VARBIN_BAR = ["坏样本率", "LIFT值", "分档KS值"]                  # 变量分箱明细：坏样本率/LIFT/KS → 数据条
_CF_OVERVIEW_SCALE = ["坏样本率"]                                    # 样本情况：坏样本率 → 色阶

# 数据条按指标着色：同一张表内并排的各指标使用不同色相，便于区分，避免满屏同色
_CF_BAR_COLORS = {
    "样本占比": "63C384",    # 绿 —— 规模/通过占比
    "金额占比": "4BACC6",    # 青 —— 金额规模
    "坏样本率": "F4796B",    # 红 —— 风险
    "坏账改善": "8E7CC3",    # 紫 —— 坏账改善
    "LIFT值": "F0A030",      # 橙 —— 区分度
    "拒绝LIFT": "F0A030",    # 橙 —— 拒绝区分度
    "分档KS值": "5B9BD5",    # 蓝 —— 分档 KS
    "风险拒绝比": "6E8FC7",   # 靛 —— 拒绝效率
}
_CF_BAR_DEFAULT_COLOR = "638EC6"
# 色阶：绿-黄-红三色（min→percentile50→max）
_CF_SCALE_COLORS = ("63BE7B", "FFEB84", "F8696B")

# 业务影响目标分析表的命中标签 → 模板展示口径
_SWAP_HIT_LABELS = {"合计": "原始", "未命中": "通过", "命中": "拒绝"}

# 分箱方法 → 中文展示名（用于「变量分箱」明细 sheet 的区块标题）
_METHOD_DISPLAY_NAMES = {
    "quantile": "等频分箱", "uniform": "等距分箱", "mdlp": "最优分箱(MDLP)",
    "chi": "卡方分箱", "tree": "决策树分箱", "cart": "CART分箱", "dt": "决策树分箱",
    "best_ks": "最优KS分箱", "best_iv": "最优IV分箱", "best_lift": "最优LIFT分箱",
    "kmeans": "聚类分箱", "monotonic": "单调分箱", "genetic": "遗传分箱",
}

# 大模块之间空 2 行，模块内各表/文本之间空 1 行
_GAP_MODULE = 2
_GAP_INNER = 1


def _method_display(method: str) -> str:
    return _METHOD_DISPLAY_NAMES.get(method, f"{method}分箱")


def _normalize_target_label(label: Any) -> str:
    """将逾期标签的三种生成形式统一为 ``Rule.report`` 的空格形式 ``MOB1 7+``。

    样本情况用 ``MOB1_7+``、分箱详情用 ``MOB1@7``、规则报告用 ``MOB1 7+``，三者
    混排会令同一报告内同一标签出现三种写法，此处统一为空格形式；非标签字符串原样返回。
    """
    text = str(label)
    if "@" in text:
        head, _, tail = text.rpartition("@")
        return f"{head} {tail}+" if head and tail.isdigit() else text
    if text.endswith("+") and "_" in text:
        head, _, tail = text.rpartition("_")
        return f"{head} {tail}" if head and tail[:-1].isdigit() else text
    return text


def _normalize_target_names(target_names: Optional[Mapping[str, str]]) -> Optional[Dict[str, str]]:
    """将 ``target_names`` 的键统一为空格形式，使任意写法的键都能匹配到标签。"""
    if not target_names:
        return target_names
    return {_normalize_target_label(key): value for key, value in target_names.items()}


def _display_target_label(label: Any, target_names: Optional[Mapping[str, str]]) -> str:
    """规整单个逾期标签：先统一为空格形式，再套用 ``target_names`` 映射。"""
    normalized = _normalize_target_label(label)
    return target_names.get(normalized, normalized) if target_names else normalized


def _rename_target_level(
    table: pd.DataFrame, target_names: Optional[Mapping[str, str]], level: int = 1
) -> pd.DataFrame:
    """规整多层列中某一层的逾期标签（统一形式 + target_names 映射），非标签值保持不变。"""
    mapping = {}
    for label in dict.fromkeys(table.columns.get_level_values(level)):
        display = _display_target_label(label, target_names)
        if display != label:
            mapping[label] = display
    return table.rename(columns=mapping, level=level) if mapping else table


def _as_text_list(value: Optional[Union[str, Sequence[str]]]) -> List[str]:
    """将文本参数规整为字符串列表，``None`` 返回空列表."""
    if value is None:
        return []
    if isinstance(value, str):
        return [value]
    return [str(item) for item in value]


def _resolve_swap_rules(
    rules: Union[str, "Rule", Sequence[Union[str, "Rule"]]],
) -> Tuple["Rule", List["Rule"], bool]:
    """解析规则集，返回 ``(整体规则, 子规则列表, 是否为多规则)``.

    单规则时整体即该规则本身；多规则时整体为各子规则按「或」组合的并集。
    """
    if isinstance(rules, (str, Rule)):
        rule_list = [rules]
    else:
        rule_list = list(rules)
    if not rule_list:
        raise ValueError("rules 不能为空")

    rule_objs = [r if isinstance(r, Rule) else Rule(r) for r in rule_list]
    multi = len(rule_objs) > 1
    if multi:
        combined = rule_objs[0]
        for r in rule_objs[1:]:
            combined = combined | r
        combined.name = "整体规则"
    else:
        combined = rule_objs[0]
    return combined, rule_objs, multi


def _resolve_swap_targets(
    data: pd.DataFrame,
    target: Optional[str],
    overdue: Optional[Union[str, List[str]]],
    dpds: Optional[Union[int, List[int]]],
    del_grey: bool,
) -> "OrderedDict[str, Tuple[pd.DataFrame, pd.Series]]":
    """解析单标签或 ``overdue + dpds`` 多标签，返回标签到 ``(子样本, 0/1标签)`` 的映射.

    标签命名与 :func:`~hscredit.report.feature_bin_stats` 保持一致（``MOB1_7+``）。
    """
    result: "OrderedDict[str, Tuple[pd.DataFrame, pd.Series]]" = OrderedDict()
    if overdue is not None:
        if dpds is None:
            raise ValueError("传入 overdue 参数时必须同时传入 dpds")
        overdue_cols = [overdue] if isinstance(overdue, str) else list(overdue)
        dpd_values = [dpds] if isinstance(dpds, (int, np.integer)) else list(dpds)
        for col in overdue_cols:
            if col not in data.columns:
                raise ValueError(f"数据集缺少逾期天数列: {col}")
            days = pd.to_numeric(data[col], errors="coerce")
            for dpd in dpd_values:
                label = f"{col}_{dpd}+"
                y = (days > dpd).astype(int)
                subset = data
                if del_grey:
                    keep = ~((days > 0) & (days <= dpd))
                    subset = data[keep]
                    y = y[keep]
                result[label] = (subset, y)
    else:
        if target is None or target not in data.columns:
            raise ValueError("必须传入数据集中存在的 target 或 overdue+dpds 参数")
        result[target] = (data, data[target].astype(int))
    return result


def _swap_sample_overview(
    targets: "OrderedDict[str, Tuple[pd.DataFrame, pd.Series]]",
    date_col: Optional[str],
    data: pd.DataFrame,
    target_names: Optional[Mapping[str, str]] = None,
) -> pd.DataFrame:
    """构建样本情况表：全量数据在各标签下的样本总数与坏样本率."""
    labels = [_display_target_label(label, target_names) for label in targets.keys()]
    columns: List[Tuple[str, str]] = [("样本情况", "数据集")]
    values: List[Any] = ["全量数据"]
    if date_col is not None and date_col in data.columns:
        parsed = pd.to_datetime(data[date_col], errors="coerce").dropna()
        if not parsed.empty:
            columns.append(("样本情况", "时间区间"))
            values.append(f"{parsed.min():%Y-%m-%d} 至 {parsed.max():%Y-%m-%d}")
    columns.extend(("样本总数", label) for label in labels)
    values.extend(int(len(subset)) for subset, _ in targets.values())
    columns.extend(("坏样本率", label) for label in labels)
    values.extend(round(float(y.mean()), 4) if len(y) else np.nan for _, y in targets.values())
    return pd.DataFrame([values], columns=pd.MultiIndex.from_tuples(columns))


def _rename_amount_caliber(table: pd.DataFrame) -> pd.DataFrame:
    """将金额口径表格的「样本总数 / 样本占比」顶层列名替换为「金额总数 / 金额占比」."""
    table = table.copy()
    mapping = {"样本总数": "金额总数", "样本占比": "金额占比"}
    if isinstance(table.columns, pd.MultiIndex):
        table.columns = pd.MultiIndex.from_tuples(
            [(mapping.get(top, top), sub) for top, sub in table.columns]
        )
    else:
        table.columns = [mapping.get(col, col) for col in table.columns]
    return table


def _write_banner(writer, worksheet, row: int, text: str, start_col: int, width: int, style: str) -> int:
    """在整行范围内套用指定样式后写入一行横幅文本，保证合并区每个单元格样式一致.

    ``style`` 取 ``"header"``（模块标题）或 ``"middle"``（正文文本），内容统一左对齐。
    先对整行套用命名样式并合并，再设置取值与左对齐（避免合并时被命名样式的居中对齐覆盖）。
    """
    end_col = start_col + max(width, 1) - 1
    for col in range(start_col, end_col + 1):
        worksheet.cell(row=row, column=col).style = style
    if end_col > start_col:
        worksheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = worksheet.cell(row=row, column=start_col)
    cell.value = writer.astype_insertvalue(text)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    return row + 1


def _autosize_columns(writer, worksheet, start_col: int = 1) -> None:
    """按内容估算并设置列宽，不重置单元格样式/数字格式（避免覆盖已设置的条件格式与百分比格式）."""
    from openpyxl.utils import get_column_letter

    for col in range(start_col, worksheet.max_column + 1):
        max_width = 8.0
        for r in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=r, column=col).value
            if value is None or value == "":
                continue
            _, eng_cnt, chi_cnt = writer.check_contain_chinese(str(value))
            content_width = (eng_cnt * writer.english_width + chi_cnt * writer.chinese_width) * writer.fontsize + 2.0
            max_width = max(max_width, content_width)
        writer.set_column_width(worksheet, get_column_letter(col), min(max_width, 50.0))


def _write_swap_text(
    writer, worksheet, row: int, texts: Optional[Union[str, Sequence[str]]], start_col: int,
) -> int:
    """写入纯文本叙述段落：列表自动加「序号、内容」，每行一条，左对齐、无填充无边框（与模板一致）."""
    items = _as_text_list(texts)
    if not items:
        return row
    numbered = isinstance(texts, (list, tuple)) and len(items) > 1
    text_font = Font(name=writer.font, size=writer.fontsize, color="000000")
    for idx, content in enumerate(items, 1):
        line = f"{idx}、{content}" if numbered else content
        cell = worksheet.cell(row=row, column=start_col)
        cell.value = writer.astype_insertvalue(line)
        cell.font = text_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        row += 1
    return row


def _table_display_width(table: pd.DataFrame, index: bool = False) -> int:
    """表格在 Excel 中实际占用的列数（含多层索引列）."""
    return table.shape[1] + (table.index.nlevels if index else 0)


def _caliber_pair_width(order_table: pd.DataFrame, amount_table: Optional[pd.DataFrame]) -> int:
    """订单口径（+并排金额口径）一组表格占用的总列数（含中间 1 列间隔）."""
    width = _table_display_width(order_table)
    if amount_table is not None:
        width += _GAP_INNER + _table_display_width(amount_table)
    return width


def _format_columns(table: pd.DataFrame) -> Tuple[list, list]:
    """按多层列名任意层级匹配，拆分出 百分比 / 千分位整数 两类数字格式列（返回完整列标签）."""
    percent_cols, count_cols = [], []
    for col in table.columns:
        names = col if isinstance(col, tuple) else (col,)
        if any(name in _SWAP_PERCENT_NAMES for name in names):
            percent_cols.append(col)
        elif any(name in _SWAP_COUNT_NAMES for name in names):
            count_cols.append(col)
    return percent_cols, count_cols


def _cf_columns(table: pd.DataFrame, names: Optional[Sequence[str]]) -> list:
    """返回表中（任意层级）列名命中 ``names`` 的完整列标签，用于条件格式列定位."""
    if not names:
        return []
    name_set = set(names)
    matched = []
    for col in table.columns:
        col_names = col if isinstance(col, tuple) else (col,)
        if any(n in name_set for n in col_names):
            matched.append(col)
    return matched


def _write_subtitle(writer, worksheet, row: int, text: str, start_col: int) -> None:
    """模块内子标题：仅占一列（不跨列合并），套用标题样式（主题色底 + 白字、左对齐）."""
    cell = worksheet.cell(row=row, column=start_col)
    cell.style = "header"
    cell.value = writer.astype_insertvalue(text)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _cf_condition_color(bar_names: Optional[Sequence[str]], scale_names: Optional[Sequence[str]]) -> dict:
    """构建 ``dataframe2excel`` 的 ``condition_color`` 字典：数据条按指标取单色、色阶取三色锚点."""
    color_map: Dict[str, Any] = {}
    for name in bar_names or []:
        color_map[name] = _CF_BAR_COLORS.get(name, _CF_BAR_DEFAULT_COLOR)
    for name in scale_names or []:
        color_map[name] = list(_CF_SCALE_COLORS)
    return color_map


def _write_swap_table(
    writer, worksheet, table: pd.DataFrame, row: int, start_col: int, title: Optional[str],
    bar_names: Optional[Sequence[str]] = None, scale_names: Optional[Sequence[str]] = None, **kwargs,
):
    """通过 dataframe2excel 写入单张表格：自动套用百分比/整数格式，并仅对 ``bar_names`` /
    ``scale_names`` 指定的关键列分别加数据条 / 色阶条件格式（按指标取不同颜色，匹配模板风格）.

    ``title`` 作为模块内子标题，仅占一列（不随表宽合并），写于表格上方一行。
    """
    from ..excel import dataframe2excel

    if title is not None:
        _write_subtitle(writer, worksheet, row, title, start_col)
        row += 1
    percent_cols, count_cols = _format_columns(table)
    params = dict(custom_format="#,##0")
    params.update(kwargs)
    params.setdefault("percent_cols", percent_cols or None)
    params.setdefault("custom_cols", count_cols or None)
    bar_cols = _cf_columns(table, bar_names)
    scale_cols = _cf_columns(table, scale_names)
    if bar_cols:
        params.setdefault("condition_cols", bar_cols)
    if scale_cols:
        params.setdefault("color_cols", scale_cols)
    color_map = _cf_condition_color(bar_names, scale_names)
    if color_map:
        params.setdefault("condition_color", color_map)
    return dataframe2excel(
        table, excel_writer=writer, sheet_name=worksheet, title=None,
        start_row=row, start_col=start_col, **params,
    )


def _add_group_databars(
    writer, worksheet, table: pd.DataFrame, end_row: int, start_col: int, group_names: Sequence[str]
) -> None:
    """对多层列下的指标分组（如稳定性的坏样本率/样本占比跨多期）整组添加一条数据条，避免逐列重复着色."""
    from openpyxl.utils import get_column_letter

    n_rows = len(table)
    if n_rows == 0:
        return
    data_first, data_last = end_row - n_rows, end_row - 1
    for group in group_names:
        idxs = [i for i, col in enumerate(table.columns) if group in (col if isinstance(col, tuple) else (col,))]
        if not idxs:
            continue
        first_letter = get_column_letter(start_col + idxs[0])
        last_letter = get_column_letter(start_col + idxs[-1])
        writer.add_conditional_formatting(
            worksheet, f"{first_letter}{data_first}", f"{last_letter}{data_last}",
            condition_color=_CF_BAR_COLORS.get(group, _CF_BAR_DEFAULT_COLOR),
        )


def _rename_hit_labels(table: pd.DataFrame) -> pd.DataFrame:
    """将业务影响目标分析表的命中标签（合计/未命中/命中）改为模板口径（原始/通过/拒绝）."""
    table = table.copy()
    table.columns = pd.MultiIndex.from_tuples(
        [(top, _SWAP_HIT_LABELS.get(sub, sub)) for top, sub in table.columns]
    )
    return table


def _drop_total_rows(table: pd.DataFrame) -> pd.DataFrame:
    """删除规则效果明细表中的「合计」行，仅保留命中/未命中两行（与模板一致）."""
    if "命中情况" in table.columns:
        table = table[table["命中情况"] != "合计"].reset_index(drop=True)
    return table


def _write_caliber_pair(
    writer, worksheet, row: int, start_col: int, prefix: str,
    order_table: pd.DataFrame, amount_table: Optional[pd.DataFrame] = None,
    bar_names: Optional[Sequence[str]] = None, **kwargs,
) -> int:
    """订单口径 / 金额口径两张表并排写入同一起始行，返回下一可用行（含模块内 1 行间隔）."""
    end_row, _ = _write_swap_table(
        writer, worksheet, order_table, row, start_col, f"{prefix}：订单口径", bar_names=bar_names, **kwargs,
    )
    if amount_table is not None:
        amount_col = start_col + _table_display_width(order_table) + 1
        amount_end, _ = _write_swap_table(
            writer, worksheet, amount_table, row, amount_col, f"{prefix}：金额口径", bar_names=bar_names, **kwargs,
        )
        end_row = max(end_row, amount_end)
    return end_row + _GAP_INNER


def _write_binning_detail_sheet(
    writer,
    binning_tables: Mapping[str, Mapping[str, pd.DataFrame]],
    features: List[str],
    methods: List[str],
    sheet_name: str = "变量分箱",
    target_names: Optional[Mapping[str, str]] = None,
) -> None:
    """将各分箱方法的明细分箱表横向并排写入「变量分箱」sheet（数据源自 feature_binning_summary）.

    每个分箱方法占一个横向区块：区块标题为方法中文名，下方为该方法下所有指标的
    长表（逾期标签纵向堆叠）拼接结果；指标名称 / 指标含义列纵向合并以便阅读。
    """
    from openpyxl.utils import get_column_letter

    worksheet = writer.get_sheet_by_name(sheet_name)
    start_col = 2
    title_row = 2                    # 子标题（分箱方法名）所在行
    header_row = title_row + 1       # 表头行
    first_data_row = header_row + 1  # 数据首行
    first_block_ref: Optional[str] = None  # 首个区块的「表头行→末行」筛选区域
    for method in methods:
        method_tables = [binning_tables[feat][method] for feat in features if method in binning_tables.get(feat, {})]
        if not method_tables:
            continue
        block = pd.concat(method_tables, ignore_index=True)
        if "逾期标签" in block.columns:
            block = block.copy()
            block["逾期标签"] = [_display_target_label(value, target_names) for value in block["逾期标签"]]
        merge_cols = [col for col in ("指标名称", "指标含义") if col in block.columns]
        end_row, _ = _write_swap_table(
            writer, worksheet, block, title_row, start_col, _method_display(method),
            fill=True, merge_column=merge_cols or None, merge=bool(merge_cols),
            bar_names=_CF_VARBIN_BAR,
        )
        block_width = _table_display_width(block)
        if first_block_ref is None:
            last_letter = get_column_letter(start_col + block_width - 1)
            first_block_ref = f"{get_column_letter(start_col)}{header_row}:{last_letter}{end_row - 1}"
        start_col += block_width + 1
    _autosize_columns(writer, worksheet)
    # 冻结表头行（向下滚动时方法名 + 表头始终可见），并在首个分箱区块上加自动筛选
    if first_block_ref is not None:
        writer.set_freeze_panes(worksheet, (first_data_row, 1))
        writer.add_auto_filter(worksheet, first_block_ref)


def swap_out_report(
    data: pd.DataFrame,
    rules: Union[str, "Rule", Sequence[Union[str, "Rule"]]],
    background: Optional[Union[str, List[str]]] = None,
    summary: Optional[Union[str, List[str]]] = None,
    describe: Optional[Union[str, List[str]]] = None,
    rule_summary: Optional[Union[str, List[str]]] = None,
    impact: Optional[Union[str, List[str]]] = None,
    target: str = "target",
    overdue: Optional[Union[str, List[str]]] = None,
    dpds: Optional[Union[int, List[int]]] = None,
    save: Optional[str] = None,
    verbose: bool = False,
    methods: Union[str, List[str]] = "quantile",
    bin_params: Optional[Union[Dict[str, Any], Dict[str, Dict[str, Any]]]] = None,
    features: Optional[List[str]] = None,
    amount: Optional[str] = None,
    date_col: Optional[str] = None,
    freq: str = "M",
    group_col: Optional[str] = None,
    current_pass_rate: float = 1.0,
    prior_rules: Optional["Rule"] = None,
    del_grey: bool = False,
    target_names: Optional[Mapping[str, str]] = None,
    theme_color: str = "2639E9",
    sheet_name: str = "策略迭代",
    **kwargs: Any,
):
    """生成拒绝规则置换（策略迭代）分析报告，输出 hscredit 美化后的 Excel 文件.

    参考「策略迭代参考模板」组织内容，输出两个工作表：

    - ``策略迭代`` 主表：迭代背景、策略迭代总结、样本描述（样本情况 + 相关系数）、
      规则变量效果（分箱详情）、业务影响情况分析（订单/金额口径并排）、规则效果分析
      （命中/未命中明细，订单/金额口径并排）、规则稳定性分析（按时间或分组对比）。
    - ``变量分箱`` 明细表：各分箱方法的完整分箱表横向并排展示。

    其中「规则变量效果」与「变量分箱」均源自 :func:`~hscredit.report.feature_binning_summary`
    的一次计算结果（``binning_summary`` 与 ``binning_tables``）。当 ``rules`` 仅传入单条
    规则时，仅展示整体效果，不再拆分子规则。

    **参数**

    :param data: 原始明细数据 DataFrame，需包含规则字段、目标/逾期字段及所需分析列
    :param rules: 拒绝规则，单条 :class:`~hscredit.core.rules.Rule`/表达式字符串，或其列表；
        多条规则时整体为各规则按「或」组合的并集，并逐条展示子规则效果
    :param background: 迭代背景文本，``str`` 或 ``list[str]``（列表自动加「序号、内容」）
    :param summary: 策略迭代总结文本，``str`` 或 ``list[str]``
    :param describe: 样本描述文本，``str`` 或 ``list[str]``
    :param rule_summary: 规则表里效果说明文本，``str`` 或 ``list[str]``
    :param impact: 业务影响情况说明文本，``str`` 或 ``list[str]``
    :param target: 目标变量列名，默认 ``"target"``，0=好样本，1=坏样本
    :param overdue: 逾期天数字段名（可选，传入时以逾期天数>DPD定义坏样本，支持多标签）
    :param dpds: 逾期定义方式，逾期天数 > DPD 为坏样本，可传入列表支持多DPD联合分析
    :param save: 报告保存路径（``.xlsx``）；为 None 时不落盘，仅返回 ExcelWriter
    :param verbose: 是否打印计算进度，默认 False
    :param methods: 分箱详情所用分箱方法，``str`` 或 ``list[str]``，默认 ``"quantile"``
    :param bin_params: 分箱参数，``dict``（所有方法统一）或 ``dict[method: dict]``（按方法名映射）
    :param features: 参与相关系数与分箱详情的指标列表，默认取所有规则引用到的字段
    :param amount: 金额字段名（可选），传入时额外输出金额口径的业务影响与规则效果
    :param date_col: 日期列名，与 ``freq`` 配合做规则稳定性分析（与 ``group_col`` 二选一）
    :param freq: 稳定性分析的时间频率，``'D'/'W'/'M'/'Q'``，默认 ``'M'``
    :param group_col: 分组字段列名，做规则稳定性分析（与 ``date_col`` 二选一）
    :param current_pass_rate: 规则执行前的当前通过率，取值 [0, 1]，默认 1.0
    :param prior_rules: 先验规则（可选），评估前先排除命中先验规则的样本
    :param del_grey: 是否删除逾期天数在 (0, DPD] 区间内的灰度样本，默认 False
    :param target_names: 逾期指标名称映射，如 ``{'MOB1 7+': 'fpd7'}``
    :param theme_color: Excel 主题色（不含 #），默认 ``"2639E9"``
    :param sheet_name: 报告工作表名称，默认 ``"策略迭代"``
    :param kwargs: 透传给 ``Rule.report`` 的其他参数（如 ``desc``、``margins`` 等）
    :return: :class:`~hscredit.excel.ExcelWriter` 实例（已写入全部内容）

    **参考样例**

    >>> from hscredit.report import swap_out_report
    >>> swap_out_report(
    ...     data,
    ...     rules=["衡枢鉴真分老客版 < 0.05", "近六个月非银多头机构数 > 30"],
    ...     background="为压降逾期，筛选近期放款老客样本进行策略迭代验证",
    ...     summary=["低分拒绝规则性价比最高", "多头机构数规则作为补充收紧尾部"],
    ...     overdue=["MOB1"], dpds=[7, 0], amount="放款金额",
    ...     date_col="放款时间", freq="M", methods=["quantile", "mdlp"],
    ...     save="策略迭代报告.xlsx",
    ... )
    """
    from ..excel import ExcelWriter

    if not isinstance(data, pd.DataFrame) or data.empty:
        raise ValueError("data 必须是非空的 DataFrame")

    combined_rule, rule_objs, multi = _resolve_swap_rules(rules)
    methods_list = [methods] if isinstance(methods, str) else list(methods)
    # 统一 target_names 的键形式（兼容 MOB1_7+ / MOB1@7 / MOB1 7+ 三种写法）
    target_names = _normalize_target_names(target_names)

    if features is None:
        features = _ordered_unique([col for rule in rule_objs for col in rule.feature_names_in_])
    features = [feat for feat in features if feat in data.columns]

    targets = _resolve_swap_targets(data, target, overdue, dpds, del_grey)
    report_kwargs = dict(target=target, overdue=overdue, dpds=dpds, del_grey=del_grey, prior_rules=prior_rules)

    # 整体 + 子规则（单规则时仅整体）
    section_rules: List[Tuple[str, "Rule"]] = [("整体", combined_rule)]
    if multi:
        section_rules += [(f"规则{i + 1}", rule) for i, rule in enumerate(rule_objs)]

    # 一次性计算多分箱方法结果：binning_summary 作为「规则变量效果」内容，
    # binning_tables 作为「变量分箱」明细 sheet 的内容来源
    binning_tables: Dict[str, Dict[str, pd.DataFrame]] = {}
    binning_summary = pd.DataFrame()
    if features:
        if verbose:
            print(f"[swap_out_report] 分箱计算中：方法={methods_list}，指标={features} ...")
        from .feature_analyzer import feature_binning_summary

        binning_tables, binning_summary = feature_binning_summary(
            data, features, methods=methods_list, bin_params=bin_params,
            target=None if overdue is not None else target,
            overdue=overdue, dpds=dpds, del_grey=del_grey, long_format=True, verbose=0,
        )
        # 统一分箱详情的逾期标签形式（MOB1@7 → MOB1 7+）并套用 target_names 映射
        binning_summary = _rename_target_level(binning_summary, target_names, level=1)

    start_col = 2
    writer = ExcelWriter(theme_color=theme_color)
    worksheet = writer.get_sheet_by_name(sheet_name)
    row = 2

    # 各模块的内容宽度：保证标题横幅与文本宽度同模块内最宽表格一致
    overview = _swap_sample_overview(targets, date_col, data, target_names)
    corr = None
    if len(features) >= 2:
        corr = data[features].apply(pd.to_numeric, errors="coerce").corr().round(4)
        corr.index.name = "相关系数"
    describe_width = max(
        _table_display_width(overview),
        _table_display_width(corr, index=True) if corr is not None else 0,
    )
    summary_width = _table_display_width(binning_summary) if not binning_summary.empty else 14
    # 各表格模块的内容宽度（用于大标题横幅从起始列铺满到该模块最右列）
    has_amount = amount is not None and amount in data.columns
    impact_pair = 14                                       # rule_target_analysis 固定 14 列
    effect_pair = 3 + len(_SWAP_EFFECT_METRICS)            # 规则详情/逾期指标/命中情况 + 各指标
    impact_width = impact_pair * 2 + _GAP_INNER if has_amount else impact_pair
    effect_width = effect_pair * 2 + _GAP_INNER if has_amount else effect_pair
    # 纯文本模块（迭代背景/策略迭代总结）横幅取报告主体最大宽度，避免短横幅突兀
    body_width = max(14, describe_width, summary_width, impact_width, effect_width)

    # —— 1. 迭代背景 ——
    if background is not None:
        row = _write_banner(writer, worksheet, row, "迭代背景", start_col, body_width, "header")
        row = _write_swap_text(writer, worksheet, row, background, start_col)
        row += _GAP_MODULE

    # —— 2. 策略迭代总结 ——
    if summary is not None:
        row = _write_banner(writer, worksheet, row, "策略迭代总结", start_col, body_width, "header")
        row = _write_swap_text(writer, worksheet, row, summary, start_col)
        row += _GAP_MODULE

    # —— 3. 样本描述（样本情况 + 相关系数）——
    row = _write_banner(writer, worksheet, row, "样本描述", start_col, describe_width, "header")
    if describe is not None:
        row = _write_swap_text(writer, worksheet, row, describe, start_col)
        row += _GAP_INNER
    # 样本情况 / 相关系数：表头左上角即为标题（样本情况 / 相关系数），无需额外标题行
    row, _ = _write_swap_table(
        writer, worksheet, overview, row, start_col, None, scale_names=_CF_OVERVIEW_SCALE,
    )
    if corr is not None:
        row += _GAP_INNER
        row, _ = _write_swap_table(writer, worksheet, corr, row, start_col, None, index=True)
    row += _GAP_MODULE

    # —— 4. 规则变量效果（分箱详情，源自 feature_binning_summary）——
    if not binning_summary.empty:
        row = _write_banner(writer, worksheet, row, "规则变量效果", start_col, summary_width, "header")
        if rule_summary is not None:
            row = _write_swap_text(writer, worksheet, row, rule_summary, start_col)
            row += _GAP_INNER
        # 分箱详情：表头左上角即为「分箱详情」，无需额外标题行
        row, _ = _write_swap_table(
            writer, worksheet, binning_summary, row, start_col, None,
            merge_column=[("分箱详情", "分箱方法")], merge=True,
            bar_names=_CF_BIN_BAR, scale_names=_CF_BIN_SCALE,
        )
        row += _GAP_MODULE

    # —— 5. 业务影响情况分析（订单 / 金额口径目标分析，并排；命中口径展示为 原始/通过/拒绝）——
    row = _write_banner(writer, worksheet, row, "业务影响情况分析", start_col, impact_width, "header")
    if impact is not None:
        row = _write_swap_text(writer, worksheet, row, impact, start_col)
        row += _GAP_INNER
    for label, rule in section_rules:
        prefix = "业务影响情况" if label == "整体" else f"{label}影响情况"
        order_rep = rule.report(data, amount=None, **report_kwargs, **kwargs)
        order_tbl = _rename_hit_labels(rule_target_analysis(
            order_rep, current_pass_rate=current_pass_rate, rule_name=rule.name,
            target_names=target_names, target_name=target,
        ))
        amount_tbl = None
        if amount is not None and amount in data.columns:
            amt_rep = rule.report(data, amount=amount, **report_kwargs, **kwargs)
            amount_tbl = _rename_amount_caliber(_rename_hit_labels(rule_target_analysis(
                amt_rep, current_pass_rate=current_pass_rate, rule_name=rule.name,
                target_names=target_names, target_name=target,
            )))
        row = _write_caliber_pair(
            writer, worksheet, row, start_col, prefix, order_tbl, amount_tbl, bar_names=_CF_IMPACT_BAR,
        )
    row += _GAP_MODULE - _GAP_INNER

    # —— 6. 规则效果分析（仅命中/未命中明细，不含合计；订单 / 金额口径并排）——
    row = _write_banner(writer, worksheet, row, "规则效果分析", start_col, effect_width, "header")
    effect_merge = ["规则详情", "逾期指标"]
    for label, rule in section_rules:
        prefix = "规则整体效果" if label == "整体" else f"{label}效果"
        order_rep = rule.report(data, amount=None, **report_kwargs, **kwargs)
        order_tbl = _drop_total_rows(rule_target_table(
            order_rep, rule_name=rule.name, target_names=target_names,
            metrics=_SWAP_EFFECT_METRICS, target_name=target,
        ))
        amount_tbl = None
        if amount is not None and amount in data.columns:
            amt_rep = rule.report(data, amount=amount, **report_kwargs, **kwargs)
            amount_tbl = _rename_amount_caliber(_drop_total_rows(rule_target_table(
                amt_rep, rule_name=rule.name, target_names=target_names,
                metrics=_SWAP_EFFECT_METRICS, target_name=target,
            )))
        row = _write_caliber_pair(
            writer, worksheet, row + 1, start_col, prefix, order_tbl, amount_tbl,
            merge_column=effect_merge, merge=True, bar_names=_CF_EFFECT_BAR,
        )
    row += _GAP_MODULE - _GAP_INNER

    # —— 7. 规则稳定性分析（按时间或分组对比）——
    if date_col is not None or group_col is not None:
        # 先逐规则计算稳定性表，按最大宽度铺横幅；全部失败时不写空模块
        stability_blocks: List[Tuple[str, pd.DataFrame]] = []
        for label, rule in section_rules:
            try:
                stability = rule_group_compare(
                    data, rule, date_col=date_col, freq=freq, group_col=group_col,
                    target=target, overdue=overdue, dpds=dpds, rule_name=rule.name,
                    target_names=target_names, prior_rules=prior_rules, del_grey=del_grey,
                )
            except Exception as exc:  # noqa: BLE001 - 稳定性切分失败不阻断整体报告
                if verbose:
                    print(f"[swap_out_report] 规则稳定性分析失败 ({label}): {exc}")
                continue
            prefix = "规则效果稳定性" if label == "整体" else f"{label}稳定性"
            stability_blocks.append((prefix, stability))
        if stability_blocks:
            stability_width = max(_table_display_width(tbl) for _, tbl in stability_blocks)
            row = _write_banner(writer, worksheet, row, "规则稳定性分析", start_col, stability_width, "header")
            for prefix, stability in stability_blocks:
                end_r, _ = _write_swap_table(
                    writer, worksheet, stability, row + 1, start_col, prefix,
                    merge_column=[("规则详情", "规则名称"), ("规则详情", "逾期指标")], merge=True,
                )
                # 坏样本率 / 样本占比 各按「指标分组」整体加一条数据条（跨各时间周期共享标尺）
                _add_group_databars(writer, worksheet, stability, end_r, start_col, _CF_STABILITY_BAR)
                row = end_r + _GAP_INNER

    _autosize_columns(writer, worksheet)

    # —— 第二个 sheet：变量分箱明细（数据源自 feature_binning_summary 的 binning_tables）——
    if binning_tables and features:
        _write_binning_detail_sheet(writer, binning_tables, features, methods_list, target_names=target_names)

    if save:
        writer.save(save)
        if verbose:
            print(f"[swap_out_report] 报告已保存至: {save}")
    return writer


__all__ = [
    "rule_report_table",
    "rule_target_analysis",
    "rule_target_table",
    "rule_group_hit_table",
    "rule_group_compare",
    "swap_out_report",
]
