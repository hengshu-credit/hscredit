"""Excel写入模块测试."""

import inspect
import os
import re
import tempfile
import zipfile
import xml.etree.ElementTree as ET
import pytest
import pandas as pd
import numpy as np
from openpyxl import load_workbook

from hscredit.excel import ExcelWriter, dataframe2excel, register_pivot_aggregation
from hscredit.utils import fonts
import hscredit.excel.writer as writer_module


def _conditional_format_colors(worksheet):
    """提取工作表数据条与色阶规则中的实际颜色。"""
    data_bar_colors = []
    color_scale_colors = []
    for rules in worksheet.conditional_formatting._cf_rules.values():
        for rule in rules:
            if rule.type == "dataBar":
                data_bar_colors.append(rule.dataBar.color.rgb)
            elif rule.type == "colorScale":
                color_scale_colors.extend(color.rgb for color in rule.colorScale.color)
    return data_bar_colors, color_scale_colors


class TestExcelWriter:
    """测试ExcelWriter类"""
    
    def setup_method(self):
        """每个测试方法前的设置"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, "test.xlsx")
    
    def teardown_method(self):
        """每个测试方法后的清理"""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    def test_init_uses_runtime_default_font(self, monkeypatch):
        """默认构造应在运行时读取字体初始化结果."""
        monkeypatch.setattr(fonts, "_default_font_name", "Alimama FangYuanTi VF")

        writer = ExcelWriter(theme_color='3f1dba')

        assert writer.theme_color == '3f1dba'
        assert writer.fontsize == 10
        assert writer.font == '阿里妈妈方圆体 VF Medium'
        content_style = next(style for style in writer.name_styles if style.name == "content")
        assert content_style.font.name == '阿里妈妈方圆体 VF Medium'

    def test_init_explicit_font_overrides_runtime_default(self, monkeypatch):
        """显式字体参数应优先于自动初始化结果."""
        monkeypatch.setattr(fonts, "_default_font_name", "Alimama FangYuanTi VF")

        writer = ExcelWriter(font="楷体")

        assert writer.font == "楷体"
        content_style = next(style for style in writer.name_styles if style.name == "content")
        assert content_style.font.name == "楷体"

    def test_condition_color_defaults_to_secondary_theme_for_data_bar(self):
        """ExcelWriter 默认应使用副主题色生成数据条。"""
        writer = ExcelWriter()
        worksheet = writer.get_sheet_by_name("Condition")

        writer.add_conditional_formatting(worksheet, "B2", "B3")

        data_bar_colors, _ = _conditional_format_colors(worksheet)
        assert writer.condition_color == "F76E6C"
        assert data_bar_colors == ["00F76E6C"]

    @pytest.mark.parametrize(
        ("call_color", "expected_color"),
        [
            (None, "00112233"),
            ("445566", "00445566"),
        ],
    )
    def test_add_conditional_formatting_prefers_call_color_over_writer_color(
        self,
        call_color,
        expected_color,
    ):
        """方法显式颜色应覆盖 ExcelWriter 级条件格式颜色。"""
        writer = ExcelWriter(condition_color="112233")
        worksheet = writer.get_sheet_by_name("Condition")

        writer.add_conditional_formatting(
            worksheet,
            "B2",
            "B3",
            condition_color=call_color,
        )

        data_bar_colors, _ = _conditional_format_colors(worksheet)
        assert data_bar_colors == [expected_color]

    @pytest.mark.parametrize("speed", ["normal", "fast"])
    def test_runtime_default_font_persists_in_written_cells(self, monkeypatch, speed):
        """普通与快速写入保存后都应保留运行时默认字体."""
        monkeypatch.setattr(fonts, "_default_font_name", "Alimama FangYuanTi VF")
        writer = ExcelWriter()
        worksheet = writer.get_sheet_by_name("Font")

        writer.insert_df2sheet(
            worksheet,
            pd.DataFrame({"特征": ["年龄"]}),
            "B2",
            speed=speed,
        )
        writer.save(self.test_file)

        loaded = load_workbook(self.test_file)
        assert loaded["Font"]["B2"].font.name == "阿里妈妈方圆体 VF Medium"
        assert loaded["Font"]["B3"].font.name == "阿里妈妈方圆体 VF Medium"
        loaded.close()
    
    def test_get_sheet_by_name(self):
        """测试获取或创建sheet"""
        writer = ExcelWriter()
        
        # 创建新sheet
        ws1 = writer.get_sheet_by_name("Sheet1")
        assert ws1.title == "Sheet1"
        
        # 获取已有sheet
        ws2 = writer.get_sheet_by_name("Sheet1")
        assert ws2 is ws1
    
    def test_insert_value(self):
        """测试插入值"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        # 插入普通值
        end_row, end_col = writer.insert_value2sheet(ws, "B2", value="测试内容")
        assert ws["B2"].value == "测试内容"
        assert end_row == 3
        assert end_col == 3

    def test_auto_width_preserves_template_column_fill(self):
        """测试自动列宽保留模板列白色填充样式。"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")

        # G列在模板中继承 C:XFD 的列维度样式，调整列宽后仍应保留该样式。
        writer.insert_value2sheet(ws, "G2", value="自动列宽测试", auto_width=True)
        assert ws.column_dimensions["G"].style == 1

        writer.save(self.test_file)
        with zipfile.ZipFile(self.test_file) as zf:
            worksheet_xmls = [
                zf.read(name).decode("utf-8")
                for name in zf.namelist()
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
            ]

        target_xml = next(xml for xml in worksheet_xmls if 'r="G2"' in xml)
        root = ET.fromstring(target_xml)
        ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        adjusted_col = next(
            col for col in root.findall("x:cols/x:col", ns)
            if col.attrib.get("min") == "7" and col.attrib.get("max") == "7"
        )
        assert adjusted_col.attrib.get("style") == "1"
        assert 'r="G3"' not in target_xml
    
    def test_insert_value_with_merge(self):
        """测试合并单元格插入"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        end_row, end_col = writer.insert_value2sheet(
            ws, "B2",
            value="合并单元格",
            end_space="D2"
        )
        
        assert ws["B2"].value == "合并单元格"
        assert end_row == 3
        #  end_col 是开区间，B2到D2是3列(B,C,D)，所以end_col=5
        assert end_col in [4, 5]  # 允许实现差异
    
    def test_insert_dataframe(self):
        """测试插入DataFrame"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        df = pd.DataFrame({
            'A': [1, 2, 3],
            'B': [4, 5, 6],
            'C': [7, 8, 9]
        })
        
        end_row, end_col = writer.insert_df2sheet(ws, df, "B2")
        
        # 检查header
        assert ws["B2"].value == 'A'
        assert ws["C2"].value == 'B'
        assert ws["D2"].value == 'C'
        
        # 检查数据
        assert ws["B3"].value == 1
        assert ws["C3"].value == 4
        assert ws["D3"].value == 7
    
    def test_insert_dataframe_with_index(self):
        """测试插入带索引的DataFrame"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        df = pd.DataFrame({
            'A': [1, 2, 3],
            'B': [4, 5, 6]
        }, index=['X', 'Y', 'Z'])
        
        end_row, end_col = writer.insert_df2sheet(ws, df, "B2", index=True)
        
        # 检查索引
        assert ws["B3"].value == 'X'
        assert ws["B4"].value == 'Y'
        assert ws["B5"].value == 'Z'

    def test_insert_dataframe_multi_header_merge_levels(self):
        """测试多层表头可选择仅合并指定层级。"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")

        df = pd.DataFrame(
            [[0.1, 0.2, 0.3]],
            columns=pd.MultiIndex.from_tuples([
                ("坏样本率", "拒绝"),
                ("坏样本率", "拒绝"),
                ("LIFT值", "拒绝"),
            ]),
        )

        writer.insert_df2sheet(ws, df, "B2", merge_header=[0])

        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        assert "B2:C2" in merged_ranges
        assert "B3:C3" not in merged_ranges
        assert ws["B3"].value == "拒绝"
        assert ws["C3"].value == "拒绝"

    def test_multi_header_does_not_write_single_cell_merge_refs(self):
        """多层表头不应在 xlsx XML 中写入单格 mergeCell 记录。"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")

        df = pd.DataFrame(
            [[10, 8, 2, 0.2]],
            index=pd.Index(["训练集"], name="数据集"),
            columns=pd.MultiIndex.from_tuples([
                ("统计详情", "样本总数"),
                ("好样本数", "MOB1@7"),
                ("坏样本数", "MOB1@7"),
                ("坏样本率", "MOB1@7"),
            ]),
        )

        writer.insert_df2sheet(ws, df, "B2", index=True)
        writer.save(self.test_file)

        with zipfile.ZipFile(self.test_file) as workbook_zip:
            merge_refs = []
            for name in workbook_zip.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    xml = workbook_zip.read(name).decode("utf-8")
                    merge_refs.extend(re.findall(r'<mergeCell ref="([^"]+)"', xml))

        assert merge_refs
        assert all(":" in ref for ref in merge_refs)

    def test_multi_header_nan_placeholder_does_not_write_reverse_merge_refs(self):
        """Multi-level header blank placeholders must not produce reverse merge ranges."""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")

        item, start, length = ExcelWriter.calc_continuous_cnt([np.nan, "指标"], 0)
        assert pd.isna(item)
        assert start == 0
        assert length == 1

        writer.insert_rows(ws, [np.nan, "指标"], 2, 2, style="header", multi_levels=True)

        assert ws["B2"].value is None or pd.isna(ws["B2"].value)
        assert [str(rng) for rng in ws.merged_cells.ranges] == []

    def test_insert_dataframe_multi_header_merge_can_be_disabled(self):
        """测试多层表头可完全关闭横向合并。"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")

        df = pd.DataFrame(
            [[0.1, 0.2]],
            columns=pd.MultiIndex.from_tuples([
                ("坏样本率", "拒绝"),
                ("坏样本率", "拒绝"),
            ]),
        )

        writer.insert_df2sheet(ws, df, "B2", merge_header=False)

        assert list(ws.merged_cells.ranges) == []
        assert ws["B2"].value == "坏样本率"
        assert ws["C2"].value == "坏样本率"
        assert ws["B3"].value == "拒绝"
        assert ws["C3"].value == "拒绝"

    def test_insert_dataframe_with_merge(self):
        """测试插入DataFrame并合并相同值"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        df = pd.DataFrame({
            'Category': ['A', 'A', 'B', 'B'],
            'Value': [1, 2, 3, 4]
        })
        
        end_row, end_col = writer.insert_df2sheet(
            ws, df, "B2",
            merge_column='Category',
            merge=True
        )
        
        # 检查合并后的值
        assert ws["B3"].value == 'A'
        assert ws["B5"].value == 'B'
    
    def test_insert_hyperlink(self):
        """测试插入超链接"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        writer.insert_value2sheet(ws, "B2", value="点击跳转")
        writer.insert_hyperlink2sheet(ws, "B2", target_space="B10")
        
        assert ws["B2"].hyperlink is not None
    
    def test_set_number_format(self):
        """测试设置数字格式"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        ws["B2"] = 0.123456
        writer.set_number_format(ws, "B2", "0.00%")
        
        assert ws["B2"].number_format == "0.00%"
    
    def test_set_column_width(self):
        """测试设置列宽"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        writer.set_column_width(ws, 'B', 20)
        
        assert ws.column_dimensions['B'].width == 20
    
    def test_set_freeze_panes(self):
        """测试设置冻结窗格"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        writer.set_freeze_panes(ws, "B2")
        
        assert ws.freeze_panes == "B2"
    
    def test_save_and_load(self):
        """测试保存和加载"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})
        writer.insert_df2sheet(ws, df, "B2")
        writer.save(self.test_file)
        
        # 重新加载验证
        loaded_wb = load_workbook(self.test_file)
        loaded_ws = loaded_wb['Test']
        
        assert loaded_ws["B2"].value == 'A'
        assert loaded_ws["B3"].value == 1
    
    def test_append_mode(self):
        """测试追加模式"""
        # 第一次写入
        writer1 = ExcelWriter()
        ws1 = writer1.get_sheet_by_name("Sheet1")
        writer1.insert_value2sheet(ws1, "B2", value="第一次")
        writer1.save(self.test_file)
        
        # 追加写入
        writer2 = ExcelWriter(mode='append')
        ws2 = writer2.get_sheet_by_name("Sheet2")
        writer2.insert_value2sheet(ws2, "B2", value="第二次")
        writer2.save(self.test_file)
        
        # 验证两个sheet都存在
        loaded_wb = load_workbook(self.test_file)
        assert 'Sheet1' in loaded_wb.sheetnames
        assert 'Sheet2' in loaded_wb.sheetnames

    def test_insert_picture_row_span_uses_ceil(self, monkeypatch):
        """测试图片占用行数使用向上取整，避免覆盖下方内容"""

        class DummyImage:
            def __init__(self, _fig):
                self.width = None
                self.height = None

        monkeypatch.setattr(writer_module, "Image", DummyImage)

        writer = ExcelWriter(system="windows")
        ws = writer.get_sheet_by_name("Test")

        end_row, end_col = writer.insert_pic2sheet(ws, "dummy.png", "B2", figsize=(600, 250))

        assert end_row == 18
        assert end_col == 10


class TestExcelChart:
    """测试原生图表插入功能"""

    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, "chart.xlsx")
        self.bin_table = pd.DataFrame({
            "分箱": [0, 1, 2, 3],
            "分箱标签": ["(-inf, 580]", "(580, 620]", "(620, 660]", "(660, inf]"],
            "样本总数": [120, 200, 180, 100],
            "好样本数": [80, 160, 165, 96],
            "坏样本数": [40, 40, 15, 4],
            "坏样本率": [0.333, 0.20, 0.083, 0.04],
        })

    def teardown_method(self):
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_insert_chart2sheet(self):
        """测试插入原生 openpyxl 图表"""
        from openpyxl.chart import BarChart, Reference

        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        writer.insert_df2sheet(ws, self.bin_table, "B2")

        chart = BarChart()
        data = Reference(ws, min_col=4, min_row=2, max_row=6)  # 样本总数列
        chart.add_data(data, titles_from_data=True)
        end_row, end_col = writer.insert_chart2sheet(ws, "J2", chart)

        assert len(ws._charts) == 1
        assert end_row > 2 and end_col > column_letter_to_index("J")

    def test_insert_bin_chart2sheet(self):
        """测试基于分箱表生成分箱图（柱状+折线双轴）"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("分箱图")
        writer.insert_df2sheet(ws, self.bin_table, "B2", fill=True)
        writer.insert_bin_chart2sheet(ws, self.bin_table, "B2", title="某特征分箱图")

        writer.save(self.test_file)

        loaded_wb = load_workbook(self.test_file)
        loaded_ws = loaded_wb["分箱图"]
        assert len(loaded_ws._charts) == 1

    def test_bin_chart_handles_missing_columns(self):
        """测试缺失部分列时不报错（仅引用存在的列）"""
        partial = self.bin_table[["分箱标签", "好样本数", "坏样本率"]]
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        writer.insert_df2sheet(ws, partial, "B2")
        # 坏样本数 缺失，应自动跳过
        writer.insert_bin_chart2sheet(ws, partial, "B2")
        assert len(ws._charts) == 1

    def test_bin_chart_axis_graph_is_consistent(self):
        """回归：双轴分箱图必须有2个分类轴+2个数值轴且 crossAx 两两一致。

        若折线与柱状共用同一条分类轴，会产生悬空的次数值轴（crossAx 指向不回指的轴），
        Excel 打开时会判定图形损坏并删除整个 drawing 部件。
        """
        import re
        import zipfile

        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("分箱图")
        writer.insert_df2sheet(ws, self.bin_table, "B2", fill=True)
        writer.insert_bin_chart2sheet(ws, self.bin_table, "B2", title="某特征分箱图")
        writer.save(self.test_file)

        with zipfile.ZipFile(self.test_file) as z:
            chart_xml = z.read("xl/charts/chart1.xml").decode("utf-8")

        # 收集每条轴的 (自身id -> crossAx 目标)
        axes = {}
        for kind in ("catAx", "valAx"):
            for m in re.finditer(r"<(\w+:)?%s>(.*?)</(\w+:)?%s>" % (kind, kind), chart_xml, re.S):
                body = m.group(0)
                # 兼容 lxml（<axId val="10"/>）与 ElementTree（<axId val="10" />，含空格）
                # 两种序列化后端，避免 CI 无 lxml 时回退导致正则失配
                sid = re.search(r'<\w*:?axId val="(\d+)"\s*/?>', body)
                cross = re.search(r'<\w*:?crossAx val="(\d+)"\s*/?>', body)
                assert sid and cross, "轴缺少 axId 或 crossAx"
                axes[sid.group(1)] = cross.group(1)

        n_cat = len(re.findall(r"<\w*:?catAx>", chart_xml))
        n_val = len(re.findall(r"<\w*:?valAx>", chart_xml))
        assert n_cat == 2 and n_val == 2, "双轴图应有2个分类轴+2个数值轴，实际 cat=%d val=%d" % (n_cat, n_val)

        # crossAx 必须两两互指：A.crossAx=B 则 B.crossAx=A
        for sid, target in axes.items():
            assert target in axes, "crossAx 目标 %s 不存在" % target
            assert axes[target] == sid, "crossAx 不一致: %s->%s 但 %s->%s" % (sid, target, target, axes[target])

    def test_bin_chart_without_header(self):
        """回归：header=False 时系列取数从数据首行起，不引用数据上方空行，文件合法"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("分箱图")
        writer.insert_df2sheet(ws, self.bin_table, "B2", header=False)
        writer.insert_bin_chart2sheet(ws, self.bin_table, "B2", header=False)
        writer.save(self.test_file)

        import zipfile
        import xml.dom.minidom as minidom
        with zipfile.ZipFile(self.test_file) as z:
            for n in z.namelist():
                if n.endswith(".xml") or n.endswith(".rels"):
                    minidom.parseString(z.read(n))  # XML 合法
        loaded_wb = load_workbook(self.test_file)
        assert len(loaded_wb["分箱图"]._charts) == 1


def column_letter_to_index(letter):
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)


class TestSparkline:
    """测试迷你图（Sparkline）功能"""

    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, "spark.xlsx")
        self.df = pd.DataFrame({
            "指标": ["A", "B", "C"],
            "m1": [3, 5, 2], "m2": [5, 4, 6], "m3": [2, 7, 3],
            "m4": [6, 3, 8], "m5": [4, 6, 5],
        })

    def teardown_method(self):
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def _read_sheet_xml(self, path):
        import zipfile
        with zipfile.ZipFile(path) as z:
            for n in z.namelist():
                if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
                    yield z.read(n).decode("utf-8")

    def test_add_sparkline_records_spec(self):
        """测试 add_sparkline 仅记录配置，不立即写入"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        writer.add_sparkline(ws, "H2", "B2:G2")
        assert len(writer._sparkline_specs) == 1
        spec = writer._sparkline_specs[0]
        # 数据区域自动补全当前 sheet 前缀
        assert spec["sparklines"][0][0] == "Test!B2:G2"

    def test_sparkline_injected_on_save(self):
        """测试保存后迷你图 XML 被注入且文件可正常打开"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("迷你图")
        writer.insert_df2sheet(ws, self.df, "B2", fill=True)
        writer.add_sparkline(ws, "H3", "C3:G3", markers=True, high_point=True, low_point=True)
        writer.add_sparkline(ws, "H4", "C4:G4", type="column")
        writer.add_sparkline(ws, "H5", "C5:G5", type="win_loss")
        writer.save(self.test_file)

        # 文件可被 openpyxl 重新打开（未损坏）
        load_workbook(self.test_file)

        # worksheet XML 含 sparklineGroups
        xmls = list(self._read_sheet_xml(self.test_file))
        assert any("sparklineGroups" in x for x in xmls)
        # win_loss -> stacked, column -> column
        joined = "".join(xmls)
        assert 'type="column"' in joined
        assert 'type="stacked"' in joined

    def test_sparkline_grouped(self):
        """测试单次调用批量生成同组迷你图"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        writer.add_sparkline(
            ws, ["I3", "I4", "I5"], ["C3:G3", "C4:G4", "C5:G5"], type="line"
        )
        assert len(writer._sparkline_specs) == 1
        assert len(writer._sparkline_specs[0]["sparklines"]) == 3

    def test_sparkline_invalid_type(self):
        """测试非法类型抛出异常"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        with pytest.raises(ValueError):
            writer.add_sparkline(ws, "H2", "B2:G2", type="pie")

    def test_sparkline_location_range_mismatch(self):
        """测试 location 与 data_range 数量不一致抛出异常"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        with pytest.raises(ValueError):
            writer.add_sparkline(ws, ["H2", "H3"], ["B2:G2"])


class TestDataframe2Excel:
    """测试dataframe2excel便捷函数"""
    
    def setup_method(self):
        """每个测试方法前的设置"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, "test.xlsx")
    
    def teardown_method(self):
        """每个测试方法后的清理"""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    def test_basic_write(self):
        """测试基本写入"""
        df = pd.DataFrame({
            'A': [1, 2, 3],
            'B': [4, 5, 6]
        })
        
        end_row, end_col = dataframe2excel(df, self.test_file)
        
        assert os.path.exists(self.test_file)
        assert end_row > 0
        assert end_col > 0

    @pytest.mark.parametrize('speed', ['normal', 'fast'])
    def test_decimal_none_preserves_float_precision(self, speed):
        """decimal=None 时不得主动截断 DataFrame 浮点值。"""
        value = 1.234567890123

        dataframe2excel(pd.DataFrame({'值': [value]}), self.test_file, decimal=None, speed=speed)

        loaded_wb = load_workbook(self.test_file, data_only=False)
        assert loaded_wb.active['B3'].value == value

    @pytest.mark.parametrize('speed', ['normal', 'fast'])
    def test_decimal_controls_float_precision(self, speed):
        """decimal 应与 ScoreCard 一样表示保留的小数位数。"""
        dataframe2excel(pd.DataFrame({'值': [1.2356]}), self.test_file, decimal=2, speed=speed)

        loaded_wb = load_workbook(self.test_file, data_only=False)
        assert loaded_wb.active['B3'].value == 1.24

    @pytest.mark.parametrize('decimal', [-1, True, 1.5, '2'])
    def test_invalid_decimal_raises_chinese_value_error(self, decimal):
        """非法 decimal 不得静默写出精度不明确的文件。"""
        with pytest.raises(ValueError, match='decimal 必须是大于等于 0 的整数或 None'):
            dataframe2excel(pd.DataFrame({'值': [1.2345]}), self.test_file, decimal=decimal)

    def test_dataframe2excel_speed_defaults_to_auto_and_replaces_fast(self):
        """公开 API 应只保留默认 auto 的 speed 参数。"""
        function_params = inspect.signature(dataframe2excel).parameters
        method_params = inspect.signature(ExcelWriter.insert_df2sheet).parameters

        assert function_params['speed'].default == 'auto'
        assert method_params['speed'].default == 'auto'
        assert 'fast' not in function_params
        assert 'fast' not in method_params

    @pytest.mark.parametrize(
        'rows, cols, index, expected',
        [
            (499, 1, False, 'normal'),
            (500, 1, False, 'fast'),
            (1, 49, False, 'normal'),
            (1, 50, False, 'fast'),
            (399, 25, False, 'normal'),
            (400, 25, False, 'fast'),
            (1, 49, True, 'fast'),
        ],
    )
    def test_auto_speed_uses_rows_effective_columns_and_cells(self, rows, cols, index, expected):
        """auto 应在三个公开边界上选择确定的写入路径。"""
        data = pd.DataFrame(np.zeros((rows, cols)))

        assert ExcelWriter._resolve_write_speed(data, 'auto', index=index) == expected

    def test_explicit_speed_overrides_auto_and_normalizes_text(self):
        """显式 normal/fast 必须覆盖大小判断，并兼容空格和大小写。"""
        small = pd.DataFrame([[1]])
        large = pd.DataFrame(np.zeros((500, 50)))

        assert ExcelWriter._resolve_write_speed(small, ' FAST ') == 'fast'
        assert ExcelWriter._resolve_write_speed(large, 'NORMAL') == 'normal'

    @pytest.mark.parametrize('speed', [None, True, 1, 'turbo'])
    def test_invalid_speed_raises_chinese_value_error(self, speed):
        """未知速度不能静默回退到任意路径。"""
        with pytest.raises(ValueError, match="speed 必须是 'auto'、'normal' 或 'fast'"):
            dataframe2excel(pd.DataFrame({'值': [1]}), self.test_file, speed=speed)

    def test_default_auto_speed_writes_small_and_large_tables(self):
        """默认 auto 选择的普通和快速路径都应写出正确内容。"""
        small_file = os.path.join(self.temp_dir, 'auto-small.xlsx')
        large_file = os.path.join(self.temp_dir, 'auto-large.xlsx')
        small = pd.DataFrame(np.zeros((10, 10)))
        large = pd.DataFrame(np.zeros((500, 2)))

        dataframe2excel(small, small_file)
        dataframe2excel(large, large_file)

        assert load_workbook(small_file).active['B3'].value == 0
        assert load_workbook(large_file).active['B3'].value == 0

    def test_fast_mode_preserves_values_order_styles_and_coordinates(self):
        """快速模式不得改变值、行列顺序、样式或返回坐标。"""
        df = pd.DataFrame(
            {'编号': ['001', '002'], '数值': [1.23456, 2.34567]},
            index=pd.Index(['甲', '乙'], name='样本'),
        )
        normal_file = os.path.join(self.temp_dir, 'normal.xlsx')
        fast_file = os.path.join(self.temp_dir, 'fast.xlsx')

        normal_end = dataframe2excel(df, normal_file, sheet_name='S', index=True, fill=True, speed='normal')
        fast_end = dataframe2excel(df, fast_file, sheet_name='S', index=True, fill=True, speed='fast')

        normal_ws = load_workbook(normal_file, data_only=False)['S']
        fast_ws = load_workbook(fast_file, data_only=False)['S']
        assert fast_end == normal_end
        for row in range(2, normal_end[0]):
            for col in range(2, normal_end[1]):
                normal_cell = normal_ws.cell(row=row, column=col)
                fast_cell = fast_ws.cell(row=row, column=col)
                assert fast_cell.value == normal_cell.value
                assert fast_cell._style == normal_cell._style
        assert fast_ws['C3'].value == '001'
        assert fast_ws['C3'].number_format == '@'

    def test_fast_mode_preserves_multiindex_headers_and_merges(self):
        """快速模式必须保留多层列名、索引顺序和合并范围。"""
        columns = pd.MultiIndex.from_tuples([('统计', '金额'), ('统计', '数量'), ('标签', '等级')])
        index = pd.MultiIndex.from_tuples(
            [('甲组', 1), ('甲组', 2), ('乙组', 1)],
            names=['客群', '序号'],
        )
        df = pd.DataFrame([[10.5, 2, 'A'], [20.5, 3, 'B'], [30.5, 4, 'C']], columns=columns, index=index)
        normal_file = os.path.join(self.temp_dir, 'multi-normal.xlsx')
        fast_file = os.path.join(self.temp_dir, 'multi-fast.xlsx')

        normal_end = dataframe2excel(df, normal_file, sheet_name='S', index=True, fill=True, speed='normal')
        fast_end = dataframe2excel(df, fast_file, sheet_name='S', index=True, fill=True, speed='fast')

        normal_ws = load_workbook(normal_file, data_only=False)['S']
        fast_ws = load_workbook(fast_file, data_only=False)['S']
        assert fast_end == normal_end
        assert sorted(map(str, fast_ws.merged_cells.ranges)) == sorted(map(str, normal_ws.merged_cells.ranges))
        for row in range(2, normal_end[0]):
            for col in range(2, normal_end[1]):
                normal_cell = normal_ws.cell(row=row, column=col)
                fast_cell = fast_ws.cell(row=row, column=col)
                assert fast_cell.value == normal_cell.value
                assert fast_cell._style == normal_cell._style

    @pytest.mark.parametrize('fill', [False, True])
    def test_fast_mode_preserves_fill_style_variants(self, fill):
        """快速路径不能弱化边框模式或填充模式。"""
        df = pd.DataFrame({'A': [1, 2, 3], 'B': ['x', 'y', 'z']})
        normal_file = os.path.join(self.temp_dir, f'fill-{fill}-normal.xlsx')
        fast_file = os.path.join(self.temp_dir, f'fill-{fill}-fast.xlsx')

        normal_end = dataframe2excel(df, normal_file, fill=fill, speed='normal')
        fast_end = dataframe2excel(df, fast_file, fill=fill, speed='fast')

        normal_ws = load_workbook(normal_file).active
        fast_ws = load_workbook(fast_file).active
        assert fast_end == normal_end
        for row in range(2, normal_end[0]):
            for col in range(2, normal_end[1]):
                assert fast_ws.cell(row, col)._style == normal_ws.cell(row, col)._style

    def test_dataframe_auto_width_does_not_snapshot_each_cell(self, monkeypatch):
        """DataFrame 自动列宽只能按列处理一次，不能在每个单元格重复快照。"""
        calls = []
        original = ExcelWriter._get_column_cells_data

        def counted(writer, worksheet, col_letter):
            calls.append(col_letter)
            return original(writer, worksheet, col_letter)

        monkeypatch.setattr(ExcelWriter, '_get_column_cells_data', counted)
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name('S')

        dataframe2excel(
            pd.DataFrame({'A': range(20), 'B': range(20)}),
            writer,
            sheet_name=ws,
            auto_width=True,
        )

        assert calls == ['B', 'C']

    def test_fast_auto_width_avoids_column_style_snapshots(self, monkeypatch):
        """快速模式应在写入时累计宽度，不能再次扫描和复制整列样式。"""
        calls = []
        original = ExcelWriter._get_column_cells_data

        def counted(writer, worksheet, col_letter):
            calls.append(col_letter)
            return original(writer, worksheet, col_letter)

        monkeypatch.setattr(ExcelWriter, '_get_column_cells_data', counted)
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name('S')

        dataframe2excel(
            pd.DataFrame({'中文列': ['短', '较长内容'], 'number': [1.2, 345.6]}),
            writer,
            sheet_name=ws,
            auto_width=True,
            speed='fast',
        )

        assert calls == []
        assert ws.column_dimensions['B'].width >= 8
        assert ws.column_dimensions['C'].width >= 8

    def test_fast_mode_preserves_formats_alignment_filter_and_width(self):
        """快速模式必须保留外层格式、对齐、筛选和自动列宽。"""
        df = pd.DataFrame({
            '特征': ['短名', '较长的中文特征名'],
            '占比': [0.12345, 0.54321],
            '金额': [1234.5, 6789.0],
            '指标': [-0.2, 0.8],
        })
        normal_file = os.path.join(self.temp_dir, 'formats-normal.xlsx')
        fast_file = os.path.join(self.temp_dir, 'formats-fast.xlsx')
        params = dict(
            sheet_name='S',
            percent_cols=['占比'],
            custom_cols=['金额'],
            custom_format='#,##0.00',
            condition_cols=['指标'],
            color_cols=['占比'],
            left_cols=['特征'],
            right_cols=['金额'],
            auto_filter=True,
            auto_width=True,
        )

        normal_end = dataframe2excel(df, normal_file, speed='normal', **params)
        fast_end = dataframe2excel(df, fast_file, speed='fast', **params)

        normal_ws = load_workbook(normal_file, data_only=False)['S']
        fast_ws = load_workbook(fast_file, data_only=False)['S']
        assert fast_end == normal_end
        assert fast_ws.auto_filter.ref == normal_ws.auto_filter.ref
        for col in range(2, normal_end[1]):
            letter = normal_ws.cell(row=2, column=col).column_letter
            assert fast_ws.column_dimensions[letter].width == pytest.approx(
                normal_ws.column_dimensions[letter].width
            )
        for row in range(2, normal_end[0]):
            for col in range(2, normal_end[1]):
                normal_cell = normal_ws.cell(row=row, column=col)
                fast_cell = fast_ws.cell(row=row, column=col)
                assert fast_cell.value == normal_cell.value
                assert fast_cell._style == normal_cell._style

        normal_rules = [
            (str(key.sqref), [rule.type for rule in rules])
            for key, rules in normal_ws.conditional_formatting._cf_rules.items()
        ]
        fast_rules = [
            (str(key.sqref), [rule.type for rule in rules])
            for key, rules in fast_ws.conditional_formatting._cf_rules.items()
        ]
        assert fast_rules == normal_rules
    
    def test_write_with_title(self):
        """测试带标题写入"""
        df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})
        
        dataframe2excel(
            df, self.test_file,
            title="测试标题",
            sheet_name="测试Sheet"
        )
        
        loaded_wb = load_workbook(self.test_file)
        ws = loaded_wb['测试Sheet']
        
        assert ws["B2"].value == "测试标题"
    
    def test_write_with_percent_format(self):
        """测试百分比格式"""
        df = pd.DataFrame({
            'feature': ['A', 'B'],
            'rate': [0.05, 0.10]
        })
        
        dataframe2excel(
            df, self.test_file,
            percent_cols=['rate']
        )
        
        loaded_wb = load_workbook(self.test_file)
        ws = loaded_wb.active
        
        # 检查数字格式
        assert ws["C3"].number_format == "0.00%"
    
    def test_write_with_condition_format(self):
        """测试条件格式"""
        df = pd.DataFrame({
            'feature': ['A', 'B', 'C'],
            'value': [1, 2, 3]
        })
        
        dataframe2excel(
            df, self.test_file,
            condition_cols=['value']
        )
        
        loaded_wb = load_workbook(self.test_file)
        ws = loaded_wb.active
        
        # 检查条件格式是否存在
        assert len(ws.conditional_formatting._cf_rules) > 0

    @pytest.mark.parametrize(
        ("call_color", "expected_color"),
        [
            (None, "00112233"),
            ("445566", "00445566"),
            ({"未命中字段": "445566"}, "00112233"),
        ],
    )
    def test_dataframe2excel_prefers_call_color_over_existing_writer_color(
        self,
        call_color,
        expected_color,
    ):
        """单次调用颜色优先，未传或字典未命中时继承已有 Writer。"""
        writer = ExcelWriter(condition_color="112233")
        worksheet = writer.get_sheet_by_name("Condition")
        data = pd.DataFrame({"数据条": [1, 2], "色阶": [2, 1]})

        dataframe2excel(
            data,
            writer,
            sheet_name=worksheet,
            condition_cols=["数据条"],
            color_cols=["色阶"],
            condition_color=call_color,
        )

        data_bar_colors, color_scale_colors = _conditional_format_colors(worksheet)
        assert data_bar_colors == [expected_color]
        assert expected_color in color_scale_colors

    @pytest.mark.parametrize(
        ("writer_params", "expected_color"),
        [
            ({}, "00F76E6C"),
            ({"condition_color": "112233"}, "00112233"),
        ],
    )
    def test_dataframe2excel_file_path_uses_writer_condition_color(
        self,
        writer_params,
        expected_color,
    ):
        """文件路径模式应使用默认或 writer_params 指定的条件格式颜色。"""
        data = pd.DataFrame({"数值": [1, 2]})

        dataframe2excel(
            data,
            self.test_file,
            condition_cols=["数值"],
            writer_params=writer_params,
        )

        loaded_wb = load_workbook(self.test_file)
        data_bar_colors, _ = _conditional_format_colors(loaded_wb.active)
        loaded_wb.close()
        assert data_bar_colors == [expected_color]

    def test_dataframe_save_existing_writer_uses_writer_condition_color(self):
        """DataFrame.save 复用 Writer 时也应继承其条件格式颜色。"""
        writer = ExcelWriter(condition_color="112233")
        worksheet = writer.get_sheet_by_name("Condition")
        data = pd.DataFrame({"数值": [1, 2]})

        data.save(
            writer,
            worksheet=worksheet,
            condition_cols=["数值"],
        )

        data_bar_colors, _ = _conditional_format_colors(worksheet)
        assert data_bar_colors == ["00112233"]
    
    def test_write_with_custom_format(self):
        """测试自定义格式"""
        df = pd.DataFrame({
            'amount': [1000, 2000, 3000]
        })
        
        dataframe2excel(
            df, self.test_file,
            custom_cols=['amount'],
            custom_format='#,##0'
        )
        
        loaded_wb = load_workbook(self.test_file)
        ws = loaded_wb.active
        
        assert ws["B3"].number_format == "#,##0"

    def test_auto_filter_persisted_with_file_path(self):
        """回归：传入文件路径且 auto_filter=True 时筛选应写入文件（修复保存顺序bug）"""
        df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
        dataframe2excel(df, self.test_file, sheet_name='S', auto_filter=True)

        loaded_wb = load_workbook(self.test_file)
        ws = loaded_wb['S']
        assert ws.auto_filter.ref is not None

    def test_row_format_does_not_overflow_extra_column(self):
        """回归：按行设置格式不应越界到数据右侧的空白列（修复off-by-one）"""
        df = pd.DataFrame({'A': [0.1, 0.2], 'B': [0.3, 0.4]}, index=['r1', 'r2'])
        dataframe2excel(
            df, self.test_file, sheet_name='S',
            index=True, percent_rows=['r1'], start_col=2, start_row=2
        )
        loaded_wb = load_workbook(self.test_file)
        ws = loaded_wb['S']
        # 数据列为 C、D；E 列为空白，不应被设置为百分比格式
        assert ws['C3'].number_format == '0.00%'
        assert ws['D3'].number_format == '0.00%'
        assert ws['E3'].number_format != '0.00%'

    def test_write_with_figures_keeps_gap_before_header(self, monkeypatch):
        """测试插图后表头会自动下移，避免被图片覆盖"""

        def fake_insert_pic2sheet(self, worksheet, fig, insert_space, figsize=(600, 250)):
            if isinstance(insert_space, str):
                row = int(''.join(ch for ch in insert_space if ch.isdigit()))
                col = 2
            else:
                row, col = insert_space
            return row + 10, col + 8

        monkeypatch.setattr(ExcelWriter, "insert_pic2sheet", fake_insert_pic2sheet)

        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})

        dataframe2excel(df, writer, sheet_name=ws, figures=["dummy.png"], start_row=2)

        assert ws["B12"].value is None
        assert ws["B13"].value == "A"


class TestUtilityFunctions:
    """测试工具函数"""
    
    def test_check_contain_chinese(self):
        """测试中文检测"""
        result, eng_cnt, chi_cnt = ExcelWriter.check_contain_chinese("测试test")
        
        assert len(result) == 6
        assert eng_cnt == 4
        assert chi_cnt == 2
    
    def test_calc_continuous_cnt(self):
        """测试连续计数"""
        list_ = ['A', 'A', 'A', 'B', 'B', 'C']
        
        item, start, length = ExcelWriter.calc_continuous_cnt(list_, 0)
        assert item == 'A' and start == 0 and length == 3
        
        item, start, length = ExcelWriter.calc_continuous_cnt(list_, 3)
        assert item == 'B' and start == 3 and length == 2
        
        item, start, length = ExcelWriter.calc_continuous_cnt(list_, 5)
        assert item == 'C' and start == 5 and length == 1
    
    def test_get_cell_space(self):
        """测试位置格式转换"""
        # 字符串转元组
        result = ExcelWriter.get_cell_space("B3")
        assert result == (2, 3)
        
        # 元组转字符串
        result = ExcelWriter.get_cell_space((2, 2))
        assert result == "B2"
    
    def test_calculate_rgba_color(self):
        """测试颜色计算"""
        result = ExcelWriter.calculate_rgba_color("FFFFFF", 0.5)
        assert result.upper() == "#FFFFFF"

        result = ExcelWriter.calculate_rgba_color("000000", 0.5)
        # 颜色计算可能有精度差异，接受相近结果
        assert result.upper() in ["#808080", "#7F7F7F"]


class TestMultiLevelIndex:
    """测试多层索引"""
    
    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, "test.xlsx")
    
    def teardown_method(self):
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    def test_multi_level_columns(self):
        """测试多层列名"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        # 创建多层列名
        columns = pd.MultiIndex.from_product([['Group1', 'Group2'], ['A', 'B']])
        df = pd.DataFrame(
            np.random.rand(3, 4),
            columns=columns
        )
        
        writer.insert_df2sheet(ws, df, "B2")
        
        # 验证多层表头
        assert ws["B2"].value == 'Group1'
        assert ws["D2"].value == 'Group2'
        assert ws["B3"].value == 'A'
        assert ws["C3"].value == 'B'
    
    def test_multi_level_index(self):
        """测试多层索引"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("Test")
        
        # 创建多层索引
        index = pd.MultiIndex.from_product([['X', 'Y'], [1, 2]])
        df = pd.DataFrame(
            np.random.rand(4, 2),
            index=index,
            columns=['A', 'B']
        )
        
        writer.insert_df2sheet(ws, df, "B2", index=True, merge_index=True)

        # 验证多层索引 - 合并单元格后可能无法直接读取值
        # 检查起始位置的值即可
        assert ws["B3"].value == 'X' or str(ws["B3"]).startswith("<MergedCell")


class TestMoveSheet:
    """测试 move_sheet 绝对位置移动（修复前 index 偏移计算错误）"""

    def _make_writer(self):
        writer = ExcelWriter()
        for n in ["A", "B", "C", "D"]:
            writer.get_sheet_by_name(n)
        return writer

    def test_move_to_index_middle(self):
        writer = self._make_writer()
        # 模板 '初始化' 仍在首位：['初始化','A','B','C','D']
        writer.move_sheet("A", index=2)
        assert writer.workbook.sheetnames.index("A") == 2

    def test_move_to_first(self):
        writer = self._make_writer()
        writer.move_sheet("C", index=0)
        assert writer.workbook.sheetnames.index("C") == 0

    def test_move_to_last(self):
        writer = self._make_writer()
        total = len(writer.workbook.sheetnames)
        writer.move_sheet("A", index=total - 1)
        assert writer.workbook.sheetnames.index("A") == total - 1

    def test_move_negative_index(self):
        writer = self._make_writer()
        total = len(writer.workbook.sheetnames)
        writer.move_sheet("A", index=-1)
        assert writer.workbook.sheetnames.index("A") == total - 1


class TestPivotTable:
    """测试数据透视表 / 数据透视图功能"""

    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, "pivot.xlsx")
        self.df = pd.DataFrame({
            "商品类别": ["数码", "服饰", "数码", "服饰", "数码", "食品"],
            "区域": ["华东", "华东", "华南", "华南", "华北", "华东"],
            "放款金额": [100, 200, 300, 400, 500, 600],
            "笔数": [1, 2, 3, 4, 5, 6],
        })

    def teardown_method(self):
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def _pivot_parts(self, path):
        import zipfile
        with zipfile.ZipFile(path) as z:
            return z.namelist()

    def _read(self, path, part):
        import zipfile
        with zipfile.ZipFile(path) as z:
            return z.read(part).decode("utf-8")

    def _assert_all_xml_wellformed(self, path):
        import zipfile
        import xml.dom.minidom as minidom
        with zipfile.ZipFile(path) as z:
            for n in z.namelist():
                if n.endswith(".xml") or n.endswith(".rels"):
                    minidom.parseString(z.read(n))  # 抛异常即不合法

    def test_records_spec_only(self):
        """insert_pivot_table2sheet 仅记录配置，不立即落地"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视表")
        writer.insert_pivot_table2sheet(
            ws, self.df, "B2", rows="商品类别", columns="区域", values=[("放款金额", "sum")]
        )
        assert len(writer._pivot_specs) == 1
        assert writer._pivot_specs[0]["name"] == "数据透视表1"

    def test_requires_values(self):
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视表")
        with pytest.raises(ValueError):
            writer.insert_pivot_table2sheet(ws, self.df, "B2", rows="商品类别")

    def test_pivot_parts_injected(self):
        """保存后注入 pivotCache/pivotTable 部件，且文件可被 openpyxl 重新打开"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视表")
        writer.insert_pivot_table2sheet(
            ws, self.df, "B2",
            rows="商品类别", columns="区域", values=[("放款金额", "sum"), ("笔数", "sum")],
        )
        writer.save(self.test_file)

        parts = self._pivot_parts(self.test_file)
        assert "xl/pivotTables/pivotTable1.xml" in parts
        assert "xl/pivotCache/pivotCacheDefinition1.xml" in parts
        assert "xl/pivotCache/pivotCacheRecords1.xml" in parts

        # 所有 XML 合法
        self._assert_all_xml_wellformed(self.test_file)

        # 文件可被 openpyxl 重新打开（未损坏）
        load_workbook(self.test_file)

    def test_pivot_wiring(self):
        """工作表 -> pivotTable、workbook -> pivotCache 关系完整"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视表")
        writer.insert_pivot_table2sheet(
            ws, self.df, "B2", rows="商品类别", values=[("放款金额", "sum")]
        )
        writer.save(self.test_file)

        # workbook 声明 pivotCaches
        wb_xml = self._read(self.test_file, "xl/workbook.xml")
        assert "<pivotCaches>" in wb_xml
        # [Content_Types] 含三类 Override
        ct = self._read(self.test_file, "[Content_Types].xml")
        assert "pivotTable+xml" in ct
        assert "pivotCacheDefinition+xml" in ct
        assert "pivotCacheRecords+xml" in ct
        # 透视表所在 sheet 的 _rels 引用 pivotTable
        parts = self._pivot_parts(self.test_file)
        sheet_rels = [p for p in parts if p.startswith("xl/worksheets/_rels/")]
        assert sheet_rels
        joined = "".join(self._read(self.test_file, p) for p in sheet_rels)
        assert "pivotTable1.xml" in joined

    def test_records_count_matches(self):
        """缓存记录数与源数据行数一致"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视表")
        writer.insert_pivot_table2sheet(
            ws, self.df, "B2", rows="商品类别", values=[("放款金额", "sum")]
        )
        writer.save(self.test_file)
        rec = self._read(self.test_file, "xl/pivotCache/pivotCacheRecords1.xml")
        assert 'count="{}"'.format(len(self.df)) in rec

    def test_unsupported_agg(self):
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视表")
        with pytest.raises(ValueError):
            writer.insert_pivot_table2sheet(
                ws, self.df, "B2", rows="商品类别", values=[("放款金额", "median")]
            )

    def test_pivot_theme_style_and_number_formats_injected(self):
        """数据透视表主题样式应适配 writer.theme_color，自定义数字格式应写入 styles.xml。"""
        writer = ExcelWriter(theme_color="3F1DBA", opacity=0.80)
        ws = writer.get_sheet_by_name("透视表")
        writer.insert_pivot_table2sheet(
            ws,
            self.df,
            "B2",
            rows="商品类别",
            values=[
                {"field": "放款金额", "agg": "sum", "name": "放款金额"},
                {
                    "field": "放款金额",
                    "agg": "sum",
                    "show_as": "全局占比",
                    "name": "金额占比",
                    "number_format": "0.000%",
                },
            ],
        )
        writer.save(self.test_file)

        styles = self._read(self.test_file, "xl/styles.xml")
        pivot = self._read(self.test_file, "xl/pivotTables/pivotTable1.xml")

        assert "HSCreditPivotStyle" in styles
        assert 'rgb="FF3F1DBA"' in styles
        assert 'formatCode="0.000%"' in styles
        assert 'name="HSCreditPivotStyle"' in pivot
        assert 'showDataAs="percentOfTotal"' in pivot
        assert 'name="金额占比"' in pivot
        assert 'numFmtId="' in pivot

    def test_pivot_filters_groups_totals_and_percent_modes(self):
        """透视表支持筛选项、多个轴字段分组、行列汇总开关和占比显示。"""
        data = self.df.assign(
            年龄=[21, 28, 35, 42, 49, 56],
            收入=[3500, 5200, 6800, 8100, 9900, 12000],
        )
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视表")
        writer.insert_pivot_table2sheet(
            ws,
            data,
            "B2",
            rows=["商品类别", "年龄"],
            columns=["收入", "区域"],
            values=[
                {"field": "笔数", "agg": "sum", "show_as": "行占比", "name": "行占比"},
                {"field": "笔数", "agg": "sum", "show_as": "组合占比", "name": "组合占比"},
            ],
            filters={"区域": ["华东", "华南"]},
            filter_items={"商品类别": ["数码", "服饰"]},
            groups={"年龄": {"start": 20, "interval": 10}, "收入": (3000, 3000)},
            show_row_totals=False,
            show_col_totals=False,
            subtotals=True,
        )
        writer.save(self.test_file)

        cache = self._read(self.test_file, "xl/pivotCache/pivotCacheDefinition1.xml")
        pivot = self._read(self.test_file, "xl/pivotTables/pivotTable1.xml")

        assert 'startNum="20" endNum="60" groupInterval="10"' in cache
        assert 'startNum="3000" endNum="15000" groupInterval="3000"' in cache
        assert '<pageFields count="1">' in pivot
        assert 'multipleItemSelectionAllowed="1"' in pivot
        assert 'showDataAs="percentOfRow"' in pivot
        assert 'pivotShowAs="percentOfParentRow"' in pivot
        assert 'rowGrandTotals="0"' in pivot
        assert 'colGrandTotals="0"' in pivot
        assert ' h="1"' in pivot

    def test_register_pivot_aggregation_alias(self):
        """公开 API 可扩展聚合方式别名。"""
        register_pivot_aggregation("业务均值", "average", "业务均值项")

        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视表")
        writer.insert_pivot_table2sheet(
            ws,
            self.df,
            "B2",
            rows="商品类别",
            values=[("放款金额", "业务均值")],
        )
        writer.save(self.test_file)

        pivot = self._read(self.test_file, "xl/pivotTables/pivotTable1.xml")
        assert 'subtotal="average"' in pivot
        assert 'name="业务均值项:放款金额"' in pivot

    def test_pivot_chart_injects_pivotsource(self):
        """数据透视图在 chart XML 注入 pivotSource，且文件合法"""
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视分析")
        writer.insert_pivot_table2sheet(
            ws, self.df, "B2", rows="商品类别", values=[("放款金额", "sum")]
        )
        writer.insert_pivot_chart2sheet(ws, "H2", chart_type="bar", title="各类别放款金额")
        writer.save(self.test_file)

        self._assert_all_xml_wellformed(self.test_file)
        load_workbook(self.test_file)

        import zipfile
        with zipfile.ZipFile(self.test_file) as z:
            charts = [n for n in z.namelist() if n.startswith("xl/charts/chart") and n.endswith(".xml")]
            assert charts
            assert any("pivotSource" in z.read(c).decode("utf-8") for c in charts)

    def test_pivot_chart_requires_table(self):
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视分析")
        with pytest.raises(ValueError):
            writer.insert_pivot_chart2sheet(ws, "H2")

    def test_long_pivot_name_source_sheet_within_limit(self):
        """回归：超长透视表名自动创建的源数据表名应 ≤31字符、保留'源数据'标识且唯一，文件合法"""
        long_name = "超长透视表名称用于触发工作表名称超过三十一个字符的边界情况测试用例"
        writer = ExcelWriter()
        ws = writer.get_sheet_by_name("透视表")
        writer.insert_pivot_table2sheet(
            ws, self.df, "B2", rows="商品类别", values=[("放款金额", "sum")], name=long_name
        )
        writer.insert_pivot_table2sheet(
            ws, self.df, "B20", rows="商品类别", values=[("放款金额", "sum")], name=long_name + "X"
        )

        src_sheets = [s for s in writer.workbook.sheetnames if "源数据" in s]
        assert len(src_sheets) == 2
        assert all(len(s) <= 31 for s in src_sheets)
        assert len(set(src_sheets)) == 2  # 唯一不重名

        writer.save(self.test_file)
        self._assert_all_xml_wellformed(self.test_file)
        load_workbook(self.test_file)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
