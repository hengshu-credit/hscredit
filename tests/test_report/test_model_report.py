"""Tests for model_report module."""

import threading
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

import hscredit.report.model_report as model_report_module
from hscredit.report.model_report import ModelReport


def _merged_range_for_row(ws, row, start_col=2):
    for cell_range in ws.merged_cells.ranges:
        if cell_range.min_row == row and cell_range.max_row == row and cell_range.min_col == start_col:
            return cell_range
    return None


def _row_for_value(ws, value, col=2):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, col).value == value:
            return row
    raise AssertionError(f"未找到单元格值: {value}")


class MockModel:
    """Minimal mock model for testing ModelReport."""

    def __init__(self, feature_names=None):
        self._feature_names = feature_names or ["f0"]
        self._coef = np.array([0.5] * len(self._feature_names))
        self._intercept = np.array([-0.5])

    def predict_proba(self, X):
        arr = np.asarray(X)
        if arr.ndim == 1:
            arr = arr.reshape(-1, 1)
        n_feat = min(arr.shape[1], len(self._coef))
        scores = arr[:, :n_feat] @ self._coef[:n_feat] + self._intercept[0]
        prob = 1 / (1 + np.exp(-scores))
        return np.column_stack([1 - prob, prob])

    def get_feature_importances(self):
        return pd.Series(dict(zip(self._feature_names, [0.5] * len(self._feature_names))))

    @property
    def feature_importances_(self):
        return np.array([0.5] * len(self._feature_names))


class ReversedClassModel(MockModel):
    """Mock model whose probability columns use classes_=[1, 0]."""

    classes_ = np.array([1, 0])

    def predict_proba(self, X):
        proba = super().predict_proba(X)
        return proba[:, ::-1]


class RankedImportanceModel(MockModel):
    """按 3:1 返回可手算重要性的模型。"""

    def __init__(self):
        super().__init__(["f0", "f1"])

    def get_feature_importances(self):
        return pd.Series({"f0": 3.0, "f1": 1.0})


class ZeroImportanceModel(RankedImportanceModel):
    def get_feature_importances(self):
        return pd.Series({"f0": 0.0, "f1": 0.0})


class ScoreConversionModel(RankedImportanceModel):
    """带已拟合概率评分转换器的普通风险模型。"""

    def __init__(self):
        from hscredit.core.models import ScoreTransformer

        super().__init__()
        self.score_transformer_ = ScoreTransformer(
            method="standard",
            lower=300,
            upper=900,
            direction="descending",
            base_odds=0.05,
            base_score=600,
            pdo=20,
            rate=2,
        ).fit(np.array([0.05, 0.1, 0.2, 0.4]))

    def predict_score(self, X):
        probability = self.predict_proba(X)[:, 1]
        return self.score_transformer_.predict(probability)


class _ScorecardLogisticModel:
    feature_names_in_ = np.array(["f0", "f1"])
    coef_ = np.array([[0.8, -0.2]])
    intercept_ = np.array([-0.3])

    def summary(self):
        return pd.DataFrame(
            {
                "变量": ["截距", "f0", "f1"],
                "系数": [-0.3, 0.8, -0.2],
            }
        )


class _CompatibleScoreCardBase:
    feature_names_ = ["f0", "f1"]
    feature_names_in_ = np.array(["f0", "f1"])
    pdo = 60
    rate = 2
    base_odds = 35
    base_score = 750
    factor = 60 / np.log(2)
    offset = 750 - factor * np.log(35)
    rules = {
        "f0": {"低": 20.0, "高": -10.0},
        "f1": {"低": 8.0, "高": -4.0},
    }

    def get_feature_importances(self):
        return pd.Series({"f0": 0.8, "f1": 0.2})

    def predict(self, X):
        values = np.asarray(X[["f0", "f1"]], dtype=float)
        return 650.0 - values[:, 0] * 5.0 - values[:, 1] * 2.0

    def predict_proba(self, X):
        score = self.predict(X)
        probability = 1 / (np.exp((score - self.offset) / self.factor) + 1)
        return np.column_stack([1 - probability, probability])


class HsCreditScoreCard(_CompatibleScoreCardBase):
    lr_model = _ScorecardLogisticModel()
    rules_ = _CompatibleScoreCardBase.rules

    def scorecard_scale(self):
        return pd.DataFrame(
            {
                "刻度项": ["base_score", "pdo"],
                "刻度值": [750, 60],
                "备注": ["基础分", "分数翻倍点数"],
            }
        )

    def score_formula(self, decimal=4):
        return {"公式": "Score = 750 - 86.5617 × ln(odds)", "A": 750, "B": 86.5617}

    def scorecard_points(self, feature_map=None):
        return pd.DataFrame(
            {
                "变量名称": ["f0", "f1"],
                "变量含义": [(feature_map or {}).get("f0", ""), (feature_map or {}).get("f1", "")],
                "变量分箱": ["低", "高"],
                "对应分数": [20.0, -4.0],
            }
        )

    @property
    def score_odds_reference(self):
        return pd.DataFrame(
            {
                "评分": [750, 690],
                "理论Odds(坏好比)": [1 / 35, 2 / 35],
                "理论逾期率": [1 / 36, 2 / 37],
            }
        )


class ScorecardPipelineScoreCard(_CompatibleScoreCardBase):
    model = _ScorecardLogisticModel()

    def scorecard_scale(self):
        return pd.DataFrame(
            {
                "刻度项": ["base_odds", "base_score", "rate", "pdo"],
                "刻度值": [35, 750, 2, 60],
                "备注": ["好坏比", "基础分", "倍率", "PDO"],
            }
        )

    def scorecard_points(self, feature_map=None):
        return pd.DataFrame(
            {
                "变量名称": ["f0", "f1"],
                "变量含义": [(feature_map or {}).get("f0", ""), (feature_map or {}).get("f1", "")],
                "变量分箱": ["低", "高"],
                "对应分数": [20.0, -4.0],
            }
        )


class ToadScoreCard(_CompatibleScoreCardBase):
    model = _ScorecardLogisticModel()

    def export(self, to_frame=False, **kwargs):
        if not to_frame:
            return self.rules
        return pd.DataFrame(
            {
                "name": ["f0", "f0", "f1", "f1"],
                "value": ["低", "高", "低", "高"],
                "score": [20.0, -10.0, 8.0, -4.0],
            }
        )


def test_model_report_excel_tables_always_use_fast_writer(tmp_path, monkeypatch):
    """模型报告由大量小表组成，不能逐表退回普通逐单元格写入。"""
    from hscredit.excel import ExcelWriter

    observed_speeds = []
    original = ExcelWriter.insert_df2sheet

    def capture_speed(self, worksheet, data, insert_space, *args, **kwargs):
        observed_speeds.append(kwargs.get("speed", "auto"))
        return original(self, worksheet, data, insert_space, *args, **kwargs)

    monkeypatch.setattr(ExcelWriter, "insert_df2sheet", capture_speed)
    X = pd.DataFrame({"f0": np.linspace(0.0, 1.0, 20)})
    y = pd.Series([0, 1] * 10)
    report = ModelReport(MockModel(["f0"]), X_train=X, y_train=y, feature_names=["f0"], n_jobs=1)

    report.to_excel(str(tmp_path / "fast-report.xlsx"), with_plots=False)

    assert observed_speeds
    assert set(observed_speeds) == {"fast"}


def test_model_report_does_not_freeze_any_worksheet(tmp_path):
    """模型报告每个 Sheet 都应保持可自由滚动，不设置冻结窗口。"""
    X = pd.DataFrame({"f0": np.linspace(0.0, 1.0, 20)})
    y = pd.Series([0, 1] * 10)
    output = tmp_path / "no-freeze-report.xlsx"
    report = ModelReport(MockModel(["f0"]), X_train=X, y_train=y, feature_names=["f0"], n_jobs=1)

    report.to_excel(str(output), with_plots=False)

    workbook = load_workbook(output)
    assert {sheet.title: sheet.freeze_panes for sheet in workbook.worksheets} == {
        "目录": None,
        "1-基本信息": None,
        "2-模型性能": None,
        "3-入模变量分析": None,
        "4-稳定性分析": None,
        "5-模型参数": None,
        "6-模型部署需求": None,
    }


def test_export_plots_runs_all_plot_groups_concurrently_and_preserves_order(tmp_path, monkeypatch):
    """完整图表必须在线程中重叠执行，并按原报告顺序返回全部路径和 PSI 表。"""
    state = {"active": 0, "max_active": 0}
    lock = threading.Lock()
    main_thread_id = threading.get_ident()
    thread_ids = set()
    calls = []

    def record_plot(kind, *, psi=False):
        def plot(*args, save=None, result=False, **kwargs):
            assert "dpi" not in kwargs
            with lock:
                state["active"] += 1
                state["max_active"] = max(state["max_active"], state["active"])
                thread_ids.add(threading.get_ident())
                calls.append((kind, Path(save).name, threading.get_ident()))
            time.sleep(0.02)
            Path(save).touch()
            with lock:
                state["active"] -= 1
            if psi and result:
                return pd.DataFrame({"指标名称": [kwargs.get("desc", "")], "分箱": ["全部"]})
            return None

        return plot

    monkeypatch.setattr("hscredit.core.viz.bin_plot", record_plot("bin"))
    monkeypatch.setattr("hscredit.core.viz.ks_plot", record_plot("ks"))
    monkeypatch.setattr("hscredit.core.viz.lift_plot", record_plot("lift"))
    monkeypatch.setattr("hscredit.core.viz.hist_plot", record_plot("hist"))
    monkeypatch.setattr("hscredit.core.viz.corr_plot", record_plot("corr"))
    monkeypatch.setattr("hscredit.core.viz.psi_plot", record_plot("psi", psi=True))

    X = pd.DataFrame(
        {
            "f0": np.linspace(0.0, 1.0, 40),
            "f1": np.linspace(1.0, 3.0, 40) ** 2,
            "target": [0, 1] * 20,
        }
    )
    report = ModelReport(
        MockModel(["f0", "f1"]),
        datasets={"train": X, "test": X.copy()},
        target="target",
        feature_names=["f0", "f1"],
        n_jobs=4,
    )

    paths, tables = report._export_plots(tmp_path)

    assert state["max_active"] >= 2
    assert len(thread_ids) >= 2
    assert all(thread_id != main_thread_id for _, _, thread_id in calls)
    assert len(calls) == 19
    assert list(paths) == [
        "model_train",
        "model_test",
        "feature_contribution",
        "feature_corr",
        "feat_bin_f0",
        "feat_hist_f0",
        "feat_psi_f0",
        "feat_bin_f1",
        "feat_hist_f1",
        "feat_psi_f1",
    ]
    assert list(tables) == ["feat_psi_f0", "feat_psi_f1"]
    assert [Path(path).name for path in paths["model_train"]] == [
        "bin_train.png",
        "ks_train.png",
        "lift_train.png",
        "hist_train.png",
    ]
    assert [Path(path).name for path in paths["model_test"]] == [
        "bin_test.png",
        "ks_test.png",
        "lift_test.png",
        "hist_test.png",
    ]
    assert sum(len(figures) for figures in paths.values()) == 20


def test_threaded_plot_context_uses_agg_canvas_without_switching_user_backend():
    """报告线程必须使用独立 Agg 画布，且不能改变调用者的交互式 backend。"""
    import matplotlib
    import matplotlib.pyplot as plt
    from hscredit.core.viz import utils as viz_utils

    original_backend = matplotlib.get_backend()
    original_subplots = plt.subplots

    def build_figure(_):
        with model_report_module._threaded_agg_rendering():
            figure, _ = viz_utils._create_subplots(figsize=(2, 1))
            viz_utils._tight_layout(figure)
        return figure.canvas.__class__.__name__

    with ThreadPoolExecutor(max_workers=2) as executor:
        canvas_names = list(executor.map(build_figure, range(2)))

    assert canvas_names == ["FigureCanvasAgg", "FigureCanvasAgg"]
    assert matplotlib.get_backend() == original_backend
    assert plt.subplots is original_subplots


def test_threaded_plot_context_does_not_patch_unrelated_pyplot_calls(monkeypatch):
    """报告线程的 Agg 上下文不能改变其他调用线程看到的 pyplot 函数。"""
    import matplotlib.pyplot as plt

    sentinel = object()

    def caller_subplots(*args, **kwargs):
        return sentinel, None

    monkeypatch.setattr(plt, "subplots", caller_subplots)
    with model_report_module._threaded_agg_rendering():
        figure, _ = plt.subplots()

    assert figure is sentinel


def test_feature_summary_inherits_model_report_parallel_configuration(monkeypatch):
    """模型报告内部 DataFrame.summary 不得退回独立的全核默认。"""
    observed = {}

    def fake_summary(frame, **kwargs):
        observed.update(kwargs)
        return pd.DataFrame({"统计值": [1.0]}, index=["f0"])

    monkeypatch.setattr(pd.DataFrame, "summary", fake_summary, raising=False)
    X = pd.DataFrame({"f0": [0.0, 1.0, 2.0, 3.0]})
    y = pd.Series([0, 0, 1, 1])
    report = ModelReport(
        MockModel(["f0"]),
        X_train=X,
        y_train=y,
        feature_names=["f0"],
        n_jobs=3,
        parallel_backend="threading",
        parallel_config={"batch_size": 2},
    )

    report._get_features_summary()

    assert observed["n_jobs"] == 3
    assert observed["parallel_backend"] == "threading"
    assert observed["parallel_config"] == {"batch_size": 2}


class RealDataContractModel:
    """Deterministic non-constant model used by the real-data report contract."""

    classes_ = np.array([0, 1])

    def __init__(self, feature_names):
        self.feature_names_in_ = np.asarray(feature_names)

    def predict_proba(self, X):
        values = np.asarray(X, dtype=float)
        logits = (values[:, 0] - 0.09) * 5 + (values[:, 1] - 60) / 30 - (values[:, 2] - 600) / 200
        probabilities = 1 / (1 + np.exp(-logits))
        return np.column_stack([1 - probabilities, probabilities])

    def get_feature_importances(self):
        return pd.Series([0.5, 0.3, 0.2], index=self.feature_names_in_)

    def get_params(self):
        return {"name": "real-data-contract-model"}


class TestModelReportTarget:
    """Test target parameter handling."""

    def test_target_str(self):
        """target as string column name."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4],
                "label": [0, 0, 1, 1],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, target="label", feature_names=["f0"])
        assert report._datasets["训练集"].y.tolist() == [0, 0, 1, 1]

    def test_target_dict_overdue_dpds(self):
        """target as dict with overdue+dpds."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4, 5, 6],
                "overdue": [0, 0, 1, 1, 1, 1],
                "dpds": [0, 2, 3, 5, 6, 10],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, target={"overdue": "overdue", "dpds": "dpds", "threshold": 3}, feature_names=["f0"])
        assert report._datasets["训练集"].y.tolist() == [0, 0, 0, 1, 1, 1]

    def test_target_dict_overdue_only(self):
        """target as dict with overdue only (no dpds): overdue col > 0 → y=1."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4],
                "overdue": [0, 0, 1, 1],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, target={"overdue": "overdue"}, feature_names=["f0"])
        # overdue > 0 → [0, 0, 1, 1]
        assert report._datasets["训练集"].y.tolist() == [0, 0, 1, 1]

    def test_datasets_y_none(self):
        """datasets dict with y=None derives y from target config."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4],
                "target": [0, 0, 1, 1],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, datasets={"train": (X, None)}, target="target", feature_names=["f0"])
        assert report._datasets["train"].y.tolist() == [0, 0, 1, 1]

    def test_datasets_y_none_dict_target(self):
        """datasets dict with y=None and dict target."""
        X_train = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4],
                "overdue": [0, 1, 1, 1],
                "dpds": [0, 1, 5, 6],
            }
        )
        X_test = pd.DataFrame(
            {
                "f0": [5, 6],
                "overdue": [1, 1],
                "dpds": [4, 8],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, datasets={"train": (X_train, None), "test": (X_test, None)}, target={"overdue": "overdue", "dpds": "dpds", "threshold": 3}, feature_names=["f0"])
        assert report._datasets["train"].y.tolist() == [0, 0, 1, 1]
        assert report._datasets["test"].y.tolist() == [1, 1]

    def test_y_proba_produced(self):
        """Model produces y_proba after init."""
        X = pd.DataFrame({"f0": [1, 2, 3, 4], "label": [0, 0, 1, 1]})
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, target="label", feature_names=["f0"])
        proba = report._datasets["训练集"].y_proba
        assert proba is not None
        assert len(proba) == 4
        assert proba.min() >= 0 and proba.max() <= 1

    def test_get_metrics(self):
        """get_metrics returns DataFrame with expected columns."""
        X = pd.DataFrame({"f0": [1, 2, 3, 4], "label": [0, 0, 1, 1]})
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, target="label", feature_names=["f0"])
        metrics = report.get_metrics()
        assert "统计项" in metrics.columns
        assert "KS" in metrics["统计项"].values
        assert "AUC" in metrics["统计项"].values
        assert "样本数" in metrics["统计项"].values
        assert "坏样本率" in metrics["统计项"].values

    def test_target_default_column_fallback(self):
        """When target=None, searches for common column names."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4],
                "flag": [0, 0, 1, 1],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, target=None, feature_names=["f0"])
        # Should find 'flag' column as fallback
        assert report._datasets["训练集"].y.tolist() == [0, 0, 1, 1]


class TestModelReportOverdueDpdsSeparate:
    """Test overdue/dpds as separate __init__ parameters (not inside target dict)."""

    def test_overdue_dpds_single_col_single_threshold(self):
        """overdue as str + dpds as int is equivalent to target='col'."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4],
                "dpds": [0, 1, 5, 10],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, overdue="dpds", dpds=3, feature_names=["f0"])
        # dpds > 3 → [0, 0, 1, 1]
        assert report._datasets["训练集"].y.tolist() == [0, 0, 1, 1]

    def test_overdue_dpds_single_col_list_thresholds(self):
        """overdue as str + dpds as list thresholds."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4, 5, 6],
                "dpds": [0, 3, 7, 15, 20, 30],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, overdue="dpds", dpds=[15, 7, 0], feature_names=["f0"])
        # dpds > 15 or > 7 or > 0:
        #   0: 0>15? 0>7? 0>0? → false → 0
        #   3: 3>15? 3>7? 3>0? → false, false, true → 1 ← FAILS: test says [0,1,...]
        # Actually dpds > 0 for all values >= 1, so:
        #   [0, 1, 1, 1, 1, 1]  (only index 0 is false for >0)
        assert report._datasets["训练集"].y.tolist() == [0, 1, 1, 1, 1, 1]

    def test_overdue_dpds_multi_col(self):
        """overdue as list of str + dpds as list."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4, 5, 6],
                "dpds_m1": [0, 0, 0, 0, 0, 0],
                "dpds_m3": [0, 0, 0, 0, 1, 1],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, overdue=["dpds_m1", "dpds_m3"], dpds=[3, 0], feature_names=["f0"])
        # dpds_m1 > 3 or > 0 → [0, 0, 0, 0, 0, 0]
        # dpds_m3 > 3 or > 0 → [0, 0, 0, 0, 1, 1]
        # any true → y=1 → [0, 0, 0, 0, 1, 1]
        assert report._datasets["训练集"].y.tolist() == [0, 0, 0, 0, 1, 1]

    def test_overdue_dpds_override_target(self):
        """overdue/dpds takes priority over target when both provided."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4],
                "label": [1, 1, 1, 1],  # would give all 1s
                "dpds": [0, 0, 5, 10],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, X_train=X, y_train=None, target="label", overdue="dpds", dpds=3, feature_names=["f0"])  # ignored because overdue/dpds provided
        # dpds > 3 → [0, 0, 1, 1]
        assert report._datasets["训练集"].y.tolist() == [0, 0, 1, 1]

    def test_overdue_dpds_with_datasets_dict(self):
        """overdue/dpds works with datasets dict."""
        X_train = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4],
                "dpds": [0, 1, 5, 10],
            }
        )
        X_test = pd.DataFrame(
            {
                "f0": [5, 6],
                "dpds": [3, 15],
            }
        )
        model = MockModel(feature_names=["f0"])
        report = ModelReport(model=model, datasets={"train": (X_train, None), "test": (X_test, None)}, overdue="dpds", dpds=3, feature_names=["f0"])
        # train: dpds > 3 → [0, 0, 1, 1]
        # test:  dpds > 3 → [0, 1]
        assert report._datasets["train"].y.tolist() == [0, 0, 1, 1]
        assert report._datasets["test"].y.tolist() == [0, 1]

    def test_overdue_dpds_auto_model_report(self):
        """auto_model_report with overdue/dpds separate parameters."""
        from hscredit.report.model_report import auto_model_report

        X = pd.DataFrame(
            {
                "f0": list(range(100)),
                "dpds": list(range(100)),
            }
        )
        model = MockModel(feature_names=["f0"])
        report = auto_model_report(
            model=model,
            X_train=X,
            overdue="dpds",
            dpds=[30, 15, 7],
            feature_names=["f0"],
            verbose=False,
            with_plots=False,
        )
        # dpds > 30 or > 15 or > 7 → dpds > 7 → rows 8-99 → 92 out of 100
        assert report._datasets["训练集"].y.sum() == 92
        assert report._datasets["训练集"].y.mean() == 0.92

    def test_overdue_dpds_equivalent_to_dict_target(self):
        """overdue/dpds as separate params should produce same y as dict target."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4, 5, 6],
                "dpds": [0, 2, 5, 8, 12, 20],
            }
        )
        model = MockModel(feature_names=["f0"])

        # via dict target
        r1 = ModelReport(model=model, X_train=X, y_train=None, target={"overdue": "dpds", "dpds": [10, 5, 0]}, feature_names=["f0"])

        # via separate params
        r2 = ModelReport(model=model, X_train=X, y_train=None, overdue="dpds", dpds=[10, 5, 0], feature_names=["f0"])

        assert r1._datasets["训练集"].y.tolist() == r2._datasets["训练集"].y.tolist()

    def test_del_grey_uses_independent_sample_base_for_each_dpd(self):
        """删除灰样本时若模型摘要或评分分箱仍共用一个总样本数，本测试必须失败。"""
        data = pd.DataFrame(
            {
                "f0": np.arange(12, dtype=float),
                "MOB1": [0, 2, 0, 5, 0, 2, 5, 8, 0, 4, 2, 8],
            }
        )

        report = ModelReport(
            MockModel(["f0"]),
            datasets={"样本": data},
            overdue="MOB1",
            dpds=[1, 3],
            del_grey=True,
            feature_names=["f0"],
            n_jobs=1,
        )

        summary = report.summary()
        assert summary.loc["MOB1@1", ("样本数", "样本")] == 12
        assert summary.loc["MOB1@3", ("样本数", "样本")] == 9
        assert summary.loc["MOB1@1", ("坏样本率", "样本")] == pytest.approx(8 / 12)
        assert summary.loc["MOB1@3", ("坏样本率", "样本")] == pytest.approx(5 / 9)

        score_table = report.get_bin_table(
            "样本",
            max_n_bins=2,
            margins=True,
            labels=report._label_names,
        )
        total = score_table.loc[score_table[("分箱详情", "分箱标签")] == "合计"].iloc[0]
        assert total[("MOB1>1", "样本总数")] == 12
        assert total[("MOB1>3", "样本总数")] == 9

    def test_del_grey_filters_single_overdue_target_dataset(self):
        """单个 overdue/DPD 口径若只支持多目标删灰，本测试必须失败。"""
        data = pd.DataFrame(
            {
                "f0": [1.0, 2.0, 3.0, 4.0],
                "MOB1": [0.0, 0.5, 1.0, 0.5],
            }
        )

        report = ModelReport(
            MockModel(["f0"]),
            datasets={"样本": data},
            overdue="MOB1",
            dpds=[0.5],
            del_grey=True,
            feature_names=["f0"],
            n_jobs=1,
        )

        dataset = report._datasets["样本"]
        assert dataset.X.index.tolist() == [0, 2]
        assert dataset.y.tolist() == [0, 1]
        metrics = report.get_metrics().set_index("统计项")
        assert metrics.loc["样本数", "样本"] == 2
        assert metrics.loc["坏样本率", "样本"] == pytest.approx(0.5)

    def test_auto_model_report_removes_grey_from_overall_time_and_group_tables(self, tmp_path):
        """自动模型报告顶部任一分布表继续计入灰样本时，本测试必须失败。"""
        from hscredit.report.model_report import auto_model_report

        data = pd.DataFrame(
            {
                "f0": np.arange(12, dtype=float),
                "MOB1": [0, 2, 0, 5, 0, 2, 5, 8, 0, 4, 2, 8],
                "放款月份": pd.to_datetime(["2024-01-01"] * 6 + ["2024-02-01"] * 6),
                "客群": ["A"] * 6 + ["B"] * 6,
            }
        )
        output = tmp_path / "model-report-del-grey.xlsx"

        report = auto_model_report(
            MockModel(["f0"]),
            datasets={"样本": data},
            overdue="MOB1",
            dpds=[1, 3],
            del_grey=True,
            feature_names=["f0"],
            excel_path=str(output),
            verbose=False,
            with_plots=False,
            date_col="放款月份",
            date_freq="M",
            group_col="客群",
            n_bins=2,
            n_jobs=1,
        )

        assert report.del_grey is True
        worksheet = load_workbook(output)["1-基本信息"]
        header_rows = [
            row
            for row in range(1, worksheet.max_row)
            if worksheet.cell(row + 1, 2).value == "数据集"
        ]
        sample_header, time_header, group_header = header_rows[:3]

        def header_columns(group_row):
            columns = {}
            current_group = None
            for column in range(2, worksheet.max_column + 1):
                group = worksheet.cell(group_row, column).value
                if group is not None:
                    current_group = group
                label = worksheet.cell(group_row + 1, column).value
                if label is not None:
                    columns[(current_group, label)] = column
            return columns

        sample_columns = header_columns(sample_header)
        sample_row = sample_header + 2
        assert worksheet.cell(sample_row, sample_columns[("样本总数", "MOB1@1")]).value == 12
        assert worksheet.cell(sample_row, sample_columns[("样本总数", "MOB1@3")]).value == 9

        time_columns = header_columns(time_header)
        time_group_col = time_columns[("统计详情", "数据分组")]
        january_row = next(
            row
            for row in range(time_header + 2, worksheet.max_row + 1)
            if worksheet.cell(row, time_group_col).value == "2024-01"
        )
        february_row = next(
            row
            for row in range(time_header + 2, worksheet.max_row + 1)
            if worksheet.cell(row, time_group_col).value == "2024-02"
        )
        assert worksheet.cell(january_row, time_columns[("样本总数", "MOB1@1")]).value == 6
        assert worksheet.cell(january_row, time_columns[("样本总数", "MOB1@3")]).value == 4
        assert worksheet.cell(february_row, time_columns[("样本总数", "MOB1@1")]).value == 6
        assert worksheet.cell(february_row, time_columns[("样本总数", "MOB1@3")]).value == 5

        group_columns = header_columns(group_header)
        group_value_col = group_columns[("统计详情", "数据分组")]
        group_a_row = next(
            row
            for row in range(group_header + 2, worksheet.max_row + 1)
            if worksheet.cell(row, group_value_col).value == "A"
        )
        group_b_row = next(
            row
            for row in range(group_header + 2, worksheet.max_row + 1)
            if worksheet.cell(row, group_value_col).value == "B"
        )
        assert worksheet.cell(group_a_row, group_columns[("样本总数", "MOB1@1")]).value == 6
        assert worksheet.cell(group_a_row, group_columns[("样本总数", "MOB1@3")]).value == 4
        assert worksheet.cell(group_b_row, group_columns[("样本总数", "MOB1@1")]).value == 6
        assert worksheet.cell(group_b_row, group_columns[("样本总数", "MOB1@3")]).value == 5

    def test_del_grey_filters_labels_and_amounts_in_top_lift(self):
        """TOP LIFT 的标签、预测和金额未同步删灰时，本测试必须失败。"""
        data = pd.DataFrame(
            {
                "f0": np.arange(12, dtype=float),
                "MOB1": [0, 2, 0, 5, 0, 2, 5, 8, 0, 4, 2, 8],
                "金额": np.ones(12),
            }
        )
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"样本": data},
            overdue="MOB1",
            dpds=[1, 3],
            del_grey=True,
            feature_names=["f0"],
            n_jobs=1,
        )

        table = report._compute_top_n_lift_table(
            percentiles=(0.5,),
            amount_col="金额",
            label="MOB1>3",
        ).set_index("统计项")

        assert table.loc["坏样本率", "TOTAL"] == pytest.approx(5 / 9)
        assert table.loc["坏样本率", "TOP 50%"] == pytest.approx(3 / 4)
        assert table.loc["LIFT值", "TOP 50%"] == pytest.approx((3 / 4) / (5 / 9))


class TestModelReportRegression:
    """Regression tests for report sections that previously exported blank."""

    @staticmethod
    def _multi_label_data():
        return pd.DataFrame(
            {
                "f0": np.arange(20),
                "MOB1": [0, 1, 3, 7, 8] * 4,
                "放款金额": np.arange(100, 120),
                "放款时间": pd.date_range("2024-01-01", periods=20, freq="D"),
            }
        )

    def test_positive_probability_respects_model_classes(self):
        X = pd.DataFrame({"f0": [1, 2, 3, 4]})
        y = pd.Series([0, 0, 1, 1])
        model = ReversedClassModel(feature_names=["f0"])

        report = ModelReport(model, X_train=X, y_train=y, feature_names=["f0"])

        expected = model.predict_proba(X)[:, 0]
        np.testing.assert_allclose(report._datasets["训练集"].y_proba, expected)

    def test_summary_reuses_per_label_metric_cache(self, monkeypatch):
        """摘要不得在 get_metrics 已缓存后再次计算同一标签和数据集指标。"""
        X = self._multi_label_data()
        calls = []
        original = model_report_module._binary_metric_worker

        def recording_worker(task):
            calls.append(len(task[0]))
            return original(task)

        monkeypatch.setattr(model_report_module, "_binary_metric_worker", recording_worker)
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"训练集": X, "测试集": X.copy()},
            overdue=["MOB1"],
            dpds=[7, 3, 0],
            feature_names=["f0"],
            n_jobs=1,
        )
        for label in report._label_names:
            report.get_metrics(label)
        calls_before_summary = len(calls)

        summary = report.summary()

        assert summary.index.tolist() == ["MOB1@7", "MOB1@3", "MOB1@0"]
        assert len(calls) == calls_before_summary

    def test_multi_label_lift_contains_values_and_amount_metrics(self):
        X = self._multi_label_data()
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"训练集": X, "测试集": X.copy()},
            overdue=["MOB1"],
            dpds=[7, 3, 0],
            feature_names=["f0"],
        )

        order_table = report._get_top_n_lift_table(labels=report._label_names)
        amount_table = report._get_top_n_lift_table(
            labels=report._label_names,
            amount_col="放款金额",
        )

        assert report.feature_names == ["f0"]
        assert not order_table.isna().any().any()
        assert not amount_table.isna().any().any()
        assert not order_table.equals(amount_table)

    def test_model_input_table_contains_ranked_importance_contract(self):
        X = pd.DataFrame(
            {
                "f0": [0.0, 1.0, 2.0, 3.0],
                "f1": [3.0, 2.0, 1.0, 0.0],
            }
        )
        y = pd.Series([0, 0, 1, 1])
        report = ModelReport(
            RankedImportanceModel(),
            X_train=X,
            y_train=y,
            feature_names=["f0", "f1"],
            n_jobs=1,
        )

        table = report._get_model_input_table({"f0": "字段零", "f1": "字段一"})

        assert table.columns.tolist() == [
            "序号",
            "入参字段",
            "字段名称",
            "特征重要性",
            "特征重要性%",
            "累积特征重要性%",
        ]
        assert table["入参字段"].tolist() == ["f0", "f1"]
        assert table["字段名称"].tolist() == ["字段零", "字段一"]
        np.testing.assert_allclose(table["特征重要性"], [3.0, 1.0])
        np.testing.assert_allclose(table["特征重要性%"], [0.75, 0.25])
        np.testing.assert_allclose(table["累积特征重要性%"], [0.75, 1.0])

    @pytest.mark.parametrize(
        ("feature_map", "expected_headers", "expected_descriptions"),
        [
            (None, ["入参字段", "Coef."], None),
            (
                {"f0": "字段零", "f1": "字段一"},
                ["入参字段", "字段名称", "Coef."],
                {"const": None, "f0": "字段零", "f1": "字段一"},
            ),
        ],
    )
    def test_auto_model_report_writes_direct_lr_summary_after_feature_importance(
        self,
        tmp_path,
        feature_map,
        expected_headers,
        expected_descriptions,
    ):
        """直接传入 LR 时，统计摘要必须紧随入模特征列表并应用字段映射。"""
        from hscredit.core.models import LogisticRegression
        from hscredit.report.model_report import auto_model_report

        rng = np.random.default_rng(20260821)
        X = pd.DataFrame(rng.normal(size=(60, 2)), columns=["f0", "f1"])
        y = pd.Series(np.tile([0, 1], 30))
        model = LogisticRegression(max_iter=500).fit(X, y)
        output = tmp_path / f"direct_lr_summary_{feature_map is not None}.xlsx"

        auto_model_report(
            model,
            X_train=X,
            y_train=y,
            feature_names=["f0", "f1"],
            excel_path=str(output),
            feature_map=feature_map,
            with_plots=False,
            verbose=False,
            n_jobs=1,
        )

        worksheet = load_workbook(output)["5-模型参数"]
        feature_heading_row = _row_for_value(worksheet, "3、入模特征列表")
        summary_heading_row = _row_for_value(worksheet, "4、逻辑回归拟合结果")
        assert summary_heading_row > feature_heading_row

        summary_header_row = next(row for row in range(summary_heading_row + 1, worksheet.max_row + 1) if worksheet.cell(row, 2).value == "入参字段")
        headers = [worksheet.cell(summary_header_row, column).value for column in range(2, 2 + len(expected_headers))]
        assert headers == expected_headers

        summary_rows = {worksheet.cell(row, 2).value: row for row in range(summary_header_row + 1, summary_header_row + 4)}
        assert set(summary_rows) == {"const", "f0", "f1"}
        if expected_descriptions is not None:
            for feature, description in expected_descriptions.items():
                assert worksheet.cell(summary_rows[feature], 3).value == description

    def test_feature_contribution_figure_uses_rank_and_two_percent_axes(self):
        X = pd.DataFrame({"f0": [0.0, 1.0], "f1": [1.0, 0.0]})
        y = pd.Series([0, 1])
        report = ModelReport(
            RankedImportanceModel(),
            X_train=X,
            y_train=y,
            feature_names=["f0", "f1"],
            n_jobs=1,
        )
        table = pd.DataFrame(
            {
                "序号": [1, 2],
                "入参字段": ["f0", "f1"],
                "特征重要性": [3.0, 1.0],
                "特征重要性%": [0.75, 0.25],
                "累积特征重要性%": [0.75, 1.0],
            }
        )

        figure = report._create_feature_contribution_figure(table)
        try:
            assert len(figure.axes) == 2
            primary, secondary = figure.axes
            assert primary.get_ylabel() == "累积特征重要性%"
            assert secondary.get_ylabel() == "特征重要性%"
            np.testing.assert_allclose(primary.lines[0].get_ydata(), [0.75, 1.0])
            np.testing.assert_allclose([patch.get_height() for patch in secondary.patches], [0.75, 0.25])
            assert [tick.get_text() for tick in primary.get_xticklabels()] == ["1", "2"]
            assert primary.get_ylim() == pytest.approx((0.0, 1.0))
            assert secondary.get_ylim() == pytest.approx((0.0, 1.0))
            figure.canvas.draw()
            renderer = figure.canvas.get_renderer()
            title_box = figure._suptitle.get_window_extent(renderer)
            legend_box = figure.legends[0].get_window_extent(renderer)
            axes_box = primary.get_window_extent(renderer)
            assert not title_box.overlaps(legend_box)
            assert legend_box.y0 >= axes_box.y1 + 8
        finally:
            plt.close(figure)

    def test_zero_importance_keeps_fields_without_fake_contribution_percentages(self):
        X = pd.DataFrame({"f0": [0.0, 1.0], "f1": [1.0, 0.0]})
        y = pd.Series([0, 1])
        report = ModelReport(
            ZeroImportanceModel(),
            X_train=X,
            y_train=y,
            feature_names=["f0", "f1"],
            n_jobs=1,
        )

        table = report._get_model_input_table()

        assert table["入参字段"].tolist() == ["f0", "f1"]
        assert table["特征重要性%"].isna().all()
        assert table["累积特征重要性%"].isna().all()

    def test_feature_contribution_figure_uses_hscredit_theme(self):
        from matplotlib.colors import to_hex

        from hscredit.core.viz.utils import BAD_RATE_COLOR, DEFAULT_COLORS

        X = pd.DataFrame({"f0": [0.0, 1.0], "f1": [1.0, 0.0]})
        y = pd.Series([0, 1])
        report = ModelReport(
            RankedImportanceModel(),
            X_train=X,
            y_train=y,
            feature_names=["f0", "f1"],
            n_jobs=1,
        )
        table = pd.DataFrame(
            {
                "序号": [1, 2],
                "入参字段": ["f0", "f1"],
                "特征重要性": [3.0, 1.0],
                "特征重要性%": [0.75, 0.25],
                "累积特征重要性%": [0.75, 1.0],
            }
        )

        figure = report._create_feature_contribution_figure(table)
        try:
            primary, secondary = figure.axes
            line = primary.lines[0]
            bars = secondary.patches
            theme_color = DEFAULT_COLORS[0].lower()

            assert to_hex(line.get_color(), keep_alpha=False).lower() == BAD_RATE_COLOR.lower()
            assert line.is_dashed()
            assert line.get_clip_on() is False
            assert to_hex(bars[0].get_facecolor(), keep_alpha=False).lower() == theme_color
            assert all(bar.get_hatch() == "/" for bar in bars)
            assert to_hex(primary.spines["left"].get_edgecolor(), keep_alpha=False).lower() == theme_color
            assert to_hex(secondary.spines["right"].get_edgecolor(), keep_alpha=False).lower() == theme_color
            assert primary.spines["top"].get_visible() is False
            assert secondary.spines["top"].get_visible() is False
            assert primary.xaxis.label.get_color().lower() == theme_color
            assert primary.yaxis.label.get_color().lower() == theme_color
            assert secondary.yaxis.label.get_color().lower() == theme_color
            assert not any(gridline.get_visible() for gridline in primary.get_ygridlines())
            assert [text.get_text() for text in primary.texts] == ["75.00%", "100.00%"]
            assert [text.get_text() for text in secondary.texts] == ["25.00%"]
            assert [text.get_text() for text in figure.texts if text.get_text()] == ["入模特征贡献"]
        finally:
            plt.close(figure)

    @pytest.mark.parametrize(
        ("feature_count", "label_limit", "expected_line_labels", "expected_bar_labels"),
        [
            (10, 10, 10, 9),
            (11, 10, 0, 0),
            (11, None, 11, 10),
            (2, 0, 0, 0),
        ],
    )
    def test_feature_contribution_label_limit_is_configurable(
        self,
        feature_count,
        label_limit,
        expected_line_labels,
        expected_bar_labels,
    ):
        feature_names = [f"f{index}" for index in range(feature_count)]
        X = pd.DataFrame(
            np.arange(feature_count * 4, dtype=float).reshape(4, feature_count),
            columns=feature_names,
        )
        y = pd.Series([0, 1, 0, 1])
        report = ModelReport(
            MockModel(feature_names),
            X_train=X,
            y_train=y,
            feature_names=feature_names,
            n_jobs=1,
        )
        importance_ratio = np.repeat(1 / feature_count, feature_count)
        table = pd.DataFrame(
            {
                "序号": np.arange(1, feature_count + 1),
                "入参字段": feature_names,
                "特征重要性": np.ones(feature_count),
                "特征重要性%": importance_ratio,
                "累积特征重要性%": importance_ratio.cumsum(),
            }
        )

        figure = report._create_feature_contribution_figure(
            table,
            label_max_features=label_limit,
        )
        try:
            primary, secondary = figure.axes
            assert len(primary.texts) == expected_line_labels
            assert len(secondary.texts) == expected_bar_labels
        finally:
            plt.close(figure)

    def test_model_parameter_sheet_places_contribution_chart_after_one_blank_column(self, tmp_path, monkeypatch):
        X = pd.DataFrame(
            {
                "f0": [0.0, 1.0, 2.0, 3.0],
                "f1": [3.0, 2.0, 1.0, 0.0],
            }
        )
        y = pd.Series([0, 0, 1, 1])
        report = ModelReport(
            RankedImportanceModel(),
            X_train=X,
            y_train=y,
            feature_names=["f0", "f1"],
            n_jobs=1,
        )
        contribution_path = tmp_path / "feature_contribution.png"
        figure = plt.figure(figsize=(2, 1))
        figure.savefig(contribution_path)
        plt.close(figure)
        monkeypatch.setattr(
            report,
            "_export_plots",
            lambda *args, **kwargs: ({"feature_contribution": [str(contribution_path)]}, {}),
        )
        output = tmp_path / "model_parameter_contribution.xlsx"

        report.to_excel(
            str(output),
            with_plots=True,
            feature_map={"f0": "字段零", "f1": "字段一"},
        )

        worksheet = load_workbook(output)["5-模型参数"]
        input_header = next(cell for row in worksheet.iter_rows() for cell in row if cell.value == "入参字段")
        headers = [worksheet.cell(input_header.row, column).value for column in range(2, 8)]
        assert headers == ["序号", "入参字段", "字段名称", "特征重要性", "特征重要性%", "累积特征重要性%"]
        assert worksheet.cell(input_header.row + 1, 6).number_format == "0.00%"
        assert worksheet.cell(input_header.row + 1, 7).number_format == "0.00%"
        assert len(worksheet._images) == 1
        assert worksheet._images[0].anchor._from.col == 8

    def test_predict_score_report_outputs_fitted_transformer_sections(self, tmp_path):
        X = pd.DataFrame(
            {
                "f0": [0.0, 1.0, 2.0, 3.0],
                "f1": [3.0, 2.0, 1.0, 0.0],
            }
        )
        y = pd.Series([0, 0, 1, 1])
        report = ModelReport(
            ScoreConversionModel(),
            X_train=X,
            y_train=y,
            feature_names=["f0", "f1"],
            method="predict_score",
            n_jobs=1,
        )
        output = tmp_path / "score_conversion_report.xlsx"

        report.to_excel(str(output), with_plots=False)

        worksheet = load_workbook(output)["5-模型参数"]
        values = [cell.value for row in worksheet.iter_rows() for cell in row]
        assert "4、评分转换器选型" in values
        assert "5、评分转换基础参数配置" in values
        assert "6、概率转评分公式" in values
        assert "standard" in values
        assert "StandardScoreTransformer" in values
        assert "base_score" in values
        assert 600 in values
        assert any(isinstance(value, str) and value.startswith("Score =") for value in values)

    @pytest.mark.parametrize(
        "scorecard_type",
        [HsCreditScoreCard, ScorecardPipelineScoreCard, ToadScoreCard],
    )
    def test_scorecard_report_normalizes_compatible_model_sections(self, tmp_path, scorecard_type):
        train_X = pd.DataFrame(
            {
                "f0": [0.0, 1.0, 2.0, 3.0, 4.0, 5.0],
                "f1": [5.0, 4.0, 3.0, 2.0, 1.0, 0.0],
            }
        )
        test_X = train_X + pd.DataFrame({"f0": [0.5] * 6, "f1": [0.25] * 6})
        train_y = pd.Series([0, 0, 0, 1, 1, 1])
        test_y = pd.Series([0, 0, 1, 0, 1, 1])
        report = ModelReport(
            scorecard_type(),
            X_train=train_X,
            y_train=train_y,
            X_test=test_X,
            y_test=test_y,
            feature_names=["f0", "f1"],
            method="predict",
            n_jobs=1,
        )
        output = tmp_path / f"{scorecard_type.__name__}.xlsx"

        report.to_excel(
            str(output),
            with_plots=False,
            feature_map={"f0": "字段零", "f1": "字段一"},
        )

        worksheet = load_workbook(output)["5-模型参数"]
        values = [cell.value for row in worksheet.iter_rows() for cell in row]
        assert "4、逻辑回归拟合结果" in values
        assert "5、评分卡基础参数配置" in values
        assert "6、评分卡转换公式" in values
        assert "7、评分卡分值表" in values
        assert "8、评分、ODDS与逾期率参考表" in values
        assert "9、评分稳定性分析" in values
        assert "变量名称" in values
        assert "变量分箱" in values
        assert "对应分数" in values
        assert "f0" in values
        assert "理论逾期率" in values
        assert any(isinstance(value, str) and value.startswith("Score =") for value in values)

    def test_excel_contains_all_sections_and_multi_label_description(self, tmp_path):
        X = self._multi_label_data()
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"训练集": X, "测试集": X.copy()},
            overdue=["MOB1"],
            dpds=[7, 3, 0],
            feature_names=["f0"],
        )
        output = tmp_path / "model_report.xlsx"

        report.to_excel(
            str(output),
            with_plots=False,
            amount_col="放款金额",
            project_desc="测试项目描述",
            data_source="测试数据源",
        )

        workbook = load_workbook(output)
        assert workbook.sheetnames == [
            "目录",
            "1-基本信息",
            "2-模型性能",
            "3-入模变量分析",
            "4-稳定性分析",
            "5-模型参数",
            "6-模型部署需求",
        ]
        contents = [cell.value for row in workbook["目录"].iter_rows() for cell in row]
        basic_info = [cell.value for row in workbook["1-基本信息"].iter_rows() for cell in row]
        performance = [cell.value for row in workbook["2-模型性能"].iter_rows() for cell in row]
        feature_sheet = workbook["3-入模变量分析"]
        summary_feature_cell = next(cell for row in feature_sheet.iter_rows() for cell in row if cell.value == "f0" and cell.hyperlink is not None)
        feature_title_cell = next(cell for row in feature_sheet.iter_rows() for cell in row if cell.value == "3.1、f0 有效性分析")

        assert "5-模型参数" in contents
        assert "6-模型部署需求" in contents
        assert "测试项目描述" in basic_info
        assert "测试数据源" in basic_info
        assert any(isinstance(value, str) and "MOB1@7:" in value for value in basic_info)
        assert "各数据集标签坏样本率" not in basic_info
        assert any(isinstance(value, float) and not np.isnan(value) for value in performance)

        basic_info_sheet = workbook["1-基本信息"]
        sample_total_header = next(cell for row in basic_info_sheet.iter_rows() for cell in row if cell.value == "样本总数")
        stats_group_row = sample_total_header.row - 1
        assert basic_info_sheet.cell(stats_group_row, sample_total_header.column - 1).value == "统计详情"
        assert any(
            cell_range.min_row == stats_group_row
            and cell_range.max_row == stats_group_row
            and cell_range.min_col == sample_total_header.column - 1
            and cell_range.max_col == sample_total_header.column
            for cell_range in basic_info_sheet.merged_cells.ranges
        )

        feature_values = [cell.value for row in feature_sheet.iter_rows() for cell in row]
        assert feature_values.count("训练集 订单口径") == 1
        assert feature_values.count("测试集 订单口径") == 1
        assert summary_feature_cell.hyperlink.location == f"#'3-入模变量分析'!{feature_title_cell.coordinate}"
        assert feature_title_cell.hyperlink.location == f"#'3-入模变量分析'!{summary_feature_cell.coordinate}"

    def test_excel_contents_sheet_adjusts_column_width(self, tmp_path):
        X = pd.DataFrame(
            {
                "f0": np.arange(20),
                "target": [0, 1] * 10,
            }
        )
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"train": X},
            target="target",
            feature_names=["f0"],
        )
        output = tmp_path / "model_report_contents_width.xlsx"

        report.to_excel(str(output), with_plots=False, model_name="VeryLongModelNameForColumnWidthCheck")

        contents_sheet = load_workbook(output)["目录"]
        assert contents_sheet.column_dimensions["B"].width > 8
        assert contents_sheet.column_dimensions["C"].width > 20
        assert contents_sheet.column_dimensions["D"].width > 30

    def test_report_sheets_apply_requested_auto_width_ranges(self, tmp_path):
        overdue_col = "一个特别长的逾期指标字段名称"
        feature_col = "一个特别长的入模特征字段名称"
        X = pd.DataFrame(
            {
                feature_col: np.arange(20),
                overdue_col: [0, 1, 3, 7, 8] * 4,
            }
        )
        report = ModelReport(
            MockModel([feature_col]),
            datasets={"数据集名称非常非常长用于验证自动列宽": X, "测试集": X.copy()},
            overdue=overdue_col,
            dpds=[7, 3],
            feature_names=[feature_col],
        )
        output = tmp_path / "model_report_auto_width.xlsx"

        report.to_excel(str(output), with_plots=False)

        workbook = load_workbook(output)
        basic = workbook["1-基本信息"]
        performance = workbook["2-模型性能"]
        feature_sheet = workbook["3-入模变量分析"]
        assert basic.column_dimensions["B"].width == pytest.approx(14.1640625)
        assert basic.column_dimensions["C"].width == pytest.approx(10.83203125)
        assert basic.column_dimensions["D"].width > 20
        assert performance.column_dimensions["C"].width > 20
        assert feature_sheet.column_dimensions["B"].width > 20

    def test_excel_title_merges_follow_actual_content_width(self, tmp_path):
        X = self._multi_label_data()
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"训练集": X, "测试集": X.copy()},
            overdue=["MOB1"],
            dpds=[7, 3, 0],
            feature_names=["f0"],
        )
        output = tmp_path / "model_report_dynamic_title_merges.xlsx"

        report.to_excel(
            str(output),
            with_plots=False,
            amount_col="放款金额",
            date_col="放款时间",
        )

        workbook = load_workbook(output)
        contents = workbook["目录"]
        basic = workbook["1-基本信息"]
        performance = workbook["2-模型性能"]
        feature_sheet = workbook["3-入模变量分析"]

        assert _merged_range_for_row(contents, 2).max_col == 4
        assert _merged_range_for_row(basic, 2).max_col == basic.max_column
        assert _merged_range_for_row(performance, 2).max_col == performance.max_column
        assert _merged_range_for_row(feature_sheet, 2).max_col == feature_sheet.max_column
        assert _merged_range_for_row(basic, 2).max_col != 35

        desc_row = _row_for_value(basic, "2、数据样本描述")
        assert _merged_range_for_row(basic, desc_row).max_col == 3

        parent_row = _row_for_value(feature_sheet, "3、入模变量有效性分析")
        child_row = _row_for_value(feature_sheet, "3.1、f0 有效性分析")
        assert _merged_range_for_row(feature_sheet, parent_row).max_col == feature_sheet.max_column
        assert _merged_range_for_row(feature_sheet, child_row).max_col == feature_sheet.max_column

    def test_excel_skips_hyperlink_when_feature_missing_from_summary(self, tmp_path, monkeypatch):
        """特征不在重要性汇总表中时（summary_row为None），应跳过超链接而不是抛异常."""
        X = self._multi_label_data()
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"训练集": X, "测试集": X.copy()},
            overdue=["MOB1"],
            dpds=[7, 3, 0],
            feature_names=["f0"],
        )

        original_summary = report._get_features_summary
        monkeypatch.setattr(
            report,
            "_get_features_summary",
            lambda: original_summary().iloc[0:0],
        )

        output = tmp_path / "model_report_missing_feature.xlsx"
        report.to_excel(str(output), with_plots=False, amount_col="放款金额")

        feature_sheet = load_workbook(output)["3-入模变量分析"]
        feature_title_cell = next(cell for row in feature_sheet.iter_rows() for cell in row if cell.value == "3.1、f0 有效性分析")
        assert feature_title_cell.hyperlink is None

    def test_time_distribution_header_groups_detail_columns(self, tmp_path):
        X = self._multi_label_data()
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"训练集": X, "测试集": X.copy()},
            overdue=["MOB1"],
            dpds=[7, 3, 0],
            feature_names=["f0"],
        )

        output = tmp_path / "model_report_time_distribution.xlsx"
        report.to_excel(str(output), with_plots=False, date_col="放款时间")

        basic = load_workbook(output)["1-基本信息"]
        data_group_header = next(cell for row in basic.iter_rows() for cell in row if cell.value == "数据分组")
        header_row = data_group_header.row
        data_set_header = next(basic.cell(header_row, col) for col in range(1, basic.max_column + 1) if basic.cell(header_row, col).value == "数据集")
        sample_total_header = next(basic.cell(header_row, col) for col in range(1, basic.max_column + 1) if basic.cell(header_row, col).value == "样本总数")

        assert basic.cell(header_row - 1, data_set_header.column).value == "统计详情"
        assert any(cell_range.min_row == header_row - 1 and cell_range.min_col == data_set_header.column and cell_range.max_row == header_row - 1 and cell_range.max_col == sample_total_header.column for cell_range in basic.merged_cells.ranges)

    def test_single_label_performance_metric_formats(self, tmp_path):
        X = pd.DataFrame({"f0": [1, 2, 3, 4, 5, 6], "target": [0, 0, 0, 1, 1, 1]})
        report = ModelReport(
            MockModel(["f0"]),
            X_train=X,
            y_train=None,
            target="target",
            feature_names=["f0"],
        )

        output = tmp_path / "single_label_report.xlsx"
        report.to_excel(str(output), with_plots=False)

        performance = load_workbook(output)["2-模型性能"]
        ks_cell = next(cell for row in performance.iter_rows() for cell in row if cell.value == "KS")
        auc_cell = next(cell for row in performance.iter_rows() for cell in row if cell.value == "AUC")
        sample_cell = next(cell for row in performance.iter_rows() for cell in row if cell.value == "样本总数")
        bad_rate_cell = next(cell for row in performance.iter_rows() for cell in row if cell.value == "坏样本率")

        assert performance.cell(ks_cell.row, ks_cell.column + 1).number_format == "0.00%"
        assert performance.cell(auc_cell.row, auc_cell.column + 1).number_format == "0.00%"
        assert performance.cell(bad_rate_cell.row, bad_rate_cell.column + 1).number_format == "0.00%"
        assert performance.cell(sample_cell.row, sample_cell.column + 1).number_format == "#,##0"
        assert isinstance(performance.cell(sample_cell.row, sample_cell.column + 1).value, int)

    def test_compare_models_summary_excel_keeps_ratio_values(self, tmp_path):
        from hscredit.report.model_report import compare_models

        X = pd.DataFrame({"f0": [1, 2, 3, 4, 5, 6]})
        y = pd.Series([0, 0, 0, 1, 1, 1])
        output = tmp_path / "compare_models.xlsx"

        result = compare_models(
            {"LR": MockModel(["f0"])},
            X,
            y,
            X_test=X,
            y_test=y,
            excel_path=str(output),
        )

        assert result.loc[("LR", "target"), ("坏样本率", "训练集")] == 0.5

        sheet = load_workbook(output)["Sheet1"]
        bad_rate_header = next(cell for row in sheet.iter_rows() for cell in row if cell.value == "坏样本率")
        train_header = next(sheet.cell(bad_rate_header.row + 1, col) for col in range(bad_rate_header.column, sheet.max_column + 1) if sheet.cell(bad_rate_header.row + 1, col).value == "训练集")
        data_cell = sheet.cell(train_header.row + 1, train_header.column)
        assert data_cell.value == 0.5
        assert data_cell.number_format == "0.00%"

    def test_export_plots_contains_feature_psi(self, tmp_path):
        X = self._multi_label_data()
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"train": X, "test": X.copy()},
            overdue="MOB1",
            dpds=3,
            feature_names=["f0"],
        )

        paths, tables = report._export_plots(tmp_path)

        assert "feat_psi_f0" in paths
        assert Path(paths["feat_psi_f0"][0]).exists()
        assert "feat_psi_f0" in tables
        assert not tables["feat_psi_f0"].empty
        assert tables["feat_psi_f0"]["预期坏样本率"].gt(0).any()
        assert tables["feat_psi_f0"]["实际坏样本率"].gt(0).any()

    def test_multi_label_tables_use_expected_layout(self, tmp_path):
        X = self._multi_label_data()
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"训练集": X, "测试集": X.copy()},
            overdue="MOB1",
            dpds=[7, 3, 0],
            feature_names=["f0"],
        )

        bin_table = report.get_bin_table(labels=report._label_names)
        feature_table = report.get_feature_bin_table("f0", labels=report._label_names)
        for table in (bin_table, feature_table):
            assert isinstance(table.columns, pd.MultiIndex)
            assert {"MOB1>7", "MOB1>3", "MOB1>0"} <= set(table.columns.get_level_values(0))
            assert "指标名称" not in table.columns.get_level_values(-1)
            assert "指标含义" not in table.columns.get_level_values(-1)

        output = tmp_path / "multi_layout.xlsx"
        report.to_excel(str(output), with_plots=False, amount_col="放款金额")
        workbook = load_workbook(output)
        performance = workbook["2-模型性能"]
        basic = workbook["1-基本信息"]

        assert performance["B7"].value == "统计项"
        assert performance["B7"].style == "header_left"
        assert [performance.cell(7, col).value for col in (3, 5, 7)] == ["MOB1>7", "MOB1>3", "MOB1>0"]
        assert [performance.cell(8, col).value for col in range(2, 8)] == ["统计指标", "训练集", "测试集", "训练集", "测试集", "训练集"]
        assert performance["B11"].value == "样本总数"
        assert performance["C9"].number_format == "0.00%"
        assert performance["C10"].number_format == "0.00%"
        assert performance["C11"].number_format == "#,##0"
        assert isinstance(performance["C11"].value, int)
        assert performance["B19"].value == "统计指标"
        assert performance.auto_filter.ref == "B20:AJ26"
        sample_total_header = next(cell for row in basic.iter_rows() for cell in row if cell.value == "样本总数")
        stats_group_row = sample_total_header.row - 1
        assert basic.cell(stats_group_row, sample_total_header.column - 1).value == "统计详情"
        assert any(
            cell_range.min_row == stats_group_row
            and cell_range.max_row == stats_group_row
            and cell_range.min_col == sample_total_header.column - 1
            and cell_range.max_col == sample_total_header.column
            for cell_range in basic.merged_cells.ranges
        )
        assert basic.cell(sample_total_header.row, sample_total_header.column - 1).value == "数据集"

    def test_feature_psi_table_uses_percent_and_condition_formats(self, tmp_path, monkeypatch):
        X = pd.DataFrame({"f0": np.arange(20), "target": [0, 1] * 10})
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"train": X, "test": X.copy()},
            target="target",
            feature_names=["f0"],
        )
        psi_table = pd.DataFrame(
            {
                "指标名称": ["f0", "f0"],
                "分箱": ["(-inf, 0]", "(0, inf)"],
                "预期样本数": [10, 10],
                "预期样本占比": [0.5, 0.5],
                "预期坏样本率": [0.2, 0.8],
                "实际样本数": [8, 12],
                "实际样本占比": [0.4, 0.6],
                "实际坏样本率": [0.25, 0.75],
                "实际% - 预期%": [-0.1, 0.1],
                "ln(实际% / 预期%)": [-0.2231, 0.1823],
                "分档PSI值": [0.02231, 0.01823],
                "总体PSI值": [0.04054, 0.04054],
            }
        )
        monkeypatch.setattr(
            report,
            "_export_plots",
            lambda *args, **kwargs: ({}, {"feat_psi_f0": psi_table}),
        )
        output = tmp_path / "feature_psi_formats.xlsx"

        report.to_excel(str(output), with_plots=True)

        worksheet = load_workbook(output)["3-入模变量分析"]
        percent_headers = {
            "预期样本占比",
            "预期坏样本率",
            "实际样本占比",
            "实际坏样本率",
            "实际% - 预期%",
            "分档PSI值",
            "总体PSI值",
        }
        header_cells = {
            cell.value: cell
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value in percent_headers
        }
        assert set(header_cells) == percent_headers
        assert all(
            worksheet.cell(cell.row + 1, cell.column).number_format == "0.00%"
            for cell in header_cells.values()
        )
        data_bar_columns = {
            cell_range.min_col
            for key, rules in worksheet.conditional_formatting._cf_rules.items()
            if any(rule.type == "dataBar" for rule in rules)
            for cell_range in key.sqref.ranges
        }
        assert header_cells["实际% - 预期%"].column in data_bar_columns
        assert header_cells["分档PSI值"].column in data_bar_columns

    def test_score_psi_reference_matches_matrix_width_and_is_left_aligned(self, tmp_path):
        X = pd.DataFrame({"f0": np.arange(20), "target": [0, 1] * 10})
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"train": X, "test": X.copy()},
            target="target",
            feature_names=["f0"],
        )
        output = tmp_path / "score_psi_reference.xlsx"

        report.to_excel(str(output), with_plots=False)

        worksheet = load_workbook(output)["4-稳定性分析"]
        reference = next(
            cell
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("PSI参考标准：")
        )
        merged_range = next(
            cell_range
            for cell_range in worksheet.merged_cells.ranges
            if cell_range.min_row == reference.row
            and cell_range.max_row == reference.row
            and cell_range.min_col == reference.column
        )
        assert merged_range.max_col == reference.column + 2
        assert reference.alignment.horizontal == "left"


@pytest.mark.skipif(
    not (Path(__file__).parents[2] / "examples" / "hscredit_yyp.xlsx").exists(),
    reason="缺少 examples/hscredit_yyp.xlsx",
)
class TestModelReportRealDataContract:
    """真实放款数据的完整 Excel 报告契约。"""

    FEATURE_NAMES = ["衡枢鉴真分老客版", "近六个月非银多头机构数", "青云24"]
    EXPECTED_SHEETS = [
        "目录",
        "1-基本信息",
        "2-模型性能",
        "3-入模变量分析",
        "4-稳定性分析",
        "5-模型参数",
        "6-模型部署需求",
    ]

    @staticmethod
    def _sheet_values(ws):
        return [cell.value for row in ws.iter_rows() for cell in row]

    @staticmethod
    def _numeric_values_after(ws, label):
        cell = next(cell for row in ws.iter_rows() for cell in row if cell.value == label)
        return [ws.cell(cell.row, col).value for col in range(cell.column + 1, ws.max_column + 1) if isinstance(ws.cell(cell.row, col).value, (int, float))]

    def test_auto_model_report_real_data_excel_contract(self, tmp_path):
        from hscredit.report.model_report import auto_model_report

        source = pd.read_excel(Path(__file__).parents[2] / "examples" / "hscredit_yyp.xlsx")
        frames = {
            "训练集": source.iloc[:500].copy(),
            "测试集": source.iloc[500:750].copy(),
            "OOT集": source.iloc[750:].copy(),
        }
        expected_labels = ["MOB1>7", "MOB1>3", "MOB1>0"]
        expected_display_labels = ["MOB1@7", "MOB1@3", "MOB1@0"]
        expected_counts = [len(frame) for frame in frames.values()]
        expected_bad_rates = {label: [float((frame["MOB1"] > threshold).mean()) for frame in frames.values()] for label, threshold in zip(expected_labels, [7, 3, 0])}
        output = tmp_path / "真实数据模型评估报告.xlsx"

        report = auto_model_report(
            model=RealDataContractModel(self.FEATURE_NAMES),
            datasets=frames,
            feature_names=self.FEATURE_NAMES,
            overdue=["MOB1"],
            dpds=[7, 3, 0],
            excel_path=str(output),
            amount_col="放款金额",
            date_col="放款时间",
            group_col="商品类别",
            with_plots=True,
            verbose=False,
            model_name="真实数据契约模型",
            project_desc="真实放款数据多标签模型验证",
            data_source="examples/hscredit_yyp.xlsx",
            loc_cols="客户编号",
        )

        assert output.exists() and output.stat().st_size > 0
        assert report._label_names == expected_labels
        for dataset_key, frame in frames.items():
            dataset = report._datasets[dataset_key]
            assert len(dataset.y) == len(frame)
            assert np.ptp(dataset.y_proba) > 0
            for label, threshold in zip(expected_labels, [7, 3, 0]):
                expected_y = (frame["MOB1"] > threshold).astype(int).to_numpy()
                np.testing.assert_array_equal(dataset.y_dict[label], expected_y)

        summary = report.summary()
        assert summary.index.tolist() == expected_display_labels
        for label, display_label in zip(expected_labels, expected_display_labels):
            for dataset_label, expected_rate in zip(["训练集", "测试集", "OOT集"], expected_bad_rates[label]):
                assert summary.loc[display_label, ("样本数", dataset_label)] == len(frames[dataset_label])
                assert summary.loc[display_label, ("坏样本率", dataset_label)] == expected_rate

        workbook = load_workbook(output, data_only=False)
        assert workbook.sheetnames == self.EXPECTED_SHEETS

        critical_headings = {
            "目录": ["模型评估报告"],
            "1-基本信息": ["一、基本信息", "1、项目目标", "2、数据样本描述", "3、数据样本统计", "4、样本分布情况"],
            "2-模型性能": ["二、模型性能评估", "1、模型性能验证指标"],
            "3-入模变量分析": ["三、入模变量分析", "1、入模变量重要性及分布情况", "2、入模变量相关性", "3、入模变量有效性分析"],
            "4-稳定性分析": ["四、模型稳定性分析", "1、评分分布统计"],
            "5-模型参数": ["五、模型选型及参数", "1、模型选型", "2、模型参数", "3、入模特征列表"],
            "6-模型部署需求": ["六、模型部署需求", "1、入模变量信息", "2、生产订单测试用例"],
        }
        for sheet_name, headings in critical_headings.items():
            values = self._sheet_values(workbook[sheet_name])
            for heading in headings:
                assert heading in values, f"{sheet_name} 缺少章节 {heading}"

        basic = workbook["1-基本信息"]
        basic_values = self._sheet_values(basic)
        assert "真实放款数据多标签模型验证" in basic_values
        assert "examples/hscredit_yyp.xlsx" in basic_values
        assert "月度分布" in basic_values
        assert "商品类别分布" in basic_values
        assert set(source["商品类别"].unique()) <= set(basic_values)
        assert all(any(isinstance(value, str) and f"{label}:" in value for value in basic_values) for label in expected_display_labels)

        performance = workbook["2-模型性能"]
        assert self._numeric_values_after(performance, "样本总数")[:9] == expected_counts * 3
        workbook_bad_rates = self._numeric_values_after(performance, "坏样本率")[:9]
        np.testing.assert_allclose(
            workbook_bad_rates,
            [rate for label in expected_labels for rate in expected_bad_rates[label]],
            atol=5e-5,
        )
        bad_rate_cell = next(cell for row in performance.iter_rows() for cell in row if cell.value == "坏样本率")
        assert all(performance.cell(bad_rate_cell.row, col).number_format == "0.00%" for col in range(bad_rate_cell.column + 1, bad_rate_cell.column + 10))

        for sheet_name in self.EXPECTED_SHEETS:
            ws = workbook[sheet_name]
            assert _merged_range_for_row(ws, 2) is not None
        for sheet_name in self.EXPECTED_SHEETS[1:]:
            assert workbook[sheet_name]["B2"].hyperlink.location == "#'目录'!B2"

        contents = workbook["目录"]
        for sheet_name in self.EXPECTED_SHEETS[1:]:
            content_cell = next(cell for row in contents.iter_rows() for cell in row if cell.value == sheet_name)
            assert content_cell.hyperlink.location == f"#'{sheet_name}'!B2"

        feature_sheet = workbook["3-入模变量分析"]
        for index, feature in enumerate(self.FEATURE_NAMES, start=1):
            summary_cell = next(cell for row in feature_sheet.iter_rows() for cell in row if cell.value == feature and cell.hyperlink is not None)
            title_cell = next(cell for row in feature_sheet.iter_rows() for cell in row if cell.value == f"3.{index}、{feature} 有效性分析")
            assert summary_cell.hyperlink.location == f"#'3-入模变量分析'!{title_cell.coordinate}"
            assert title_cell.hyperlink.location == f"#'3-入模变量分析'!{summary_cell.coordinate}"

        assert all(workbook[sheet_name].freeze_panes is None for sheet_name in self.EXPECTED_SHEETS)
        for sheet_name in ["2-模型性能", "6-模型部署需求"]:
            ref = workbook[sheet_name].auto_filter.ref
            assert ref and ":" in ref
        assert workbook["2-模型性能"]._images
        assert feature_sheet._images

        deployment = workbook["6-模型部署需求"]
        deployment_values = self._sheet_values(deployment)
        assert "客户编号" in deployment_values
        assert all(feature in deployment_values for feature in self.FEATURE_NAMES)
        assert "模型分数" in deployment_values
        model_score = next(cell for row in deployment.iter_rows() for cell in row if cell.value == "模型分数")
        scores = [deployment.cell(row, model_score.column).value for row in range(model_score.row + 1, model_score.row + 6)]
        assert all(isinstance(value, (int, float)) for value in scores)
        assert len(set(scores)) > 1

        formula_errors = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"}
        assert not [(ws.title, cell.coordinate, cell.value) for ws in workbook.worksheets for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value in formula_errors]

    def test_optional_plot_insert_failure_is_logged(self, tmp_path, monkeypatch, caplog):
        X = pd.DataFrame({"f0": range(20), "target": [0, 1] * 10})
        report = ModelReport(MockModel(["f0"]), datasets={"train": X}, target="target", feature_names=["f0"])
        missing_plot = tmp_path / "missing-model-plot.png"
        monkeypatch.setattr(report, "_export_plots", lambda *args, **kwargs: ({"model_train": [str(missing_plot)]}, {}))

        report.to_excel(str(tmp_path / "plot-warning.xlsx"), with_plots=True)

        assert str(missing_plot) in caplog.text
        assert "2-模型性能" in caplog.text

    def test_optional_plot_generation_failure_is_logged_with_context(self, tmp_path, monkeypatch, caplog):
        X = pd.DataFrame({"f0": range(20), "target": [0, 1] * 10})
        report = ModelReport(MockModel(["f0"]), datasets={"train": X}, target="target", feature_names=["f0"])

        def fail_bin_plot(*args, **kwargs):
            raise RuntimeError("injected model bin plot failure")

        monkeypatch.setattr("hscredit.core.viz.bin_plot", fail_bin_plot)
        report._export_plots(tmp_path)

        assert "模型评分分箱图" in caplog.text
        assert "train" in caplog.text
        assert str(tmp_path / "bin_train.png") in caplog.text
        assert "injected model bin plot failure" in caplog.text

    def test_required_feature_effectiveness_table_failure_surfaces(self, tmp_path, monkeypatch):
        X = pd.DataFrame({"f0": range(20), "target": [0, 1] * 10})
        report = ModelReport(MockModel(["f0"]), datasets={"train": X}, target="target", feature_names=["f0"])

        def fail_feature_table(*args, **kwargs):
            raise ValueError("injected feature table failure")

        monkeypatch.setattr(report, "get_feature_bin_table", fail_feature_table)

        with pytest.raises(RuntimeError, match=r"特征=f0.*数据集=train") as exc_info:
            report.to_excel(str(tmp_path / "required-section.xlsx"), with_plots=False)

        assert isinstance(exc_info.value.__cause__, ValueError)
        assert "injected feature table failure" in str(exc_info.value.__cause__)

    def test_visibility_flags_hide_lift_and_importance_sections(self, tmp_path):
        X = pd.DataFrame({"f0": range(20), "target": [0, 1] * 10})
        report = ModelReport(MockModel(["f0"]), datasets={"train": X}, target="target", feature_names=["f0"])
        output = tmp_path / "visibility-flags.xlsx"

        report.to_excel(
            str(output),
            with_plots=False,
            show_lift=False,
            show_importance=False,
        )

        workbook = load_workbook(output)
        feature_values = self._sheet_values(workbook["3-入模变量分析"])
        assert not any(isinstance(value, str) and "入模变量重要性及分布情况" in value for value in feature_values)
        assert any(isinstance(value, str) and "入模变量相关性" in value for value in feature_values)
        assert any(isinstance(value, str) and "入模变量有效性分析" in value for value in feature_values)

    def test_export_plots_respects_show_lift(self, tmp_path, monkeypatch):
        X = pd.DataFrame({"f0": range(20), "target": [0, 1] * 10})
        report = ModelReport(MockModel(["f0"]), datasets={"train": X}, target="target", feature_names=["f0"])

        def fail_if_called(*args, **kwargs):
            raise AssertionError("show_lift=False 时不应生成 LIFT 曲线")

        monkeypatch.setattr("hscredit.core.viz.lift_plot", fail_if_called)
        paths, _ = report._export_plots(tmp_path, show_lift=False)

        assert all("lift_" not in path for figures in paths.values() for path in figures)

    def test_auto_model_report_forwards_visibility_flags(self, tmp_path, monkeypatch):
        from hscredit.report.model_report import auto_model_report

        X = pd.DataFrame({"f0": range(20)})
        y = pd.Series([0, 1] * 10)
        captured = {}

        def capture_to_excel(self, filepath, **kwargs):
            captured.update(kwargs)
            return filepath

        monkeypatch.setattr(ModelReport, "to_excel", capture_to_excel)
        auto_model_report(
            MockModel(["f0"]),
            X_train=X,
            y_train=y,
            feature_names=["f0"],
            excel_path=str(tmp_path / "forwarding.xlsx"),
            show_lift=False,
            show_importance=False,
            feature_contribution_label_max_features=3,
            verbose=False,
        )

        assert captured["show_lift"] is False
        assert captured["show_importance"] is False
        assert captured["feature_contribution_label_max_features"] == 3

    def test_required_directory_hyperlink_failure_surfaces(self, tmp_path, monkeypatch):
        from hscredit.excel import ExcelWriter

        X = pd.DataFrame({"f0": range(20), "target": [0, 1] * 10})
        report = ModelReport(MockModel(["f0"]), datasets={"train": X}, target="target", feature_names=["f0"])

        def fail_hyperlink(*args, **kwargs):
            raise ValueError("injected hyperlink failure")

        monkeypatch.setattr(ExcelWriter, "insert_hyperlink2sheet", fail_hyperlink)
        with pytest.raises(RuntimeError, match="目录链接") as exc_info:
            report.to_excel(str(tmp_path / "required-link.xlsx"), with_plots=False)

        assert isinstance(exc_info.value.__cause__, ValueError)

    def test_required_amount_score_table_failure_surfaces(self, tmp_path, monkeypatch):
        X = pd.DataFrame(
            {
                "f0": range(20),
                "target": [0, 1] * 10,
                "放款金额": np.linspace(1000, 2000, 20),
            }
        )
        report = ModelReport(MockModel(["f0"]), datasets={"train": X}, target="target", feature_names=["f0"])
        original_get_bin_table = report.get_bin_table

        def fail_amount_table(*args, **kwargs):
            if kwargs.get("amount_col") is not None:
                raise ValueError("injected amount table failure")
            return original_get_bin_table(*args, **kwargs)

        monkeypatch.setattr(report, "get_bin_table", fail_amount_table)
        with pytest.raises(RuntimeError, match=r"金额口径评分分箱.*数据集=train") as exc_info:
            report.to_excel(
                str(tmp_path / "required-amount-table.xlsx"),
                with_plots=False,
                amount_col="放款金额",
            )

        assert isinstance(exc_info.value.__cause__, ValueError)


class TestModelReportCustomDatasetKeys:
    """datasets 传入自定义键（如 {'建模集': df, 'OOT': df}）时的回退逻辑."""

    @staticmethod
    def _custom_datasets():
        train = pd.DataFrame(
            {
                "f0": np.arange(20),
                "MOB1": [0, 1, 3, 7, 8] * 4,
                "放款金额": np.arange(100, 120),
            }
        )
        return {"建模集": train, "OOT": train.copy()}

    def _make_report(self):
        return ModelReport(
            MockModel(["f0"]),
            datasets=self._custom_datasets(),
            overdue=["MOB1"],
            dpds=[7, 3, 0],
            feature_names=["f0"],
        )

    def test_train_test_key_fallback(self):
        report = self._make_report()
        assert report._train_key == "建模集"
        assert report._test_key == "OOT"
        assert report._resolve_dataset_key("train") == "建模集"
        assert report._resolve_dataset_key("test") == "OOT"
        assert report._resolve_dataset_key("建模集") == "建模集"

    def test_resolve_dataset_key_missing(self):
        report = self._make_report()
        with pytest.raises(KeyError, match="不存在"):
            report._resolve_dataset_key("unknown")

    def test_feature_importance_custom_keys(self):
        importance = self._make_report().get_feature_importance()
        assert not importance.empty
        assert {"特征重要性", "IV", "KS", "PSI"} <= set(importance.columns)

    def test_bin_table_default_dataset_custom_keys(self):
        report = self._make_report()
        assert not report.get_bin_table(max_n_bins=5).empty
        assert not report.get_feature_bin_table("f0", max_n_bins=5).empty

    def test_describe_corr_summary_custom_keys(self):
        report = self._make_report()
        assert not report.get_features_describe().empty
        assert not report.get_features_corr().empty
        assert not report._get_features_summary().empty

    def test_print_report_custom_keys(self, capsys):
        self._make_report().print_report(n_bins=5)
        out = capsys.readouterr().out
        assert "建模集" in out
        assert "OOT" in out

    def test_to_excel_custom_keys(self, tmp_path):
        report = self._make_report()
        output = tmp_path / "model_report_custom_keys.xlsx"
        report.to_excel(str(output), with_plots=False, amount_col="放款金额")

        workbook = load_workbook(output)
        assert "4-稳定性分析" in workbook.sheetnames
        stability = [cell.value for row in workbook["4-稳定性分析"].iter_rows() for cell in row]
        assert any(isinstance(v, str) and "评分漂移分析（vs 建模集）" in v for v in stability)


class TestModelReportDatasetNaming:
    """数据集统一命名规则：list → 数据集N；X_train/X_test/X_oot → 训练集/测试集/跨时间验证集；dict → key."""

    @staticmethod
    def _data(n=6):
        return pd.DataFrame({"f0": np.arange(n), "target": [0, 1] * (n // 2)})

    def test_list_datasets_named_sequentially(self):
        report = ModelReport(
            MockModel(["f0"]),
            datasets=[self._data(), self._data(), self._data()],
            target="target",
            feature_names=["f0"],
        )
        assert list(report._datasets.keys()) == ["数据集1", "数据集2", "数据集3"]
        assert [ds.label for ds in report._datasets.values()] == ["数据集1", "数据集2", "数据集3"]

    def test_xy_params_named_train_test_oot(self):
        report = ModelReport(
            MockModel(["f0"]),
            X_train=self._data(),
            X_test=self._data(),
            X_oot=self._data(),
            target="target",
            feature_names=["f0"],
        )
        assert list(report._datasets.keys()) == ["训练集", "测试集", "跨时间验证集"]

    def test_xy_params_sklearn_style_y_priority(self):
        """sklearn 风格：显式传入的 y 优先于 target 列."""
        X = self._data()
        y = pd.Series([1, 1, 1, 0, 0, 0])
        report = ModelReport(
            MockModel(["f0"]),
            X_train=X,
            y_train=y,
            X_oot=X,
            y_oot=y,
            target="target",
            feature_names=["f0"],
        )
        assert report._datasets["训练集"].y.tolist() == [1, 1, 1, 0, 0, 0]
        assert report._datasets["跨时间验证集"].y.tolist() == [1, 1, 1, 0, 0, 0]

    def test_dict_uses_key_as_name(self):
        report = ModelReport(
            MockModel(["f0"]),
            datasets={"建模集": self._data(), "OOT": self._data()},
            target="target",
            feature_names=["f0"],
        )
        assert list(report._datasets.keys()) == ["建模集", "OOT"]
        assert [ds.label for ds in report._datasets.values()] == ["建模集", "OOT"]

    def test_overdue_alone_ignores_target(self):
        """传入 overdue（dpds 缺省为 0）时直接忽略 target."""
        X = pd.DataFrame(
            {
                "f0": [1, 2, 3, 4],
                "label": [1, 1, 1, 1],
                "dpds": [0, 0, 5, 10],
            }
        )
        report = ModelReport(
            MockModel(["f0"]),
            X_train=X,
            y_train=None,
            target="label",
            overdue="dpds",
            feature_names=["f0"],
        )
        # dpds > 0 → [0, 0, 1, 1]，而非 target 列的全 1
        assert report._datasets["训练集"].y.tolist() == [0, 0, 1, 1]

    def test_no_datasets_raises(self):
        with pytest.raises(ValueError, match="未提供任何数据集"):
            ModelReport(MockModel(["f0"]), target="target", feature_names=["f0"])


class NumpyFeatureModel(MockModel):
    """feature_names_in_ 为 numpy 数组的模拟模型（模拟 XGBoost/LightGBM）."""

    def __init__(self, feature_names):
        super().__init__(feature_names)
        self.feature_names_in_ = np.asarray(feature_names)


class TestModelReportNumpyInputs:
    """numpy 类型入参（ndarray / np.str_）兼容性回归测试."""

    @staticmethod
    def _data(n=8):
        return pd.DataFrame(
            {
                "f0": np.arange(n, dtype=float),
                "f1": np.arange(n, dtype=float) * 0.5,
                "target": [0, 1] * (n // 2),
            }
        )

    def test_feature_names_numpy_array(self):
        """feature_names 传 numpy 数组时不应触发真值歧义报错."""
        report = ModelReport(
            NumpyFeatureModel(["f0", "f1"]),
            X_train=self._data(),
            target="target",
            feature_names=np.array(["f0", "f1"]),
        )
        assert report.feature_names == ["f0", "f1"]
        assert all(type(f) is str for f in report.feature_names)

    def test_model_feature_names_in_numpy(self):
        """model.feature_names_in_ 为 np.str_ 数组时特征校验与筛选正常."""
        report = ModelReport(
            NumpyFeatureModel(["f0", "f1"]),
            X_train=self._data(),
            target="target",
            feature_names=["f0", "f1"],
        )
        assert report.feature_names == ["f0", "f1"]

    def test_np_str_columns_normalized(self):
        """DataFrame 列名为 np.str_ 时统一规整为 Python str."""
        df = self._data()
        df.columns = list(np.asarray(df.columns))  # np.str_ 列名
        report = ModelReport(
            NumpyFeatureModel(["f0", "f1"]),
            X_train=df,
            target="target",
            feature_names=np.array(["f0", "f1"]),
        )
        train_X = report._datasets["训练集"].X
        assert all(type(c) is str for c in train_X.columns)
        assert report.feature_names == ["f0", "f1"]

    def test_numpy_X_with_numpy_feature_names(self):
        """X 为 ndarray 且 feature_names 为 ndarray 时正常构建数据集."""
        df = self._data()
        report = ModelReport(
            NumpyFeatureModel(["f0", "f1"]),
            X_train=df[["f0", "f1"]].to_numpy(),
            y_train=df["target"],
            feature_names=np.array(["f0", "f1"]),
        )
        assert list(report._datasets["训练集"].X.columns) == ["f0", "f1"]
        assert report.feature_names == ["f0", "f1"]

    def test_add_dataset_numpy_feature_names(self):
        """add_dataset 的 feature_names 传 numpy 数组时不应真值歧义报错."""
        report = ModelReport(
            NumpyFeatureModel(["f0", "f1"]),
            X_train=self._data(),
            target="target",
            feature_names=["f0", "f1"],
        )
        report.add_dataset("验证集", "验证集", self._data(), feature_names=np.array(["f0", "f1"]))
        assert "验证集" in report._datasets


class TestModelReportFeatureNamesResolution:
    """feature_names 解析优先级回归测试."""

    def test_feature_names_from_model_when_not_passed(self):
        """未传 feature_names 时直接从模型 feature_names_in_ 获取（顺序以模型为准，排除非入模字段）."""
        df = pd.DataFrame(
            {
                "f1": [1.0, 2.0, 3.0, 4.0],
                "f0": [2.0, 1.0, 0.5, 0.2],
                "extra": [9, 9, 9, 9],  # 非入模字段
                "target": [0, 0, 1, 1],
            }
        )
        # 模型特征顺序与 X 列顺序不同
        report = ModelReport(NumpyFeatureModel(["f0", "f1"]), X_train=df, target="target")
        assert report.feature_names == ["f0", "f1"]
        assert all(type(f) is str for f in report.feature_names)

    def test_explicit_feature_names_filtered_by_model_required(self):
        """显式传入 feature_names 时按模型入模特征过滤，保留传入顺序."""
        df = pd.DataFrame(
            {
                "f0": [1.0, 2.0, 3.0, 4.0],
                "f1": [2.0, 1.0, 0.5, 0.2],
                "extra": [9, 9, 9, 9],
                "target": [0, 0, 1, 1],
            }
        )
        report = ModelReport(
            NumpyFeatureModel(["f0", "f1"]),
            X_train=df,
            target="target",
            feature_names=["f1", "f0", "extra"],
        )
        assert report.feature_names == ["f1", "f0"]

    def test_no_model_attrs_uses_dataset_columns(self):
        """模型无 feature_names_in_ / feature_names_ 时回退到数据集全部列名（含 target 列）."""
        df = pd.DataFrame(
            {
                "f0": [1.0, 2.0, 3.0, 4.0],
                "f1": [2.0, 1.0, 0.5, 0.2],
                "target": [0, 0, 1, 1],
            }
        )
        report = ModelReport(MockModel(), X_train=df, target="target")
        assert report.feature_names == ["f0", "f1", "target"]
