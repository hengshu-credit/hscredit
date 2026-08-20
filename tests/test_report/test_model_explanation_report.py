"""ModelReport 模型解释集成测试。"""

import pandas as pd
from openpyxl import load_workbook
from sklearn.datasets import make_classification
from sklearn.ensemble import RandomForestClassifier

from hscredit.report import ModelReport


def _report(enabled=False):
    values, y = make_classification(n_samples=45, n_features=4, random_state=23)
    X = pd.DataFrame(values, columns=["甲", "乙", "丙", "丁"])
    model = RandomForestClassifier(n_estimators=8, max_depth=3, random_state=23).fit(X, y)
    config = {"enabled": enabled, "data": X.head(12), "background_data": X.head(20), "n_bootstrap": 4}
    return ModelReport(model, X_train=X, y_train=y, n_jobs=1, explain_config=config)


def test_model_report_explanation_is_opt_in_and_structured():
    disabled = _report(False)
    assert disabled.explain_config["enabled"] is False
    assert "模型解释" not in disabled.to_dict()
    enabled = _report(True)
    explanation = enabled.get_model_explanation()
    assert {"元信息", "全局解释", "稳定性", "代表样本", "样本解释", "原因码"} <= set(explanation)
    assert "模型解释" in enabled.to_dict()


def test_model_report_appends_seventh_sheet(tmp_path):
    report = _report(True)
    output = report.to_excel(tmp_path / "解释报告.xlsx", with_plots=False)
    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames[:7] == ["目录", "1-基本信息", "2-模型性能", "3-入模变量分析", "4-稳定性分析", "5-模型参数", "6-模型部署需求"]
    assert workbook.sheetnames[7] == "7-模型解释"
