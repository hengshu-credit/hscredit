"""Rule.save Excel 导出回归测试。"""

import pandas as pd
import pytest
from openpyxl import load_workbook

from hscredit.core.rules import Rule
from hscredit.excel import ExcelWriter


def test_rule_save_writes_path_and_returns_writer(tmp_path):
    """PathLike 输入自动保存，并透传列格式参数。"""
    report = pd.DataFrame({"规则名称": ["成年"], "命中率": [0.25]})
    output = tmp_path / "rule.xlsx"

    writer = Rule.save(
        report,
        output,
        sheet_name="规则报告",
        excel_params={"percent_cols": ["命中率"]},
    )

    assert isinstance(writer, ExcelWriter)
    assert output.exists()
    workbook = load_workbook(output)
    worksheet = workbook["规则报告"]
    assert worksheet["B2"].value == "规则名称"
    assert worksheet["B3"].value == "成年"
    assert worksheet["C3"].value == 0.25
    assert worksheet["C3"].number_format == "0.00%"
    workbook.close()


def test_rule_save_reuses_writer_without_auto_save(monkeypatch):
    """已有 writer 被复用，并由调用方决定最终保存时机。"""
    report = pd.DataFrame({"规则名称": ["成年"]})
    writer = ExcelWriter()
    saved = []
    monkeypatch.setattr(writer, "save", lambda *args, **kwargs: saved.append((args, kwargs)))

    returned = Rule.save(report, writer, sheet_name="规则报告")

    assert returned is writer
    assert saved == []
    assert writer.workbook["规则报告"]["B3"].value == "成年"


def test_rule_save_preserves_multiindex_columns_index_and_order(tmp_path):
    """多层表头、多层索引和原始顺序在导出中保持不变。"""
    columns = pd.MultiIndex.from_tuples([("样本", "数量"), ("指标", "命中率")])
    index = pd.MultiIndex.from_tuples(
        [("规则B", "整体"), ("规则A", "分组1")],
        names=["规则", "分组"],
    )
    report = pd.DataFrame([[10, 0.2], [5, 0.4]], columns=columns, index=index)
    original = report.copy(deep=True)
    writer = ExcelWriter()

    returned = Rule.save(
        report,
        writer,
        sheet_name="多层级",
        excel_params={"index": True, "merge_index": False},
    )

    assert returned is writer
    pd.testing.assert_frame_equal(report, original)
    output = tmp_path / "multi.xlsx"
    writer.save(str(output))
    workbook = load_workbook(output)
    worksheet = workbook["多层级"]
    assert [worksheet.cell(2, column).value for column in range(2, 6)] == [None, None, "样本", "指标"]
    assert [worksheet.cell(3, column).value for column in range(2, 6)] == ["规则", "分组", "数量", "命中率"]
    assert [worksheet.cell(4, column).value for column in range(2, 6)] == ["规则B", "整体", 10, 0.2]
    assert [worksheet.cell(5, column).value for column in range(2, 6)] == ["规则A", "分组1", 5, 0.4]
    workbook.close()


def test_rule_save_explicit_arguments_win_over_conflicting_excel_params(tmp_path):
    """excel_params 不能覆盖显式 report、writer 和 sheet_name。"""
    report = pd.DataFrame({"规则名称": ["成年"]})
    output = tmp_path / "explicit.xlsx"

    Rule.save(
        report,
        output,
        sheet_name="显式名称",
        excel_params={
            "data": pd.DataFrame({"错误": [1]}),
            "excel_writer": tmp_path / "wrong.xlsx",
            "sheet_name": "错误名称",
        },
    )

    workbook = load_workbook(output)
    assert "显式名称" in workbook.sheetnames
    assert "错误名称" not in workbook.sheetnames
    assert workbook["显式名称"]["B3"].value == "成年"
    workbook.close()
    assert not (tmp_path / "wrong.xlsx").exists()


def test_rule_save_rejects_invalid_writer_type():
    """非法 writer 类型返回稳定的中文错误。"""
    report = pd.DataFrame({"规则名称": ["成年"]})

    with pytest.raises(TypeError, match="excel_writer 必须是路径或 ExcelWriter 对象"):
        Rule.save(report, 1)
