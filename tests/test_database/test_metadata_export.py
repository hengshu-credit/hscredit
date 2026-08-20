"""数据库元数据宽表与 Excel 导出测试。"""

from pathlib import Path

import openpyxl
import pytest

from hscredit.database import Database, register_adapter
from hscredit.database.adapters.base import BaseDatabaseAdapter
from hscredit.database.metadata import METADATA_COLUMNS_ZH, MetadataInspection
from hscredit.exceptions import ValidationError

RAW_METADATA_ROW = {
    "database_type": "mysql",
    "catalog": "def",
    "database": "risk_db",
    "schema": None,
    "table_name": "events",
    "qualified_name": "risk_db.events",
    "table_type": "BASE TABLE",
    "table_comment": "事件表",
    "table_engine": "InnoDB",
    "column_name": "id",
    "ordinal_position": 1,
    "data_type": "bigint",
    "full_data_type": "bigint unsigned",
    "pandas_dtype": "Int64",
    "nullable": "YES",
    "default_value": None,
    "primary_key": True,
    "unique_key": "PRI",
    "partition_key": None,
    "sort_key": False,
    "bucket_key": None,
    "column_comment": "事件编号",
}


class ObservableMetadataAdapter(BaseDatabaseAdapter):
    database_type = "observable_metadata"

    def __init__(self, *, connect_kwargs, pool_options, adapter_options):
        super().__init__(
            connect_kwargs=connect_kwargs,
            pool_options=pool_options,
            adapter_options=adapter_options,
        )
        self.received_targets = None

    def inspect_schema(self, targets):
        self.received_targets = targets
        return MetadataInspection(
            rows=[RAW_METADATA_ROW],
            errors=[{"目标": "secret.hidden", "错误": "权限不足"}],
        )


@pytest.fixture(autouse=True)
def register_metadata_adapter():
    register_adapter("observable_metadata", ObservableMetadataAdapter, replace=True)


@pytest.fixture
def database():
    return Database("observable_metadata")


def test_metadata_headers_are_chinese_and_values_are_unchanged(database):
    frame = database.export_schema(targets=["risk_db", "risk_db.events"])

    assert frame.columns.tolist() == METADATA_COLUMNS_ZH
    assert frame.loc[0, "表类型"] == "BASE TABLE"
    assert frame.loc[0, "是否可空"] == "YES"
    assert bool(frame.loc[0, "是否主键"]) is True
    assert frame.loc[0, "是否唯一键"] == "PRI"
    assert frame.attrs["错误"] == [{"目标": "secret.hidden", "错误": "权限不足"}]
    assert [target.parts for target in database.adapter.received_targets] == [
        ("risk_db",),
        ("risk_db", "events"),
    ]


def test_missing_metadata_values_are_none_not_guessed(database):
    frame = database.export_schema()

    assert frame.loc[0, "模式名"] is None
    assert frame.loc[0, "是否分区键"] is None


def test_metadata_xlsx_uses_dataframe2excel_defaults(database, tmp_path):
    output = tmp_path / "数据库表结构.xlsx"

    frame = database.export_schema(output=output)

    workbook = openpyxl.load_workbook(output)
    assert workbook.sheetnames == ["表结构"]
    worksheet = workbook["表结构"]
    visible_values = [cell.value for row in worksheet.iter_rows(min_row=1, max_row=5) for cell in row]
    assert "数据库表结构" in visible_values
    assert "数据库类型" in visible_values
    assert "events" in [cell.value for row in worksheet.iter_rows() for cell in row]
    assert frame.loc[0, "字段名"] == "id"


def test_excel_params_override_defaults(database, tmp_path):
    output = tmp_path / "custom.XLSX"

    database.export_schema(
        output=output,
        excel_params={
            "sheet_name": "字段清单",
            "title": "原始数据库元数据",
            "start_row": 1,
            "start_col": 1,
        },
    )

    workbook = openpyxl.load_workbook(output)
    assert workbook.sheetnames == ["字段清单"]
    assert workbook["字段清单"]["A1"].value == "原始数据库元数据"


@pytest.mark.parametrize("suffix", [".csv", ".tsv", ".xls", ".json"])
def test_metadata_rejects_non_xlsx_output(database, tmp_path, suffix):
    output = Path(tmp_path) / f"schema{suffix}"

    with pytest.raises(ValidationError, match="仅支持 .xlsx"):
        database.export_schema(output=output)

    assert not output.exists()


@pytest.mark.parametrize("targets", ["", ["risk_db", ""], ["risk..events"]])
def test_metadata_rejects_empty_or_malformed_targets(database, targets):
    with pytest.raises(ValidationError, match="目标"):
        database.export_schema(targets=targets)


def test_exact_table_target_must_exist():
    class EmptyAdapter(ObservableMetadataAdapter):
        def inspect_schema(self, targets):
            self.received_targets = targets
            return MetadataInspection(rows=[], errors=[])

    register_adapter("empty_metadata", EmptyAdapter, replace=True)
    database = Database("empty_metadata")

    from hscredit.database.exceptions import DatabaseMetadataError

    with pytest.raises(DatabaseMetadataError, match="未找到精确指定的数据库表"):
        database.export_schema(targets=["risk.events"])
