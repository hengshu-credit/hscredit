"""hsreport 完整 Excel 报告操作测试。"""

from pathlib import Path

import matplotlib
import pytest
from openpyxl import Workbook, load_workbook
from sklearn.linear_model import LogisticRegression

matplotlib.use("Agg")

from hscredit.skills_runtime import execute_skill
from hscredit.skills_runtime.errors import SkillExecutionError
from hscredit.skills_runtime.operations import reports as report_operations


def _base_request(tmp_path, operation, parameters, name, inputs=None):
    return {
        "version": "1",
        "operation": operation,
        "inputs": inputs or {"data": {"kind": "object_ref", "ref": "data:credit"}},
        "parameters": parameters,
        "output": {"directory": str(tmp_path), "name": name, "overwrite": False},
        "environment": {"mode": "current", "install_missing": False},
    }


def _workbook(result):
    artifact = next(item for item in result["artifacts"] if item["type"] == "excel")
    path = Path(artifact["path"])
    assert path.is_file()
    return load_workbook(path, read_only=True)


def test_auto_feature_analysis_writes_a_real_workbook(tmp_path, credit_frame):
    """防止特征报告适配器只返回坐标而不生成 Excel。"""
    request = _base_request(
        tmp_path,
        "auto_feature_analysis",
        {
            "features": ["score", "age"],
            "target": "target",
            "date": "apply_date",
            "amount": "amount",
            "pictures": [],
            "n_jobs": 1,
        },
        "feature_report",
    )

    result = execute_skill("hsreport", request, objects={"data:credit": credit_frame})
    workbook = _workbook(result)

    assert workbook.sheetnames == ["分析报告"]
    assert result["summary"]["end_row"] > 2
    assert result["summary"]["end_col"] > 2


def test_auto_feature_analysis_combines_overdue_labels_and_exposes_feature_summary(
    monkeypatch,
    tmp_path,
    credit_frame,
):
    """多标签分析应单次执行，并在运行摘要中返回标签组合和变量关键指标。"""
    calls = []

    def fake_auto_feature_analysis(
        data,
        features=None,
        overdue=None,
        dpds=None,
        excel_writer=None,
        output_dir=None,
        pictures=None,
        n_jobs=1,
    ):
        calls.append({"features": features, "overdue": overdue, "dpds": dpds})
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "分析报告"
        worksheet.cell(row=1, column=1, value="2、变量综合统计")
        headers = [
            "特征名",
            "字段类型",
            "样本数",
            "缺失率",
            "唯一值数",
            "IV",
            "KS",
            "趋势",
            "PSI",
            "平均值",
        ]
        for column, header in enumerate(headers, start=1):
            worksheet.cell(row=3, column=column, value=header)
        values = ["feature_a", "数值型", 100, 0.05, 80, 0.42, 0.31, "ascending", 0.02, 12.5]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=4, column=column, value=value)
        workbook.save(excel_writer)
        workbook.close()
        return 4, len(headers)

    monkeypatch.setattr(report_operations, "auto_feature_analysis", fake_auto_feature_analysis)
    request = _base_request(
        tmp_path,
        "auto_feature_analysis",
        {
            "features": ["feature_a"],
            "overdue": ["overdue_a", "overdue_b"],
            "dpds": [0, 7],
            "pictures": [],
            "n_jobs": 1,
        },
        "multi_label_feature_report",
    )

    result = execute_skill("hsreport", request, objects={"data:credit": credit_frame})

    assert calls == [
        {
            "features": ["feature_a"],
            "overdue": ["overdue_a", "overdue_b"],
            "dpds": [0, 7],
        }
    ]
    assert len([artifact for artifact in result["artifacts"] if artifact["type"] == "excel"]) == 1
    assert result["summary"]["label_combinations"] == [
        {"overdue": "overdue_a", "dpd": 0},
        {"overdue": "overdue_a", "dpd": 7},
        {"overdue": "overdue_b", "dpd": 0},
        {"overdue": "overdue_b", "dpd": 7},
    ]
    assert result["summary"]["feature_summary"] == {
        "rows": 1,
        "columns": ["特征名", "字段类型", "样本数", "缺失率", "唯一值数", "IV", "KS", "趋势", "PSI"],
        "preview": [
            {
                "特征名": "feature_a",
                "字段类型": "数值型",
                "样本数": 100,
                "缺失率": 0.05,
                "唯一值数": 80,
                "IV": 0.42,
                "KS": 0.31,
                "趋势": "ascending",
                "PSI": 0.02,
            }
        ],
    }


def test_auto_model_report_writes_a_real_workbook(tmp_path, credit_frame):
    """防止模型和命名数据集装配错误。"""
    X = credit_frame[["score", "age"]]
    y = credit_frame["target"]
    model = LogisticRegression(max_iter=200).fit(X, y)
    request = _base_request(
        tmp_path,
        "auto_model_report",
        {
            "datasets": {"训练集": "train"},
            "target": "target",
            "feature_names": ["score", "age"],
            "with_plots": False,
            "verbose": False,
            "n_jobs": 1,
        },
        "model_report",
        inputs={
            "model": {"kind": "object_ref", "ref": "model:lr"},
            "train": {"kind": "object_ref", "ref": "data:credit"},
        },
    )

    result = execute_skill(
        "hsreport",
        request,
        objects={"model:lr": model, "data:credit": credit_frame},
    )
    workbook = _workbook(result)

    assert "1-基本信息" in workbook.sheetnames
    assert "2-模型性能" in workbook.sheetnames
    assert result["summary"]["datasets"] == ["训练集"]


def test_swap_out_report_writes_strategy_and_binning_sheets(tmp_path, credit_frame):
    """防止策略报告丢失变量分箱明细或规则文本。"""
    request = _base_request(
        tmp_path,
        "swap_out_report",
        {
            "rules": ["score < 560", "age < 25"],
            "target": "target",
            "features": ["score", "age"],
            "amount": "amount",
            "date_col": "apply_date",
            "methods": "quantile",
            "n_jobs": 1,
        },
        "strategy_report",
    )

    result = execute_skill("hsreport", request, objects={"data:credit": credit_frame})
    workbook = _workbook(result)

    assert "策略迭代" in workbook.sheetnames
    assert "变量分箱" in workbook.sheetnames
    assert result["summary"]["rules"] == 2


def test_failed_report_preserves_cause_and_publishes_no_workbook(tmp_path, credit_frame):
    """防止失败报告被误标成功或留下半成品。"""
    request = _base_request(
        tmp_path,
        "auto_feature_analysis",
        {
            "features": ["score"],
            "target": "missing_target",
            "pictures": [],
            "n_jobs": 1,
        },
        "broken_report",
    )

    with pytest.raises(SkillExecutionError) as exc_info:
        execute_skill("hsreport", request, objects={"data:credit": credit_frame})

    assert exc_info.value.code == "HSCREDIT_EXECUTION_FAILED"
    assert exc_info.value.cause is not None
    assert not (tmp_path / "broken_report.xlsx").exists()
    assert not list(tmp_path.glob(".hscredit-skill-*"))
