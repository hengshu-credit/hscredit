"""hsbin 表格分析操作测试。"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from hscredit.skills_runtime import execute_skill


def _request(tmp_path, operation, parameters, name):
    return {
        "version": "1",
        "operation": operation,
        "inputs": {"data": {"kind": "object_ref", "ref": "data:credit"}},
        "parameters": parameters,
        "output": {"directory": str(tmp_path), "name": name, "overwrite": False},
        "environment": {"mode": "current", "install_missing": False},
    }


def _workbook(result):
    artifact = next(item for item in result["artifacts"] if item["type"] == "excel")
    path = Path(artifact["path"])
    assert path.is_file()
    return load_workbook(path, read_only=True)


def test_feature_bin_stats_executes_real_hscredit_and_writes_excel(tmp_path, credit_frame):
    """防止适配器返回空壳结果或跳过真实分箱统计。"""
    request = _request(
        tmp_path,
        "feature_bin_stats",
        {
            "feature": "score",
            "target": "target",
            "method": "quantile",
            "max_n_bins": 4,
            "n_jobs": 1,
        },
        "score_stats",
    )

    result = execute_skill("hsbin", request, objects={"data:credit": credit_frame})

    assert result["status"] == "success"
    assert result["summary"]["rows"] >= 2
    assert result["summary"]["preview"]
    assert _workbook(result).sheetnames == ["分箱统计"]


def test_feature_bin_stats_combines_overdue_labels_and_exposes_combinations(tmp_path, credit_frame):
    """多标签分箱应单次执行，并在现有表格摘要中追加全部标签组合。"""
    data = credit_frame.copy()
    data["overdue_a"] = data["target"] * 10
    data["overdue_b"] = data["target"] * 20
    request = _request(
        tmp_path,
        "feature_bin_stats",
        {
            "feature": "score",
            "overdue": ["overdue_a", "overdue_b"],
            "dpds": [0, 7],
            "method": "quantile",
            "max_n_bins": 4,
            "n_jobs": 1,
        },
        "multi_label_stats",
    )

    result = execute_skill("hsbin", request, objects={"data:credit": data})

    assert result["summary"]["rows"] >= 2
    assert result["summary"]["preview"]
    assert result["summary"]["label_combinations"] == [
        {"overdue": "overdue_a", "dpd": 0},
        {"overdue": "overdue_a", "dpd": 7},
        {"overdue": "overdue_b", "dpd": 0},
        {"overdue": "overdue_b", "dpd": 7},
    ]
    assert _workbook(result).sheetnames == ["分箱统计"]


def test_benchmark_binning_methods_writes_method_comparison(tmp_path, credit_frame):
    """防止方法对比操作调用错误的目标列或遗漏结果。"""
    request = _request(
        tmp_path,
        "benchmark_binning_methods",
        {
            "feature": "score",
            "overdue": "MOB1",
            "dpds": [0],
            "hscredit_methods": ["quantile", "tree"],
            "max_n_bins": 4,
            "long_format": True,
            "n_jobs": 1,
        },
        "binning_benchmark",
    )

    result = execute_skill("hsbin", request, objects={"data:credit": credit_frame})

    assert result["summary"]["rows"] == 2
    assert _workbook(result).sheetnames == ["方法对比"]


def test_benchmark_binning_methods_combines_overdue_labels_and_uses_chinese_columns(tmp_path, credit_frame):
    """技能入口丢失多逾期组合元数据或返回英文指标名时，本测试必须失败。"""
    data = credit_frame.copy()
    data["MOB3"] = data["target"] * 10
    request = _request(
        tmp_path,
        "benchmark_binning_methods",
        {
            "feature": "score",
            "overdue": ["MOB1", "MOB3"],
            "dpds": [0, 3],
            "hscredit_methods": ["quantile"],
            "prebinning": None,
            "lift_refine": False,
            "long_format": True,
            "n_jobs": 1,
        },
        "multi_label_benchmark",
    )

    result = execute_skill("hsbin", request, objects={"data:credit": data})

    assert result["summary"]["rows"] == 4
    assert result["summary"]["label_combinations"] == [
        {"overdue": "MOB1", "dpd": 0},
        {"overdue": "MOB1", "dpd": 3},
        {"overdue": "MOB3", "dpd": 0},
        {"overdue": "MOB3", "dpd": 3},
    ]
    assert "分箱方法" in result["summary"]["preview"][0]
    assert "逾期字段" in result["summary"]["preview"][0]
    assert _workbook(result).sheetnames == ["方法对比"]


def test_benchmark_binning_methods_forwards_method_specific_kwargs_through_skill(tmp_path, credit_frame):
    """技能适配器拦截底层合法 kwargs 时，本测试必须失败。"""
    request = _request(
        tmp_path,
        "benchmark_binning_methods",
        {
            "feature": "score",
            "overdue": "MOB1",
            "dpds": [3],
            "hscredit_methods": ["quantile"],
            "prebinning": None,
            "lift_refine": False,
            "quantiles": [0, 0.2, 0.8, 1],
            "force_numerical": True,
            "long_format": True,
            "n_jobs": 1,
        },
        "benchmark_kwargs",
    )

    result = execute_skill("hsbin", request, objects={"data:credit": credit_frame})

    assert result["status"] == "success"
    assert result["summary"]["rows"] == 1
    assert len(result["summary"]["preview"][0]["切分点"]) == 2


def test_benchmark_binning_methods_exposes_default_label_combinations(tmp_path, credit_frame):
    """默认逾期字段或阈值未进入技能摘要时，本测试必须失败。"""
    request = _request(
        tmp_path,
        "benchmark_binning_methods",
        {
            "feature": "score",
            "hscredit_methods": ["quantile"],
            "prebinning": None,
            "lift_refine": False,
            "n_jobs": 1,
        },
        "benchmark_defaults",
    )

    result = execute_skill("hsbin", request, objects={"data:credit": credit_frame})

    assert result["summary"]["rows"] == 1
    assert ["分箱详情", "分箱方法"] in result["summary"]["columns"]
    assert ["MOB1_3+", "综合评分"] in result["summary"]["columns"]
    assert ["MOB1_0+", "综合评分"] in result["summary"]["columns"]
    assert result["summary"]["preview"][0]["分箱详情 / 分箱方法"] == "hscredit-quantile"
    assert result["summary"]["label_combinations"] == [
        {"overdue": "MOB1", "dpd": 3},
        {"overdue": "MOB1", "dpd": 0},
    ]


def test_benchmark_binning_methods_skill_exposes_quality_metrics_in_long_format(tmp_path, credit_frame):
    """技能入口应接受 long_format 并在摘要中展示中文质量指标。"""
    request = _request(
        tmp_path,
        "benchmark_binning_methods",
        {
            "feature": "score",
            "overdue": "MOB1",
            "dpds": [3],
            "hscredit_methods": ["quantile"],
            "prebinning": None,
            "lift_refine": False,
            "long_format": True,
            "n_jobs": 1,
        },
        "benchmark_quality_metrics",
    )

    result = execute_skill("hsbin", request, objects={"data:credit": credit_frame})

    preview = result["summary"]["preview"][0]
    assert result["status"] == "success"
    assert pd.isna(preview["错误信息"])
    assert {"LIFT二次项系数", "综合评分", "LIFT序列", "坏样本率序列"}.issubset(preview)


def test_feature_binning_summary_writes_summary_and_detail_sheets(tmp_path, credit_frame):
    """防止跨方法摘要丢失嵌套分箱明细。"""
    request = _request(
        tmp_path,
        "feature_binning_summary",
        {
            "feature": ["score", "age"],
            "methods": ["quantile"],
            "target": "target",
            "max_n_bins": 4,
            "n_jobs": 1,
        },
        "binning_summary",
    )

    result = execute_skill("hsbin", request, objects={"data:credit": credit_frame})
    workbook = _workbook(result)

    assert result["summary"]["rows"] == 2
    assert workbook.sheetnames[0] == "分箱摘要"
    assert any(name.startswith("score-quantile") for name in workbook.sheetnames)
    assert any(name.startswith("age-quantile") for name in workbook.sheetnames)


def test_multi_label_binning_summary_operations_expose_combinations(tmp_path, credit_frame):
    """跨方法及分组摘要都应保留多标签组合元数据。"""
    data = credit_frame.copy()
    data["overdue_a"] = data["target"] * 10
    data["overdue_b"] = data["target"] * 20
    expected = [
        {"overdue": "overdue_a", "dpd": 0},
        {"overdue": "overdue_a", "dpd": 7},
        {"overdue": "overdue_b", "dpd": 0},
        {"overdue": "overdue_b", "dpd": 7},
    ]
    cases = [
        (
            "feature_binning_summary",
            {
                "feature": "score",
                "methods": "quantile",
                "overdue": ["overdue_a", "overdue_b"],
                "dpds": [0, 7],
                "max_n_bins": 4,
                "n_jobs": 1,
            },
        ),
        (
            "feature_group_binning_summary",
            {
                "feature": "score",
                "methods": "quantile",
                "group_col": "segment",
                "overdue": ["overdue_a", "overdue_b"],
                "dpds": [0, 7],
                "max_n_bins": 4,
                "n_jobs": 1,
            },
        ),
    ]

    for operation, parameters in cases:
        result = execute_skill(
            "hsbin",
            _request(tmp_path, operation, parameters, f"{operation}_multi_label"),
            objects={"data:credit": data},
        )

        assert result["summary"]["rows"] >= 1
        assert result["summary"]["label_combinations"] == expected
        assert len([artifact for artifact in result["artifacts"] if artifact["type"] == "excel"]) == 1


def test_feature_group_binning_summary_writes_group_details(tmp_path, credit_frame):
    """防止分组摘要复算分箱规则或丢失分组明细。"""
    request = _request(
        tmp_path,
        "feature_group_binning_summary",
        {
            "feature": "score",
            "methods": "quantile",
            "group_col": "segment",
            "target": "target",
            "max_n_bins": 4,
            "n_jobs": 1,
        },
        "group_binning_summary",
    )

    result = execute_skill("hsbin", request, objects={"data:credit": credit_frame})
    workbook = _workbook(result)

    assert result["summary"]["rows"] == 2
    assert workbook.sheetnames[0] == "分组摘要"
    assert any(name.startswith("score-quantile-") for name in workbook.sheetnames)
