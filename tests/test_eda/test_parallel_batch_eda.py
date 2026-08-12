"""批量 EDA 统一并行接口与精确输出回归。"""

import numpy as np
import pandas as pd
import pytest

from hscredit.core.eda import (
    batch_iv_analysis,
    batch_psi_analysis,
    concentration_analysis,
    feature_drift_report,
    feature_importance_ranking,
    feature_stability_over_time,
    outlier_detection,
    population_monitoring_report,
    population_profile,
    population_shift_analysis,
    rare_category_detection,
    score_drift_report,
    segment_drift_analysis,
    stability_report,
    time_psi_tracking,
    feature_cross_segment_effectiveness,
    feature_group_analysis,
    population_stability_monitor,
    psi_cross_analysis,
)


@pytest.fixture
def batch_eda_data():
    rng = np.random.RandomState(2026)
    size = 180
    dates = pd.date_range("2025-01-01", periods=size, freq="D")
    x1 = rng.normal(size=size)
    x2 = rng.gamma(shape=2.0, scale=1.5, size=size)
    target = (x1 + 0.25 * x2 + rng.normal(scale=0.4, size=size) > 0.5).astype(int)
    category = np.where(np.arange(size) % 17 == 0, "稀有", np.where(x1 > 0, "甲", "乙"))
    return pd.DataFrame({"x1": x1, "x2": x2, "category": category, "target": target, "date": dates})


@pytest.mark.parametrize("backend", ["threading", "loky"])
def test_batch_iv_and_importance_parallel_match_serial_exactly(batch_eda_data, backend):
    features = ["x1", "x2"]
    serial_iv = batch_iv_analysis(batch_eda_data, features, "target", n_jobs=1)
    parallel_iv = batch_iv_analysis(
        batch_eda_data,
        features,
        "target",
        n_jobs=2,
        parallel_backend=backend,
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel_iv, serial_iv, check_exact=True)

    serial_ranking = feature_importance_ranking(batch_eda_data, features, "target", n_jobs=1)
    parallel_ranking = feature_importance_ranking(
        batch_eda_data,
        features,
        "target",
        n_jobs=2,
        parallel_backend=backend,
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel_ranking, serial_ranking, check_exact=True)


@pytest.mark.parametrize("backend", ["threading", "loky"])
def test_stability_batch_parallel_matches_serial_exactly(batch_eda_data, backend):
    data = batch_eda_data.copy()
    data["month"] = data["date"].dt.to_period("M").astype(str)
    periods = sorted(data["month"].unique())
    common = dict(features=["x1", "x2"], date_col="month", base_period=periods[0], compare_periods=periods[1:])
    serial = batch_psi_analysis(data, n_jobs=1, **common)
    parallel = batch_psi_analysis(
        data,
        n_jobs=2,
        parallel_backend=backend,
        parallel_config={"batch_size": 1},
        **common,
    )
    pd.testing.assert_frame_equal(parallel, serial, check_exact=True)

    serial_tracking = time_psi_tracking(batch_eda_data, ["x1", "x2"], "date", n_jobs=1)
    parallel_tracking = time_psi_tracking(
        batch_eda_data,
        ["x1", "x2"],
        "date",
        n_jobs=2,
        parallel_backend=backend,
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel_tracking, serial_tracking, check_exact=True)

    serial_report = stability_report(batch_eda_data, ["x1", "x2"], "date", n_jobs=1)
    parallel_report = stability_report(
        batch_eda_data,
        ["x1", "x2"],
        "date",
        n_jobs=2,
        parallel_backend=backend,
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel_report, serial_report, check_exact=True)


@pytest.mark.parametrize(
    "function,args,kwargs",
    [
        (outlier_detection, (), {"features": ["x1", "x2"]}),
        (rare_category_detection, (), {"features": ["category"], "threshold": 0.1}),
        (concentration_analysis, (), {"features": ["x1", "x2"]}),
        (feature_stability_over_time, (["x1", "x2"], "date"), {}),
    ],
)
def test_feature_batch_eda_threading_matches_serial_exactly(batch_eda_data, function, args, kwargs):
    serial = function(batch_eda_data, *args, n_jobs=1, **kwargs)
    parallel = function(
        batch_eda_data,
        *args,
        n_jobs=2,
        parallel_backend="threading",
        parallel_config={"batch_size": 1},
        **kwargs,
    )
    pd.testing.assert_frame_equal(parallel, serial, check_exact=True)


def test_drift_reports_keep_exact_results_with_parallel_configuration(batch_eda_data):
    base = batch_eda_data.iloc[:90]
    target = batch_eda_data.iloc[90:]
    serial = feature_drift_report(base, target, features=["x1", "x2"], n_jobs=1)
    parallel = feature_drift_report(
        base,
        target,
        features=["x1", "x2"],
        n_jobs=2,
        parallel_backend="loky",
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel, serial, check_exact=True)

    serial_score = score_drift_report(base["x1"], target["x1"], base["target"], target["target"], n_jobs=1)
    parallel_score = score_drift_report(
        base["x1"],
        target["x1"],
        base["target"],
        target["target"],
        n_jobs=2,
        parallel_backend="threading",
        parallel_config={"batch_size": 1},
    )
    assert serial_score.keys() == parallel_score.keys()
    for key in serial_score:
        if isinstance(serial_score[key], pd.DataFrame):
            pd.testing.assert_frame_equal(parallel_score[key], serial_score[key], check_exact=True)
        else:
            assert parallel_score[key] == serial_score[key]


@pytest.mark.parametrize("backend", ["threading", "loky"])
def test_population_batch_analysis_matches_serial_exactly(batch_eda_data, backend):
    serial_profile = population_profile(
        batch_eda_data,
        ["x1", "x2"],
        segment_col="category",
        target="target",
        n_jobs=1,
    )
    parallel_profile = population_profile(
        batch_eda_data,
        ["x1", "x2"],
        segment_col="category",
        target="target",
        n_jobs=2,
        parallel_backend=backend,
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel_profile, serial_profile, check_exact=True)

    base, target = batch_eda_data.iloc[:90], batch_eda_data.iloc[90:]
    serial_shift = population_shift_analysis(base, target, ["x1", "x2"], target="target", n_jobs=1)
    parallel_shift = population_shift_analysis(
        base,
        target,
        ["x1", "x2"],
        target="target",
        n_jobs=2,
        parallel_backend=backend,
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel_shift, serial_shift, check_exact=True)


def test_segment_and_cross_effectiveness_parallel_match_serial(batch_eda_data):
    common = dict(
        df=batch_eda_data,
        date_col="date",
        segment_col="category",
        features=["x1", "x2"],
        target="target",
    )
    serial_drift = segment_drift_analysis(n_jobs=1, **common)
    parallel_drift = segment_drift_analysis(
        n_jobs=2,
        parallel_backend="threading",
        parallel_config={"batch_size": 1},
        **common,
    )
    pd.testing.assert_frame_equal(parallel_drift, serial_drift, check_exact=True)

    serial_effect = feature_cross_segment_effectiveness(
        batch_eda_data,
        ["x1", "x2"],
        "target",
        "category",
        min_segment_size=5,
        n_jobs=1,
    )
    parallel_effect = feature_cross_segment_effectiveness(
        batch_eda_data,
        ["x1", "x2"],
        "target",
        "category",
        min_segment_size=5,
        n_jobs=2,
        parallel_backend="threading",
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel_effect, serial_effect, check_exact=True)


def test_population_monitoring_excel_values_match_serial(batch_eda_data, tmp_path):
    from openpyxl import load_workbook

    base = batch_eda_data.iloc[:60]
    comparisons = [batch_eda_data.iloc[60:120], batch_eda_data.iloc[120:]]
    serial_path = tmp_path / "serial_population.xlsx"
    parallel_path = tmp_path / "parallel_population.xlsx"
    common = dict(
        df_base=base,
        df_compare_list=comparisons,
        compare_labels=["中期", "后期"],
        features=["x1", "x2"],
        target="target",
        top_drift_n=2,
    )
    population_monitoring_report(output_path=str(serial_path), n_jobs=1, **common)
    population_monitoring_report(
        output_path=str(parallel_path),
        n_jobs=2,
        parallel_backend="threading",
        parallel_config={"batch_size": 1},
        **common,
    )
    serial_book = load_workbook(serial_path, data_only=False)
    parallel_book = load_workbook(parallel_path, data_only=False)
    assert serial_book.sheetnames == parallel_book.sheetnames
    for sheet_name in serial_book.sheetnames:
        serial_sheet = serial_book[sheet_name]
        parallel_sheet = parallel_book[sheet_name]
        assert [row for row in serial_sheet.values] == [row for row in parallel_sheet.values]


def test_generate_report_forwards_parallel_config_and_preserves_sections(batch_eda_data, monkeypatch):
    import hscredit.core.eda.report as report_module

    observed = []

    def fake_feature_summary(*args, **kwargs):
        observed.append(("summary", kwargs["n_jobs"], kwargs["parallel_backend"], kwargs["parallel_config"]))
        return pd.DataFrame({"特征名": ["x1"]})

    def fake_batch_iv(*args, **kwargs):
        observed.append(("iv", kwargs["n_jobs"], kwargs["parallel_backend"], kwargs["parallel_config"]))
        return pd.DataFrame({"特征名": ["x1"], "IV值": [0.1]})

    monkeypatch.setattr(report_module, "feature_summary", fake_feature_summary)
    monkeypatch.setattr(report_module, "batch_iv_analysis", fake_batch_iv)
    config = {"batch_size": 1}
    report = report_module.generate_report(
        batch_eda_data,
        target="target",
        features=["x1", "x2"],
        date_col="date",
        n_jobs=3,
        parallel_backend="threading",
        parallel_config=config,
    )
    assert observed == [
        ("summary", 3, "threading", config),
        ("iv", 3, "threading", config),
    ]
    assert list(report)[:4] == ["1.数据基础信息", "2.缺失值分析", "3.特征描述统计", "4.数据质量问题"]


def test_overview_population_and_group_analysis_parallel_match_serial(batch_eda_data):
    expected = batch_eda_data.iloc[:90]
    actual = batch_eda_data.iloc[90:]
    serial_monitor = population_stability_monitor(
        expected,
        actual,
        ["x1", "x2"],
        date_col="date",
        metrics=["占比", "样本数", "绝对变化率"],
        n_jobs=1,
    )
    parallel_monitor = population_stability_monitor(
        expected,
        actual,
        ["x1", "x2"],
        date_col="date",
        metrics=["占比", "样本数", "绝对变化率"],
        n_jobs=2,
        parallel_backend="threading",
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel_monitor, serial_monitor, check_exact=True)

    serial_group = feature_group_analysis(
        batch_eda_data,
        features=["x1", "x2"],
        group_cols="category",
        y="target",
        y_stats=["逾期率", "坏样本数"],
        n_jobs=1,
    )
    parallel_group = feature_group_analysis(
        batch_eda_data,
        features=["x1", "x2"],
        group_cols="category",
        y="target",
        y_stats=["逾期率", "坏样本数"],
        n_jobs=2,
        parallel_backend="threading",
        parallel_config={"batch_size": 1},
    )
    pd.testing.assert_frame_equal(parallel_group, serial_group, check_exact=True)


def test_psi_cross_parallel_matches_serial_exactly(batch_eda_data):
    serial = psi_cross_analysis(
        batch_eda_data,
        ["x1", "x2"],
        group_col="category",
        n_jobs=1,
    )
    parallel = psi_cross_analysis(
        batch_eda_data,
        ["x1", "x2"],
        group_col="category",
        n_jobs=2,
        parallel_backend="loky",
        parallel_config={"batch_size": 1},
    )
    assert serial.keys() == parallel.keys()
    for feature in serial:
        pd.testing.assert_frame_equal(parallel[feature], serial[feature], check_exact=True)
