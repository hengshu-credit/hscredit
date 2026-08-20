"""模型解释可执行示例测试。"""

import subprocess
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

REPOSITORY_ROOT = Path(__file__).resolve().parents[2]


def test_interpretability_example_runs_and_writes_explanation_sheet(tmp_path):
    source = tmp_path / "样例数据.xlsx"
    output = tmp_path / "模型解释报告.xlsx"
    rng = np.random.default_rng(29)
    frame = pd.DataFrame(
        {
            "衡枢鉴真分老客版": rng.normal(600, 80, 80),
            "近六个月非银多头机构数": rng.integers(0, 8, 80),
            "青云24": rng.normal(0, 1, 80),
            "FPD": np.tile([0, 1], 40),
        }
    )
    frame.to_excel(source, index=False)
    completed = subprocess.run(
        [
            sys.executable,
            "examples/27_model_interpretability.py",
            "--input",
            str(source),
            "--output",
            str(output),
            "--max-samples",
            "20",
            "--bootstrap",
            "3",
        ],
        cwd=REPOSITORY_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    assert completed.returncode == 0, completed.stderr
    assert "7-模型解释" in load_workbook(output, read_only=True).sheetnames
