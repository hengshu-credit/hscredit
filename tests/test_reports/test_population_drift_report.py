import pandas as pd
from openpyxl import load_workbook

from hscredit.report.population_drift_report import population_drift_report


def _read_feature_rows(worksheet):
    rows = {}
    for row_idx in range(3, worksheet.max_row + 1):
        feature_name = worksheet.cell(row=row_idx, column=2).value
        if feature_name:
            rows[feature_name] = {
                "value": worksheet.cell(row=row_idx, column=3).value,
                "rating": worksheet.cell(row=row_idx, column=4).value,
            }
    return rows


def test_population_drift_report_generates_expected_sheets(tmp_path):
    expected = pd.DataFrame(
        {
            "age": [23, 25, 27, 30, 35, 38, 42, 45],
            "const_feature": [1] * 8,
            "score": [610, 625, 640, 655, 670, 690, 710, 730],
            "target": [0, 0, 0, 1, 0, 1, 1, 1],
        }
    )
    actual = pd.DataFrame(
        {
            "age": [24, 26, 29, 33, 37, 41, 46, 52],
            "const_feature": [1] * 8,
            "score": [600, 618, 636, 648, 666, 684, 708, 742],
            "target": [0, 0, 1, 1, 0, 1, 1, 1],
        }
    )
    output = tmp_path / "population_drift_report.xlsx"

    result = population_drift_report(
        expected=expected,
        actual=actual,
        features=["age", "const_feature", "missing_feature"],
        target_col="target",
        score_col="score",
        output=str(output),
    )

    assert result == str(output)
    assert output.exists()

    workbook = load_workbook(output)
    assert {"PSI总览", "特征分布对比", "逾期率对比", "评分分布对比"}.issubset(workbook.sheetnames)

    psi_sheet = workbook["PSI总览"]
    assert psi_sheet["B2"].value == "特征名"
    psi_rows = _read_feature_rows(psi_sheet)
    assert "age" in psi_rows
    assert "const_feature" in psi_rows
    assert "missing_feature" not in psi_rows
    assert psi_rows["const_feature"]["value"] == 0
    assert psi_rows["const_feature"]["rating"] == "稳定"

    dist_sheet = workbook["特征分布对比"]
    assert dist_sheet["B2"].value == "特征名"

    badrate_sheet = workbook["逾期率对比"]
    assert badrate_sheet["B2"].value == "特征名"

    score_sheet = workbook["评分分布对比"]
    assert score_sheet["B2"].value == "特征名"
    score_features = {
        score_sheet.cell(row=row_idx, column=2).value
        for row_idx in range(3, score_sheet.max_row + 1)
        if score_sheet.cell(row=row_idx, column=2).value
    }
    assert score_features == {"score"}