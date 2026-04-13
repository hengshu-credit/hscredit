import pandas as pd
from openpyxl import load_workbook

from hscredit.core.rules import Rule
import hscredit.report.rule_analysis as rule_analysis_module
from hscredit.report.rule_analysis import multi_label_rule_analysis, ruleset_analysis


def test_ruleset_analysis_returns_expected_rows():
    data = pd.DataFrame(
        {
            "age": [18, 22, 30, 17],
            "target": [0, 1, 0, 1],
        }
    )

    result = ruleset_analysis(data, [Rule("age >= 21")], target="target")

    assert list(result["分箱"]) == ["原始样本", "age >= 21", "剩余样本", "所有规则"]


def test_ruleset_analysis_handles_multiindex_rule_report_columns():
    data = pd.DataFrame(
        {
            "age": [18, 22, 30, 17],
            "MOB1": [0, 4, 1, 7],
        }
    )

    result = ruleset_analysis(data, [Rule("age >= 21")], overdue="MOB1", dpds=[3, 1])

    assert isinstance(result.columns, pd.MultiIndex)
    assert ("分箱详情", "分箱") in result.columns
    assert ("MOB1 3+", "坏样本率") in result.columns
    assert ("MOB1 1+", "LIFT值") in result.columns
    assert list(result[("分箱详情", "分箱")]) == ["原始样本", "age >= 21", "剩余样本", "所有规则"]


def test_multi_label_rule_analysis_writes_excel(tmp_path, monkeypatch):
    class DummyMiner:
        def __init__(self, **kwargs):
            self.kwargs = kwargs

        def fit(self, df, features):
            self.df = df
            self.features = features

        def get_report(self):
            return pd.DataFrame(
                {
                    "规则类型": ["单特征规则"],
                    "规则": ["age > 20"],
                    "覆盖率": [0.5],
                }
            )

        def get_effectiveness_matrix(self):
            return pd.DataFrame(
                {
                    "规则": ["age > 20"],
                    "短期标签": [1.8],
                }
            )

    monkeypatch.setattr(rule_analysis_module, "MultiLabelRuleMiner", DummyMiner)

    output = tmp_path / "rule_analysis.xlsx"
    result = multi_label_rule_analysis(
        df=pd.DataFrame({"age": [18, 22], "label_a": [0, 1]}),
        features=["age"],
        labels={"短期标签": "label_a"},
        output_path=str(output),
    )

    assert result == str(output)
    workbook = load_workbook(output)
    assert {"规则汇总", "有效性矩阵", "规则分类统计"}.issubset(workbook.sheetnames)